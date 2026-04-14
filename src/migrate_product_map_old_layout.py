#!/usr/bin/env python3
"""
One-time migration: reshape legacy Product Map (CATEGORY + NOTES before Charm Shop)
to the current 8-column layout (NOTES last), widen PHOTO column A to stop image overlap,
and re-anchor thumbnails.

Used for data/supplier_catalog_OLD.xlsx and similar legacy files.
"""

from __future__ import annotations

import argparse
from copy import copy
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage

# Run from repo root with: python -m src.migrate_product_map_old_layout
from generate_shopping_route import (
    CATALOG_SHEET,
    PHOTO_PX,
    PHOTO_COL_W,
    PRODUCT_MAP_NUM_COLS,
    ensure_catalog_column_layout,
    extract_photos_from_xlsx,
)
from supplier_catalog_backup import backup_supplier_catalog_before_write

# Target column widths (match data/supplier_catalog.xlsx Product Map)
_PRODUCT_MAP_WIDTHS: dict[str, float] = {
    "A": PHOTO_COL_W,
    "B": 50.7265625,
    "C": 16.0,
    "D": 13.0,
    "E": 12.0,
    "F": 10.7265625,
    "G": 18.0,
    "H": 16.0,
}

_HEADER_TEXT = [
    "PHOTO",
    "PRODUCT TITLE",
    "SHOP NAME",
    "STALL",
    "PRICE (¥)",
    "Charm Shop",
    "Charm Code",
    "NOTES",
]


def _is_total_row(ws, row: int) -> bool:
    a = ws.cell(row, 1).value
    if a and isinstance(a, str) and a.strip().upper().startswith("TOTAL:"):
        return True
    b = ws.cell(row, 2).value
    return bool(b and isinstance(b, str) and b.strip().upper().startswith("TOTAL:"))


def _is_product_row(ws, row: int) -> bool:
    if _is_total_row(ws, row):
        return False
    b = ws.cell(row, 2).value
    if not b or not isinstance(b, str):
        return False
    t = b.strip()
    if t.startswith("TOTAL:") or t == "Unknown Product":
        return False
    return True


def _copy_cell_style(dst, src) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def _apply_header_from_template(ws_tgt, ws_tpl) -> None:
    """Copy row-1 header values and styles from a current-layout workbook."""
    for c in range(1, PRODUCT_MAP_NUM_COLS + 1):
        dst = ws_tgt.cell(1, c)
        src = ws_tpl.cell(1, c)
        dst.value = src.value or _HEADER_TEXT[c - 1]
        _copy_cell_style(dst, src)


def _apply_column_widths(ws) -> None:
    for letter, w in _PRODUCT_MAP_WIDTHS.items():
        ws.column_dimensions[letter].width = w


def migrate_product_map(
    catalog_path: Path,
    *,
    template_path: Path | None = None,
) -> None:
    """
    Transform Product Map in *catalog_path* to 8-column layout:
    drop CATEGORY, order ... PRICE, Charm Shop, Charm Code, NOTES.
    """
    if not catalog_path.is_file():
        raise FileNotFoundError(catalog_path)

    tpl = template_path
    if tpl is None:
        tpl = catalog_path.parent / "supplier_catalog.xlsx"
    if not tpl.is_file():
        tpl = None

    wb = openpyxl.load_workbook(catalog_path)
    ws = wb[CATALOG_SHEET]

    h3 = str(ws.cell(1, 3).value or "").strip().lower()
    if h3 != "category":
        # Already migrated — still widen PHOTO column and run catalog layout fixes.
        _apply_column_widths(ws)
        ensure_catalog_column_layout(wb)
        wb.save(catalog_path)
        return

    row_photos = extract_photos_from_xlsx(
        catalog_path, sheet_name=CATALOG_SHEET, photo_col_idx=0
    )

    backup_supplier_catalog_before_write(catalog_path, "migrate_product_map_layout")

    # Unmerge TOTAL (and any merge spanning col 3) before delete_cols
    for rng in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(rng))
        except ValueError:
            pass

    ws.delete_cols(3)

    max_r = ws.max_row
    for r in range(2, max_r + 1):
        if not _is_product_row(ws, r):
            continue
        # After CATEGORY removal: F=NOTES, G=Charm Shop → F=Charm Shop, G=Charm Code, H=NOTES
        notes_val = ws.cell(r, 6).value
        charm_shop_val = ws.cell(r, 7).value
        ws.cell(r, 6).value = charm_shop_val
        ws.cell(r, 7).value = None
        ws.cell(r, 8).value = notes_val

    if tpl is not None:
        wb_tpl = openpyxl.load_workbook(tpl, data_only=False)
        if CATALOG_SHEET in wb_tpl.sheetnames:
            _apply_header_from_template(ws, wb_tpl[CATALOG_SHEET])
        wb_tpl.close()
    else:
        for c, text in enumerate(_HEADER_TEXT, 1):
            ws.cell(1, c).value = text

    _apply_column_widths(ws)

    # Clear embedded images; re-add from ZIP extraction (column A width fixed)
    while ws._images:
        ws._images.pop()
    for row_num, data in sorted(row_photos.items()):
        if row_num < 2 or row_num > max_r:
            continue
        try:
            xl_img = XLImage(BytesIO(data))
            xl_img.width = PHOTO_PX
            xl_img.height = PHOTO_PX
            xl_img.anchor = f"A{row_num}"
            ws.add_image(xl_img)
        except Exception:
            pass

    ensure_catalog_column_layout(wb)
    wb.save(catalog_path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Migrate legacy Product Map to current 8-column layout.")
    ap.add_argument(
        "catalog",
        nargs="?",
        default="data/supplier_catalog_OLD.xlsx",
        help="Path to catalog .xlsx (default: data/supplier_catalog_OLD.xlsx)",
    )
    ap.add_argument(
        "--template",
        default="data/supplier_catalog.xlsx",
        help="Current-layout catalog for header styling (default: data/supplier_catalog.xlsx)",
    )
    args = ap.parse_args()
    root = Path(__file__).resolve().parents[1]
    cat = Path(args.catalog)
    if not cat.is_absolute():
        cat = root / cat
    tpl = Path(args.template)
    if not tpl.is_absolute():
        tpl = root / tpl
    migrate_product_map(cat, template_path=tpl if tpl.is_file() else None)


if __name__ == "__main__":
    main()
