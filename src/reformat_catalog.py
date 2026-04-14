#!/usr/bin/env python3
"""
reformat_catalog.py
Reformats supplier_catalog.xlsx:
  - Sorts Product Map rows by floor (2F → 4F → 5F → unknown), then stall, shop, title
  - Reduces row height from 80pt → 46pt (matching shopping_route.xlsx)
  - Applies clean alternating row colours grouped by stall
  - Adds thin borders and consistent alignment
  - Re-embeds all product photos at their new row positions
  - Renames "MALL / LOCATION" header to "STALL"
  - Updates the TOTAL row count
"""

from __future__ import annotations

import argparse
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from supplier_catalog_backup import backup_supplier_catalog_before_write

sys.stdout.reconfigure(encoding="utf-8")

CATALOG_FILE = "supplier_catalog.xlsx"  # default; use --catalog for organized layout
PHOTO_PX     = 58     # thumbnail size in pixels (matches main script)
ROW_HEIGHT   = 46.0   # row height in points (matches shopping_route.xlsx)
# Excel column A width (char units); keep ≥ PHOTO_PX/7 + margin to avoid overlap into col B
PHOTO_COL_W = 26.0

# ---------- Styles ----------
_THIN   = Side(style="thin", color="C0C0C0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HDR_FILL  = PatternFill("solid", fgColor="2C3E50")
_HDR_FONT  = Font("Calibri", bold=True, color="FFFFFF", size=12)
_BODY      = Font("Calibri", size=11)
_BODY_BOLD = Font("Calibri", bold=True, size=11)
_WARN_FILL = PatternFill("solid", fgColor="FFF3CD")   # amber: missing supplier
_PRICE_FILL= PatternFill("solid", fgColor="FFF9E6")   # light yellow: price TBD

_GROUP_FILLS = [
    PatternFill("solid", fgColor="EBF2FA"),  # light blue  (group 0)
    PatternFill("solid", fgColor="FFFFFF"),  # white       (group 1)
]

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP   = Alignment(vertical="center", wrap_text=True)

# ---------- Supplier dropdown named ranges ----------
_SUPPLIER_SHOP_NAMED_RANGE  = "SupplierShopNames"
_SUPPLIER_STALL_NAMED_RANGE = "SupplierStalls"
_SUPPLIER_LIST_MAX_ROW      = 500

# ---------- Floor helpers (same logic as generate_shopping_route.py) ----------

def _floor_order(stall: str) -> int:
    """Return a numeric floor for sort ordering. 999 = unknown."""
    if not stall or stall.strip() in ("", "\u2014", "???"):
        return 999
    s = stall.strip()
    if re.match(r"^A2", s, re.IGNORECASE):
        return 2
    m = re.match(r"^(\d)", s)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d)[A-Za-z]", s)
    if m:
        return int(m.group(1))
    return 999


# ---------- Step 1 – extract product photos from the original file ----------

_R   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_A   = "http://schemas.openxmlformats.org/drawingml/2006/main"


def extract_images(xlsx_path: Path) -> dict[int, bytes]:
    """Return {excel_row_1based: image_bytes} for all images anchored in col A."""
    result: dict[int, bytes] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            all_files = set(zf.namelist())
            rels_path = "xl/drawings/_rels/drawing1.xml.rels"
            if rels_path not in all_files:
                return result
            rid_to_path: dict[str, str] = {}
            for rel in ET.parse(zf.open(rels_path)).getroot():
                rid    = rel.get("Id", "")
                target = rel.get("Target", "")
                rid_to_path[rid] = "xl/media/" + target.split("/")[-1]

            drawing_path = "xl/drawings/drawing1.xml"
            if drawing_path not in all_files:
                return result
            for anchor in ET.parse(zf.open(drawing_path)).getroot():
                a_tag = anchor.tag.split("}")[-1]
                if a_tag not in ("oneCellAnchor", "twoCellAnchor"):
                    continue
                fr = anchor.find(f"{{{_XDR}}}from")
                if fr is None:
                    continue
                col_e = fr.find(f"{{{_XDR}}}col")
                row_e = fr.find(f"{{{_XDR}}}row")
                if col_e is None or row_e is None:
                    continue
                if int(col_e.text) != 0:          # column A only
                    continue
                excel_row = int(row_e.text) + 1   # 0-based → 1-based
                blip = anchor.find(f".//{{{_A}}}blip")
                if blip is None:
                    continue
                img_path = rid_to_path.get(blip.get(f"{{{_R}}}embed", ""), "")
                if img_path and img_path in all_files:
                    result[excel_row] = zf.read(img_path)
    except Exception as e:
        print(f"  Warning: image extraction error: {e}")
    return result


# ---------- Step 2 – read catalog data ----------

def read_product_map(xlsx_path: Path) -> list[dict]:
    """Read all product rows from the Product Map sheet (skip TOTAL rows)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Product Map"]
    h3 = str(ws.cell(1, 3).value or "").strip().lower()
    h7 = str(ws.cell(1, 7).value or "").strip().lower()
    has_category = h3 == "category"
    legacy_notes_first = has_category and h7 == "notes"
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        title = row[1]
        if not title or not isinstance(title, str):
            continue
        t = title.strip()
        if t.startswith("TOTAL:") or t == "Unknown Product":
            continue
        if has_category:
            if legacy_notes_first:
                notes = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                charm_shop = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                charm_code = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            else:
                charm_shop = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                charm_code = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                notes = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            category = str(row[2]).strip() if row[2] else ""
            shop_name = str(row[3]).strip() if row[3] else ""
            stall = str(row[4]).strip() if row[4] else ""
            price = (
                str(row[5]).strip()
                if len(row) > 5 and row[5] and str(row[5]).strip() not in ("None", "")
                else ""
            )
        else:
            category = ""
            shop_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            stall = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            price = (
                str(row[4]).strip()
                if len(row) > 4 and row[4] and str(row[4]).strip() not in ("None", "")
                else ""
            )
            charm_shop = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            charm_code = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            notes = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        rows.append({
            "orig_row":  i,
            "title":     t,
            "category":  category,
            "shop_name": shop_name,
            "stall":     stall,
            "price":     price,
            "notes":     notes,
            "charm_shop": charm_shop,
            "charm_code": charm_code,
        })
    wb.close()
    return rows


# ---------- Step 3 – sort ----------

def sort_rows(rows: list[dict]) -> list[dict]:
    def _key(r: dict):
        stall = r["stall"]
        sort_key = r["category"] or r["shop_name"] or "\uffff"
        return (
            _floor_order(stall),
            stall or "\uffff",
            sort_key,
            r["title"],
        )
    return sorted(rows, key=_key)


# ---------- Step 3b – read existing Suppliers data and sync from Product Map ----------

def read_suppliers(xlsx_path: Path) -> list[list]:
    """Read all data rows from the Suppliers sheet (preserving all columns)."""
    rows: list[list] = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Suppliers" not in wb.sheetnames:
            wb.close()
            return rows
        ws = wb["Suppliers"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                rows.append(list(row))
        wb.close()
    except Exception as e:
        print(f"  Warning: could not read Suppliers sheet: {e}")
    return rows


def build_supplier_rows(
    product_rows: list[dict],
    existing_rows: list[list],
) -> list[list]:
    """
    Return a merged, sorted list of supplier rows for the Suppliers sheet.

    Starts from existing_rows (preserving ID, Mall, Floor, Address, Contact, Notes),
    then appends any unique (Shop Name, Stall) pairs from product_rows that are not
    already present.  The final list is sorted by floor order then stall then shop name.

    Suppliers sheet columns (1-based): ID, Shop Name, Mall, Floor, Stall, Address, Contact, Notes
    """
    # Build a lookup from (shop, stall) -> existing row (list of up to 8 values)
    existing_map: dict[tuple[str, str], list] = {}
    max_id = 0
    for row in existing_rows:
        shop  = str(row[1] if len(row) > 1 and row[1] else "").strip()
        stall = str(row[4] if len(row) > 4 and row[4] else "").strip()
        if shop or stall:
            existing_map[(shop, stall)] = row
        vid = row[0] if row else None
        if vid is not None:
            try:
                max_id = max(max_id, int(vid))
            except (TypeError, ValueError):
                pass

    # Collect unique (shop, stall) pairs from Product Map (preserve first-seen order)
    seen: set[tuple[str, str]] = set()
    pairs: list[tuple[str, str]] = []
    for r in product_rows:
        shop  = r.get("shop_name", "").strip()
        stall = r.get("stall", "").strip()
        if not shop and not stall:
            continue
        key = (shop, stall)
        if key not in seen:
            seen.add(key)
            pairs.append(key)

    # Sort by floor order, then stall, then shop name
    pairs.sort(key=lambda t: (_floor_order(t[1]), t[1].lower(), t[0].lower()))

    # Build final rows: start from existing map, add any new pairs
    result: list[list] = []
    for shop, stall in pairs:
        key = (shop, stall)
        if key in existing_map:
            result.append(list(existing_map[key]))
        else:
            max_id += 1
            # [ID, Shop Name, Mall, Floor, Stall, Address, Contact, Notes]
            result.append([max_id, shop, None, None, stall, None, None, None])

    return result


# ---------- Step 4 – write the reformatted workbook ----------

def write_catalog(
    xlsx_path: Path,
    sorted_rows: list[dict],
    row_images: dict[int, bytes],
    supplier_rows: list[list] | None = None,
) -> None:
    wb = openpyxl.Workbook()

    # ---- Suppliers sheet (first / sheet1) ----
    ws_sup = wb.active
    ws_sup.title = "Suppliers"
    sup_headers = ["ID", "Shop Name", "Mall", "Floor", "Stall", "Address", "Contact", "Notes"]
    for ci, h in enumerate(sup_headers, 1):
        c = ws_sup.cell(1, ci, h)
        c.fill      = _HDR_FILL
        c.font      = _HDR_FONT
        c.alignment = _CENTER
        c.border    = _BORDER
    ws_sup.row_dimensions[1].height = 22
    for ci, w in enumerate([8, 22, 14, 8, 12, 35, 20, 30], 1):
        ws_sup.column_dimensions[get_column_letter(ci)].width = w
    ws_sup.freeze_panes = "A2"

    if supplier_rows:
        for ri, row_data in enumerate(supplier_rows, 2):
            # Pad row to at least 8 columns so zip never drops values
            padded = list(row_data) + [None] * max(0, len(sup_headers) - len(row_data))
            for ci, val in enumerate(padded[:len(sup_headers)], 1):
                cell = ws_sup.cell(ri, ci, val)
                cell.border    = _BORDER
                cell.alignment = _CENTER if ci in (1, 4, 5) else _WRAP
                cell.font      = Font("Calibri", size=11)

    # ---- Product Map sheet (second / sheet2) ----
    ws = wb.create_sheet("Product Map")

    # Header row
    pm_headers = [
        "PHOTO", "PRODUCT TITLE", "SHOP NAME", "STALL",
        "PRICE (¥)", "Charm Shop", "Charm Code", "NOTES",
    ]
    for ci, h in enumerate(pm_headers, 1):
        c = ws.cell(1, ci, h)
        c.fill      = _HDR_FILL
        c.font      = _HDR_FONT
        c.alignment = _CENTER
        c.border    = _BORDER
    ws.row_dimensions[1].height = 22

    # Column widths (A–H; no CATEGORY; same relative widths as 9-col minus old col C)
    for col, w in zip(
        "ABCDEFGH",
        [PHOTO_COL_W, 50.7, 16.0, 12.0, 10.7, 16.0, 16.0, 30.7],
    ):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "2C3E50"

    # Data rows
    current_stall = object()  # sentinel
    group_idx     = -1

    for offset, dr in enumerate(sorted_rows):
        row_num = offset + 2   # data starts at row 2

        # Toggle group fill on stall change
        stall = dr["stall"]
        if stall != current_stall:
            current_stall = stall
            group_idx     = (group_idx + 1) % 2

        missing = not dr["shop_name"] and not stall
        row_fill = _WARN_FILL if missing else _GROUP_FILLS[group_idx]

        # A – photo placeholder (no text value; image embedded below)
        ws.cell(row_num, 1).border = _BORDER

        # B – product title (bold)
        _c(ws, row_num, 2, dr["title"],      row_fill, _BODY_BOLD, _WRAP)
        # C – shop name
        _c(ws, row_num, 3, dr["shop_name"] or None, row_fill, _BODY, _WRAP)
        # D – stall (centred)
        _c(ws, row_num, 4, stall or None,    row_fill, _BODY, _CENTER)
        # E – price (yellow fill if empty)
        price_val = dr["price"] if dr["price"] else None
        _c(ws, row_num, 5, price_val, _PRICE_FILL if not price_val else row_fill, _BODY, _CENTER)
        # F – Charm Shop, G – Charm Code, H – NOTES
        _c(ws, row_num, 6, dr.get("charm_shop") or None, row_fill, _BODY, _CENTER)
        _c(ws, row_num, 7, dr.get("charm_code") or None, row_fill, _BODY, _CENTER)
        _c(ws, row_num, 8, dr["notes"] or None, row_fill, _BODY, _WRAP)

        ws.row_dimensions[row_num].height = ROW_HEIGHT

        # Embed photo
        orig_row = dr["orig_row"]
        if orig_row in row_images:
            try:
                xl_img        = XLImage(BytesIO(row_images[orig_row]))
                xl_img.width  = PHOTO_PX
                xl_img.height = PHOTO_PX
                xl_img.anchor = f"A{row_num}"
                ws.add_image(xl_img)
            except Exception as e:
                print(f"  Warning: photo embed failed row {row_num}: {e}")

    # TOTAL row
    n = len(sorted_rows)
    total_row = n + 2
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    tc = ws.cell(total_row, 1, f"TOTAL: {n} products")
    tc.fill      = _HDR_FILL
    tc.font      = Font("Calibri", bold=True, size=11, color="FFFFFF")
    tc.alignment = _CENTER
    tc.border    = _BORDER
    ws.row_dimensions[total_row].height = 20

    # ---- Supplier dropdown validations (SHOP NAME / STALL) ----
    _apply_supplier_dropdowns(wb, ws)

    if "Product Map" in wb.sheetnames:
        wb.active = wb["Product Map"]

    wb.save(xlsx_path)


def _c(ws, row, col, value, fill, font, alignment):
    """Helper: write a cell with fill/font/alignment/border."""
    cell            = ws.cell(row, col, value)
    cell.fill       = fill
    cell.font       = font
    cell.alignment  = alignment
    cell.border     = _BORDER


def _apply_supplier_dropdowns(wb, ws_pm) -> None:
    """
    Attach SHOP NAME (col C) and STALL (col D) dropdown lists to the Product Map.

    Reads the Suppliers sheet header row to find the actual column positions so
    the function works regardless of layout.  Uses workbook-level named ranges
    as formula1 — the same mechanism used by the Charm Shop and Charm Code
    dropdowns in generate_shopping_route.py.
    """
    if "Suppliers" not in wb.sheetnames:
        print("  Warning: Suppliers sheet not found — supplier DVs skipped.")
        return

    ws_sup = wb["Suppliers"]

    # Detect actual column letters from the header row.
    shop_col_letter  = "B"   # standard layout: ID=A, Shop Name=B
    stall_col_letter = "E"   # standard layout: Stall=E
    for ci in range(1, 20):
        hdr = str(ws_sup.cell(1, ci).value or "").strip().lower()
        if hdr == "shop name":
            shop_col_letter  = get_column_letter(ci)
        elif hdr == "stall":
            stall_col_letter = get_column_letter(ci)

    shop_ref  = f"'Suppliers'!${shop_col_letter}$2:${shop_col_letter}${_SUPPLIER_LIST_MAX_ROW}"
    stall_ref = f"'Suppliers'!${stall_col_letter}$2:${stall_col_letter}${_SUPPLIER_LIST_MAX_ROW}"

    # Always overwrite named ranges so column positions stay current.
    wb.defined_names[_SUPPLIER_SHOP_NAMED_RANGE] = DefinedName(
        name=_SUPPLIER_SHOP_NAMED_RANGE, attr_text=shop_ref,
    )
    wb.defined_names[_SUPPLIER_STALL_NAMED_RANGE] = DefinedName(
        name=_SUPPLIER_STALL_NAMED_RANGE, attr_text=stall_ref,
    )

    max_row = max(ws_pm.max_row, 2)

    dv_shop = DataValidation(
        type="list",
        formula1=_SUPPLIER_SHOP_NAMED_RANGE,   # named range name (same as CharmShopNames)
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Value not in the Suppliers tab. You can still type it manually.",
        errorTitle="Shop Name not in list",
    )
    ws_pm.add_data_validation(dv_shop)
    dv_shop.add(f"C2:C{max_row + 500}")

    dv_stall = DataValidation(
        type="list",
        formula1=_SUPPLIER_STALL_NAMED_RANGE,  # named range name (same as CharmCodes)
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Value not in the Suppliers tab. You can still type it manually.",
        errorTitle="Stall not in list",
    )
    ws_pm.add_data_validation(dv_stall)
    dv_stall.add(f"D2:D{max_row + 500}")

    print(f"  Supplier DVs: SHOP NAME C2:C{max_row + 500} ← {shop_ref}")
    print(f"  Supplier DVs: STALL      D2:D{max_row + 500} ← {stall_ref}")


# ---------- Main ----------

def main() -> None:
    ap = argparse.ArgumentParser(description="Reformat supplier_catalog.xlsx (sort, style, re-embed photos)")
    ap.add_argument("--catalog", default=CATALOG_FILE, help="Catalog .xlsx path (default: supplier_catalog.xlsx)")
    args = ap.parse_args()
    path = Path(args.catalog)
    if not path.exists():
        print(f"ERROR: {CATALOG_FILE} not found")
        sys.exit(1)

    print(f"Reading images from {CATALOG_FILE} ...")
    row_images = extract_images(path)
    print(f"  {len(row_images)} embedded photos found")

    print("Reading product data ...")
    rows = read_product_map(path)
    print(f"  {len(rows)} product rows read")

    print("Reading existing suppliers data ...")
    existing_supplier_rows = read_suppliers(path)
    print(f"  {len(existing_supplier_rows)} existing supplier rows read")

    print("Sorting by floor order ...")
    sorted_rows = sort_rows(rows)

    # Print floor distribution
    floor_counts: dict[int, int] = {}
    for r in sorted_rows:
        f = _floor_order(r["stall"])
        floor_counts[f] = floor_counts.get(f, 0) + 1
    for f in sorted(floor_counts):
        label = f"{f}F" if f != 999 else "unknown"
        print(f"  {label}: {floor_counts[f]} products")

    print("Building supplier list from Product Map ...")
    supplier_rows = build_supplier_rows(rows, existing_supplier_rows)
    print(f"  {len(supplier_rows)} unique supplier entries")

    print(f"Writing reformatted catalog ...")
    backup_supplier_catalog_before_write(path, "reformat_catalog")
    write_catalog(path, sorted_rows, row_images, supplier_rows)

    print(f"\nDone. {CATALOG_FILE} updated:")
    print(f"  - {len(sorted_rows)} products")
    print(f"  - Row height: {ROW_HEIGHT}pt (was 80pt)")
    print(f"  - Sorted: lowest to highest floor")
    print(f"  - {len(row_images)} photos re-embedded")
    print(f"  - {len(supplier_rows)} supplier rows (SHOP NAME & STALL dropdowns updated)")


if __name__ == "__main__":
    main()
