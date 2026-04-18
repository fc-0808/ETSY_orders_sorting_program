#!/usr/bin/env python3
"""
generate_shopping_route.py

Parses Etsy order PDFs, cross-references each item against the supplier
catalog, and produces ``shopping_route.xlsx`` -- a purchasing guide
organised by supplier location for efficient in-person shopping.

Also updates ``supplier_catalog.xlsx`` with any products not yet catalogued
(adds them with their product photo so you can fill in supplier info later).
A duplicate guard prevents the same title from being appended more than once
across multiple runs.  The Shopping Route places such items in a dedicated
blue "Awaiting Supplier Info" section so they are never buried in the table.

**Charms (two layers):**

1. **Charm Library** sheet — master list of physical charm designs: **A** photo, **B** stable
   code, **C** SKU (sortable label), **D** default charm shop.  One row per distinct charm
   you stock.  Disk assets ``data/charm_images/<Code>.png`` (etc.) override embedded
   photos when generating routes.  New codes use zero-padded suffixes (e.g. ``CH-00001``;
   width follows existing rows/files).

2. **Product Map** — each product row is a phone case / listing variant.  **Column F**
   (*Charm Shop*) = which stall supplies the charm for *that product*.  **Column G**
   (*Charm Code*) = optional link to **Charm Library** column **B**.  **Column H** = *NOTES*.

   **Discontinuing a product:** use the UI button or ``--mark-product-discontinued``.
   The row is **moved** from Product Map to the **Discontinued Products** sheet
   (with a timestamp and photo), so the catalog stays clean. The duplicate guard
   checks both sheets, preventing re-addition.

   **When column G (Charm Code) is set** to a library code, the shopping route (and HTML) **Charm**
   section **aggregates by charm code** — one row per unique charm with the library photo,
   code, SKU, charm shop, and total quantity across all orders.  Folder file wins over embed.

   **When Charm Code is blank**, the order is shown in a separate **"Awaiting Charm Code"**
   sub-section with the product photo and a prompt to assign a code in the catalog.

   Typical setup: add every distinct charm to the library once; for each catalog product
   that ships with that charm, pick the matching **G** (shop) and **H** (code).  **G** can
   match **D** from the library as a default, but you may override **G** per product if
   sourcing differs.

Unless ``--no-charm-manifest``, each run writes ``data/charm_manifest.json``; use
``--export-charm-manifest`` for that step only.

Usage
-----
    python generate_shopping_route.py                         # auto-discover PDFs
    python generate_shopping_route.py order1.pdf order2.pdf   # explicit PDFs
    python generate_shopping_route.py --threshold 60          # lower match bar
    python generate_shopping_route.py --no-catalog-update     # skip catalog write

Adding a new batch of today's orders
-------------------------------------
Each day you download new order PDFs, run:

    python generate_shopping_route.py --new-batch

The script auto-discovers all *.pdf files in the current directory.  Any PDF
whose filename was already processed in a previous run is automatically
skipped (its orders are already in the cache).  Only brand-new PDFs are
parsed and merged on top of the existing shopping route.  After processing,
every ingested PDF filename is recorded in orders_cache.json so it is never
re-parsed on future runs.

If you delete ``shopping_route.xlsx`` from the output folder (or it is absent)
and run again **without** ``--refresh-catalog``, prior orders are **not** read
from ``orders_cache.json`` or any leftover path: the script rebuilds solely
from the PDFs currently in ``input/`` (same idea as ``--reset``).  Use
``--refresh-catalog`` when you need to regenerate the Excel from the cache
after deleting only the output file.

You may also pass the PDF files explicitly instead of auto-discovery:

    python generate_shopping_route.py --new-batch file1.pdf file2.pdf ...

Post-shopping cleanup
---------------------
After returning from a shopping trip and updating the status dropdowns in
shopping_route.xlsx, run:

    python generate_shopping_route.py --purge-purchased

This reads every status you have already entered in the Excel file and applies
**independent section-level purging**:

  • Case / Grip section  (supplier floors) and
  • Charm section        (separate building)

are evaluated separately.  If the Case/Grip section is fully purchased (every
present Case/Grip component is "Purchased") it is stripped from the item's
style even if the Charm is still pending -- and vice versa.  An item is only
removed from the route entirely when *both* sections are fully purchased.

Items with any component still "Out of Stock" or "Pending" are kept.
"Out of Production" is treated as complete (same as Purchased); those items
are purged and recorded to out_of_production_log.csv so they never reappear.
The cache (orders_cache.json) is updated to match, so partially-purged items
re-appear on the next run with only their remaining section.
"""

from __future__ import annotations

import argparse
import base64
import csv
import fnmatch
import json
import logging
import os
import re
import secrets
import shutil
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

import fitz          # pymupdf -- for JPEG extraction
import pdfplumber
import openpyxl
try:
    from deep_translator import GoogleTranslator as _GoogleTranslator
    _DEEP_TRANSLATOR_AVAILABLE = True
except ImportError:
    _DEEP_TRANSLATOR_AVAILABLE = False
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import EMU_to_pixels, pixels_to_EMU, points_to_pixels
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from rapidfuzz import fuzz, process

from supplier_catalog_backup import (
    backup_supplier_catalog_before_write,
    catalog_backup_dir,
    list_supplier_catalog_backups,
    restore_supplier_catalog,
)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

CATALOG_FILE      = "supplier_catalog.xlsx"
CATALOG_SHEET     = "Product Map"
SUPPLIERS_SHEET   = "Suppliers"
DISCONTINUED_SHEET = "Discontinued Products"
CHARM_SHOPS_SHEET = "Charm Shops"          # separate sheet for charm-shop reference data
CHARM_LIBRARY_SHEET = "Charm Library"    # one row per physical charm (shared photo + optional default shop)
# Optional on-disk charm assets (same codes as Product Map col H / Charm Library col B).
# With --project-dir: ``<project>/data/charm_images/``.  Filenames: ``<Charm Code>.png`` (or .jpg, .jpeg, .webp).
CHARM_IMAGES_DIR_NAME = "charm_images"
# Charm Library column B drives Product Map col H (Charm Code) dropdowns (named range CharmCodes).
# Named range ``CharmCodes`` and Product Map validation span (room for large libraries).
CHARM_CODES_LIST_MAX_ROW = 10000
# When no ``PREFIX`` + digits codes exist yet, new codes use at least this many digits
# (e.g. CH-00001).  If CH-001 … already exist, new codes extend that width and grow with N.
CHARM_CODE_NUMERIC_MIN_WIDTH = 5
# Charm Library column C — human-readable SKU / short product label (was "Display Name").
CHARM_LIBRARY_COL_SKU_HEADER = "SKU"

# Merged help rows — marker substring must appear in A1 text so layout/import skips the row.
CHARM_LIBRARY_INSTRUCTION_MARKER = "charm library — quick guide"
CHARM_LIBRARY_INSTRUCTION_TEXT = (
    "Charm Library — quick guide\n\n"
    "• One charm per row.  Column B = stable code (e.g. CH-00001);  "
    "Column C = SKU / short label;  Column D = Default Charm Shop.\n\n"
    "• Photo: column A, or file data/charm_images/<Code>.png  "
    "(folder file takes priority when both exist).\n\n"
    "• Product Map → column F = Charm Shop (where to buy);  "
    "column G = Charm Code (same as column B here).  "
    "Leave G blank to use the order PDF photo.\n\n"
    "• To insert a new charm between existing ones:\n"
    "  1. Import the photo via the Python UI (it appends at the end).\n"
    "  2. Cut its row here and paste it where it belongs "
    "(right-click a row number → Insert Cut Cells).\n"
    "  3. Open the Python UI → Tab 2 → Section C → click "
    "\"Renumber charm codes\" — codes are reassigned by row order "
    "and Product Map references update automatically.\n\n"
    "• After editing: save this workbook, then run the route generator "
    "(--refresh-catalog)."
)
# Version tag embedded in the text — bump this string to force an upgrade on next run.
_CHARM_LIBRARY_INSTRUCTION_VERSION = "v2-reorder"
CHARM_SHOPS_INSTRUCTION_MARKER = "charm shops — quick guide"
CHARM_SHOPS_INSTRUCTION_TEXT = (
    "Charm Shops — quick guide\n\n"
    "• Edit the rows above (shop name + stall).\n\n"
    "• Save the workbook, then run the route generator."
)

OUTPUT_FILE       = "shopping_route.xlsx"
CHARM_MANIFEST_FILE = "charm_manifest.json"   # default under data/ with --project-dir
CACHE_FILE        = "orders_cache.json"
OOP_LOG_FILE      = "out_of_production_log.csv"   # append-only log of purged OOP items
ZH_TRANS_CACHE    = "translations_zh_cache.json"   # persisted product-title translations
MATCH_THRESHOLD   = 65

# Default charm shops (pre-populated when creating the Charm Shops sheet for the first time).
# Edit directly in supplier_catalog.xlsx → Charm Shops tab to add/remove shops.
_DEFAULT_CHARM_SHOPS = [
    ("彩虹",     "2D21",    ""),
    ("有米UMI",  "2D02",    ""),
    ("長金飾品", "2D04",    ""),
    ("一樂潮品", "2C666",   ""),
    ("小艾飾品", "2D41-43", ""),
]

# Fuzzy-match false-positive thresholds
# ----------------------------------------
# Phone-case product titles share many generic tokens ("Case", "Charm", "Cover",
# "iPhone 17 16 15 14 13 Pro Max", "Kawaii", "Cute", "Gift", …).  token_sort_ratio
# therefore produces spuriously high scores for completely different products.
# Two separate thresholds guard against the two failure modes:
#
#  EMPTY_ENTRY_MATCH_THRESHOLD (90)
#    Applies when the best catalog match has NO location info at all (shop_name
#    AND stall both empty).  If the match score is below 90 the match is treated
#    as a false positive and the product gets its own new amber row in the catalog.
#    False-positive ceiling observed in production: 66.  Gap to true matches: large.
#
#  FILLED_ENTRY_MATCH_THRESHOLD (85)
#    Applies when the best catalog match DOES have location info (shop_name or
#    stall filled in).  A false positive here is more harmful: the product would
#    be sent to the WRONG supplier on the shopping trip.  If the score is below 85
#    the match is rejected and the product is treated as unmatched (gets its own
#    amber catalog row so the correct supplier can be recorded).
#    False-positive ceiling observed in production: 83.1.
#    Lowest true match observed in production: 93.7.  Gap: ~10.6 points.
EMPTY_ENTRY_MATCH_THRESHOLD  = 90
FILLED_ENTRY_MATCH_THRESHOLD = 85

# Catalog completeness threshold
# ----------------------------------------
# Controls whether a matched item still gets its OWN row in supplier_catalog.xlsx.
#
# Phone-case suppliers often sell the same product in multiple colour / style
# variants with near-identical titles (e.g. "Pink Miffy … Quicksand Gift" vs
# "Purple Miffy … Quicksand Gift").  These score 93-95 on token_sort_ratio —
# well above FILLED_ENTRY_MATCH_THRESHOLD — so they are routed to the correct
# stall.  However, each variant is a *distinct product* and must have its own
# catalog row so the user can track it individually (price, category, etc.).
#
# Any match whose score falls below SAME_PRODUCT_THRESHOLD is therefore treated
# as a *variant* of the matched entry: the item is still routed to that
# supplier's stall (correct location), but update_catalog() also appends a new
# amber row pre-filled with the inferred shop_name / stall so the user only
# needs to add Category and Price.
#
# Rationale for 97: exact same product title → 100; one-word phone-model
# variation ("13" added/removed) → ~97; colour/style variant → 93–95.
SAME_PRODUCT_THRESHOLD = 97

# Photo display dimensions used in both the route Excel and the catalog
PHOTO_PX    = 155   # square thumbnail px (enlarged for better visibility)
ROW_HEIGHT  = 120.0 # Excel row height in points (~160 px at 96 dpi)
# ~7 px per unit at Calibri 11 — keep width ≥ PHOTO_PX/7 + margin so images do not spill into col B
PHOTO_COL_W = 26.0  # Excel column width in character units

# Chinese-version overrides — larger rows/photos so images are clearly readable
ZH_ROW_HEIGHT  = 210.0  # ~280 px at 96 dpi
ZH_PHOTO_PX    = 265    # square thumbnail px for ZH sheet
# Excel column width is ~7 px per "character" (Calibri 11); keep col ≥ image px / 7 + margin
# or embedded images spill into 供应商 (next column).
ZH_PHOTO_COL_W = 44.0   # must cover ZH_PHOTO_PX — was 34 and caused overlap into col C

# Charm Library — column A / row height match ZH route; embedded photos are sized to
# *fill that cell box* (minus CHARM_LIB_CELL_PAD_PX), not a fixed square, so they are
# not tiny thumbnails floating in a large row.
CHARM_LIB_ROW_HEIGHT    = ZH_ROW_HEIGHT
CHARM_LIB_COL_A_WIDTH   = ZH_PHOTO_COL_W
CHARM_LIB_CELL_PAD_PX   = 4   # small inset from cell edges (gridlines / anti-aliasing)

# Charm SKU vision (optional) — OpenAI-compatible ``/chat/completions`` (OpenAI, OpenRouter,
# Azure proxy, Ollama ``/v1``, LM Studio, etc.); stdlib HTTP only.
_OPENAI_DEFAULT_BASE = "https://api.openai.com/v1"
_CHARM_VISION_COOLDOWN_SEC = 0.35   # light spacing between API calls on bulk import

# Env resolution order: provider-neutral first, then common provider vars (backward compatible).
_CHARM_VISION_API_KEY_ENVS = (
    "CHARM_VISION_API_KEY",
    "OPENAI_API_KEY",
    "OPENROUTER_API_KEY",
)
_CHARM_VISION_BASE_URL_ENVS = ("CHARM_VISION_BASE_URL", "OPENAI_BASE_URL")


def _resolve_charm_vision_api_key(cli_key: str | None) -> str:
    k = (cli_key or "").strip()
    if k:
        return k
    for name in _CHARM_VISION_API_KEY_ENVS:
        v = (os.environ.get(name) or "").strip()
        if v:
            return v
    return ""


def _resolve_charm_vision_base_url(cli_url: str | None) -> str:
    u = (cli_url or "").strip()
    if u:
        return u
    for name in _CHARM_VISION_BASE_URL_ENVS:
        v = (os.environ.get(name) or "").strip()
        if v:
            return v
    return _OPENAI_DEFAULT_BASE


def _charm_vision_base_allows_empty_key(base_url: str) -> bool:
    """Local OpenAI-compatible servers often accept requests with no ``Authorization`` header."""
    try:
        host = (urllib.parse.urlparse((base_url or "").strip()).hostname or "").lower()
    except ValueError:
        return False
    return host in ("localhost", "127.0.0.1", "::1")

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s  %(message)s")
log = logging.getLogger("shopping_route")

# Purchase-status options shown in the dropdown (order matters)
STATUS_OPTIONS = ["Pending", "Purchased", "Out of Stock", "Out of Production"]

# Simplified-Chinese status options (same order)
ZH_STATUS_OPTIONS = ["待处理", "已购买", "缺货", "停产"]

# Simplified-Chinese translation table
_ZH: dict[str, str] = {
    # Sheet / tab names
    "Shopping Route":   "购物路线",
    "Orders Detail":    "订单明细",
    "Summary":          "汇总",
    # Route sheet column headers
    "Photo":            "图片",
    "Floor":            "楼层",
    "Supplier":         "供应商",
    "Stall":            "摊位",
    "Product":          "产品",
    "Items to Purchase": "待购项",
    "Case":             "手机壳",
    "Grip":             "支架",
    "Charm":            "挂件",
    "Phone Model":      "手机型号",
    "Qty":              "数量",
    "Order #":          "订单号",
    "Etsy Shop":        "Etsy店铺",
    # Orders Detail extra headers
    "Buyer":            "买家",
    "Ship To":          "收货人",
    "Country":          "国家",
    "Order Date":       "订单日期",
    "Match %":          "匹配度",
    # Summary stat labels
    "Total orders":                          "订单总数",
    "Total line items":                      "商品行数",
    "Total quantity":                        "总数量",
    "Ready (supplier + location)":           "就绪（供应商+位置齐全）",
    "In catalog \u2013 needs supplier info": "目录中\u2013待补供应商信息",
    "Not in catalog (unmatched)":            "不在目录中（未匹配）",
    "Items per Supplier":                    "各供应商商品数",
    "Items per Etsy Shop":                   "各Etsy店铺商品数",
    "Items":                                 "商品数",
    "Orders":                                "订单数",
    # Status cell values (dropdown + written into cells)
    "Pending":           "待处理",
    "Purchased":         "已购买",
    "Out of Stock":      "缺货",
    "Out of Production": "停产",
    "N/A":               "不适用",
    "case only":         "仅手机壳",
    "grip only":         "仅支架",
    "case, grip":        "手机壳、支架",
    # Charm section strings
    "Charms to Purchase":    "待购挂件",
    "Separate Building":     "独立楼栋",
    "charm(s) needed across": "个挂件，涉及",
    "order(s)":              "个订单",
    "Charm shops":           "挂件店铺",
    "No charm shops configured": "未配置挂件店铺",
    "Private Notes":         "私信备注",
}


def _t(key: str, lang: str) -> str:
    """Return the Simplified-Chinese translation of *key* when lang=='zh',
    otherwise return the key unchanged."""
    return _ZH.get(key, key) if lang == "zh" else key


# ---------------------------------------------------------------------------
# Product-title translation helpers (used for the Chinese shopping route)
# ---------------------------------------------------------------------------

def _load_trans_cache(path: Path) -> dict[str, str]:
    """Load the persisted product-title translation cache from disk."""
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_trans_cache(path: Path, cache: dict[str, str]) -> None:
    """Persist the translation cache to disk."""
    try:
        path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        log.warning("Could not save translation cache: %s", e)


def _translate_title(title: str, cache: dict[str, str], cache_path: Path) -> str:
    """Translate *title* to Simplified Chinese, using the on-disk cache.

    Falls back to the original English title if deep_translator is not
    installed or if the network request fails.
    """
    if not title:
        return title
    if title in cache:
        return cache[title]
    if not _DEEP_TRANSLATOR_AVAILABLE:
        log.warning("deep_translator not installed – product titles will remain in English. "
                    "Run: pip install deep-translator")
        return title
    try:
        result = _GoogleTranslator(source="en", target="zh-CN").translate(title)
        cache[title] = result or title
        _save_trans_cache(cache_path, cache)
        return cache[title]
    except Exception as e:
        log.warning("Translation failed for %r: %s", title[:50], e)
        return title

# Row highlight fills/fonts per status (applied via conditional formatting)
_STATUS_FILLS = {
    "Purchased":         PatternFill("solid", fgColor="C6EFCE"),  # green
    "Out of Stock":      PatternFill("solid", fgColor="FFEB9C"),  # amber
    "Out of Production": PatternFill("solid", fgColor="FFC7CE"),  # red
}
_STATUS_FONTS = {
    "Purchased":         Font("Calibri", size=10, color="276221"),
    "Out of Stock":      Font("Calibri", size=10, color="7D4E00"),
    "Out of Production": Font("Calibri", size=10, color="9C0006"),
}
_ITEMS_TO_PURCHASE_FONT = Font("Calibri", size=10, bold=True)

# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------


@dataclass
class OrderItem:
    title:       str
    quantity:    int   = 1
    phone_model: str   = ""
    style:       str   = ""
    photo_bytes: bytes | None = None   # raw JPEG bytes extracted from PDF


@dataclass
class Order:
    order_number:   str = ""
    etsy_shop:      str = ""
    buyer_name:     str = ""
    buyer_username: str = ""
    ship_to_name:   str = ""
    ship_to_country: str = ""
    order_date:     str = ""
    private_notes:  str = ""
    items: list[OrderItem] = field(default_factory=list)


@dataclass
class CatalogEntry:
    product_title: str = ""
    category:      str = ""
    shop_name:     str = ""
    stall:         str = ""
    price:         str = ""
    notes:         str = ""
    # Foreign key → Charm Shops tab: which charm shop supplies the charm for
    # this product.  Empty string means no charm, or charm shop not yet assigned.
    charm_shop:    str = ""
    # Foreign key → Charm Library tab: stable code for the physical charm (shared photo).
    charm_code:    str = ""


@dataclass
class CharmLibraryEntry:
    """One row in the Charm Library sheet — a reusable charm SKU."""
    code:               str              # column B, unique key
    sku:                str = ""         # column C — short label / stock SKU
    default_charm_shop: str = ""         # optional; same names as Charm Shops col A
    notes:              str = ""
    photo_bytes:        bytes | None = None


@dataclass
class CharmShop:
    """One entry in the Charm Shops reference sheet."""
    shop_name: str = ""
    stall:     str = ""
    notes:     str = ""


@dataclass
class ResolvedItem:
    order:       Order
    item:        OrderItem
    supplier:    CatalogEntry | None = None
    match_score: float = 0.0


# ---------------------------------------------------------------------------
# PDF parsing -- column-aware extraction for Etsy two-column label slips
# ---------------------------------------------------------------------------
#
# Layout: Left column (x < 200): shipping address, shop name, order date
#         Right column (x >= 200): product title, qty, model, style + photo
# Images sort top->bottom and match items 1-to-1 in that order.

_COL_BOUNDARY     = 200   # px dividing left address block from right product block
_LEFT_META_MAX_X  = 100   # px — order metadata (order#, address, shop, date) lives at x≈36
_LEFT_PN_MIN_X    = 100   # px — Private notes content lives at x≈174 (separate sub-column)
_LEFT_PN_MAX_Y    = 500   # px — crop footer ("Do the green thing", etc. at y≈700+)
_Y_TOLERANCE      = 3.0   # px -- merge words within this gap into one line

_QUANTITY_RE          = re.compile(r"^Quantity:\s*(\d+)")
_MODEL_RE             = re.compile(r"^(?:Phone|iPhone)\s+Model:\s*(.+)", re.IGNORECASE)
_STYLE_RE             = re.compile(r"^Styles?:\s*(.+)")
_CURRENCY_RE          = re.compile(r"^[A-Z]{3}$")
_ORDER_RE             = re.compile(r"^Order\s+#(\d+)")
_SCHEDULED_RE         = re.compile(r"^Scheduled\s+to\s+(?:ship|dispatch)\s+by", re.IGNORECASE)
_PRIVATE_NOTES_RE     = re.compile(r"^Private\s+notes?$", re.IGNORECASE)
# Some shops (e.g. LUVEKASEofficial, LUVKASEofficial) append a currency code
# like " HKD" at the end of the first title line because their PDF lays out a
# price tag on the same horizontal baseline as the product text.  Strip it so
# the title stays clean for fuzzy matching.
_TRAILING_CURRENCY_RE = re.compile(r"\s+[A-Z]{3}$")


def _words_to_lines(words: list[dict], x_min: float, x_max: float) -> list[str]:
    """Group PDF words within an x-band into text lines, sorted top to bottom."""
    filtered = [w for w in words if x_min <= w["x0"] < x_max]
    if not filtered:
        return []
    filtered.sort(key=lambda w: (w["top"], w["x0"]))

    lines: list[tuple[float, list[str]]] = []
    for w in filtered:
        text = w["text"].strip()
        if not text:
            continue
        if lines and abs(w["top"] - lines[-1][0]) <= _Y_TOLERANCE:
            lines[-1][1].append(text)
        else:
            lines.append((w["top"], [text]))

    return [" ".join(parts) for _, parts in lines]


def _extract_private_notes(lines: list[str]) -> str:
    """Return the Private notes content from lines in the inner-right sub-column (x ≈ 174).

    Scans for the "Private notes" header, then collects all following lines as
    the note text.  Stops at the first known footer pattern so page footers like
    "Do the green thing …" are never included.
    """
    i, n = 0, len(lines)
    while i < n:
        if _PRIVATE_NOTES_RE.match(lines[i]):
            i += 1
            note_parts: list[str] = []
            while i < n:
                note_parts.append(lines[i])
                i += 1
            return " ".join(note_parts).strip()
        i += 1
    return ""


def parse_pdf(path: Path) -> list[Order]:
    """Extract all orders (with product photos) from a single Etsy order PDF."""
    fitz_doc = fitz.open(str(path))
    xref_to_jpeg: dict[int, bytes] = {}
    for xref in range(1, fitz_doc.xref_length()):
        try:
            img = fitz_doc.extract_image(xref)
            if img and img.get("ext") == "jpeg":
                xref_to_jpeg[xref] = img["image"]
        except Exception:
            pass
    fitz_doc.close()

    orders: list[Order] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(keep_blank_chars=True, extra_attrs=["top"])

            # Order metadata lives at x ≈ 36 (strictly left).  Using x < 100
            # keeps it separate from the "Private notes" content at x ≈ 174,
            # preventing Y-proximity merges like "Shop Private notes" or
            # "Y2KASEofficial need case".
            left  = _words_to_lines(words, 0,                _LEFT_META_MAX_X)
            right = _words_to_lines(words, _COL_BOUNDARY,    9999)

            # Inner-right sub-column: "Private notes" header + note content.
            # Crop to y < _LEFT_PN_MAX_Y to exclude page footers (y ≈ 700+).
            pn_words  = [w for w in words
                         if _LEFT_PN_MIN_X <= w["x0"] < _COL_BOUNDARY
                         and w.get("top", 0) < _LEFT_PN_MAX_Y]
            left_pn   = _words_to_lines(pn_words, _LEFT_PN_MIN_X, _COL_BOUNDARY)
            pn_text   = _extract_private_notes(left_pn)

            order = _parse_left_column(left)
            if not order:
                continue
            items, right_private_notes = _parse_right_column(right)
            if not items:
                continue
            # Inner-right sub-column is the primary source for private notes;
            # right column (x ≥ 200) is a fallback for unusual PDF layouts.
            order.private_notes = pn_text or right_private_notes

            # Assign photos: images are sorted top->bottom matching item order
            page_imgs = sorted(page.images, key=lambda i: i["top"])
            for idx, item in enumerate(items):
                if idx < len(page_imgs):
                    xref = getattr(page_imgs[idx]["stream"], "objid", None)
                    if xref is not None:
                        item.photo_bytes = xref_to_jpeg.get(xref)
                    if item.photo_bytes is None:
                        try:
                            item.photo_bytes = page_imgs[idx]["stream"].rawdata or None
                        except Exception:
                            pass

            order.items = items
            orders.append(order)

    return orders


def _parse_left_column(lines: list[str]) -> Order | None:
    """Parse order metadata from the left column of one page."""
    if not lines:
        return None
    o = Order()
    i, n = 0, len(lines)

    m = _ORDER_RE.match(lines[i]) if i < n else None
    if not m:
        return None
    o.order_number = m.group(1)
    i += 1

    if i < n:
        bm = re.match(r"(.+?)\s*\((.+?)\)", lines[i])
        if bm:
            o.buyer_name     = bm.group(1).strip()
            o.buyer_username = bm.group(2)
        i += 1

    if i < n and re.match(r"(?:Ship|Deliver)\s+to", lines[i]):
        i += 1

    addr: list[str] = []
    while i < n and not _SCHEDULED_RE.match(lines[i]):
        addr.append(lines[i])
        i += 1
    o.ship_to_name    = addr[0]  if addr          else ""
    o.ship_to_country = addr[-1] if len(addr) > 1 else ""

    if i < n and _SCHEDULED_RE.match(lines[i]):
        i += 1
    if i < n:
        i += 1  # scheduled date value

    if i < n and lines[i] == "Shop":
        i += 1
    if i < n:
        o.etsy_shop = lines[i]
        i += 1

    if i < n and lines[i] == "Order date":
        i += 1
    if i < n:
        o.order_date = lines[i]

    return o


def _parse_right_column(lines: list[str]) -> tuple[list[OrderItem], str]:
    """Parse product items and optional private notes from the right column of one page.

    Returns (items, private_notes) where private_notes is everything that
    follows the "Private notes" section header (if present).
    """
    items: list[OrderItem] = []
    private_notes = ""
    i, n = 0, len(lines)

    while i < n:
        # "Private notes" section header — collect all remaining lines as notes
        if _PRIVATE_NOTES_RE.match(lines[i]):
            i += 1
            note_parts: list[str] = []
            while i < n:
                note_parts.append(lines[i])
                i += 1
            private_notes = " ".join(note_parts).strip()
            break

        title_parts: list[str] = []
        while i < n:
            if _QUANTITY_RE.match(lines[i]):
                break
            # Stop title collection at the Private notes header
            if _PRIVATE_NOTES_RE.match(lines[i]):
                break
            if _CURRENCY_RE.match(lines[i]):
                i += 1
                continue
            # Strip a trailing currency code (e.g. " HKD") that some shops
            # place on the same PDF baseline as the product title text.
            cleaned = _TRAILING_CURRENCY_RE.sub("", lines[i]).strip()
            if cleaned:
                title_parts.append(cleaned)
            i += 1

        if not title_parts:
            break

        item = OrderItem(title=" ".join(title_parts))

        if i < n and (qm := _QUANTITY_RE.match(lines[i])):
            item.quantity = int(qm.group(1))
            i += 1

        if i < n and (mm := _MODEL_RE.match(lines[i])):
            item.phone_model = mm.group(1).strip()
            i += 1

        if i < n and (sm := _STYLE_RE.match(lines[i])):
            item.style = sm.group(1).strip()
            i += 1

        if i < n and _CURRENCY_RE.match(lines[i]):
            i += 1

        items.append(item)

    return items, private_notes


# ---------------------------------------------------------------------------
# Supplier catalog -- load
# ---------------------------------------------------------------------------


def load_catalog(path: Path) -> list[CatalogEntry]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[CATALOG_SHEET]
    h3 = str(ws.cell(1, 3).value or "").strip().lower()
    h7 = str(ws.cell(1, 7).value or "").strip().lower()
    has_category = h3 == "category"
    legacy_notes_first = has_category and h7 == "notes"

    entries: list[CatalogEntry] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[1]
        if not title or not isinstance(title, str):
            continue
        if title.startswith("TOTAL:") or title == "Unknown Product":
            continue
        if has_category:
            if legacy_notes_first:
                notes = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
                charm_shop = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
                charm_code = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
            else:
                charm_shop = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
                charm_code = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
                notes = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
            cat_v = str(row[2]).strip() if row[2] is not None else ""
            shop_v = str(row[3]).strip() if row[3] is not None else ""
            stall_v = str(row[4]).strip() if row[4] is not None else ""
            price_v = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        else:
            # 8-column layout: B title, C shop, D stall, E price, F/G charm, H notes
            cat_v = ""
            shop_v = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            stall_v = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            price_v = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
            charm_shop = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
            charm_code = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
            notes = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
        entries.append(CatalogEntry(
            product_title = title.strip(),
            category   = cat_v,
            shop_name  = shop_v,
            stall      = stall_v,
            price      = price_v,
            notes      = notes,
            charm_shop = charm_shop,
            charm_code = charm_code,
        ))

    wb.close()
    log.info("Loaded %d products from catalog", len(entries))
    return entries


# ---------------------------------------------------------------------------
# Charm shop catalog -- init + load
# ---------------------------------------------------------------------------
#
# Columns F–G of the Product Map are FOREIGN KEYs into Charm Shops / Charm Library.
# F stores the charm shop name; G stores the charm library code when set.
#
# Relationship:  one charm shop  →  many products' charms  (many-to-one)
#
# The named range  CharmShopNames  (= 'Charm Shops'!$A$2:$A$100)  drives an
# Excel dropdown on column F so the user only has to pick a shop name from a
# list; the stall / floor are resolved automatically at generation time.
#
# Product Map: 8 columns A–H (no CATEGORY).  F/G = Charm Shop / Charm Code; H = NOTES.
# ---------------------------------------------------------------------------

PRODUCT_MAP_NUM_COLS = 8

# Column F — Charm Shop (1-based); column G — Charm Code; column H — NOTES
_CHARM_COL_IDX   = 6
_CHARM_COL_LETTER = "F"
_CHARM_NAMED_RANGE = "CharmShopNames"

_CHARM_CODE_COL_IDX    = 7
_CHARM_CODE_COL_LETTER = "G"
_CHARM_CODES_NAMED_RANGE = "CharmCodes"

# Supplier dropdown named ranges (Product Map C/D → Suppliers B/E)
_SUPPLIER_SHOP_NAMED_RANGE  = "SupplierShopNames"
_SUPPLIER_STALL_NAMED_RANGE = "SupplierStalls"
_SUPPLIER_LIST_MAX_ROW      = 500


def _unmerge_any_spanning_column(ws, col_idx: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col <= col_idx <= rng.max_col:
            try:
                ws.unmerge_cells(str(rng))
            except ValueError:
                pass


def _fix_product_map_total_merge(ws, end_col: int | None = None) -> None:
    """Ensure the TOTAL row spans only A–end_col (default: PRODUCT_MAP_NUM_COLS)."""
    if end_col is None:
        end_col = PRODUCT_MAP_NUM_COLS
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.startswith("TOTAL:"):
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row <= r <= rng.max_row and rng.min_col <= 1 <= rng.max_col:
                    try:
                        ws.unmerge_cells(str(rng))
                    except ValueError:
                        pass
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
            break


def _strip_product_map_discontinued_column_j(ws) -> None:
    """Remove legacy column J (Discontinued / Y markers) from the Product Map."""
    if ws.max_column < 10:
        return
    h10 = str(ws.cell(1, 10).value or "").strip().lower()
    if "discontinu" not in h10 and h10 not in ("y", "n", "yes", "no"):
        return
    _unmerge_any_spanning_column(ws, 10)
    ws.delete_cols(10)
    log.info("Product Map: removed legacy column J (Discontinued).")


def _migrate_product_map_column_order(ws) -> None:
    """
    Intermediate migration for 9-column layout (CATEGORY still in col C):
    old  → G=NOTES, H=Charm Shop, I=Charm Code.
    fixed → G=Charm Shop, H=Charm Code, I=NOTES.

    After this runs, _migrate_remove_category_column_product_map() removes col C
    (CATEGORY) so the final 8-column layout lands at F=Charm Shop, G=Charm Code, H=NOTES.
    Column-9 writes here are intentional — CATEGORY is still present at this point.
    """
    h7 = str(ws.cell(1, 7).value or "").strip().lower()
    if h7 == "charm shop":
        return
    if h7 != "notes":
        return
    max_r = ws.max_row
    for r in range(2, max_r + 1):
        notes_v = ws.cell(r, 7).value
        cs_v = ws.cell(r, 8).value
        cc_v = ws.cell(r, 9).value   # col 9 valid here: CATEGORY still in col 3
        ws.cell(r, 7, cs_v)
        ws.cell(r, 8, cc_v)
        ws.cell(r, 9, notes_v)       # col 9 valid here: CATEGORY still in col 3
    ws.cell(1, 7, "Charm Shop")
    ws.cell(1, 8, "Charm Code")
    ws.cell(1, 9, "NOTES")           # col 9 valid here: CATEGORY still in col 3
    log.info("Product Map: rotated G/H/I column order (pre-CATEGORY-removal step).")


def _migrate_discontinued_products_column_order(ws) -> None:
    """
    Intermediate migration for 10-column Discontinued Products (CATEGORY in col C):
    old  → G=NOTES, H=Charm Shop, I=Charm Code, J=Disc Date.
    fixed → G=Charm Shop, H=Charm Code, I=NOTES, J=Disc Date.

    After this runs, _migrate_remove_category_column_discontinued() removes col C
    so the final 9-column layout lands at F=Charm Shop, G=Charm Code, H=NOTES, I=Disc Date.
    Column-9/10 writes here are intentional — CATEGORY is still present at this point.
    """
    h7 = str(ws.cell(1, 7).value or "").strip().lower()
    if h7 == "charm shop":
        return
    if "note" not in h7:
        return
    for r in range(2, ws.max_row + 1):
        notes_v = ws.cell(r, 7).value
        cs_v = ws.cell(r, 8).value
        cc_v = ws.cell(r, 9).value   # col 9/10 valid here: CATEGORY still in col 3
        dt_v = ws.cell(r, 10).value
        ws.cell(r, 7, cs_v)
        ws.cell(r, 8, cc_v)
        ws.cell(r, 9, notes_v)       # col 9/10 valid here: CATEGORY still in col 3
        ws.cell(r, 10, dt_v)
    ws.cell(1, 7, "CHARM SHOP")
    ws.cell(1, 8, "CHARM CODE")
    ws.cell(1, 9, "NOTES")           # col 9/10 valid here: CATEGORY still in col 3
    log.info("%s: rotated G/H/I/J column order (pre-CATEGORY-removal step).", DISCONTINUED_SHEET)


def _migrate_remove_category_column_product_map(ws) -> None:
    """Remove legacy CATEGORY column (C) so Product Map is A–H only."""
    h3 = str(ws.cell(1, 3).value or "").strip().lower()
    if h3 != "category":
        return
    _unmerge_any_spanning_column(ws, 3)
    ws.delete_cols(3)
    log.info("Product Map: removed CATEGORY column (now %d columns A–H).", PRODUCT_MAP_NUM_COLS)


def _migrate_remove_category_column_discontinued(ws) -> None:
    """Remove CATEGORY from Discontinued Products to match Product Map."""
    h3 = str(ws.cell(1, 3).value or "").strip().lower()
    if h3 != "category":
        return
    _unmerge_any_spanning_column(ws, 3)
    ws.delete_cols(3)
    log.info("%s: removed CATEGORY column.", DISCONTINUED_SHEET)


def ensure_catalog_column_layout(wb: openpyxl.Workbook) -> None:
    """Migrate legacy column J + NOTES/G/H order + CATEGORY removal on Product Map.

    After any structural migration that deletes or inserts columns, existing
    data-validation sqref ranges become stale (shifted to wrong columns).
    We clear them here so the next DV-application pass starts clean.
    """
    if CATALOG_SHEET in wb.sheetnames:
        wpm = wb[CATALOG_SHEET]
        _strip_product_map_discontinued_column_j(wpm)
        _migrate_product_map_column_order(wpm)
        _strip_product_map_discontinued_column_j(wpm)
        _migrate_remove_category_column_product_map(wpm)
        _fix_product_map_total_merge(wpm, PRODUCT_MAP_NUM_COLS)
        _clear_product_map_validations(wpm)
    if DISCONTINUED_SHEET in wb.sheetnames:
        ws_disc = wb[DISCONTINUED_SHEET]
        _migrate_discontinued_products_column_order(ws_disc)
        _migrate_remove_category_column_discontinued(ws_disc)
        ws_disc.column_dimensions["A"].width = PHOTO_COL_W


def set_supplier_catalog_active_to_product_map(wb: openpyxl.Workbook) -> None:
    """Select Product Map so Excel opens on that tab after the next save."""
    if CATALOG_SHEET in wb.sheetnames:
        wb.active = wb[CATALOG_SHEET]


def _charm_codes_range_ref() -> str:
    """Workbook reference for all charm codes (Charm Library col B)."""
    return f"'{CHARM_LIBRARY_SHEET}'!$B$2:$B${CHARM_CODES_LIST_MAX_ROW}"


def _ensure_charm_codes_named_range(wb) -> bool:
    """
    Create or upgrade ``CharmCodes`` to span B2:B{CHARM_CODES_LIST_MAX_ROW}.
    Returns True if the workbook was modified.
    """
    ref    = _charm_codes_range_ref()
    changed = False
    if _CHARM_CODES_NAMED_RANGE not in wb.defined_names:
        wb.defined_names[_CHARM_CODES_NAMED_RANGE] = DefinedName(
            name=_CHARM_CODES_NAMED_RANGE,
            attr_text=ref,
        )
        changed = True
    elif wb.defined_names[_CHARM_CODES_NAMED_RANGE].attr_text != ref:
        wb.defined_names[_CHARM_CODES_NAMED_RANGE].attr_text = ref
        changed = True
    return changed


# Styling for the Charm Shop column header in the Product Map
_CAT_CHARM_HDR_FILL = PatternFill("solid", fgColor="5B1A6B")   # same purple as route sheet
_CAT_CHARM_HDR_FONT = Font("Calibri", bold=True, color="FFFFFF", size=12)
# Light-lavender fill for cells awaiting a charm-shop assignment
_CAT_CHARM_PENDING_FILL = PatternFill("solid", fgColor="EFD9FC")


def _clear_product_map_validations(ws_pm) -> None:
    """Remove **all** data-validation rules from the Product Map.

    This prevents stale DVs (left behind by column migrations that shift
    sqref ranges) from accumulating.  The caller must re-apply fresh DVs
    via ``_refresh_all_product_map_validations`` after clearing.
    """
    ws_pm.data_validations.dataValidation.clear()


def _apply_charm_column_validation(wb, ws_pm) -> None:
    """Named range ``CharmShopNames`` + dropdown on Product Map column F (Charm Shop)."""
    if _CHARM_NAMED_RANGE not in wb.defined_names:
        wb.defined_names[_CHARM_NAMED_RANGE] = DefinedName(
            name=_CHARM_NAMED_RANGE,
            attr_text=f"'{CHARM_SHOPS_SHEET}'!$A$2:$A$100",
        )
    dv = DataValidation(
        type="list",
        formula1=_CHARM_NAMED_RANGE,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error=(
            "Please pick a charm shop from the dropdown, "
            "or leave blank if this product has no charm."
        ),
        errorTitle="Invalid Charm Shop",
    )
    ws_pm.add_data_validation(dv)
    max_row = max(ws_pm.max_row, 2)
    dv.add(f"{_CHARM_COL_LETTER}2:{_CHARM_COL_LETTER}{max_row + 500}")


def _apply_charm_code_column_validation(wb, ws_pm) -> None:
    """Named range ``CharmCodes`` + dropdown on Product Map column G (Charm Code)."""
    _ensure_charm_codes_named_range(wb)
    dv = DataValidation(
        type="list",
        formula1=_CHARM_CODES_NAMED_RANGE,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error=(
            "Pick a charm code from the Charm Library sheet, "
            "or leave blank to use the order PDF photo."
        ),
        errorTitle="Invalid Charm Code",
    )
    ws_pm.add_data_validation(dv)
    max_row = max(ws_pm.max_row, 2)
    dv.add(
        f"{_CHARM_CODE_COL_LETTER}2:{_CHARM_CODE_COL_LETTER}{max_row + 500}"
    )


def _apply_charm_library_default_shop_validation(wb, ws_lib) -> None:
    """Optional dropdown on Charm Library column D → Charm Shops names."""
    if _CHARM_NAMED_RANGE not in wb.defined_names:
        return
    dv = DataValidation(
        type="list",
        formula1=_CHARM_NAMED_RANGE,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Pick a shop from the Charm Shops tab or leave blank.",
        errorTitle="Invalid shop",
    )
    ws_lib.add_data_validation(dv)
    max_row = max(ws_lib.max_row, 2)
    dv.add(f"D2:D{max_row + 500}")


def _apply_supplier_column_validations(wb, ws_pm) -> None:
    """
    Attach SHOP NAME (col C) and STALL (col D) dropdown lists to the Product Map.

    The dropdown source is the Suppliers sheet.  Column positions are read
    dynamically from the Suppliers header row so the function works regardless
    of whether the sheet has an ID column or a different layout.

    Mechanism mirrors the working Charm Shop / Charm Code dropdowns exactly:
      1. Create workbook-level named ranges pointing to the source columns.
      2. Use those named-range names as formula1 in DataValidation objects.
    """
    if SUPPLIERS_SHEET not in wb.sheetnames:
        log.warning("_apply_supplier_column_validations: '%s' sheet not found – skipped.", SUPPLIERS_SHEET)
        return

    ws_sup = wb[SUPPLIERS_SHEET]

    # Discover actual column letters by reading the header row.
    # Defaults match the standard layout: ID=A, Shop Name=B, Mall=C, Floor=D, Stall=E.
    shop_col_letter  = "B"
    stall_col_letter = "E"
    for ci in range(1, 20):
        hdr = str(ws_sup.cell(1, ci).value or "").strip().lower()
        if hdr == "shop name":
            shop_col_letter  = get_column_letter(ci)
        elif hdr == "stall":
            stall_col_letter = get_column_letter(ci)

    shop_ref  = f"'{SUPPLIERS_SHEET}'!${shop_col_letter}$2:${shop_col_letter}${_SUPPLIER_LIST_MAX_ROW}"
    stall_ref = f"'{SUPPLIERS_SHEET}'!${stall_col_letter}$2:${stall_col_letter}${_SUPPLIER_LIST_MAX_ROW}"

    # (Re-)create named ranges – always overwrite so column positions stay current.
    wb.defined_names[_SUPPLIER_SHOP_NAMED_RANGE] = DefinedName(
        name=_SUPPLIER_SHOP_NAMED_RANGE, attr_text=shop_ref,
    )
    wb.defined_names[_SUPPLIER_STALL_NAMED_RANGE] = DefinedName(
        name=_SUPPLIER_STALL_NAMED_RANGE, attr_text=stall_ref,
    )

    max_row = max(ws_pm.max_row, 2)

    dv_shop = DataValidation(
        type="list",
        formula1=_SUPPLIER_SHOP_NAMED_RANGE,   # named range name – same as CharmShopNames
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Value not in the Suppliers tab. You can still type it manually.",
        errorTitle="Shop Name not in list",
    )
    ws_pm.add_data_validation(dv_shop)
    dv_shop.add(f"C2:C{max_row + 500}")

    dv_stall = DataValidation(
        type="list",
        formula1=_SUPPLIER_STALL_NAMED_RANGE,  # named range name – same as CharmCodes
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Value not in the Suppliers tab. You can still type it manually.",
        errorTitle="Stall not in list",
    )
    ws_pm.add_data_validation(dv_stall)
    dv_stall.add(f"D2:D{max_row + 500}")

    log.info(
        "Supplier DVs applied — SHOP NAME C2:C%d ← %s | STALL D2:D%d ← %s",
        max_row + 500, shop_ref, max_row + 500, stall_ref,
    )


# Thin wrappers kept so existing call-sites compile without change.
def _apply_supplier_shop_name_validation(wb, ws_pm) -> None:
    _apply_supplier_column_validations(wb, ws_pm)


def _apply_supplier_stall_validation(wb, ws_pm) -> None:
    pass  # work is done by _apply_supplier_column_validations above


def _refresh_all_product_map_validations(wb, ws_pm) -> None:
    """Clear every DV from the Product Map, then re-apply all four dropdowns.

    Column migrations (e.g. CATEGORY removal) shift existing DV sqref ranges,
    leaving stale dropdowns on the wrong columns (the original cause of shop
    names appearing in the Charm Code column).  Clearing first guarantees only
    the correct DVs survive.
    """
    _clear_product_map_validations(ws_pm)
    _apply_charm_column_validation(wb, ws_pm)
    _apply_charm_code_column_validation(wb, ws_pm)
    _apply_supplier_column_validations(wb, ws_pm)


def _charm_library_image_anchor_col_a(img) -> bool:
    """True if an openpyxl drawing image is anchored in column A (photo column)."""
    a = getattr(img, "anchor", None)
    if isinstance(a, str):
        return bool(re.match(r"^A\d", a.strip(), re.IGNORECASE))
    fr = getattr(a, "_from", None)
    return fr is not None and getattr(fr, "col", None) == 0


def _excel_column_width_chars_to_px(char_width: float) -> int:
    """Approximate interior width in pixels (96 dpi) from Excel column width in characters."""
    if char_width is None or char_width <= 0:
        char_width = CHARM_LIB_COL_A_WIDTH
    return max(int(char_width * 7 + 8), 1)


def _charm_library_photo_fill_pixels(ws_lib, row: int) -> tuple[int, int]:
    """
    Target pixel size for a Charm Library photo so it fills column A × *row*
    (stretching slightly if needed — matches user expectation of using the whole box).
    """
    cw = ws_lib.column_dimensions["A"].width
    w_px = _excel_column_width_chars_to_px(cw if cw is not None else CHARM_LIB_COL_A_WIDTH)
    rh = ws_lib.row_dimensions[row].height
    if rh is None:
        rh = CHARM_LIB_ROW_HEIGHT
    h_px = points_to_pixels(rh)
    pad = CHARM_LIB_CELL_PAD_PX
    return max(w_px - pad, 48), max(h_px - pad, 48)


def _openpyxl_image_ext_pixels(img) -> tuple[int, int]:
    """Pixel (cx, cy) Excel uses for a drawing — OneCellAnchor.ext, not img.width alone."""
    anc = getattr(img, "anchor", None)
    if hasattr(anc, "ext") and anc.ext is not None:
        try:
            return EMU_to_pixels(anc.ext.width), EMU_to_pixels(anc.ext.height)
        except Exception:
            pass
    try:
        return int(img.width), int(img.height)
    except (TypeError, ValueError):
        return 0, 0


def _sync_openpyxl_image_display_ext(img, tw: int, th: int) -> None:
    """
    openpyxl's drawing writer only copies ``img.width`` / ``img.height`` into
    ``OneCellAnchor.ext`` when *anchor* is still a cell string.  Files loaded
    from disk already have ``OneCellAnchor`` objects — then ``ext`` stays whatever
    was saved (e.g. 155×155) while Python attributes may differ; **Excel renders ext**.
    """
    tw, th = max(int(tw), 1), max(int(th), 1)
    img.width = tw
    img.height = th
    anc = getattr(img, "anchor", None)
    if hasattr(anc, "ext") and anc.ext is not None:
        anc.ext.width = pixels_to_EMU(tw)
        anc.ext.height = pixels_to_EMU(th)


def _ensure_charm_library_sheet_layout(ws_lib) -> bool:
    """
    Widen column A, raise data row heights, and rescale embedded photos in column A
    to fill each cell (minus padding) so images are not small thumbs in oversized rows.
    Idempotent; returns whether any change was applied.
    """
    changed = False
    cur_w = ws_lib.column_dimensions["A"].width
    if cur_w is None or cur_w < CHARM_LIB_COL_A_WIDTH - 0.01:
        ws_lib.column_dimensions["A"].width = CHARM_LIB_COL_A_WIDTH
        changed = True

    for r in range(2, ws_lib.max_row + 1):
        if _charm_library_instruction_row(ws_lib, r):
            continue
        b_val = str(ws_lib.cell(r, 2).value or "").strip()
        if not b_val or b_val.casefold() == "charm code":
            continue
        rh = ws_lib.row_dimensions[r].height
        cur_h = rh if rh is not None else 15.0
        if cur_h < CHARM_LIB_ROW_HEIGHT - 0.01:
            ws_lib.row_dimensions[r].height = CHARM_LIB_ROW_HEIGHT
            changed = True

    for img in list(getattr(ws_lib, "_images", []) or []):
        if not _charm_library_image_anchor_col_a(img):
            continue
        row = _anchor_row(img)
        if row is None:
            continue
        tw, th = _charm_library_photo_fill_pixels(ws_lib, row)
        ew, eh = _openpyxl_image_ext_pixels(img)
        if ew != tw or eh != th:
            _sync_openpyxl_image_display_ext(img, tw, th)
            changed = True

    return changed


def _style_charm_instruction_cell(cell) -> None:
    """Readable help text: spaced lines, top-aligned, slightly larger than body."""
    cell.font = Font("Calibri", size=10, italic=True, color="555555")
    cell.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True,
    )


def _charm_library_instruction_row(ws_lib, r: int) -> bool:
    """True if row *r* is the merged Charm Library help row (current or legacy wording)."""
    a1 = ws_lib.cell(r, 1).value
    if not a1 or not isinstance(a1, str):
        return False
    t = a1.lower()
    return (
        CHARM_LIBRARY_INSTRUCTION_MARKER in t
        or "one row per physical charm" in t
    )


_CHARM_LIB_NOTE_HEIGHT = 188   # row height (points) for the instruction cell

def _upgrade_charm_library_instruction_note(ws_lib) -> bool:
    """
    Replace outdated Charm Library help text with the current
    ``CHARM_LIBRARY_INSTRUCTION_TEXT``.

    Triggers an upgrade when:
    * The row was written by a legacy build (no marker, or old marker wording).
    * The current text is missing the reorder/renumber instructions
      (detected by the ``_CHARM_LIBRARY_INSTRUCTION_VERSION`` tag).
    """
    for r in range(2, min(ws_lib.max_row, 24) + 1):
        v = ws_lib.cell(r, 1).value
        if not isinstance(v, str):
            continue
        low = v.lower()

        # Legacy build: no marker yet
        if "one row per physical charm" in low and CHARM_LIBRARY_INSTRUCTION_MARKER not in low:
            c = ws_lib.cell(r, 1, CHARM_LIBRARY_INSTRUCTION_TEXT)
            _style_charm_instruction_cell(c)
            ws_lib.row_dimensions[r].height = _CHARM_LIB_NOTE_HEIGHT
            return True

        if CHARM_LIBRARY_INSTRUCTION_MARKER not in low:
            continue

        # Any version that predates the reorder/renumber instructions
        needs_update = (
            _CHARM_LIBRARY_INSTRUCTION_VERSION not in low
            or "insert cut cells" not in low
            or "section c" not in low
        )
        if needs_update:
            c = ws_lib.cell(r, 1, CHARM_LIBRARY_INSTRUCTION_TEXT)
            _style_charm_instruction_cell(c)
            ws_lib.row_dimensions[r].height = _CHARM_LIB_NOTE_HEIGHT
            return True
    return False


def _upgrade_charm_library_sku_header(ws_lib) -> bool:
    """Rename legacy column C header ``Display Name`` → ``SKU``."""
    h = ws_lib.cell(1, 3).value
    if isinstance(h, str) and h.strip() == "Display Name":
        cell = ws_lib.cell(1, 3, CHARM_LIBRARY_COL_SKU_HEADER)
        cell.font = Font("Calibri", bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor="5B1A6B")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="C0C0C0"),
            right=Side(style="thin", color="C0C0C0"),
            top=Side(style="thin", color="C0C0C0"),
            bottom=Side(style="thin", color="C0C0C0"),
        )
        return True
    return False


def _upgrade_charm_shops_instruction_note(ws_cs) -> bool:
    """Replace legacy Charm Shops footer with CHARM_SHOPS_INSTRUCTION_TEXT."""
    for r in range(2, ws_cs.max_row + 1):
        v = ws_cs.cell(r, 1).value
        if not isinstance(v, str):
            continue
        low = v.lower()
        if (
            "add, remove or edit rows above" in low
            and CHARM_SHOPS_INSTRUCTION_MARKER not in low
        ):
            c = ws_cs.cell(r, 1, CHARM_SHOPS_INSTRUCTION_TEXT)
            _style_charm_instruction_cell(c)
            ws_cs.row_dimensions[r].height = 72
            return True
    return False


def init_charm_shops_sheet(path: Path) -> None:
    """
    Idempotent initialisation of all charm-related infrastructure in
    *supplier_catalog.xlsx*.  Called automatically on every program run.

    Tasks (each skipped if already in place):
      1.  Create the ``Charm Shops`` reference sheet (pre-populated with the
          default 5 shops).
      2.  Add the ``Charm Shop`` column (G) header to the Product Map, styled
          with a purple header to match the Shopping Route charm section.
      3.  Create the workbook-level named range ``CharmShopNames`` that points
          to ``'Charm Shops'!$A$2:$A$100``.
      4.  Attach a dropdown data-validation on column G of the Product Map that
          only accepts values present in the Charm Shops tab.
    """
    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    changed = False

    # ------------------------------------------------------------------ #
    # 1.  Charm Shops reference tab                                       #
    # ------------------------------------------------------------------ #
    if CHARM_SHOPS_SHEET not in wb.sheetnames:
        ws_cs = wb.create_sheet(CHARM_SHOPS_SHEET)
        ws_cs.sheet_properties.tabColor = "7B2D8B"

        hdr_fill = PatternFill("solid", fgColor="5B1A6B")
        hdr_font = Font("Calibri", bold=True, color="FFFFFF", size=12)
        thin     = Side(style="thin", color="C0C0C0")
        bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, h in enumerate(["Shop Name", "Stall", "Notes"], 1):
            cell           = ws_cs.cell(1, ci, h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bdr
        ws_cs.row_dimensions[1].height = 20

        row_fills = [
            PatternFill("solid", fgColor="F3E9FB"),
            PatternFill("solid", fgColor="FFFFFF"),
        ]
        for ridx, (shop, stall, notes) in enumerate(_DEFAULT_CHARM_SHOPS, start=2):
            fill = row_fills[ridx % 2]
            for ci, val in enumerate([shop, stall, notes], 1):
                cell           = ws_cs.cell(ridx, ci, val)
                cell.font      = Font("Calibri", size=11)
                cell.fill      = fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border    = bdr
            ws_cs.row_dimensions[ridx].height = 18

        ws_cs.column_dimensions["A"].width = 18
        ws_cs.column_dimensions["B"].width = 14
        ws_cs.column_dimensions["C"].width = 30
        ws_cs.freeze_panes = "A2"

        note_row = len(_DEFAULT_CHARM_SHOPS) + 3
        ws_cs.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row,   end_column=3,
        )
        note_cell = ws_cs.cell(note_row, 1, CHARM_SHOPS_INSTRUCTION_TEXT)
        _style_charm_instruction_cell(note_cell)
        ws_cs.row_dimensions[note_row].height = 72

        changed = True
        log.info(
            "Created '%s' sheet in %s with %d default charm shop(s).",
            CHARM_SHOPS_SHEET, path.name, len(_DEFAULT_CHARM_SHOPS),
        )

    # ------------------------------------------------------------------ #
    # 2.  Product Map  ─  column F (Charm Shop) header                    #
    # ------------------------------------------------------------------ #
    if CATALOG_SHEET in wb.sheetnames:
        ws_pm = wb[CATALOG_SHEET]
        if ws_pm.cell(1, _CHARM_COL_IDX).value != "Charm Shop":
            cell           = ws_pm.cell(1, _CHARM_COL_IDX, "Charm Shop")
            cell.fill      = _CAT_CHARM_HDR_FILL
            cell.font      = _CAT_CHARM_HDR_FONT
            cell.alignment = _CAT_CENTER
            cell.border    = _BORDER
            ws_pm.column_dimensions[_CHARM_COL_LETTER].width = 18
            changed = True
            log.info(
                "Added 'Charm Shop' column (G) to '%s' in %s.",
                CATALOG_SHEET, path.name,
            )

    # ------------------------------------------------------------------ #
    # 3 + 4.  Named range  +  dropdown validation                         #
    #                                                                      #
    # NOTE: only Charm Shop (F) + Supplier (C/D) dropdowns are applied     #
    # here; Charm Code (G) is added by init_charm_library_sheet which      #
    # always runs after this function. init_charm_library_sheet calls      #
    # _refresh_all_product_map_validations which clears stale DVs first.   #
    # ------------------------------------------------------------------ #
    if CATALOG_SHEET in wb.sheetnames:
        ws_pm = wb[CATALOG_SHEET]
        _apply_charm_column_validation(wb, ws_pm)
        _apply_supplier_column_validations(wb, ws_pm)
        changed = True

    if CHARM_SHOPS_SHEET in wb.sheetnames:
        if _upgrade_charm_shops_instruction_note(wb[CHARM_SHOPS_SHEET]):
            changed = True

    if changed:
        backup_supplier_catalog_before_write(path, "init_charm_shops")
        set_supplier_catalog_active_to_product_map(wb)
        wb.save(path)


def init_charm_library_sheet(path: Path) -> None:
    """
    Idempotent initialisation of the Charm Library and Product Map column G (Charm Code).

    Creates the ``Charm Library`` sheet (photo + code + metadata), the workbook
    named range ``CharmCodes`` for dropdown validation, column **Charm Code** on
    the Product Map, and validation hooks.  Requires ``Charm Shops`` / named
    range ``CharmShopNames`` to exist first (run ``init_charm_shops_sheet``).
    """
    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    changed = False

    if CHARM_LIBRARY_SHEET not in wb.sheetnames:
        ws_lib = wb.create_sheet(CHARM_LIBRARY_SHEET)
        ws_lib.sheet_properties.tabColor = "6B4C9A"

        hdr_fill = PatternFill("solid", fgColor="5B1A6B")
        hdr_font = Font("Calibri", bold=True, color="FFFFFF", size=12)
        thin     = Side(style="thin", color="C0C0C0")
        bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, h in enumerate(
            [
                "Photo",
                "Charm Code",
                CHARM_LIBRARY_COL_SKU_HEADER,
                "Default Charm Shop",
                "Notes",
            ],
            1,
        ):
            cell           = ws_lib.cell(1, ci, h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bdr
        ws_lib.row_dimensions[1].height = 20
        ws_lib.column_dimensions["A"].width = CHARM_LIB_COL_A_WIDTH
        ws_lib.column_dimensions["B"].width = 22
        ws_lib.column_dimensions["C"].width = 28
        ws_lib.column_dimensions["D"].width = 18
        ws_lib.column_dimensions["E"].width = 36
        ws_lib.freeze_panes = "A2"

        note_row = 4
        ws_lib.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row,   end_column=5,
        )
        note_cell = ws_lib.cell(note_row, 1, CHARM_LIBRARY_INSTRUCTION_TEXT)
        _style_charm_instruction_cell(note_cell)
        ws_lib.row_dimensions[note_row].height = _CHARM_LIB_NOTE_HEIGHT

        changed = True
        log.info("Created '%s' sheet in %s.", CHARM_LIBRARY_SHEET, path.name)

    if _ensure_charm_codes_named_range(wb):
        changed = True

    if CHARM_LIBRARY_SHEET in wb.sheetnames:
        ws_lib = wb[CHARM_LIBRARY_SHEET]
        if _upgrade_charm_library_instruction_note(ws_lib):
            changed = True
        if _upgrade_charm_library_sku_header(ws_lib):
            changed = True
        if _ensure_charm_library_sheet_layout(ws_lib):
            changed = True
        _apply_charm_library_default_shop_validation(wb, ws_lib)

    if CATALOG_SHEET in wb.sheetnames:
        ws_pm = wb[CATALOG_SHEET]
        if ws_pm.cell(1, _CHARM_CODE_COL_IDX).value != "Charm Code":
            cell           = ws_pm.cell(1, _CHARM_CODE_COL_IDX, "Charm Code")
            cell.fill      = _CAT_CHARM_HDR_FILL
            cell.font      = _CAT_CHARM_HDR_FONT
            cell.alignment = _CAT_CENTER
            cell.border    = _BORDER
            ws_pm.column_dimensions[_CHARM_CODE_COL_LETTER].width = 16
            changed = True
            log.info(
                "Added 'Charm Code' column (H) to '%s' in %s.",
                CATALOG_SHEET, path.name,
            )
        _refresh_all_product_map_validations(wb, ws_pm)
        changed = True

    # ------------------------------------------------------------------ #
    # 5.  Discontinued Products sheet (archive for removed products)      #
    # ------------------------------------------------------------------ #
    if DISCONTINUED_SHEET not in wb.sheetnames:
        _ensure_discontinued_sheet(wb)
        changed = True

    if changed:
        backup_supplier_catalog_before_write(path, "init_charm_library")
        set_supplier_catalog_active_to_product_map(wb)
        wb.save(path)


def load_charm_library(path: Path) -> dict[str, CharmLibraryEntry]:
    """
    Load charm rows from ``Charm Library`` and return a mapping
    ``charm_code → CharmLibraryEntry`` (photo bytes from embedded images).
    """
    if not path.exists():
        return {}
    try:
        row_photos = extract_photos_from_xlsx(
            path, sheet_name=CHARM_LIBRARY_SHEET, photo_col_idx=0
        )
    except Exception as exc:
        log.warning("Charm Library photo extraction skipped: %s", exc)
        row_photos = {}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if CHARM_LIBRARY_SHEET not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb[CHARM_LIBRARY_SHEET]
        by_code: dict[str, CharmLibraryEntry] = {}
        for r_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not code or code.lower() == "charm code":
                continue
            sku_val = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            def_shop = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            notes = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            ent = CharmLibraryEntry(
                code=code,
                sku=sku_val,
                default_charm_shop=def_shop,
                notes=notes,
                photo_bytes=row_photos.get(r_num),
            )
            if code in by_code:
                log.warning(
                    "Duplicate Charm Code %r in %s — using last row wins",
                    code, CHARM_LIBRARY_SHEET,
                )
            by_code[code] = ent
        wb.close()
        n_img = sum(1 for e in by_code.values() if e.photo_bytes)
        log.info(
            "Loaded %d charm(s) from '%s' (%d with photos)",
            len(by_code), CHARM_LIBRARY_SHEET, n_img,
        )
        return by_code
    except Exception as exc:
        log.warning("Could not load charm library from %s: %s", path.name, exc)
        return {}


def load_charm_shops(path: Path) -> list[CharmShop]:
    """
    Load charm shop entries from the ``Charm Shops`` sheet of
    *supplier_catalog.xlsx*.  Returns an empty list (with a warning) if the
    sheet is absent.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if CHARM_SHOPS_SHEET not in wb.sheetnames:
            wb.close()
            log.warning(
                "No '%s' sheet found in %s. "
                "Run with --init-charm-shops to create it.",
                CHARM_SHOPS_SHEET, path.name,
            )
            return []
        ws = wb[CHARM_SHOPS_SHEET]
        shops: list[CharmShop] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name  = str(row[0] or "").strip() if row[0]               else ""
            stall = str(row[1] or "").strip() if len(row) > 1 and row[1] else ""
            notes = str(row[2] or "").strip() if len(row) > 2 and row[2] else ""
            # Both shop name AND stall are required for a valid entry.
            # This prevents instructional/note rows (which have no stall) from
            # being treated as shops.
            if name and stall:
                shops.append(CharmShop(shop_name=name, stall=stall, notes=notes))
        wb.close()
        log.info(
            "Loaded %d charm shop(s) from '%s'", len(shops), CHARM_SHOPS_SHEET
        )
        return shops
    except Exception as exc:
        log.warning("Could not load charm shops from %s: %s", path.name, exc)
        return []


# ---------------------------------------------------------------------------
# Supplier catalog -- update (add new products with photos)
# ---------------------------------------------------------------------------


def _needs_catalog_entry(r: ResolvedItem) -> bool:
    """
    Return True when an item is NOT reliably matched to a supplier location and
    should therefore be shown in the unmatched / awaiting-info section of the
    shopping route.

    This controls ROUTING only.  For catalog-completeness (whether the item
    needs its own row in supplier_catalog.xlsx) see _needs_own_catalog_row().

    Three cases qualify:

    1. Truly unmatched – no catalog entry found at all (r.supplier is None).

    2. False-positive against an EMPTY catalog entry (shop_name AND stall both
       empty, score < EMPTY_ENTRY_MATCH_THRESHOLD).
       Risk: item lands in "Awaiting Supplier Info" blue section but has no
       corresponding catalog row to fill in.  Adding its own row fixes the UX.

    3. False-positive against a FILLED catalog entry (shop_name or stall present,
       score < FILLED_ENTRY_MATCH_THRESHOLD).
       Risk: item inherits the WRONG supplier / stall and goes to the wrong shop.
       Threshold calibrated against the observed gap between the highest
       false-positive score (83.1) and the lowest genuine match score (93.7).
    """
    if r.supplier is None:
        return True
    # False positive against an empty-info catalog entry
    if (not r.supplier.shop_name and not r.supplier.stall
            and r.match_score < EMPTY_ENTRY_MATCH_THRESHOLD):
        return True
    # False positive against a filled catalog entry → wrong supplier risk
    if ((r.supplier.shop_name or r.supplier.stall)
            and r.match_score < FILLED_ENTRY_MATCH_THRESHOLD):
        return True
    return False


def _needs_own_catalog_row(r: ResolvedItem) -> bool:
    """
    Return True when this item should receive its own row in supplier_catalog.xlsx.

    This is STRICTER than _needs_catalog_entry() and controls catalog
    completeness independently of routing.

    Why the distinction matters
    ---------------------------
    Suppliers frequently sell the same product in multiple colour / style
    variants with near-identical titles, e.g.:
      Catalog:  "Pink Miffy MagSafe Case … Kawaii Pastel Quicksand Gift"
      New order: "Purple Miffy MagSafe Case … Kawaii Pastel Quicksand Gift"

    These score ~94 on token_sort_ratio — well above FILLED_ENTRY_MATCH_THRESHOLD
    — so the routing is correct (both go to the same stall).  But each variant
    is a distinct product that needs its own catalog row for independent pricing
    and category tracking.

    SAME_PRODUCT_THRESHOLD (97) is the cutoff: scores above it are treated as
    the exact same product (no new row needed); scores below it trigger a new
    amber row that is pre-filled with the inferred supplier info so the user
    only has to add Category and Price.
    """
    if r.supplier is None:
        return True
    # Empty catalog entry: apply normal routing threshold
    if not r.supplier.shop_name and not r.supplier.stall:
        return r.match_score < EMPTY_ENTRY_MATCH_THRESHOLD
    # Filled catalog entry: need a score >= SAME_PRODUCT_THRESHOLD to suppress
    # a new row; anything below means "similar but distinct product"
    return r.match_score < SAME_PRODUCT_THRESHOLD


def update_catalog(path: Path, resolved: list[ResolvedItem]) -> int:
    """
    Append any unmatched items to the Product Map sheet with their product photo
    so the user can fill in supplier info (Shop Name, Stall, Price) later.
    Returns the number of rows added.

    Each new row receives:
      • Amber fill across all cells  – draws the eye to items needing attention
      • Bold product title           – consistent with the rest of the catalog
      • Yellow fill on the Price cell – secondary "fill me in" cue
      • Actionable note in column H  – tells the user exactly which columns to fill
      • ROW_HEIGHT set unconditionally – whether or not a photo is available
      • A border on all seven cells  – visually consistent with sorted rows

    The TOTAL row at the bottom of the sheet is located and its count updated.
    If no TOTAL row exists one is created.

    Duplicate guard: titles already present in the catalog (normalized exact-match)
    are silently skipped, preventing repeated additions across runs.

    IMPORTANT: rows are *appended* (never inserted) so existing image anchors are
    never shifted – inserting rows with openpyxl silently corrupts all photos.
    """
    # Use the stricter _needs_own_catalog_row so that colour / style variants
    # (high token_sort_ratio but different specific product) each get their own
    # row even when they match a filled catalog entry well enough to be routed.
    needs_row = [r for r in resolved if _needs_own_catalog_row(r)]
    if not needs_row:
        log.info("Catalog update: all products already catalogued, nothing to add")
        return 0

    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    ws = wb[CATALOG_SHEET]

    # ------------------------------------------------------------------
    # Single-pass scan: collect existing catalog titles (duplicate guard)
    # + locate the TOTAL row + count existing products.
    # ------------------------------------------------------------------
    existing_titles: set[str] = set()   # normalized titles already in catalog
    total_row_num: int | None = None
    product_count: int = 0
    for r_num in range(2, ws.max_row + 1):
        a_val = ws.cell(r_num, 1).value
        b_val = ws.cell(r_num, 2).value
        if a_val and isinstance(a_val, str) and a_val.startswith("TOTAL:"):
            total_row_num = r_num
        elif b_val and isinstance(b_val, str) and not b_val.startswith("TOTAL:"):
            existing_titles.add(_normalize(b_val))
            product_count += 1

    existing_titles |= _load_discontinued_titles(wb)

    # Deduplicate within this batch AND against what is already in the catalog.
    # Using normalized exact-match keeps distinct (but similar) products separate.
    seen_titles: set[str] = set()
    to_add: list[ResolvedItem] = []
    skipped_existing = 0
    for r in needs_row:
        key = _normalize(r.item.title)
        if key in existing_titles:
            log.debug("Catalog update: skipping '%s' – title already in catalog",
                      r.item.title[:70])
            skipped_existing += 1
            continue
        if key not in seen_titles:
            seen_titles.add(key)
            to_add.append(r)

    if skipped_existing:
        log.info(
            "Catalog update: skipped %d item(s) whose title already exists in catalog",
            skipped_existing,
        )

    if not to_add:
        log.info("Catalog update: no new products to add (all already catalogued)")
        return 0

    # ------------------------------------------------------------------
    # Remove existing TOTAL row so new products are never appended after it.
    # The TOTAL will be recreated at the end after all products are added.
    # ------------------------------------------------------------------
    if total_row_num is not None:
        try:
            ws.unmerge_cells(
                start_row=total_row_num, start_column=1,
                end_row=total_row_num,   end_column=PRODUCT_MAP_NUM_COLS,
            )
        except (KeyError, ValueError):
            pass
        for ci in range(1, PRODUCT_MAP_NUM_COLS + 1):
            cell = ws.cell(total_row_num, ci)
            cell.value = None
            cell.fill  = PatternFill()
            cell.font  = Font()
            cell.border = Border()
        total_row_num = None  # will be recreated below

    # ------------------------------------------------------------------
    # Append one row per new product  (image-safe: no insert_rows)
    # ------------------------------------------------------------------
    added = 0
    for r in to_add:
        row = ws.max_row + 1

        # Determine whether we have inferred supplier info from a high-confidence
        # partial match (variant product from the same stall).  If so, pre-fill
        # Shop Name and Stall so the user only needs to add Price (and charm if needed).
        inferred_shop  = ""
        inferred_stall = ""
        is_variant     = False
        if (r.supplier
                and (r.supplier.shop_name or r.supplier.stall)
                and r.match_score >= FILLED_ENTRY_MATCH_THRESHOLD):
            inferred_shop  = r.supplier.shop_name or ""
            inferred_stall = r.supplier.stall     or ""
            is_variant     = True   # product variant; location known, details TBD

        # A – photo placeholder (no text; image anchored below)
        ws.cell(row, 1).border = _BORDER
        ws.cell(row, 1).fill   = _CAT_WARN_FILL

        # B – product title  (bold, amber background)
        _cat_cell(ws, row, 2, r.item.title,  _CAT_WARN_FILL, _CAT_BODY_BOLD, _CAT_WRAP)

        # C – shop name  (pre-filled if variant, otherwise empty)
        _cat_cell(ws, row, 3, inferred_shop or None,  _CAT_WARN_FILL, _CAT_BODY, _CAT_WRAP)

        # D – stall  (pre-filled if variant, otherwise empty)
        _cat_cell(ws, row, 4, inferred_stall or None, _CAT_WARN_FILL, _CAT_BODY, _CAT_CENTER)

        # E – price  (always empty; distinct yellow fill = price TBD)
        _cat_cell(ws, row, 5, None,          _CAT_PRICE_FILL, _CAT_BODY,     _CAT_CENTER)

        # Determine if this product has a charm component (from the order style).
        item_has_charm = _style_has(r.item.style)[2]

        # F – charm shop assignment (foreign key → Charm Shops tab)
        if item_has_charm:
            _cat_cell(ws, row, 6, None,
                      _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, row, 6, None,
                      _NA_FILL, _NA_FONT, _CAT_CENTER)

        # G – Charm Library code (optional FK → Charm Library sheet)
        if item_has_charm:
            _cat_cell(ws, row, 7, None,
                      _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, row, 7, None,
                      _NA_FILL, _NA_FONT, _CAT_CENTER)

        # H – actionable note
        if is_variant:
            note = (
                f"NEW \u2500 Variant matched to {inferred_shop or '?'} / "
                f"{inferred_stall or '?'} \u2502  Verify location, "
                "fill Price \u00a5 (E)"
            )
            if item_has_charm:
                note += "  \u2502  Assign Charm Shop (F)"
        else:
            note = (
                "NEW \u2500 Fill in: Shop Name (C)  \u2502  Stall (D)  "
                "\u2502  Price \u00a5 (E)"
            )
            if item_has_charm:
                note += "  \u2502  Charm Shop (F)"
        _cat_cell(ws, row, 8, note, _CAT_WARN_FILL, _CAT_WARN_FONT, _CAT_WRAP)

        ws.row_dimensions[row].height = ROW_HEIGHT   # 46 pt, always

        if r.item.photo_bytes:
            try:
                xl_img        = XLImage(BytesIO(r.item.photo_bytes))
                xl_img.width  = PHOTO_PX
                xl_img.height = PHOTO_PX
                xl_img.anchor = f"A{row}"
                ws.add_image(xl_img)
            except Exception as e:
                log.warning("Could not embed catalog photo row %d: %s", row, e)

        added        += 1
        product_count += 1

    if added:
        # Always create a fresh TOTAL row at the very end (old one was cleared above).
        total_row_num = ws.max_row + 1
        ws.merge_cells(
            start_row=total_row_num, start_column=1,
            end_row=total_row_num,   end_column=PRODUCT_MAP_NUM_COLS,
        )
        tc           = ws.cell(total_row_num, 1, f"TOTAL: {product_count} products")
        tc.fill      = _CAT_HDR_FILL
        tc.font      = Font("Calibri", bold=True, size=11, color="FFFFFF")
        tc.alignment = _CAT_CENTER
        tc.border    = _BORDER
        ws.row_dimensions[total_row_num].height = 20

        sync_suppliers_from_product_map(wb)
        _refresh_all_product_map_validations(wb, ws)
        set_supplier_catalog_active_to_product_map(wb)
        backup_supplier_catalog_before_write(path, "append_new_products")
        wb.save(path)
        log.info(
            "Catalog updated: +%d new product(s)  (catalog total: %d)",
            added, product_count,
        )

    return added


# ---------------------------------------------------------------------------
# Suppliers sheet – unique Shop Name + Stall from Product Map
# ---------------------------------------------------------------------------

_SUPPLIERS_HEADER_ROW = (
    "ID",
    "Shop Name",
    "Mall",
    "Floor",
    "Stall",
    "Address",
    "Contact",
    "Notes",
)


def _ensure_suppliers_sheet(wb: openpyxl.Workbook):
    """Return the Suppliers worksheet, creating it (with headers) at index 0 if missing."""
    if SUPPLIERS_SHEET in wb.sheetnames:
        return wb[SUPPLIERS_SHEET]
    ws = wb.create_sheet(SUPPLIERS_SHEET, 0)
    for ci, h in enumerate(_SUPPLIERS_HEADER_ROW, 1):
        ws.cell(1, ci, h)
    return ws


def sync_suppliers_from_product_map(wb: openpyxl.Workbook) -> int:
    """
    Append rows on *Suppliers* for each distinct (Shop Name, Stall) pair that appears
    on *Product Map* columns C/D (SHOP NAME / STALL), skipping pairs already listed on *Suppliers*.
    Existing supplier rows and columns C–D–F–H (Mall, Floor, …) are left untouched.

    Only pairs with **both** shop and stall non-empty after stripping are considered.

    Returns the number of new rows appended.
    """
    if CATALOG_SHEET not in wb.sheetnames:
        log.warning("sync_suppliers: no %r sheet — skipped", CATALOG_SHEET)
        return 0

    ws_pm = wb[CATALOG_SHEET]
    ws_sup = _ensure_suppliers_sheet(wb)

    # Existing (shop, stall) keys on Suppliers + highest numeric ID in column A
    existing: set[tuple[str, str]] = set()
    max_id = 0
    for r in range(2, ws_sup.max_row + 1):
        shop_e = str(ws_sup.cell(r, 2).value or "").strip()
        stall_e = str(ws_sup.cell(r, 5).value or "").strip()
        if shop_e and stall_e:
            existing.add((shop_e, stall_e))
        vid = ws_sup.cell(r, 1).value
        if vid is not None:
            try:
                max_id = max(max_id, int(vid))
            except (TypeError, ValueError):
                pass

    # Unique pairs from Product Map (preserve first-seen order, then sort for stability)
    seen_pm: set[tuple[str, str]] = set()
    pairs: list[tuple[str, str]] = []
    for r_num in range(2, ws_pm.max_row + 1):
        b_val = ws_pm.cell(r_num, 2).value
        if not b_val or not isinstance(b_val, str):
            continue
        title = b_val.strip()
        if title.startswith("TOTAL:") or title == "Unknown Product":
            continue
        shop = str(ws_pm.cell(r_num, 3).value or "").strip()
        stall = str(ws_pm.cell(r_num, 4).value or "").strip()
        if not shop or not stall:
            continue
        key = (shop, stall)
        if key in seen_pm:
            continue
        seen_pm.add(key)
        pairs.append(key)

    pairs.sort(
        key=lambda t: (_stall_floor(t[1]), t[1].lower(), t[0].lower()),
    )

    added = 0
    for shop, stall in pairs:
        if (shop, stall) in existing:
            continue
        max_id += 1
        row = ws_sup.max_row + 1
        ws_sup.cell(row, 1, max_id)
        ws_sup.cell(row, 2, shop)
        ws_sup.cell(row, 5, stall)
        existing.add((shop, stall))
        added += 1

    if added:
        log.info(
            "Suppliers sheet: appended %d row(s) (unique shop+stall from Product Map)",
            added,
        )
    return added


def clear_product_map_charm_codes(path: Path) -> int:
    """
    Clear Product Map column G (*Charm Code*) on every product data row.

    Use when column G contains incorrect text (e.g. charm shop names duplicated
    from column F). Valid codes must match the **Charm Library** column B list;
    users re-select via the dropdown after clearing.
    """
    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    ws = wb[CATALOG_SHEET]
    col = _CHARM_CODE_COL_IDX
    cleared = 0
    for r in range(2, ws.max_row + 1):
        b_val = ws.cell(r, 2).value
        if not b_val or not isinstance(b_val, str):
            continue
        title = b_val.strip()
        if title.startswith("TOTAL:") or title == "Unknown Product":
            continue
        cell = ws.cell(r, col)
        if cell.value is None:
            continue
        cell.value = None
        cleared += 1
    if cleared:
        log.info("Product Map: cleared Charm Code (column G) on %d row(s).", cleared)
    _refresh_all_product_map_validations(wb, ws)
    set_supplier_catalog_active_to_product_map(wb)
    backup_supplier_catalog_before_write(path, "clear_charm_codes")
    wb.save(path)
    return cleared


# ---------------------------------------------------------------------------
# Supplier catalog -- rebuild (re-sort all rows after user fills in info)
# ---------------------------------------------------------------------------


def _anchor_row(img) -> int | None:
    """Return the 1-based worksheet row that an openpyxl image is anchored to."""
    anchor = img.anchor
    if isinstance(anchor, str):
        # Plain string like "A5" – strip the column letter(s)
        digits = anchor.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz")
        try:
            return int(digits)
        except ValueError:
            return None
    # OneCellAnchor / TwoCellAnchor from a loaded workbook – ._from is 0-based
    if hasattr(anchor, "_from"):
        return anchor._from.row + 1
    return None


def _is_catalog_row_filled(row: dict) -> bool:
    """A catalog row is considered filled when shop_name or stall has been entered."""
    return bool(row["shop_name"] or row["stall"])


def _is_auto_note(text: str) -> bool:
    """Return True for auto-generated 'NEW – Fill in…' notes that should be cleared."""
    t = text.strip()
    return t.startswith("NEW") or t.startswith("NEW \u2500") or t.startswith("NEW --")


def rebuild_catalog(path: Path) -> None:
    """
    Re-sort supplier_catalog.xlsx so that every row whose Shop Name / Stall is
    already filled in appears first, ordered by ascending floor number then stall
    code → shop name → product title (the same order used in shopping_route.xlsx).
    Amber rows still awaiting supplier info are placed at the end sorted by title.

    Three bugs from the previous append-only approach are fixed here:
      1. Auto-generated "NEW – Fill in…" notes are stripped from filled rows.
      2. Sort uses _stall_floor() so rows group by physical floor, not alphabet.
      3. The sheet is fully recreated (remove + create) instead of relying on
         delete_rows(), which is unreliable when merged cells are present and
         leaves ghost blank rows behind.

    A timestamped backup is written under data/supplier_catalog_backups/ before changes.
    """
    backup_supplier_catalog_before_write(path, "rebuild_catalog")

    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    ws = wb[CATALOG_SHEET]

    # ------------------------------------------------------------------ #
    # Step 1 – Capture sheet metadata to re-apply after recreation        #
    # ------------------------------------------------------------------ #
    col_widths: dict[str, float] = {
        letter: ws.column_dimensions[letter].width
        for letter in ws.column_dimensions
    }
    freeze_panes = ws.freeze_panes
    tab_color    = ws.sheet_properties.tabColor

    # Preserve header row values (row 1); 8 columns A–H (no CATEGORY).
    hdr_values = [ws.cell(1, c).value for c in range(1, PRODUCT_MAP_NUM_COLS + 1)]
    while len(hdr_values) < PRODUCT_MAP_NUM_COLS:
        hdr_values.append(None)
    hdr_values = hdr_values[:PRODUCT_MAP_NUM_COLS]
    hdr_values[0] = "PHOTO"
    hdr_values[1] = "PRODUCT TITLE"
    hdr_values[2] = "SHOP NAME"
    hdr_values[3] = "STALL"
    hdr_values[4] = "PRICE (¥)"
    if str(hdr_values[5] or "").strip() != "Charm Shop":
        hdr_values[5] = "Charm Shop"
    if str(hdr_values[6] or "").strip() != "Charm Code":
        hdr_values[6] = "Charm Code"
    if str(hdr_values[7] or "").strip().upper() not in ("NOTES", "NOTE"):
        hdr_values[7] = "NOTES"

    # ------------------------------------------------------------------ #
    # Step 2 – Extract photo bytes keyed by their anchor row              #
    # ------------------------------------------------------------------ #
    row_photos: dict[int, bytes] = {}
    for img in list(ws._images):
        r = _anchor_row(img)
        if r is None:
            continue
        try:
            if hasattr(img, "_data"):
                data = img._data()
            elif hasattr(img, "ref") and hasattr(img.ref, "read"):
                img.ref.seek(0)
                data = img.ref.read()
            else:
                data = None
            if data:
                row_photos[r] = data
        except Exception as e:
            log.warning("Could not extract photo from row %d: %s", r, e)

    # ------------------------------------------------------------------ #
    # Step 3 – Read every data row (skip header and TOTAL)                #
    # ------------------------------------------------------------------ #
    rows_data: list[dict] = []
    for r_num in range(2, ws.max_row + 1):
        b_val = ws.cell(r_num, 2).value
        if not b_val or not isinstance(b_val, str):
            continue
        if b_val.strip().startswith("TOTAL:"):
            continue
        raw_notes = (
            str(ws.cell(r_num, 8).value or "").strip() if ws.max_column >= 8 else ""
        )
        rows_data.append({
            "title":      b_val.strip(),
            "shop_name":  str(ws.cell(r_num, 3).value or "").strip(),
            "stall":      str(ws.cell(r_num, 4).value or "").strip(),
            "price":      str(ws.cell(r_num, 5).value or "").strip(),
            # Strip auto-generated "NEW – Fill in…" templates; keep genuine user notes
            "notes":      "" if _is_auto_note(raw_notes) else raw_notes,
            "charm_shop": str(ws.cell(r_num, 6).value or "").strip()
                          if ws.max_column >= 6 else "",
            "charm_code": str(ws.cell(r_num, 7).value or "").strip()
                          if ws.max_column >= 7 else "",
            "photo":      row_photos.get(r_num),
        })

    if not rows_data:
        log.warning("Catalog rebuild: no data rows found – aborting")
        return

    # ------------------------------------------------------------------ #
    # Step 4 – Sort                                                        #
    #   Filled rows: ascending floor → stall code → shop name → title     #
    #   (mirrors the exact sort key used by _sheet_route)                  #
    #   Unfilled amber rows: alphabetical by title, placed at the end      #
    # ------------------------------------------------------------------ #
    filled = sorted(
        [r for r in rows_data if     _is_catalog_row_filled(r)],
        key=lambda r: (
            _stall_floor(r["stall"]),
            r["stall"].lower(),
            r["shop_name"].lower(),
            r["title"].lower(),
        ),
    )
    unfilled = sorted(
        [r for r in rows_data if not _is_catalog_row_filled(r)],
        key=lambda r: r["title"].lower(),
    )
    sorted_rows = filled + unfilled

    # ------------------------------------------------------------------ #
    # Step 5 – Replace the sheet entirely                                  #
    #   Using remove + create_sheet avoids delete_rows() which is broken  #
    #   when the sheet contains merged cells (TOTAL row) and leaves ghost  #
    #   blank rows behind.                                                 #
    # ------------------------------------------------------------------ #
    sheet_idx = wb.sheetnames.index(CATALOG_SHEET)
    wb.remove(ws)
    ws = wb.create_sheet(CATALOG_SHEET, sheet_idx)

    # Restore column widths, freeze, and tab colour
    for letter, width in col_widths.items():
        if width:
            ws.column_dimensions[letter].width = width
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = PHOTO_COL_W

    # ------------------------------------------------------------------ #
    # Step 6 – Write header row (row 1)                                   #
    # ------------------------------------------------------------------ #
    ws.row_dimensions[1].height = 22
    for ci, val in enumerate(hdr_values, 1):
        cell           = ws.cell(1, ci, val)
        cell.fill      = _CAT_HDR_FILL
        cell.font      = _CAT_HDR_FONT
        cell.alignment = _CAT_CENTER
        cell.border    = _BORDER
    # Columns F/G (Charm Shop / Charm Code) — purple headers
    h_hdr           = ws.cell(1, _CHARM_COL_IDX)
    h_hdr.fill      = _CAT_CHARM_HDR_FILL
    h_hdr.font      = _CAT_CHARM_HDR_FONT
    i_hdr           = ws.cell(1, _CHARM_CODE_COL_IDX)
    i_hdr.fill      = _CAT_CHARM_HDR_FILL
    i_hdr.font      = _CAT_CHARM_HDR_FONT

    # ------------------------------------------------------------------ #
    # Step 7 – Write sorted data rows                                     #
    # ------------------------------------------------------------------ #
    _FILLED_FILL = PatternFill("solid", fgColor="FFFFFF")
    product_count = 0
    for i, rd in enumerate(sorted_rows):
        r          = i + 2          # row 1 is the header
        is_amber   = not _is_catalog_row_filled(rd)
        body_fill  = _CAT_WARN_FILL  if is_amber else _FILLED_FILL
        price_fill = _CAT_PRICE_FILL if is_amber else _FILLED_FILL

        # Column A – photo placeholder cell (image anchored separately)
        ws.cell(r, 1).border = _BORDER
        ws.cell(r, 1).fill   = body_fill

        _cat_cell(ws, r, 2, rd["title"],              body_fill,  _CAT_BODY_BOLD, _CAT_WRAP)
        _cat_cell(ws, r, 3, rd["shop_name"] or None,  body_fill,  _CAT_BODY,      _CAT_WRAP)
        _cat_cell(ws, r, 4, rd["stall"]     or None,  body_fill,  _CAT_BODY,      _CAT_CENTER)
        _cat_cell(ws, r, 5, rd["price"]     or None,  price_fill, _CAT_BODY,      _CAT_CENTER)

        # Column F – Charm Shop (foreign key → Charm Shops tab)
        charm_shop_val = rd.get("charm_shop") or None
        if charm_shop_val:
            _cat_cell(ws, r, 6, charm_shop_val, _FILLED_FILL, _CAT_BODY, _CAT_CENTER)
        elif "charm" in rd["title"].lower():
            _cat_cell(ws, r, 6, None, _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, r, 6, None, _NA_FILL, _NA_FONT, _CAT_CENTER)

        # Column G – Charm Library code
        cc_val = (rd.get("charm_code") or "").strip()
        if cc_val:
            _cat_cell(ws, r, 7, cc_val, body_fill, _CAT_BODY, _CAT_CENTER)
        elif "charm" in rd["title"].lower():
            _cat_cell(ws, r, 7, None, _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, r, 7, None, _NA_FILL, _NA_FONT, _CAT_CENTER)

        # Column H – NOTES (amber template or user notes)
        if is_amber:
            note_text = (
                rd["notes"] or
                "NEW \u2500 Fill in: Shop Name (C)  \u2502  Stall (D)"
                "  \u2502  Price \u00a5 (E)  \u2502  Charm Shop (F)"
            )
            _cat_cell(ws, r, 8, note_text, _CAT_WARN_FILL, _CAT_WARN_FONT, _CAT_WRAP)
        else:
            _cat_cell(ws, r, 8, rd["notes"] or None, _FILLED_FILL, _CAT_BODY, _CAT_WRAP)

        ws.row_dimensions[r].height = ROW_HEIGHT

        if rd["photo"]:
            try:
                xl_img        = XLImage(BytesIO(rd["photo"]))
                xl_img.width  = PHOTO_PX
                xl_img.height = PHOTO_PX
                xl_img.anchor = f"A{r}"
                ws.add_image(xl_img)
            except Exception as e:
                log.warning("Photo re-embed failed at A%d: %s", r, e)

        product_count += 1

    # ------------------------------------------------------------------ #
    # Step 8 – TOTAL row  (merged across all 8 columns)                    #
    # ------------------------------------------------------------------ #
    total_row = product_count + 2
    ws.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row,   end_column=PRODUCT_MAP_NUM_COLS,   # A–H
    )
    tc           = ws.cell(total_row, 1, f"TOTAL: {product_count} products")
    tc.fill      = _CAT_HDR_FILL
    tc.font      = Font("Calibri", bold=True, size=11, color="FFFFFF")
    tc.alignment = _CAT_CENTER
    tc.border    = _BORDER
    ws.row_dimensions[total_row].height = 20

    # ------------------------------------------------------------------ #
    # Step 9 – Re-apply column F/G/H widths + dropdown validation         #
    # ------------------------------------------------------------------ #
    ws.column_dimensions[_CHARM_COL_LETTER].width = 18
    ws.column_dimensions[_CHARM_CODE_COL_LETTER].width = 16
    ws.column_dimensions["H"].width = 30.7

    sync_suppliers_from_product_map(wb)
    _refresh_all_product_map_validations(wb, ws)
    set_supplier_catalog_active_to_product_map(wb)
    wb.save(path)
    log.info(
        "Catalog rebuilt: %d products  (%d with supplier info, %d still awaiting)",
        product_count, len(filled), len(unfilled),
    )


# ---------------------------------------------------------------------------
# Product Map — mark discontinued (supplier no longer sells)
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class ProductMapPickerRow:
    """One Product Map row for UI pickers (photos + routing context)."""

    row_num: int
    title: str
    shop_name: str
    stall: str
    price: str = ""
    charm_shop: str = ""
    charm_code: str = ""
    notes: str = ""


def list_product_map_rows_for_picker(path: Path) -> list[ProductMapPickerRow]:
    """
    Return structured rows for each Product Map product (skips TOTAL).
    Used by the discontinued-product dialog (thumbnails + shop / stall).
    """
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[CATALOG_SHEET]
    except KeyError:
        wb.close()
        return []
    out: list[ProductMapPickerRow] = []
    for r_num in range(2, ws.max_row + 1):
        b_val = ws.cell(r_num, 2).value
        if not b_val or not isinstance(b_val, str):
            continue
        title = b_val.strip()
        if title.startswith("TOTAL:") or title == "Unknown Product":
            continue
        shop       = str(ws.cell(r_num, 3).value or "").strip()
        stall      = str(ws.cell(r_num, 4).value or "").strip()
        price      = str(ws.cell(r_num, 5).value or "").strip()
        charm_shop = str(ws.cell(r_num, 6).value or "").strip()
        charm_code = str(ws.cell(r_num, 7).value or "").strip()
        notes      = str(ws.cell(r_num, 8).value or "").strip()
        out.append(
            ProductMapPickerRow(
                row_num=r_num,
                title=title,
                shop_name=shop,
                stall=stall,
                price=price,
                charm_shop=charm_shop,
                charm_code=charm_code,
                notes=notes,
            )
        )
    wb.close()
    return out


def list_product_map_data_rows(path: Path) -> list[tuple[int, str, bool]]:
    """
    Return ``(excel_row, product_title, is_discontinued)`` for each Product Map
    data row (skips TOTAL). Legacy compatibility wrapper.
    """
    return [
        (r.row_num, r.title, False)
        for r in list_product_map_rows_for_picker(path)
    ]


def update_product_map_cells(
    path: Path,
    row_num: int,
    *,
    shop_name: str | None = None,
    stall: str | None = None,
    charm_shop: str | None = None,
    charm_code: str | None = None,
) -> None:
    """
    Write one or more editable fields on a single Product Map row.

    Only non-``None`` parameters are written; the rest are left untouched.
    After writing, named-range validations are refreshed and the workbook is
    saved.

    Product Map 8-column layout
    ---------------------------
    A Photo |     B Title | C Shop Name | D Stall | E Price | F Charm Shop | G Charm Code | H Notes
    """
    backup_supplier_catalog_before_write(path, "update_cells")
    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    ws = wb[CATALOG_SHEET]

    if row_num < 2 or row_num > ws.max_row:
        wb.close()
        raise ValueError(f"Row {row_num} is out of range (2 .. {ws.max_row}).")

    if shop_name is not None:
        ws.cell(row_num, 3).value = shop_name or None
    if stall is not None:
        ws.cell(row_num, 4).value = stall or None
    if charm_shop is not None:
        ws.cell(row_num, 6).value = charm_shop or None
    if charm_code is not None:
        ws.cell(row_num, 7).value = charm_code or None

    _refresh_all_product_map_validations(wb, ws)
    set_supplier_catalog_active_to_product_map(wb)
    wb.save(path)
    wb.close()


def resolve_product_map_row_for_discontinue(path: Path, query: str) -> tuple[int, str]:
    """
    Find a single Product Map row from *query*: exact normalized title match first,
    else unique case-insensitive substring match. Raises ``ValueError`` if none
    or ambiguous.
    """
    qn = _normalize(query)
    if not qn:
        raise ValueError("Empty title query.")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[CATALOG_SHEET]
    except KeyError:
        wb.close()
        raise ValueError(f"Sheet '{CATALOG_SHEET}' not found.") from None
    exact: list[tuple[int, str]] = []
    partial: list[tuple[int, str]] = []
    q_lo = query.strip().lower()
    for r_num in range(2, ws.max_row + 1):
        b_val = ws.cell(r_num, 2).value
        if not b_val or not isinstance(b_val, str):
            continue
        title = b_val.strip()
        if title.startswith("TOTAL:") or title == "Unknown Product":
            continue
        tn = _normalize(title)
        if tn == qn:
            exact.append((r_num, title))
        elif q_lo and q_lo in title.lower():
            partial.append((r_num, title))
    wb.close()
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        lines = "\n".join(f"  row {r}: {t[:100]}" for r, t in exact[:25])
        raise ValueError(f"Multiple rows share the same normalized title:\n{lines}")
    if len(partial) == 1:
        return partial[0]
    if not partial:
        raise ValueError("No product row matched that title.")
    lines = "\n".join(f"  row {r}: {t[:100]}" for r, t in partial[:25])
    raise ValueError(
        "Several products match — type a longer unique phrase, pick a row in the app, "
        f"or use --mark-product-discontinued-row:\n{lines}"
    )


_DISC_HDR_FILL = PatternFill("solid", fgColor="7F1D1D")
_DISC_HDR_FONT = Font("Calibri", bold=True, color="FFFFFF", size=11)
_DISC_BODY     = Font("Calibri", size=10)
_DISC_BODY_BOLD = Font("Calibri", bold=True, size=10)
_DISC_ROW_FILL = PatternFill("solid", fgColor="FEF2F2")
_DISC_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DISC_WRAP     = Alignment(vertical="center", wrap_text=True)

_DISC_HEADERS = [
    "PHOTO", "PRODUCT TITLE", "SHOP NAME", "STALL",
    "PRICE", "CHARM SHOP", "CHARM CODE", "NOTES",
    "DISCONTINUED DATE",
]

_DISC_COL_WIDTHS = {
    "A": PHOTO_COL_W,  # match Product Map — wide enough for PHOTO_PX embeds
    "B": 48.0, "C": 14.0, "D": 14.0, "E": 10.0,
    "F": 9.0, "G": 14.0, "H": 26.0, "I": 20.0,
}


def _ensure_discontinued_sheet(wb) -> None:
    """Create the ``Discontinued Products`` sheet if absent."""
    if DISCONTINUED_SHEET in wb.sheetnames:
        return
    ws = wb.create_sheet(DISCONTINUED_SHEET)
    ws.sheet_properties.tabColor = "7F1D1D"
    for ci, h in enumerate(_DISC_HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.fill = _DISC_HDR_FILL
        c.font = _DISC_HDR_FONT
        c.alignment = _DISC_CENTER
        c.border = _BORDER
    ws.row_dimensions[1].height = 22
    for letter, w in _DISC_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = "A2"
    log.info("Created '%s' sheet.", DISCONTINUED_SHEET)


def _load_discontinued_titles(wb) -> set[str]:
    """Return normalized titles from the Discontinued Products sheet (for duplicate guard)."""
    if DISCONTINUED_SHEET not in wb.sheetnames:
        return set()
    ws = wb[DISCONTINUED_SHEET]
    titles: set[str] = set()
    for r in range(2, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if b and isinstance(b, str):
            t = b.strip()
            if t and not t.startswith("TOTAL:"):
                titles.add(_normalize(t))
    return titles


def load_discontinued_titles(path: Path) -> set[str]:
    """Public helper: normalized titles from Discontinued Products sheet (read-only)."""
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = _load_discontinued_titles(wb)
    wb.close()
    return result


def mark_product_map_discontinued_by_row(path: Path, row_num: int, *, marker: str = "Y") -> str:
    """
    Move the product at *row_num* from Product Map to the ``Discontinued Products``
    sheet, then **rebuild** the Product Map sheet from scratch (minus the removed
    row) so that every photo is re-anchored at the correct position.

    Using ``ws.delete_rows()`` is intentionally avoided because openpyxl does not
    reliably shift image anchors when rows are deleted, causing photos to become
    misaligned with their product data.
    """
    backup_supplier_catalog_before_write(path, "mark_discontinued")
    init_charm_shops_sheet(path)
    init_charm_library_sheet(path)

    # Extract ALL photos from the saved-on-disk file (reliable ZIP-level read).
    row_photos = extract_photos_from_xlsx(path, sheet_name=CATALOG_SHEET, photo_col_idx=0)

    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    ws = wb[CATALOG_SHEET]
    if row_num < 2 or row_num > ws.max_row:
        raise ValueError(f"Row {row_num} is not a valid data row.")
    b_val = ws.cell(row_num, 2).value
    if not b_val or not isinstance(b_val, str):
        raise ValueError(f"Row {row_num} has no product title in column B.")
    title = b_val.strip()
    if title.startswith("TOTAL:") or title == "Unknown Product":
        raise ValueError(f"Row {row_num} is not a product row.")

    # ---- Copy the product to Discontinued Products ----
    _ensure_discontinued_sheet(wb)
    ws_disc = wb[DISCONTINUED_SHEET]

    disc_row = ws_disc.max_row + 1
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    src_vals = [ws.cell(row_num, c).value for c in range(1, PRODUCT_MAP_NUM_COLS + 1)]
    while len(src_vals) < PRODUCT_MAP_NUM_COLS:
        src_vals.append(None)

    ws_disc.cell(disc_row, 1).border = _BORDER
    ws_disc.cell(disc_row, 1).fill = _DISC_ROW_FILL
    for ci, val in enumerate(src_vals[1:], 2):
        v = str(val).strip() if val is not None else None
        if v and v.lower() == "none":
            v = None
        is_title = ci == 2
        _cat_cell(ws_disc, disc_row, ci, v, _DISC_ROW_FILL,
                  _DISC_BODY_BOLD if is_title else _DISC_BODY,
                  _DISC_WRAP if is_title else _DISC_CENTER)
    _cat_cell(ws_disc, disc_row, 9, stamp, _DISC_ROW_FILL, _DISC_BODY, _DISC_CENTER)

    ws_disc.row_dimensions[disc_row].height = ROW_HEIGHT

    photo_data = row_photos.get(row_num)
    if photo_data:
        try:
            xl_img = XLImage(BytesIO(photo_data))
            xl_img.width = PHOTO_PX
            xl_img.height = PHOTO_PX
            xl_img.anchor = f"A{disc_row}"
            ws_disc.add_image(xl_img)
        except Exception as e:
            log.warning("Could not embed photo in %s row %d: %s", DISCONTINUED_SHEET, disc_row, e)

    ws_disc.column_dimensions["A"].width = PHOTO_COL_W

    # ---- Rebuild the Product Map sheet (minus the deleted row) ----
    # Capture sheet metadata.
    col_widths = {
        letter: ws.column_dimensions[letter].width
        for letter in ws.column_dimensions
    }
    freeze_panes = ws.freeze_panes
    tab_color    = ws.sheet_properties.tabColor
    hdr_values   = [ws.cell(1, c).value for c in range(1, PRODUCT_MAP_NUM_COLS + 1)]

    # Read every data row EXCEPT the one being discontinued and any TOTAL rows.
    remaining: list[dict] = []
    for r in range(2, ws.max_row + 1):
        if r == row_num:
            continue
        rv = ws.cell(r, 2).value
        if not rv or not isinstance(rv, str):
            continue
        if rv.strip().startswith("TOTAL:"):
            continue
        remaining.append({
            "title":      rv.strip(),
            "shop_name":  str(ws.cell(r, 3).value or "").strip(),
            "stall":      str(ws.cell(r, 4).value or "").strip(),
            "price":      str(ws.cell(r, 5).value or "").strip(),
            "charm_shop": str(ws.cell(r, 6).value or "").strip() if ws.max_column >= 6 else "",
            "charm_code": str(ws.cell(r, 7).value or "").strip() if ws.max_column >= 7 else "",
            "notes":      str(ws.cell(r, 8).value or "").strip() if ws.max_column >= 8 else "",
            "photo":      row_photos.get(r),
            "is_filled":  bool(ws.cell(r, 3).value and ws.cell(r, 4).value),
        })

    # Remove the old sheet and recreate it.
    sheet_idx = wb.sheetnames.index(CATALOG_SHEET)
    wb.remove(ws)
    ws = wb.create_sheet(CATALOG_SHEET, sheet_idx)

    for letter, width in col_widths.items():
        if width:
            ws.column_dimensions[letter].width = width
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = PHOTO_COL_W

    # Header row.
    ws.row_dimensions[1].height = 22
    for ci, val in enumerate(hdr_values, 1):
        cell           = ws.cell(1, ci, val)
        cell.fill      = _CAT_HDR_FILL
        cell.font      = _CAT_HDR_FONT
        cell.alignment = _CAT_CENTER
        cell.border    = _BORDER
    h_hdr           = ws.cell(1, _CHARM_COL_IDX)
    h_hdr.fill      = _CAT_CHARM_HDR_FILL
    h_hdr.font      = _CAT_CHARM_HDR_FONT
    i_hdr           = ws.cell(1, _CHARM_CODE_COL_IDX)
    i_hdr.fill      = _CAT_CHARM_HDR_FILL
    i_hdr.font      = _CAT_CHARM_HDR_FONT

    _FILLED_FILL = PatternFill("solid", fgColor="FFFFFF")
    product_count = 0
    for rd in remaining:
        r = product_count + 2
        is_amber  = not rd["is_filled"]
        body_fill = _CAT_WARN_FILL if is_amber else _FILLED_FILL
        price_fill = _CAT_PRICE_FILL if is_amber else _FILLED_FILL

        ws.cell(r, 1).border = _BORDER
        ws.cell(r, 1).fill   = body_fill
        _cat_cell(ws, r, 2, rd["title"],              body_fill,  _CAT_BODY_BOLD, _CAT_WRAP)
        _cat_cell(ws, r, 3, rd["shop_name"] or None,  body_fill,  _CAT_BODY,      _CAT_WRAP)
        _cat_cell(ws, r, 4, rd["stall"]     or None,  body_fill,  _CAT_BODY,      _CAT_CENTER)
        _cat_cell(ws, r, 5, rd["price"]     or None,  price_fill, _CAT_BODY,      _CAT_CENTER)

        charm_shop_val = rd.get("charm_shop") or None
        if charm_shop_val:
            _cat_cell(ws, r, 6, charm_shop_val, _FILLED_FILL, _CAT_BODY, _CAT_CENTER)
        elif "charm" in rd["title"].lower():
            _cat_cell(ws, r, 6, None, _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, r, 6, None, _NA_FILL, _NA_FONT, _CAT_CENTER)

        cc_val = (rd.get("charm_code") or "").strip()
        if cc_val:
            _cat_cell(ws, r, 7, cc_val, body_fill, _CAT_BODY, _CAT_CENTER)
        elif "charm" in rd["title"].lower():
            _cat_cell(ws, r, 7, None, _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, r, 7, None, _NA_FILL, _NA_FONT, _CAT_CENTER)

        notes_val = rd.get("notes") or None
        if is_amber and not notes_val:
            notes_val = (
                "NEW \u2500 Fill in: Shop Name (C)  \u2502  Stall (D)"
                "  \u2502  Price \u00a5 (E)  \u2502  Charm Shop (F)"
            )
            _cat_cell(ws, r, 8, notes_val, _CAT_WARN_FILL, _CAT_WARN_FONT, _CAT_WRAP)
        else:
            _cat_cell(ws, r, 8, notes_val, body_fill if not is_amber else _CAT_WARN_FILL,
                      _CAT_BODY if not is_amber else _CAT_WARN_FONT, _CAT_WRAP)

        ws.row_dimensions[r].height = ROW_HEIGHT

        if rd["photo"]:
            try:
                xl_img        = XLImage(BytesIO(rd["photo"]))
                xl_img.width  = PHOTO_PX
                xl_img.height = PHOTO_PX
                xl_img.anchor = f"A{r}"
                ws.add_image(xl_img)
            except Exception as e:
                log.warning("Photo re-embed failed at A%d: %s", r, e)

        product_count += 1

    # TOTAL row.
    total_row = product_count + 2
    ws.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row,   end_column=PRODUCT_MAP_NUM_COLS,
    )
    tc           = ws.cell(total_row, 1, f"TOTAL: {product_count} products")
    tc.fill      = _CAT_HDR_FILL
    tc.font      = Font("Calibri", bold=True, size=11, color="FFFFFF")
    tc.alignment = _CAT_CENTER
    tc.border    = _BORDER
    ws.row_dimensions[total_row].height = 20

    # Re-apply column widths + dropdown validations (sheet was recreated).
    ws.column_dimensions[_CHARM_COL_LETTER].width = 18
    ws.column_dimensions[_CHARM_CODE_COL_LETTER].width = 16
    ws.column_dimensions["H"].width = 30.7
    _refresh_all_product_map_validations(wb, ws)

    set_supplier_catalog_active_to_product_map(wb)
    wb.save(path)
    log.info("Discontinued product moved to '%s': row %d — %s (at %s). "
             "Product Map rebuilt with %d products.",
             DISCONTINUED_SHEET, disc_row, title[:80], stamp, product_count)
    return title


def mark_product_map_discontinued(path: Path, title_query: str) -> tuple[int, str]:
    """Resolve *title_query* to one row and move it to Discontinued Products."""
    row_num, title = resolve_product_map_row_for_discontinue(path, title_query)
    mark_product_map_discontinued_by_row(path, row_num)
    return row_num, title


def update_product_map_photo(path: Path, row_num: int, photo_bytes: bytes) -> None:
    """
    Replace the embedded photo for Product Map row *row_num* with *photo_bytes*.

    Because openpyxl cannot remove or replace existing worksheet images in-place,
    the Product Map sheet is fully rebuilt (all rows + photos re-embedded) using
    the same technique as :func:`mark_product_map_discontinued_by_row`.

    *photo_bytes* must be valid image data (JPEG, PNG, WEBP, …).  The caller is
    responsible for any pre-processing (resize, format conversion, etc.).
    """
    if not photo_bytes:
        raise ValueError("photo_bytes must be non-empty.")

    backup_supplier_catalog_before_write(path, "update_photo")
    # Extract current photos so we can re-embed every other row unchanged.
    row_photos = extract_photos_from_xlsx(path, sheet_name=CATALOG_SHEET, photo_col_idx=0)
    # Inject the new photo for the target row.
    row_photos[row_num] = photo_bytes

    wb = openpyxl.load_workbook(path)
    ensure_catalog_column_layout(wb)
    ws = wb[CATALOG_SHEET]

    if row_num < 2 or row_num > ws.max_row:
        wb.close()
        raise ValueError(f"Row {row_num} is not a valid data row (2 .. {ws.max_row}).")
    b_val = ws.cell(row_num, 2).value
    if not b_val or not isinstance(b_val, str):
        wb.close()
        raise ValueError(f"Row {row_num} has no product title in column B.")

    # Capture sheet metadata before destroying the sheet.
    col_widths   = {letter: ws.column_dimensions[letter].width for letter in ws.column_dimensions}
    freeze_panes = ws.freeze_panes
    tab_color    = ws.sheet_properties.tabColor
    hdr_values   = [ws.cell(1, c).value for c in range(1, PRODUCT_MAP_NUM_COLS + 1)]

    # Read all data rows, preserving order and values.
    data_rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        rv = ws.cell(r, 2).value
        if not rv or not isinstance(rv, str):
            continue
        if rv.strip().startswith("TOTAL:"):
            continue
        data_rows.append({
            "orig_row":   r,
            "title":      rv.strip(),
            "shop_name":  str(ws.cell(r, 3).value or "").strip(),
            "stall":      str(ws.cell(r, 4).value or "").strip(),
            "price":      str(ws.cell(r, 5).value or "").strip(),
            "charm_shop": str(ws.cell(r, 6).value or "").strip() if ws.max_column >= 6 else "",
            "charm_code": str(ws.cell(r, 7).value or "").strip() if ws.max_column >= 7 else "",
            "notes":      str(ws.cell(r, 8).value or "").strip() if ws.max_column >= 8 else "",
            "photo":      row_photos.get(r),
            "is_filled":  bool(ws.cell(r, 3).value and ws.cell(r, 4).value),
        })

    # Rebuild the sheet from scratch so photo anchors are always correct.
    sheet_idx = wb.sheetnames.index(CATALOG_SHEET)
    wb.remove(ws)
    ws = wb.create_sheet(CATALOG_SHEET, sheet_idx)

    for letter, width in col_widths.items():
        if width:
            ws.column_dimensions[letter].width = width
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = PHOTO_COL_W

    # Header row.
    ws.row_dimensions[1].height = 22
    for ci, val in enumerate(hdr_values, 1):
        cell           = ws.cell(1, ci, val)
        cell.fill      = _CAT_HDR_FILL
        cell.font      = _CAT_HDR_FONT
        cell.alignment = _CAT_CENTER
        cell.border    = _BORDER
    ws.cell(1, _CHARM_COL_IDX).fill      = _CAT_CHARM_HDR_FILL
    ws.cell(1, _CHARM_COL_IDX).font      = _CAT_CHARM_HDR_FONT
    ws.cell(1, _CHARM_CODE_COL_IDX).fill = _CAT_CHARM_HDR_FILL
    ws.cell(1, _CHARM_CODE_COL_IDX).font = _CAT_CHARM_HDR_FONT

    _FILLED_FILL = PatternFill("solid", fgColor="FFFFFF")
    product_count = 0
    for rd in data_rows:
        r         = product_count + 2
        is_amber  = not rd["is_filled"]
        body_fill = _CAT_WARN_FILL if is_amber else _FILLED_FILL
        price_fill = _CAT_PRICE_FILL if is_amber else _FILLED_FILL

        ws.cell(r, 1).border = _BORDER
        ws.cell(r, 1).fill   = body_fill
        _cat_cell(ws, r, 2, rd["title"],              body_fill,  _CAT_BODY_BOLD, _CAT_WRAP)
        _cat_cell(ws, r, 3, rd["shop_name"] or None,  body_fill,  _CAT_BODY,      _CAT_WRAP)
        _cat_cell(ws, r, 4, rd["stall"]     or None,  body_fill,  _CAT_BODY,      _CAT_CENTER)
        _cat_cell(ws, r, 5, rd["price"]     or None,  price_fill, _CAT_BODY,      _CAT_CENTER)

        charm_shop_val = rd.get("charm_shop") or None
        if charm_shop_val:
            _cat_cell(ws, r, 6, charm_shop_val, _FILLED_FILL, _CAT_BODY, _CAT_CENTER)
        elif "charm" in rd["title"].lower():
            _cat_cell(ws, r, 6, None, _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, r, 6, None, _NA_FILL, _NA_FONT, _CAT_CENTER)

        cc_val = (rd.get("charm_code") or "").strip()
        if cc_val:
            _cat_cell(ws, r, 7, cc_val, body_fill, _CAT_BODY, _CAT_CENTER)
        elif "charm" in rd["title"].lower():
            _cat_cell(ws, r, 7, None, _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
        else:
            _cat_cell(ws, r, 7, None, _NA_FILL, _NA_FONT, _CAT_CENTER)

        notes_val = rd.get("notes") or None
        if is_amber and not notes_val:
            notes_val = (
                "NEW \u2500 Fill in: Shop Name (C)  \u2502  Stall (D)"
                "  \u2502  Price \u00a5 (E)  \u2502  Charm Shop (F)"
            )
            _cat_cell(ws, r, 8, notes_val, _CAT_WARN_FILL, _CAT_WARN_FONT, _CAT_WRAP)
        else:
            _cat_cell(ws, r, 8, notes_val,
                      body_fill if not is_amber else _CAT_WARN_FILL,
                      _CAT_BODY if not is_amber else _CAT_WARN_FONT, _CAT_WRAP)

        ws.row_dimensions[r].height = ROW_HEIGHT

        if rd["photo"]:
            try:
                xl_img        = XLImage(BytesIO(rd["photo"]))
                xl_img.width  = PHOTO_PX
                xl_img.height = PHOTO_PX
                xl_img.anchor = f"A{r}"
                ws.add_image(xl_img)
            except Exception as e:
                log.warning("Photo re-embed failed at A%d: %s", r, e)

        product_count += 1

    # TOTAL row.
    total_row = product_count + 2
    ws.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row,   end_column=PRODUCT_MAP_NUM_COLS,
    )
    tc           = ws.cell(total_row, 1, f"TOTAL: {product_count} products")
    tc.fill      = _CAT_HDR_FILL
    tc.font      = Font("Calibri", bold=True, size=11, color="FFFFFF")
    tc.alignment = _CAT_CENTER
    tc.border    = _BORDER
    ws.row_dimensions[total_row].height = 20

    ws.column_dimensions[_CHARM_COL_LETTER].width     = 18
    ws.column_dimensions[_CHARM_CODE_COL_LETTER].width = 16
    ws.column_dimensions["H"].width = 30.7
    _refresh_all_product_map_validations(wb, ws)

    set_supplier_catalog_active_to_product_map(wb)
    wb.save(path)
    wb.close()
    log.info("Product Map photo updated for row %d.", row_num)


# ---------------------------------------------------------------------------
# Order cache -- persist resolved items across runs (handles deleted source PDFs)
# ---------------------------------------------------------------------------


def _resolved_to_dict(r: ResolvedItem) -> dict:
    """Serialise a ResolvedItem to a JSON-compatible dict (photos as base64)."""
    return {
        "order_number":    r.order.order_number,
        "etsy_shop":       r.order.etsy_shop,
        "buyer_name":      r.order.buyer_name,
        "buyer_username":  r.order.buyer_username,
        "ship_to_name":    r.order.ship_to_name,
        "ship_to_country": r.order.ship_to_country,
        "order_date":      r.order.order_date,
        "private_notes":   r.order.private_notes,
        "title":           r.item.title,
        "quantity":        r.item.quantity,
        "phone_model":     r.item.phone_model,
        "style":           r.item.style,
        "photo_b64":       base64.b64encode(r.item.photo_bytes).decode()
                           if r.item.photo_bytes else None,
        "shop_name":       r.supplier.shop_name  if r.supplier else "",
        "stall":           r.supplier.stall       if r.supplier else "",
        "category":        r.supplier.category    if r.supplier else "",
        "price":           r.supplier.price       if r.supplier else "",
        "notes":           r.supplier.notes       if r.supplier else "",
        "charm_shop":      r.supplier.charm_shop  if r.supplier else "",
        "charm_code":      r.supplier.charm_code  if r.supplier else "",
        "match_score":     r.match_score,
    }


def _dict_to_resolved(d: dict) -> ResolvedItem:
    """Deserialise a cached dict back into a ResolvedItem."""
    order = Order(
        order_number    = d["order_number"],
        etsy_shop       = d["etsy_shop"],
        buyer_name      = d.get("buyer_name",      ""),
        buyer_username  = d.get("buyer_username",  ""),
        ship_to_name    = d.get("ship_to_name",    ""),
        ship_to_country = d.get("ship_to_country", ""),
        order_date      = d.get("order_date",      ""),
        private_notes   = d.get("private_notes",   ""),
    )
    photo_bytes = base64.b64decode(d["photo_b64"]) if d.get("photo_b64") else None
    item = OrderItem(
        title       = d["title"],
        quantity    = d.get("quantity",    1),
        phone_model = d.get("phone_model", ""),
        style       = d.get("style",       ""),
        photo_bytes = photo_bytes,
    )
    order.items = [item]
    supplier = None
    if d.get("shop_name"):
        supplier = CatalogEntry(
            product_title = d["title"],
            category  = d.get("category", ""),
            shop_name = d["shop_name"],
            stall     = d.get("stall",    ""),
            price     = d.get("price",    ""),
            notes     = d.get("notes",    ""),
            charm_shop = d.get("charm_shop", ""),
            charm_code = d.get("charm_code", ""),
        )
    return ResolvedItem(order=order, item=item, supplier=supplier,
                        match_score=d.get("match_score", 0.0))


def save_cache(
    path: Path,
    resolved: list[ResolvedItem],
    processed_pdfs: set[str] | None = None,
) -> None:
    """Write all resolved items (including photo bytes) to a JSON cache file.

    ``processed_pdfs`` is a set of PDF base-filenames that have already been
    fully ingested.  It is persisted alongside the order items so that
    ``--new-batch`` can skip re-parsing them on future runs.
    """
    data = {
        "processed_pdfs": sorted(processed_pdfs or []),
        "items": [_resolved_to_dict(r) for r in resolved],
    }
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    log.info("Cache saved: %d items -> %s", len(resolved), path.name)


def load_cache(path: Path) -> tuple[list[ResolvedItem], set[str]]:
    """Load previously cached resolved items.

    Returns ``(items, processed_pdfs)`` where *processed_pdfs* is the set of
    PDF filenames that were already ingested in a prior run.
    Returns ``([], set())`` if the cache is absent or corrupt.
    """
    if not path.exists():
        return [], set()
    try:
        data           = json.loads(path.read_text(encoding="utf-8"))
        items          = [_dict_to_resolved(d) for d in data.get("items", [])]
        processed_pdfs = set(data.get("processed_pdfs", []))
        log.info("Cache loaded: %d prior order(s) from %s", len(items), path.name)
        return items, processed_pdfs
    except Exception as e:
        log.warning("Cache load failed (%s) -- starting fresh", e)
        return [], set()


def extract_photos_from_xlsx(
    xlsx_path: Path,
    sheet_name: str = "Shopping Route",
    photo_col_idx: int = 1,     # 0-based column index of the Photo column (B = 1)
) -> dict[int, bytes]:
    """
    Extract embedded images from *sheet_name* in an xlsx file.

    Returns {excel_row_1based: jpeg_bytes} for every image anchored to
    *photo_col_idx*.  Works with both oneCellAnchor and twoCellAnchor.

    Because .xlsx files are ZIP archives we can read images directly without
    relying on openpyxl's (limited) image-read support.
    """
    _R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    _XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    _A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    def _zip_path(target: str, base_dir: str = 'xl') -> str:
        """
        Normalise a relationship Target value to a bare zip entry path.
        Handles absolute (/xl/...), relative (../...), and plain names.
        """
        t = target.strip()
        if t.startswith('/'):
            return t[1:]                         # absolute: strip leading /
        if t.startswith('../'):
            parent = base_dir.rsplit('/', 1)[0]  # go up one dir
            return parent + '/' + t[3:]
        if '/' not in t:
            return base_dir + '/' + t            # bare filename in same dir
        if not t.startswith('xl/'):
            return 'xl/' + t
        return t

    result: dict[int, bytes] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            all_files = set(zf.namelist())

            # ---- Step 1: find the worksheet rId for sheet_name ----
            wb_tree  = ET.parse(zf.open('xl/workbook.xml'))
            sheet_rId: str | None = None
            for elem in wb_tree.getroot().iter():
                if elem.tag.endswith('}sheet') or elem.tag == 'sheet':
                    if elem.get('name') == sheet_name:
                        sheet_rId = elem.get(f'{{{_R}}}id') or elem.get('r:id')
                        break
            if not sheet_rId:
                return result

            # ---- Step 2: rId → worksheet XML zip path ----
            wb_rels = 'xl/_rels/workbook.xml.rels'
            if wb_rels not in all_files:
                return result
            sheet_file: str | None = None
            for rel in ET.parse(zf.open(wb_rels)).getroot():
                if rel.get('Id') == sheet_rId:
                    sheet_file = _zip_path(rel.get('Target', ''), 'xl')
                    break
            if not sheet_file or sheet_file not in all_files:
                return result

            # ---- Step 3: worksheet rels → drawing XML zip path ----
            ws_dir  = sheet_file.rsplit('/', 1)[0]   # e.g. 'xl/worksheets'
            ws_name = sheet_file.rsplit('/', 1)[1]
            ws_rels = f'{ws_dir}/_rels/{ws_name}.rels'
            if ws_rels not in all_files:
                return result
            drawing_file: str | None = None
            for rel in ET.parse(zf.open(ws_rels)).getroot():
                if 'drawing' in rel.get('Type', '').lower():
                    drawing_file = _zip_path(rel.get('Target', ''), ws_dir)
                    break
            if not drawing_file or drawing_file not in all_files:
                return result

            # ---- Step 4: drawing rels → rId to image zip path ----
            dr_dir  = drawing_file.rsplit('/', 1)[0]  # e.g. 'xl/drawings'
            dr_name = drawing_file.rsplit('/', 1)[1]
            dr_rels = f'{dr_dir}/_rels/{dr_name}.rels'
            if dr_rels not in all_files:
                return result
            rId_to_img: dict[str, str] = {}
            for rel in ET.parse(zf.open(dr_rels)).getroot():
                rid      = rel.get('Id', '')
                img_path = _zip_path(rel.get('Target', ''), dr_dir)
                rId_to_img[rid] = img_path

            # ---- Step 5: parse drawing anchors → row → bytes ----
            d_root = ET.parse(zf.open(drawing_file)).getroot()
            for anchor in d_root:
                a_tag = anchor.tag.split('}')[-1]
                if a_tag not in ('oneCellAnchor', 'twoCellAnchor'):
                    continue
                fr = anchor.find(f'{{{_XDR}}}from')
                if fr is None:
                    continue
                col_e = fr.find(f'{{{_XDR}}}col')
                row_e = fr.find(f'{{{_XDR}}}row')
                if col_e is None or row_e is None:
                    continue
                if int(col_e.text) != photo_col_idx:
                    continue
                excel_row = int(row_e.text) + 1   # 0-based → 1-based

                blip = anchor.find(f'.//{{{_A}}}blip')
                if blip is None:
                    continue
                r_embed   = blip.get(f'{{{_R}}}embed', '')
                img_path  = rId_to_img.get(r_embed, '')
                if img_path and img_path in all_files:
                    result[excel_row] = zf.read(img_path)

    except Exception as e:
        log.warning("extract_photos_from_xlsx(%s): %s", xlsx_path.name, e)

    if result:
        log.info("Extracted %d embedded photo(s) from %s", len(result), xlsx_path.name)
    return result


def load_items_from_xlsx(xlsx_path: Path) -> list[ResolvedItem]:
    """
    Read order items directly from an existing shopping_route.xlsx Shopping Route
    sheet and reconstruct ResolvedItem objects.  Used as a fallback on the very
    first run after upgrading to the cache system (or if the cache is deleted).

    Note: product photos cannot be recovered from the Excel file, so photo_bytes
    will be None for orders loaded via this path.
    """
    if not xlsx_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Shopping Route" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Shopping Route"]
    except Exception as e:
        log.warning("Could not read existing orders from %s: %s", xlsx_path.name, e)
        return []

    # Extract embedded photos keyed by Excel row number (1-based)
    row_photos = extract_photos_from_xlsx(xlsx_path)

    items: list[ResolvedItem] = []
    charm_order_nums: set[str] = set()
    _items_vals = {"case only", "grip only", "case, grip", "\u2014",
                  "仅手机壳", "仅支架", "手机壳、支架"}

    for row_num, row_vals in enumerate(
        ws.iter_rows(min_row=5, values_only=True), start=5
    ):
        if not row_vals or len(row_vals) < 12:
            continue
        is_new_format = (
            len(row_vals) >= 14
            or (len(row_vals) >= 13 and str(row_vals[6] or "").strip() in _items_vals)
        )
        if is_new_format:
            order_cell = row_vals[12] if len(row_vals) > 12 else None
            case_idx, grip_idx, qty_idx, phone_idx, etsy_idx = 7, 8, 11, 10, 13
        else:
            order_cell = row_vals[11] if len(row_vals) > 11 else None
            case_idx, grip_idx, qty_idx, phone_idx, etsy_idx = 6, 7, 10, 9, 12
        if not order_cell or not isinstance(order_cell, str):
            continue

        order_str = order_cell.strip()

        if order_str.startswith("~C:"):
            continue
        if order_str.startswith("~?#") or order_str.startswith("~#"):
            _prefix_len = 3 if order_str.startswith("~?#") else 2
            charm_onum = order_str[_prefix_len:].strip()
            if charm_onum.isdigit():
                charm_order_nums.add(charm_onum)
            continue

        order_num = order_str.lstrip("#").strip()
        if not order_num.isdigit():
            continue

        case_val = str(row_vals[case_idx]).strip() if row_vals[case_idx] else "N/A"
        grip_val = str(row_vals[grip_idx]).strip() if row_vals[grip_idx] else "N/A"
        style_parts = []
        if case_val != "N/A": style_parts.append("Case")
        if grip_val != "N/A": style_parts.append("Grip")
        style = "+".join(style_parts) if style_parts else "Case"

        # Private Notes: col 15 (index 14) in new EN 15-col format
        _pn_idx = 14
        _pn_raw = (
            str(row_vals[_pn_idx]).strip()
            if is_new_format and len(row_vals) > _pn_idx and row_vals[_pn_idx]
            else ""
        )
        order = Order(
            order_number  = order_num,
            etsy_shop     = str(row_vals[etsy_idx]).strip() if len(row_vals) > etsy_idx and row_vals[etsy_idx] else "",
            private_notes = _pn_raw,
        )
        item = OrderItem(
            title       = str(row_vals[5]).strip() if row_vals[5] else "",
            quantity    = int(row_vals[qty_idx]) if isinstance(row_vals[qty_idx], (int, float)) else 1,
            phone_model = str(row_vals[phone_idx]).strip() if len(row_vals) > phone_idx and row_vals[phone_idx] else "",
            style       = style,
            photo_bytes = row_photos.get(row_num),
        )
        order.items = [item]

        shop_name = str(row_vals[3]).strip() if row_vals[3] else ""
        stall     = str(row_vals[4]).strip() if row_vals[4] else ""
        # Treat "--" as empty for supplier/stall
        if shop_name in ("--", "???"):
            shop_name = ""
        if stall in ("--", "???"):
            stall = ""
        supplier = None
        # Create a supplier entry whenever there is any location info, so
        # floor-based sorting still works even with incomplete catalog data.
        if shop_name or stall:
            supplier = CatalogEntry(
                product_title = item.title,
                shop_name     = shop_name,
                stall         = stall,
            )

        items.append(ResolvedItem(order=order, item=item, supplier=supplier,
                                  match_score=0.0))

    # Patch charm component back into style for orders identified in the
    # charm section.  The cache normally carries the full style string, so
    # this only matters when falling back to load_items_from_xlsx.
    if charm_order_nums:
        for r in items:
            if (r.order.order_number in charm_order_nums
                    and "charm" not in r.item.style.lower()):
                r.item.style = (r.item.style + "+Charm") if r.item.style else "Charm"

    wb.close()
    if items:
        log.info("Loaded %d prior order(s) from existing %s (cache not found)",
                 len(items), xlsx_path.name)
    return items


def load_existing_statuses(xlsx_path: Path) -> dict[tuple[str, str], str]:
    """
    Read the Shopping Route sheet from an existing output file and return every
    non-default component status so they can be re-applied after a re-generate.

    Returns {(order_number, component): status_string}
    where component is one of 'case', 'grip', 'charm'.
    Only 'Purchased', 'Out of Stock', and 'Out of Production' are preserved
    (Pending and N/A are defaults that will be restored automatically).
    """
    if not xlsx_path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Shopping Route" not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb["Shopping Route"]
    except Exception as e:
        log.warning("Could not read existing statuses from %s: %s", xlsx_path.name, e)
        return {}

    keep = {"Purchased", "Out of Stock", "Out of Production"}
    _items_vals = {"case only", "grip only", "case, grip", "\u2014",
                  "仅手机壳", "仅支架", "手机壳、支架"}
    statuses: dict[tuple[str, str], str] = {}

    for row_vals in ws.iter_rows(min_row=5, values_only=True):
        if not row_vals or len(row_vals) < 12:
            continue
        # Detect format: new has "Items to Purchase" col 7 (index 6)
        is_new_format = (
            len(row_vals) >= 14
            or (len(row_vals) >= 13 and str(row_vals[6] or "").strip() in _items_vals)
        )
        if is_new_format:
            order_cell = row_vals[12] if len(row_vals) > 12 else None
            case_col, grip_col, charm_col = 7, 8, 9
        else:
            order_cell = row_vals[11] if len(row_vals) > 11 else None
            case_col, grip_col, charm_col = 6, 7, 8
        if not order_cell or not isinstance(order_cell, str):
            continue

        order_str = order_cell.strip()

        # col 6 = Product title — used to distinguish items within the same order
        title_val  = str(row_vals[5]).strip() if row_vals[5] else ""
        norm_title = _normalize(title_val)[:50]

        # Aggregated charm row: ~C:<charm_code>
        if order_str.startswith("~C:"):
            charm_code_key = order_str[3:].strip()
            if charm_code_key and len(row_vals) > charm_col:
                val = str(row_vals[charm_col]).strip() if row_vals[charm_col] else ""
                if val in keep:
                    statuses[(charm_code_key, "", "charm_agg")] = val
            continue

        # Legacy per-order charm row: ~#<order_num> (backward compat)
        if order_str.startswith("~#") or order_str.startswith("~?#"):
            prefix_len = 3 if order_str.startswith("~?#") else 2
            charm_order_num = order_str[prefix_len:].strip()
            if charm_order_num.isdigit() and len(row_vals) > charm_col:
                val = str(row_vals[charm_col]).strip() if row_vals[charm_col] else ""
                if val in keep:
                    statuses[(charm_order_num, norm_title, "charm")] = val
            continue

        # Regular case/grip supplier row: "#XXXXXXX"
        order_num = order_str.lstrip("#").strip()
        if not order_num.isdigit():
            continue
        for comp, col_idx in (("case", case_col), ("grip", grip_col)):
            val = str(row_vals[col_idx]).strip() if row_vals[col_idx] else ""
            if val in keep:
                statuses[(order_num, norm_title, comp)] = val

    wb.close()
    if statuses:
        log.info("Preserved %d non-default status value(s) from existing file",
                 len(statuses))
    return statuses


# ---------------------------------------------------------------------------
# Shopping-route helpers
# ---------------------------------------------------------------------------


def _stall_floor(stall: str) -> int:
    """
    Parse the floor number from a stall code for ascending-floor sort order.

    Conventions observed in the catalog:
      A2xxx / A2-xx  -> 2nd floor (A-block)
      4Cxx  / 4Dxx   -> 4th floor
      5Xxx  / 5Cxx   -> 5th floor
      Chinese text containing "4D" / "5C" etc. -> parse embedded digit

    Returns 999 for unknown stalls so they sort to the very end.
    """
    if not stall or stall in ("\u2014", "???", ""):
        return 999
    if re.match(r"^A2", stall, re.IGNORECASE):
        return 2
    m = re.match(r"^(\d)", stall)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d)[A-Za-z]", stall)
    if m:
        return int(m.group(1))
    return 999


def _style_has(style: str) -> tuple[bool, bool, bool]:
    """Return (has_case, has_grip, has_charm) booleans from a style string."""
    s = style.lower()
    # "stand" / "kickstand" (e.g. "Case+Stand+Charm", "Kickstand Cover") = grip
    has_grip = "grip" in s or "stand" in s
    return "case" in s, has_grip, "charm" in s


def _style_flags(style: str) -> tuple[str, str, str]:
    """
    Return checkmark strings for the Case/Grip/Charm columns (Orders Detail).
    Empty string means the component is not included.
    """
    has_case, has_grip, has_charm = _style_has(style)
    return (
        "\u2713" if has_case  else "",
        "\u2713" if has_grip  else "",
        "\u2713" if has_charm else "",
    )


def _section_complete(status: str | None) -> bool:
    """
    True if the component status indicates procurement is complete.
    Both "Purchased" and "Out of Production" are terminal — no further action.
    """
    return status in ("Purchased", "Out of Production")


def _items_to_purchase(
    has_case: bool,
    has_grip: bool,
    case_status: str | None,
    grip_status: str | None,
    lang: str = "en",
) -> str:
    """
    Return a concise label for the "Items to Purchase" column.

    Components needing purchase: Pending, Out of Stock, or Out of Production.
    Returns: "case only", "grip only", "case, grip", or "—" when nothing needed.
    """
    needs_action = {"Pending", "Out of Stock", "Out of Production"}
    case_needs = has_case and (case_status or "Pending") in needs_action
    grip_needs = has_grip and (grip_status or "Pending") in needs_action
    if case_needs and grip_needs:
        return _t("case, grip", lang) if lang == "zh" else "case, grip"
    if case_needs:
        return _t("case only", lang) if lang == "zh" else "case only"
    if grip_needs:
        return _t("grip only", lang) if lang == "zh" else "grip only"
    return "\u2014"  # em dash


def _get_oop_components_being_purged(
    order_num: str,
    title: str,
    style: str,
    statuses: dict[tuple[str, str, str], str],
) -> list[tuple[str, str, str]]:
    """
    Return [(order_num, title, component), ...] for components marked
    "Out of Production" that are part of a section being purged.
    Used to append records to out_of_production_log.csv before removal.
    """
    has_case, has_grip, has_charm = _style_has(style)
    norm_title = _normalize(title)[:50]
    result: list[tuple[str, str, str]] = []

    case_status = statuses.get((order_num, norm_title, "case"))
    grip_status = statuses.get((order_num, norm_title, "grip"))
    charm_status = statuses.get((order_num, norm_title, "charm"))

    case_done = (not has_case) or _section_complete(case_status)
    grip_done = (not has_grip) or _section_complete(grip_status)
    charm_done = (not has_charm) or _section_complete(charm_status)

    case_grip_present = has_case or has_grip
    charm_present = has_charm
    case_grip_section_done = case_grip_present and case_done and grip_done
    charm_section_done = charm_present and charm_done

    if case_grip_section_done:
        if has_case and case_status == "Out of Production":
            result.append((order_num, title, "Case"))
        if has_grip and grip_status == "Out of Production":
            result.append((order_num, title, "Grip"))
    if charm_section_done and has_charm and charm_status == "Out of Production":
        result.append((order_num, title, "Charm"))

    return result


def _append_to_oop_log(
    records: list[tuple[str, str, str, str]],
    log_path: Path,
) -> None:
    """
    Append Out-of-Production records to the CSV log file.
    records: [(order_num, product_title, component, etsy_shop), ...]
    Creates the file with headers if it does not exist.
    """
    if not records:
        return
    try:
        file_exists = log_path.exists()
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Order #", "Product Title", "Component", "Etsy Shop", "Date Recorded"])
            today = date.today().isoformat()
            for order_num, product_title, component, etsy_shop in records:
                writer.writerow([order_num, product_title, component, etsy_shop, today])
        log.info("Recorded %d Out-of-Production item(s) to %s", len(records), log_path.name)
    except OSError as e:
        log.warning("Could not write to Out-of-Production log %s: %s", log_path.name, e)


def _compute_remaining_style(
    order_num: str,
    title: str,
    style: str,
    statuses: dict[tuple[str, str, str], str],
) -> str | None:
    """
    Determine the remaining (not-yet-purchased) style string after applying
    independent section-level purge logic.

    The shopping route splits a multi-component order into two independent
    procurement sections that are visited on separate shopping trips:

      • Case / Grip section  — purchased at the supplier floors
      • Charm section        — purchased at a completely separate building

    Each section is evaluated independently:
      - A section is "done" when every component *within that section* that is
        present in the order style is marked "Purchased" or "Out of Production".
      - If a section is done, its components are stripped from the returned style
        so the next re-generate omits that section entirely.
      - If a section is not yet done, it is retained in full (even if some
        components within it are already complete — the row stays visible
        until the whole section is done).

    Returns:
      None            – all sections complete → caller should remove item.
      str (modified)  – at least one section was stripped; item stays but only
                        the remaining section(s) will appear in the next run.
      str (original)  – nothing complete / style unchanged; item stays as-is.

    Notes:
      - A component absent from *statuses* is treated as "Pending" because only
        non-default values are stored in the dict.
      - "Out of Production" is treated as complete (same as Purchased) so those
        items are purged and do not reappear on the next shopping run.
      - Items whose style contains no recognisable components are kept as-is
        (safety guard against malformed style strings).
    """
    has_case, has_grip, has_charm = _style_has(style)
    norm_title = _normalize(title)[:50]

    # ── Case / Grip section ──────────────────────────────────────────────────
    # Present when the order includes at least one of Case or Grip.
    # Complete when every present component is "Purchased" or "Out of Production".
    case_grip_present = has_case or has_grip
    case_done = (not has_case) or _section_complete(statuses.get((order_num, norm_title, "case")))
    grip_done = (not has_grip) or _section_complete(statuses.get((order_num, norm_title, "grip")))
    case_grip_section_done = case_grip_present and case_done and grip_done

    # ── Charm section ────────────────────────────────────────────────────────
    # Present when the order includes a Charm component.
    # Complete when the charm status is "Purchased" or "Out of Production".
    charm_present = has_charm
    charm_done = _section_complete(statuses.get((order_num, norm_title, "charm")))
    charm_section_done = charm_present and charm_done

    # Safety: style has no known components — leave untouched
    if not case_grip_present and not charm_present:
        return style

    # Build the list of components that still need attention
    remaining: list[str] = []
    if has_case and not case_grip_section_done:
        remaining.append("Case")
    if has_grip and not case_grip_section_done:
        remaining.append("Grip")
    if has_charm and not charm_section_done:
        remaining.append("Charm")

    if not remaining:
        return None  # every section is complete → drop item entirely

    return "+".join(remaining)


# ---------------------------------------------------------------------------
# Fuzzy matching
# ---------------------------------------------------------------------------


def _normalize(text: str) -> str:
    text = text.replace("|", ",")
    return re.sub(r"\s+", " ", text).strip().lower()


def _route_item_sort_key(r: ResolvedItem) -> tuple[str, str, str]:
    """Within one supplier stop: stack rows with the same product title together.

    Order is independent of Case/Grip/Charm status so \"Out of Stock\" and
    \"Pending\" lines for the same product appear consecutively.
    """
    return (
        _normalize(r.item.title),
        r.item.phone_model or "",
        r.order.order_number,
    )


def match_items(
    orders: list[Order],
    catalog: list[CatalogEntry],
    threshold: int,
) -> list[ResolvedItem]:
    catalog_titles = [_normalize(e.product_title) for e in catalog]
    resolved: list[ResolvedItem] = []

    for order in orders:
        for item in order.items:
            norm = _normalize(item.title)
            result = process.extractOne(
                norm,
                catalog_titles,
                scorer=fuzz.token_sort_ratio,
                score_cutoff=threshold,
            )
            if result:
                _, score, idx = result
                resolved.append(ResolvedItem(order, item, catalog[idx], score))
            else:
                resolved.append(ResolvedItem(order, item, None, 0.0))

    matched = sum(1 for r in resolved if r.supplier)
    log.info(
        "Matched %d / %d items (threshold %d%%)", matched, len(resolved), threshold
    )
    return resolved


_CHARM_IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".webp")


def _disk_charm_files_index(root: Path) -> dict[str, tuple[str, str]]:
    """
    Map charm code (filename stem) → (path relative to *root* with /, basename).
    One file per stem; extension priority left-to-right in _CHARM_IMAGE_EXTENSIONS.
    """
    best: dict[str, tuple[str, str]] = {}
    for ext in _CHARM_IMAGE_EXTENSIONS:
        for path in root.rglob(f"*{ext}"):
            if not path.is_file():
                continue
            st = path.stem
            if not st or st.lower() in ("charm code", "photo"):
                continue
            if st in best:
                continue
            best[st] = (path.relative_to(root).as_posix(), path.name)
    return best


def _atomic_write_text(path: Path, text: str, *, encoding: str = "utf-8") -> None:
    """Write *text* to *path* via a same-directory ``*.tmp`` rename (crash-safe)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    try:
        tmp.write_text(text, encoding=encoding)
        tmp.replace(path)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise


def export_charm_manifest(
    catalog_path: Path,
    charm_images_dir: Path | None,
    output_path: Path,
    *,
    route_snapshot: dict | None = None,
) -> int:
    """
    Emit JSON merging Charm Library rows and files under *charm_images_dir*
    (including subfolders).  Intended for websites, imports, or audits.

    Writes atomically so consumers never read a half-written JSON file.
    """
    charms: dict[str, dict] = {}

    if catalog_path.exists():
        try:
            wb = openpyxl.load_workbook(catalog_path, read_only=True, data_only=True)
            if CHARM_LIBRARY_SHEET in wb.sheetnames:
                ws = wb[CHARM_LIBRARY_SHEET]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    if not code or code.lower() == "charm code":
                        continue
                    charms[code] = {
                        "code": code,
                        "sku": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                        "default_charm_shop": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                        "notes": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                        "sources": ["library"],
                    }
            wb.close()
        except Exception as exc:
            log.warning("Manifest: catalog read failed: %s", exc)

    img_root = charm_images_dir
    if img_root and img_root.is_dir():
        for st, (rel, fname) in _disk_charm_files_index(img_root).items():
            rec = charms.get(st)
            if rec:
                if "disk" not in rec["sources"]:
                    rec["sources"].append("disk")
                rec["image_file"] = fname
                rec["image_relative"] = rel
            else:
                charms[st] = {
                    "code": st,
                    "sku": "",
                    "default_charm_shop": "",
                    "notes": "",
                    "sources": ["disk"],
                    "image_file": fname,
                    "image_relative": rel,
                }

    manifest: dict = {
        "version": 1,
        "schema": "charm_manifest",
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "charm_images_dir": str(img_root.resolve()) if img_root else None,
        "supplier_catalog": str(catalog_path.resolve()) if catalog_path.exists() else None,
        "convention": (
            "Stable codes: PREFIX + zero-padded digits (e.g. CH-00001). "
            "New sequences use at least five digits when no prior PREFIX+digits codes exist; "
            "otherwise padding matches existing rows/files and widens past 9999 as needed. "
            "Filename = <code>.png|.jpg|.jpeg|.webp; subfolders allowed. "
            "Charm Library column C (SKU) holds the short label or stock code."
        ),
        "charm_codes_range_rows": CHARM_CODES_LIST_MAX_ROW,
        "charms": sorted(charms.values(), key=lambda x: x["code"].lower()),
    }
    if route_snapshot:
        manifest["route_snapshot"] = route_snapshot
    _atomic_write_text(
        output_path,
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    log.info("Wrote charm manifest  ->  %s  (%d charm(s))", output_path, len(charms))
    return len(charms)


def _sanitize_charm_code_for_filename(code: str) -> str:
    """Filename stem for disk assets; strip and replace Windows-forbidden characters."""
    s = code.strip()
    for ch in '<>:"/\\|?*':
        s = s.replace(ch, "_")
    return s


def charm_photo_bytes_from_folder(code: str, root: Path | None) -> bytes | None:
    """
    Load ``<root>/<Charm Code>.<ext>`` (png/jpg/jpeg/webp), then any same-named
    file under *root* subfolders (``rglob``).  Returns None if not found.
    """
    if not root or not root.is_dir():
        return None
    stem = _sanitize_charm_code_for_filename(code)
    if not stem:
        return None

    def _read(p: Path) -> bytes | None:
        try:
            return p.read_bytes()
        except OSError as exc:
            log.warning("Could not read charm image %s: %s", p, exc)
            return None

    for ext in _CHARM_IMAGE_EXTENSIONS:
        direct = root / f"{stem}{ext}"
        if direct.is_file():
            return _read(direct)

    for ext in _CHARM_IMAGE_EXTENSIONS:
        matches = sorted(root.rglob(f"{stem}{ext}"))
        if len(matches) > 1:
            log.warning(
                "Multiple files for charm code %r — using %s",
                code, matches[0],
            )
        if matches:
            data = _read(matches[0])
            if data:
                return data
    return None


def charm_photo_path_for_code(code: str, root: Path | None) -> Path | None:
    """First on-disk path for *code* under *root* (same rules as ``charm_photo_bytes_from_folder``)."""
    if not root or not root.is_dir():
        return None
    stem = _sanitize_charm_code_for_filename(code)
    if not stem:
        return None
    for ext in _CHARM_IMAGE_EXTENSIONS:
        direct = root / f"{stem}{ext}"
        if direct.is_file():
            return direct
    for ext in _CHARM_IMAGE_EXTENSIONS:
        matches = sorted(root.rglob(f"{stem}{ext}"))
        if matches:
            return matches[0]
    return None


def _image_mime_from_bytes(data: bytes) -> str:
    """Best-effort MIME for vision API when extension is unknown (embedded xlsx images)."""
    if len(data) >= 8 and data[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if len(data) >= 2 and data[:2] == b"\xff\xd8":
        return "image/jpeg"
    if len(data) >= 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    return "image/png"


def fill_charm_library_vision_sku(
    catalog_path: Path,
    charm_images_dir: Path,
    *,
    openai_api_key: str,
    openai_model: str,
    openai_base_url: str,
    overwrite: bool = False,
    dry_run: bool = False,
) -> tuple[int, list[str]]:
    """
    For each **Charm Library** row with a code in column B, if **SKU** (C)
    is empty (unless *overwrite*), set C from the vision API using the image on disk
    (``data/charm_images/<Code>.png`` etc.) when present, otherwise the embedded photo
    for that row.

    Returns ``(n_updated, log_lines)``.
    """
    lines: list[str] = []
    oa_key = openai_api_key.strip()
    oa_base = openai_base_url.strip() or _OPENAI_DEFAULT_BASE
    if (
        not oa_key
        and not dry_run
        and not _charm_vision_base_allows_empty_key(oa_base)
    ):
        lines.append(
            "[error] Vision API key missing — set CHARM_VISION_API_KEY or OPENAI_API_KEY, "
            "or use a local OpenAI-compatible base URL (Ollama / LM Studio) that needs no token."
        )
        return 0, lines
    if not catalog_path.exists():
        lines.append(f"[error] Catalog not found: {catalog_path}")
        return 0, lines

    try:
        row_photos = extract_photos_from_xlsx(
            catalog_path, sheet_name=CHARM_LIBRARY_SHEET, photo_col_idx=0
        )
    except Exception as exc:
        log.warning("Charm Library embed extraction: %s", exc)
        row_photos = {}

    wb = openpyxl.load_workbook(catalog_path)
    updated = 0
    try:
        if CHARM_LIBRARY_SHEET not in wb.sheetnames:
            lines.append(
                f"[error] Sheet {CHARM_LIBRARY_SHEET!r} missing — run --init-charm-shops first."
            )
            return 0, lines
        ws_lib = wb[CHARM_LIBRARY_SHEET]
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        for r in range(2, ws_lib.max_row + 1):
            if _charm_library_instruction_row(ws_lib, r):
                continue
            code = str(ws_lib.cell(r, 2).value or "").strip()
            if not code or code.casefold() == "charm code":
                continue
            cur = str(ws_lib.cell(r, 3).value or "").strip()
            if cur and not overwrite:
                continue

            disk_path = charm_photo_path_for_code(code, charm_images_dir)
            if disk_path and disk_path.is_file():
                try:
                    data = disk_path.read_bytes()
                except OSError as exc:
                    lines.append(f"[skip] {code}: could not read {disk_path}: {exc}")
                    continue
                mime = _charm_image_mime_type(disk_path.suffix) or _image_mime_from_bytes(data)
            else:
                data = row_photos.get(r)
                if not data:
                    lines.append(
                        f"[skip] {code}: no image on disk under {charm_images_dir.name!r} "
                        f"and no embed in row {r}"
                    )
                    continue
                mime = _image_mime_from_bytes(data)

            if dry_run:
                lines.append(f"  [dry-run] would vision-fill C for {code} (row {r})")
                updated += 1
                continue

            sku_text = openai_vision_charm_sku(
                data,
                mime,
                api_key=oa_key,
                model=openai_model,
                base_url=oa_base,
            )
            time.sleep(_CHARM_VISION_COOLDOWN_SEC)
            if not sku_text:
                lines.append(f"  [warn] vision returned nothing for {code}")
                continue

            _cat_cell(ws_lib, r, 3, sku_text, _CAT_WARN_FILL, _CAT_BODY, _CAT_CENTER)
            old_note = ws_lib.cell(r, 5).value
            tag = (
                f"AI suggested SKU ({openai_model}), {stamp}.\n\n"
                "Verify column C."
            )
            if isinstance(old_note, str) and old_note.strip():
                ws_lib.cell(r, 5, f"{old_note.strip()}\n\n—\n\n{tag}")
            else:
                ws_lib.cell(r, 5, tag)
            lines.append(f"  {code}  ->  {sku_text!r}")
            updated += 1

        if updated and not dry_run:
            _ensure_charm_library_sheet_layout(ws_lib)
            _ensure_charm_codes_named_range(wb)
            if CATALOG_SHEET in wb.sheetnames:
                _refresh_all_product_map_validations(wb, wb[CATALOG_SHEET])
            _apply_charm_library_default_shop_validation(wb, ws_lib)
            set_supplier_catalog_active_to_product_map(wb)
            backup_supplier_catalog_before_write(catalog_path, "fill_charm_sku_vision")
            wb.save(catalog_path)
            lines.append(f"[ok] Updated SKU for {updated} row(s) in {catalog_path.name}.")
        elif dry_run and updated:
            lines.append(f"[dry-run] Would update {updated} row(s); catalog not saved.")
        elif not updated:
            lines.append("[ok] No rows needed updates (empty column C only, unless --overwrite).")

    finally:
        wb.close()

    return updated, lines


def _resolve_charm_photo_bytes(
    r: ResolvedItem,
    charm_library: dict[str, CharmLibraryEntry] | None,
    charm_images_dir: Path | None = None,
) -> bytes | None:
    """Resolve charm photo: folder first, then Charm Library embed, else None."""
    if not r.supplier:
        return None
    code = (r.supplier.charm_code or "").strip()
    if not code:
        return None
    from_disk = charm_photo_bytes_from_folder(code, charm_images_dir)
    if from_disk:
        return from_disk
    if charm_library:
        ent = charm_library.get(code)
        if ent and ent.photo_bytes:
            return ent.photo_bytes
    return None


# ---------------------------------------------------------------------------
# Excel generation -- shared styles
# ---------------------------------------------------------------------------

_THIN   = Side(style="thin", color="C0C0C0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT   = Font("Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font("Calibri", bold=True, size=16, color="1F4E79")
_SUB_FONT   = Font("Calibri", size=11, color="555555", italic=True)
_BODY       = Font("Calibri", size=10)
_BODY_BOLD  = Font("Calibri", bold=True, size=10)
_WARN_FILL  = PatternFill("solid", fgColor="FFF3CD")
_WARN_FONT  = Font("Calibri", size=10, color="856404")
_SEC_FONT   = Font("Calibri", bold=True, size=12, color="1F4E79")

_GROUP_FILLS = [
    PatternFill("solid", fgColor="EBF2FA"),
    PatternFill("solid", fgColor="FFFFFF"),
]

# Styling for N/A component cells (component not ordered)
_NA_FILL = PatternFill("solid", fgColor="EFEFEF")
_NA_FONT = Font("Calibri", size=9, color="AAAAAA", italic=True)

# Styling for "in catalog but supplier info not yet filled" rows
_NEEDSINFO_FILL = PatternFill("solid", fgColor="D9EAF7")   # pale blue
_NEEDSINFO_FONT = Font("Calibri", size=10, color="1F4E79")  # dark navy

# ---------------------------------------------------------------------------
# Charm section styles (Shopping Route -- separate building block)
# ---------------------------------------------------------------------------
_CHARM_BANNER_FILL  = PatternFill("solid", fgColor="3D1359")  # deep purple  – section banner
_CHARM_SHOPS_FILL   = PatternFill("solid", fgColor="EFD9FC")  # light lavender – shop-list row
_CHARM_HDR_FILL     = PatternFill("solid", fgColor="5B1A6B")  # mid purple   – sub-header
_CHARM_GROUP_FILLS  = [
    PatternFill("solid", fgColor="F8F0FD"),   # very light lavender (odd rows)
    PatternFill("solid", fgColor="FFFFFF"),   # white                (even rows)
]
_CHARM_NA_HDR_FONT  = Font("Calibri", bold=True, color="CCAACC", size=11)  # muted for N/A hdr cells

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP   = Alignment(vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# Catalog-specific styles  (supplier_catalog.xlsx)
# The catalog uses Calibri 11 / dark-navy 2C3E50 header, which differs from
# the shopping-route sheet (Calibri 10 / blue 1F4E79).
# ---------------------------------------------------------------------------
_CAT_HDR_FILL   = PatternFill("solid", fgColor="2C3E50")
_CAT_HDR_FONT   = Font("Calibri", bold=True, color="FFFFFF", size=12)
_CAT_BODY       = Font("Calibri", size=11)
_CAT_BODY_BOLD  = Font("Calibri", bold=True, size=11)
_CAT_WARN_FILL  = PatternFill("solid", fgColor="FFF3CD")   # amber  – needs attention
_CAT_WARN_FONT  = Font("Calibri", size=10, italic=True, color="7D4E00")
_CAT_PRICE_FILL = PatternFill("solid", fgColor="FFF9E6")   # yellow – price TBD
_CAT_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CAT_WRAP       = Alignment(vertical="center", wrap_text=True)


def _style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER


def _style_row(ws, row: int, cols: int, *, fill=None, font=None) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font      = font or _BODY
        cell.alignment = _WRAP
        cell.border    = _BORDER
        if fill:
            cell.fill = fill


def _embed_photo(ws, photo_bytes: bytes | None, row: int, col: int,
                 photo_px: int = PHOTO_PX) -> None:
    """Embed a JPEG thumbnail anchored to (row, col) of ws."""
    if not photo_bytes:
        return
    try:
        xl_img = XLImage(BytesIO(photo_bytes))
        xl_img.width  = photo_px
        xl_img.height = photo_px
        xl_img.anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(xl_img)
    except Exception as e:
        log.warning("Photo embed failed at %s%d: %s", get_column_letter(col), row, e)


def _cat_cell(ws, row: int, col: int, value, fill, font, alignment) -> None:
    """Write one styled cell in the supplier catalog sheet."""
    cell           = ws.cell(row, col, value)
    cell.fill      = fill
    cell.font      = font
    cell.alignment = alignment
    cell.border    = _BORDER


def _numeric_suffixes_for_prefix(prefix: str, strings: Iterable[str]) -> list[int]:
    """
    Return integers *N* for strings matching ``<prefix><N>`` (case-insensitive).
    *prefix* should end with ``-`` (e.g. ``CH-`` → matches ``CH-00001``).
    """
    pfx = prefix if prefix.endswith("-") else prefix + "-"
    pat = re.compile(rf"^{re.escape(pfx)}(\d+)$", re.IGNORECASE)
    out: list[int] = []
    for s in strings:
        m = pat.match(str(s).strip())
        if m:
            out.append(int(m.group(1)))
    return out


def _max_numeric_digit_width_for_prefix(prefix: str, strings: Iterable[str]) -> int:
    """
    Longest digit run among strings matching ``PREFIX`` + digits only.
    Used so new codes keep the same visual width as existing library/disk names.
    """
    pfx = prefix if prefix.endswith("-") else prefix + "-"
    pat = re.compile(rf"^{re.escape(pfx)}(\d+)$", re.IGNORECASE)
    w = 0
    for s in strings:
        m = pat.match(str(s).strip())
        if m:
            w = max(w, len(m.group(1)))
    return w


def _resolve_charm_code_numeric_width(
    prefix: str,
    lib_codes: Iterable[str],
    disk_stems: Iterable[str],
    batch_last_n: int,
) -> int:
    """
    Zero-pad width for allocating ``PREFIX`` + numeric codes.

    • Fresh prefix (no matching rows/files): at least ``CHARM_CODE_NUMERIC_MIN_WIDTH``.
    • Existing PREFIX+digits: at least the widest digit run already in use.
    • Always wide enough for *batch_last_n* (and Python ``format`` widens further if needed).
    """
    pfx = _normalise_charm_import_prefix(prefix)
    stems = set(lib_codes) | set(disk_stems)
    existing_w = _max_numeric_digit_width_for_prefix(pfx, stems)
    need = len(str(max(1, batch_last_n)))
    if existing_w == 0:
        return max(CHARM_CODE_NUMERIC_MIN_WIDTH, need)
    return max(existing_w, need)


def _normalise_charm_import_prefix(prefix: str) -> str:
    p = (prefix or "CH-").strip() or "CH-"
    if not p.endswith("-"):
        p = p + "-"
    return p


def _charm_image_mime_type(ext: str) -> str | None:
    ext = ext.lower()
    return {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }.get(ext)


def _clean_vision_sku_text(text: str) -> str:
    t = (text or "").strip()
    t = re.sub(r"\s+", " ", t)
    t = t.strip("\"'«»“”")
    return t[:200] if len(t) > 200 else t


def openai_vision_charm_sku(
    image_bytes: bytes,
    media_type: str,
    *,
    api_key: str,
    model: str,
    base_url: str,
    timeout: float = 90.0,
) -> str | None:
    """
    One short catalog **SKU** / label from an image via OpenAI-compatible Chat Completions
    (vision ``image_url``). Returns cleaned text or ``None`` on failure.  Uses stdlib HTTP only.
    *api_key* may be empty for some local servers (Ollama, LM Studio).
    """
    url = base_url.rstrip("/") + "/chat/completions"
    b64 = base64.standard_b64encode(image_bytes).decode("ascii")
    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            "You assign short catalog SKUs / labels for bead and phone-charm strap products. "
                            "Reply with exactly one short phrase (max 12 words): colours, main beads or "
                            "characters, and style (e.g. pastel, pearl, y2k). No quotation marks, no markdown, "
                            "no leading 'This is' or similar — label text only."
                        ),
                    },
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:{media_type};base64,{b64}"},
                    },
                ],
            }
        ],
        "max_tokens": 100,
    }
    body = json.dumps(payload).encode("utf-8")
    headers: dict[str, str] = {"Content-Type": "application/json"}
    if (api_key or "").strip():
        headers["Authorization"] = f"Bearer {api_key.strip()}"
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        err_body = ""
        try:
            err_body = exc.read().decode("utf-8", errors="replace")[:500]
        except Exception:
            pass
        log.warning("Charm vision API HTTP %s: %s", exc.code, err_body or exc.reason)
        return None
    except (urllib.error.URLError, OSError) as exc:
        log.warning("Charm vision API request failed: %s", exc)
        return None
    try:
        data = json.loads(raw)
        text = data["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        log.warning("Charm vision API parse error: %s", exc)
        return None
    return _clean_vision_sku_text(str(text)) or None


def import_charm_screenshot_assets(
    images_dir: Path,
    catalog_path: Path,
    *,
    patterns: list[str] | None = None,
    dry_run: bool = False,
    prefix: str = "CH-",
    vision_sku: bool = False,
    openai_api_key: str | None = None,
    openai_model: str = "gpt-4o-mini",
    openai_base_url: str | None = None,
) -> tuple[int, list[str]]:
    """
    Find files under *images_dir* matching any glob in *patterns*, assign sequential codes
    ``<prefix>`` + zero-padded digits (width from existing library/disk and
    ``CHARM_CODE_NUMERIC_MIN_WIDTH`` for new sequences), rename on disk, append
    **Charm Library** rows with embedded photos.

    With *vision_sku* and credentials (API key and/or local OpenAI-compatible base URL),
    **SKU** (C) is filled via Chat Completions vision (review results).  Dry-run never
    calls the API.

    Returns ``(n_imported, log_lines)``.
    """
    lines: list[str] = []
    prefix_n = _normalise_charm_import_prefix(prefix)
    oa_base = (openai_base_url or _OPENAI_DEFAULT_BASE).strip()
    oa_key = (openai_api_key or "").strip()

    pat_list = [p.strip() for p in (patterns or []) if p and str(p).strip()]
    if not pat_list:
        pat_list = ["Screenshot*.png"]

    if not images_dir.is_dir():
        lines.append(f"[skip] Charm images folder not found: {images_dir}")
        return 0, lines

    seen_paths: set[Path] = set()
    candidates: list[Path] = []
    for p in images_dir.rglob("*"):
        if not p.is_file():
            continue
        if not any(fnmatch.fnmatch(p.name, pat) for pat in pat_list):
            continue
        key = p.resolve()
        if key in seen_paths:
            continue
        seen_paths.add(key)
        candidates.append(p)
    candidates.sort(key=lambda x: str(x).casefold())
    if not candidates:
        lines.append(
            f"[ok] No files match {pat_list!r} under {images_dir} — nothing to import."
        )
        return 0, lines

    if vision_sku and dry_run:
        lines.append(
            "[info] SKU (vision API): skipped in dry-run (no API calls)."
        )

    vision_can_call = bool(oa_key) or _charm_vision_base_allows_empty_key(oa_base)

    lib_codes: set[str] = set()
    lib_lower: set[str] = set()
    if catalog_path.exists():
        try:
            for k in load_charm_library(catalog_path).keys():
                lib_codes.add(str(k).strip())
                lib_lower.add(str(k).strip().casefold())
        except Exception as exc:
            lines.append(f"[warn] Could not read existing charm library: {exc}")

    disk_idx = _disk_charm_files_index(images_dir)
    disk_lower = {k.casefold() for k in disk_idx}
    nums = _numeric_suffixes_for_prefix(prefix_n, lib_codes)
    nums.extend(_numeric_suffixes_for_prefix(prefix_n, disk_idx.keys()))
    next_n = max(nums, default=0) + 1
    batch_last_n = next_n + len(candidates) - 1
    width = _resolve_charm_code_numeric_width(
        prefix_n, lib_codes, disk_idx.keys(), batch_last_n
    )
    lines.append(
        f"[info] Charm codes: zero-pad width={width} for {prefix_n}… "
        f"(next index {next_n}, batch up to {batch_last_n})"
    )

    def _alloc_code(ext: str, parent: Path) -> str | None:
        """Next unused code: not in library, no on-disk stem, no file at *parent*."""
        nonlocal next_n
        for _ in range(500_000):
            code = f"{prefix_n}{next_n:0{width}d}"
            next_n += 1
            if code.casefold() in lib_lower:
                continue
            if code.casefold() in disk_lower:
                continue
            dest_try = parent / f"{code}{ext}"
            if dest_try.exists():
                continue
            return code
        return None

    if not catalog_path.exists():
        lines.append(f"[error] Catalog not found: {catalog_path}")
        return 0, lines

    wb = openpyxl.load_workbook(catalog_path)
    imported = 0
    try:
        if CHARM_LIBRARY_SHEET not in wb.sheetnames:
            lines.append(
                f"[error] Sheet {CHARM_LIBRARY_SHEET!r} missing — run --init-charm-shops first."
            )
            return 0, lines
        ws_lib = wb[CHARM_LIBRARY_SHEET]

        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        for src in candidates:
            ext = src.suffix.lower()
            if ext not in _CHARM_IMAGE_EXTENSIONS:
                lines.append(
                    f"[skip] {src} — extension not in {_CHARM_IMAGE_EXTENSIONS}"
                )
                continue
            parent = src.parent
            code = _alloc_code(ext, parent)
            if code is None:
                lines.append(f"[error] Could not allocate a charm code for {src}")
                continue
            dest = parent / f"{code}{ext}"
            if dest.resolve() == src.resolve():
                lines.append(f"[skip] {src.name} — already named {code}")
                continue

            rel = src.relative_to(images_dir).as_posix()
            lines.append(f"  {rel}  ->  {dest.relative_to(images_dir).as_posix()}  ({code})")

            if dry_run:
                imported += 1
                lib_codes.add(code)
                lib_lower.add(code.casefold())
                disk_idx[code] = (dest.relative_to(images_dir).as_posix(), dest.name)
                disk_lower.add(code.casefold())
                continue

            src.rename(dest)
            lib_codes.add(code)
            lib_lower.add(code.casefold())
            rel_d = dest.relative_to(images_dir).as_posix()
            disk_idx[code] = (rel_d, dest.name)
            disk_lower.add(code.casefold())

            data = dest.read_bytes()
            sku_text: str | None = None
            mime = _charm_image_mime_type(ext)
            if vision_sku and vision_can_call and mime:
                sku_text = openai_vision_charm_sku(
                    data,
                    mime,
                    api_key=oa_key,
                    model=openai_model,
                    base_url=oa_base,
                )
                if sku_text:
                    lines.append(f"    sku: {sku_text}")
                else:
                    lines.append(f"    [warn] vision returned no SKU for {code}")
                time.sleep(_CHARM_VISION_COOLDOWN_SEC)

            row = ws_lib.max_row + 1
            ws_lib.cell(row, 1).border = _BORDER
            ws_lib.cell(row, 1).fill = _CAT_WARN_FILL
            _cat_cell(ws_lib, row, 2, code, _CAT_WARN_FILL, _CAT_BODY_BOLD, _CAT_CENTER)
            _cat_cell(ws_lib, row, 3, sku_text, _CAT_WARN_FILL, _CAT_BODY, _CAT_CENTER)
            _cat_cell(ws_lib, row, 4, None, _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)
            if vision_sku and vision_can_call:
                if sku_text:
                    note = (
                        f"Imported {stamp}.\n\n"
                        f"• SKU from AI ({openai_model}) — verify column C.\n\n"
                        f"• Optional: Default shop (D)."
                    )
                else:
                    note = (
                        f"Imported {stamp}.\n\n"
                        f"• Fill SKU (C).\n\n"
                        f"• Optional: Default shop (D)."
                    )
            else:
                note = (
                    f"Imported {stamp}.\n\n"
                    f"• Fill SKU (C).\n\n"
                    f"• Optional: Default shop (D)."
                )
            _cat_cell(ws_lib, row, 5, note, _CAT_WARN_FILL, _CAT_WARN_FONT, _CAT_WRAP)
            ws_lib.row_dimensions[row].height = CHARM_LIB_ROW_HEIGHT
            try:
                xl_img = XLImage(BytesIO(data))
                tw, th = _charm_library_photo_fill_pixels(ws_lib, row)
                xl_img.anchor = f"A{row}"
                _sync_openpyxl_image_display_ext(xl_img, tw, th)
                ws_lib.add_image(xl_img)
            except Exception as exc:
                log.warning("Charm import: could not embed %s: %s", dest, exc)

            imported += 1

        if not dry_run and imported:
            _ensure_charm_library_sheet_layout(ws_lib)
            _ensure_charm_codes_named_range(wb)
            if CATALOG_SHEET in wb.sheetnames:
                _refresh_all_product_map_validations(wb, wb[CATALOG_SHEET])
            _apply_charm_library_default_shop_validation(wb, ws_lib)
            set_supplier_catalog_active_to_product_map(wb)
            backup_supplier_catalog_before_write(catalog_path, "import_charm_images")
            wb.save(catalog_path)
            lines.append(
                f"[ok] Saved {imported} row(s) to {catalog_path.name} (Charm Library)."
            )
        elif dry_run and imported:
            lines.append(
                f"[dry-run] Would import {imported} file(s); catalog not modified."
            )
    finally:
        wb.close()

    return imported, lines


# ---------------------------------------------------------------------------
# Charm Library -- shared helper: apply rename_map to any sheet column G
# ---------------------------------------------------------------------------


def _apply_charm_rename_to_sheet(ws, rename_map: dict[str, str]) -> int:
    """
    Scan *ws* column G (Charm Code, index ``_CHARM_CODE_COL_IDX``) and replace
    every value that appears as a key in *rename_map* with the corresponding
    new code.

    Comparison is **case-insensitive** so that codes typed manually (or with
    inconsistent capitalisation) are still matched and updated correctly.
    Only cells whose value actually changes are written.

    Returns the number of cells updated.
    """
    if not rename_map:
        return 0
    lower_map: dict[str, str] = {k.casefold(): v for k, v in rename_map.items()}
    updated = 0
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, _CHARM_CODE_COL_IDX)
        if cell.value is None:
            continue
        raw = str(cell.value).strip()
        if not raw:
            continue
        new_val = lower_map.get(raw.casefold())
        if new_val is not None and raw != new_val:
            cell.value = new_val
            updated += 1
    return updated


# ---------------------------------------------------------------------------
# Charm Library -- reorder rows then renumber (visual drag-drop result)
# ---------------------------------------------------------------------------


def reorder_charm_library_rows(
    catalog_path: Path,
    new_code_order: list[str],
    *,
    charm_images_dir: Path | None = None,
    dry_run: bool = False,
) -> tuple[int, list[str]]:
    """
    Rewrite Charm Library rows to match *new_code_order*, then renumber.

    *new_code_order* is the desired sequence of **existing** charm codes as
    arranged by the user in the reorder dialog.  Any codes present in the
    library but absent from *new_code_order* are appended at the end in their
    original relative order.  Unknown codes are silently ignored.

    Steps
    -----
    1. Load all charm data (code, SKU, shop, notes, photo bytes) via
       ``load_charm_library``.
    2. Open workbook, locate all current data rows, strip them (cell values +
       embedded images) in one pass.
    3. Re-write data rows in *new_code_order* starting at the original
       ``first_data_row``, re-embedding photos.
    4. Assign new sequential codes (CH-00001, CH-00002 …) matching position.
    5. Update Product Map column G references for any codes that changed.
    6. Rename on-disk charm images (two-pass, collision-safe).
    7. Refresh named ranges, validations, save.

    Returns ``(n_reordered, log_lines)``.
    With ``dry_run=True`` nothing is written — only the planned mapping is logged.
    """
    lines: list[str] = []

    if not catalog_path.exists():
        lines.append(f"[error] Catalog not found: {catalog_path}")
        return 0, lines

    # ---- 1. Load all charm data (including embedded photo bytes) ----
    try:
        entries: dict[str, CharmLibraryEntry] = load_charm_library(catalog_path)
    except Exception as exc:
        lines.append(f"[error] Could not read Charm Library: {exc}")
        return 0, lines

    if not entries:
        lines.append("[ok] Charm Library is empty — nothing to reorder.")
        return 0, lines

    # ---- 2. Build final order: honour new_code_order, append any remainder ----
    known      = set(entries.keys())
    seen: set[str] = set()
    final_order: list[str] = []
    for code in new_code_order:
        if code in known and code not in seen:
            final_order.append(code)
            seen.add(code)
    for code in entries:          # dict preserves insertion (original) order
        if code not in seen:
            final_order.append(code)
            seen.add(code)

    n_total = len(final_order)

    # ---- 3. Determine new codes (sequential, same prefix/width as existing) ----
    all_old_codes = list(entries.keys())
    pfx = _normalise_charm_import_prefix("CH-")
    for code in all_old_codes:
        m = re.match(r"^([A-Za-z]+-+)(\d+)$", code.strip())
        if m:
            pfx = m.group(1).upper()
            break
    existing_w = _max_numeric_digit_width_for_prefix(pfx, all_old_codes)
    need       = len(str(n_total))
    width      = max(existing_w or 0, need, CHARM_CODE_NUMERIC_MIN_WIDTH)

    rename_map: dict[str, str] = {}   # old_code -> new_code
    for seq, code in enumerate(final_order, start=1):
        new_code = f"{pfx}{seq:0{width}d}"
        if code != new_code:
            rename_map[code] = new_code

    # ---- 4. Log / dry-run ----
    lines.append(
        f"[info] {n_total} charm(s) in new order.  "
        f"Prefix: {pfx!r}, width: {width}."
    )
    if rename_map:
        lines.append(f"[info] {len(rename_map)} code(s) change:")
        for old, new in rename_map.items():
            lines.append(f"    {old}  \u2192  {new}")
    else:
        lines.append("[info] No code changes needed (order already sequential).")

    if dry_run:
        lines.append("[dry-run] No changes written.")
        return n_total, lines

    backup_supplier_catalog_before_write(catalog_path, "reorder_charm_library")
    # ---- 5. Open workbook, strip current data rows ----
    wb = openpyxl.load_workbook(catalog_path)
    try:
        if CHARM_LIBRARY_SHEET not in wb.sheetnames:
            lines.append(
                f"[error] Sheet {CHARM_LIBRARY_SHEET!r} missing — run --init-charm-shops."
            )
            return 0, lines

        ws_lib = wb[CHARM_LIBRARY_SHEET]

        # Locate all current charm data rows
        current_rows: list[tuple[int, str]] = []
        for r in range(2, ws_lib.max_row + 1):
            if _charm_library_instruction_row(ws_lib, r):
                continue
            b_val = ws_lib.cell(r, 2).value
            if not b_val:
                continue
            code = str(b_val).strip()
            if not code or code.casefold() == "charm code":
                continue
            current_rows.append((r, code))

        if not current_rows:
            lines.append("[ok] No charm data rows found — nothing to reorder.")
            return 0, lines

        first_data_row = current_rows[0][0]
        last_data_row  = current_rows[-1][0]

        # Remove all images anchored to column A in data rows
        data_row_set = {r for r, _ in current_rows}
        ws_lib._images = [
            img for img in list(getattr(ws_lib, "_images", []) or [])
            if not (
                _charm_library_image_anchor_col_a(img)
                and _anchor_row(img) in data_row_set
            )
        ]

        # Clear cell values in every row that was or will be a data row
        clear_end = max(last_data_row, first_data_row + n_total - 1)
        for r in range(first_data_row, clear_end + 1):
            for c in range(1, 6):
                ws_lib.cell(r, c).value = None

        # ---- 6. Write rows in new order ----
        for seq, code in enumerate(final_order, start=1):
            r      = first_data_row + seq - 1
            entry  = entries[code]
            nc     = rename_map.get(code, code)   # new code for this row

            # Column A — photo cell (styling only; image added below)
            ws_lib.cell(r, 1).border = _BORDER
            ws_lib.cell(r, 1).fill   = _CAT_WARN_FILL

            # Column B — Charm Code
            _cat_cell(ws_lib, r, 2, nc,
                      _CAT_WARN_FILL, _CAT_BODY_BOLD, _CAT_CENTER)

            # Column C — SKU
            _cat_cell(ws_lib, r, 3, entry.sku or None,
                      _CAT_WARN_FILL, _CAT_BODY, _CAT_CENTER)

            # Column D — Default Charm Shop
            _cat_cell(ws_lib, r, 4, entry.default_charm_shop or None,
                      _CAT_CHARM_PENDING_FILL, _CAT_BODY, _CAT_CENTER)

            # Column E — Notes
            _cat_cell(ws_lib, r, 5, entry.notes or None,
                      _CAT_WARN_FILL, _CAT_WARN_FONT, _CAT_WRAP)

            ws_lib.row_dimensions[r].height = CHARM_LIB_ROW_HEIGHT

            # Embed photo
            if entry.photo_bytes:
                try:
                    xl_img = XLImage(BytesIO(entry.photo_bytes))
                    tw, th = _charm_library_photo_fill_pixels(ws_lib, r)
                    xl_img.anchor = f"A{r}"
                    _sync_openpyxl_image_display_ext(xl_img, tw, th)
                    ws_lib.add_image(xl_img)
                except Exception as exc:
                    lines.append(
                        f"[warn] Could not embed photo for {code}: {exc}"
                    )

        # ---- 7. Update Product Map column G + Discontinued Products column G ----
        pm_updated   = 0
        disc_updated = 0
        if rename_map:
            if CATALOG_SHEET in wb.sheetnames:
                pm_updated = _apply_charm_rename_to_sheet(
                    wb[CATALOG_SHEET], rename_map
                )
            if DISCONTINUED_SHEET in wb.sheetnames:
                disc_updated = _apply_charm_rename_to_sheet(
                    wb[DISCONTINUED_SHEET], rename_map
                )
        lines.append(
            f"[ok] Updated {pm_updated} Product Map + "
            f"{disc_updated} Discontinued Products Charm Code reference(s)."
        )

        # ---- 8. Housekeeping + save ----
        _ensure_charm_codes_named_range(wb)
        _ensure_charm_library_sheet_layout(ws_lib)
        if CATALOG_SHEET in wb.sheetnames:
            _refresh_all_product_map_validations(wb, wb[CATALOG_SHEET])
        _apply_charm_library_default_shop_validation(wb, ws_lib)
        set_supplier_catalog_active_to_product_map(wb)
        wb.save(catalog_path)
        lines.append(f"[ok] Saved workbook: {catalog_path.name}")

    finally:
        wb.close()

    # ---- 9. Rename disk files (two-pass, collision-safe) ----
    if rename_map and charm_images_dir and charm_images_dir.is_dir():
        temp_to_final: dict[Path, Path] = {}
        for old_code, new_code in rename_map.items():
            for ext in _CHARM_IMAGE_EXTENSIONS:
                old_path = charm_images_dir / f"{old_code}{ext}"
                if not old_path.exists():
                    continue
                token    = secrets.token_hex(4)
                tmp_path = old_path.with_name(
                    f"__reorder_tmp_{token}_{new_code}{ext}"
                )
                try:
                    old_path.rename(tmp_path)
                    temp_to_final[tmp_path] = charm_images_dir / f"{new_code}{ext}"
                    lines.append(f"[disk] {old_code}{ext}  \u2192  {new_code}{ext}")
                except OSError as exc:
                    lines.append(
                        f"[warn] Could not stage rename for {old_path.name}: {exc}"
                    )
        for tmp_path, final_path in temp_to_final.items():
            try:
                tmp_path.rename(final_path)
            except OSError as exc:
                lines.append(
                    f"[warn] Rename {tmp_path.name} \u2192 {final_path.name}: {exc}"
                )

    return n_total, lines


# ---------------------------------------------------------------------------
# Charm Library -- renumber (reassign codes to match current row order)
# ---------------------------------------------------------------------------


def renumber_charm_library(
    catalog_path: Path,
    *,
    charm_images_dir: Path | None = None,
    prefix: str = "CH-",
    dry_run: bool = False,
) -> tuple[int, list[str]]:
    """
    Renumber all Charm Library codes so they match the current top-to-bottom
    row order of the sheet.

    Typical workflow
    ----------------
    1. Import new charm photos (appended at end, e.g. CH-00006, CH-00007).
    2. Open ``supplier_catalog.xlsx`` → Charm Library, cut/paste rows to put
       similar charms next to each other.
    3. Run ``--renumber-charms`` — codes are reassigned CH-00001, CH-00002 …
       in the new row order, and every Product Map Charm Code reference is
       updated automatically.

    What this function does
    -----------------------
    * Reads rows of the Charm Library sheet in their physical order.
    * Builds an ``old_code → new_code`` mapping (purely sequential).
    * The prefix and zero-pad width are inferred from existing codes
      (falls back to *prefix* and ``CHARM_CODE_NUMERIC_MIN_WIDTH``).
    * Writes the new codes back to Charm Library column B.
    * Updates matching cells in Product Map column G (Charm Code foreign key).
    * Renames ``<old>.ext → <new>.ext`` files in *charm_images_dir* using a
      two-pass strategy (old → temp → new) to avoid collisions.
    * Saves the workbook.

    Returns ``(n_renumbered, log_lines)``.
    With ``dry_run=True`` nothing is written; the mapping is just printed.
    """
    lines: list[str] = []

    if not catalog_path.exists():
        lines.append(f"[error] Catalog not found: {catalog_path}")
        return 0, lines

    wb = openpyxl.load_workbook(catalog_path)
    try:
        if CHARM_LIBRARY_SHEET not in wb.sheetnames:
            lines.append(
                f"[error] Sheet {CHARM_LIBRARY_SHEET!r} missing — run --init-charm-shops first."
            )
            return 0, lines

        ws_lib = wb[CHARM_LIBRARY_SHEET]

        # ---- 1. Collect (row_number, old_code) in sheet order ----
        ordered: list[tuple[int, str]] = []
        for r in range(2, ws_lib.max_row + 1):
            if _charm_library_instruction_row(ws_lib, r):
                continue
            b_val = ws_lib.cell(r, 2).value
            if not b_val:
                continue
            code = str(b_val).strip()
            if not code or code.casefold() == "charm code":
                continue
            ordered.append((r, code))

        if not ordered:
            lines.append("[ok] Charm Library is empty — nothing to renumber.")
            return 0, lines

        # ---- 2. Determine prefix and zero-pad width ----
        all_old_codes = [c for _, c in ordered]

        # Infer prefix from the first code that looks like PREFIX-NNNNN.
        pfx = _normalise_charm_import_prefix(prefix)
        for code in all_old_codes:
            m = re.match(r"^([A-Za-z]+-+)(\d+)$", code.strip())
            if m:
                pfx = m.group(1).upper()
                break

        existing_w = _max_numeric_digit_width_for_prefix(pfx, all_old_codes)
        n_total    = len(ordered)
        need       = len(str(n_total))
        width      = max(existing_w or 0, need, CHARM_CODE_NUMERIC_MIN_WIDTH)

        lines.append(
            f"[info] {n_total} charm(s) found.  Prefix: {pfx!r},  "
            f"zero-pad width: {width}."
        )

        # ---- 3. Build old → new mapping (skip codes already correct) ----
        rename_map: dict[str, str] = {}   # old_code -> new_code
        for seq, (_, old_code) in enumerate(ordered, start=1):
            new_code = f"{pfx}{seq:0{width}d}"
            if old_code != new_code:
                rename_map[old_code] = new_code

        if not rename_map:
            lines.append(
                "[ok] Charm codes are already in sequential order — no changes needed."
            )
            return 0, lines

        lines.append(f"[info] {len(rename_map)} code(s) will be renumbered:")
        for old, new in rename_map.items():
            lines.append(f"    {old}  →  {new}")

        if dry_run:
            lines.append("[dry-run] No changes written.")
            return len(rename_map), lines

        backup_supplier_catalog_before_write(catalog_path, "renumber_charms")
        # ---- 4. Rewrite Charm Library column B ----
        for r, old_code in ordered:
            if old_code in rename_map:
                cell = ws_lib.cell(r, 2)
                cell.value = rename_map[old_code]

        # ---- 5. Update Product Map column G + Discontinued Products column G ----
        pm_updated   = _apply_charm_rename_to_sheet(
            wb[CATALOG_SHEET], rename_map
        ) if CATALOG_SHEET in wb.sheetnames else 0
        disc_updated = _apply_charm_rename_to_sheet(
            wb[DISCONTINUED_SHEET], rename_map
        ) if DISCONTINUED_SHEET in wb.sheetnames else 0
        lines.append(
            f"[ok] Updated {pm_updated} Product Map + "
            f"{disc_updated} Discontinued Products Charm Code reference(s)."
        )

        # ---- 6. Housekeeping and save ----
        _ensure_charm_codes_named_range(wb)
        if CATALOG_SHEET in wb.sheetnames:
            _refresh_all_product_map_validations(wb, wb[CATALOG_SHEET])
        set_supplier_catalog_active_to_product_map(wb)
        wb.save(catalog_path)
        lines.append(f"[ok] Saved workbook: {catalog_path.name}")

    finally:
        wb.close()

    # ---- 7. Rename disk files (two-pass to avoid collisions) ----
    if charm_images_dir and charm_images_dir.is_dir():
        # Pass 1 — rename old_code.ext → __renumber_tmp_XXXX_new_code.ext
        temp_to_final: dict[Path, Path] = {}
        for old_code, new_code in rename_map.items():
            for ext in _CHARM_IMAGE_EXTENSIONS:
                old_path = charm_images_dir / f"{old_code}{ext}"
                if not old_path.exists():
                    continue
                token    = secrets.token_hex(4)
                tmp_path = old_path.with_name(
                    f"__renumber_tmp_{token}_{new_code}{ext}"
                )
                try:
                    old_path.rename(tmp_path)
                    temp_to_final[tmp_path] = charm_images_dir / f"{new_code}{ext}"
                    lines.append(
                        f"[disk] {old_code}{ext}  →  {new_code}{ext}"
                    )
                except OSError as exc:
                    lines.append(
                        f"[warn] Could not stage rename for {old_path.name}: {exc}"
                    )

        # Pass 2 — rename temp → final
        for tmp_path, final_path in temp_to_final.items():
            try:
                tmp_path.rename(final_path)
            except OSError as exc:
                lines.append(
                    f"[warn] Could not complete rename {tmp_path.name} → "
                    f"{final_path.name}: {exc}"
                )
    elif charm_images_dir and not charm_images_dir.is_dir():
        lines.append(
            f"[info] Charm images folder not found ({charm_images_dir}) — "
            "disk files not renamed."
        )

    return len(rename_map), lines


# ---------------------------------------------------------------------------
# Sheet 1 -- Shopping Route (floor-ascending, per-component status tracking)
# ---------------------------------------------------------------------------

def _sheet_route(ws, items: list[ResolvedItem],
                 statuses: dict[tuple[str, str], str] | None = None,
                 lang: str = "en",
                 title_fn=None,
                 charm_shops: list[CharmShop] | None = None,
                 charm_library: dict[str, CharmLibraryEntry] | None = None,
                 charm_images_dir: Path | None = None) -> None:
    ws.title = _t("Shopping Route", lang)
    ws.sheet_properties.tabColor = "1F4E79"

    # EN: 15 cols (with Etsy + Private Notes). ZH: compact 8 cols.
    _zh_route_compact = lang == "zh"
    if _zh_route_compact:
        HDRS = [
            "#", _t("Photo", lang), _t("Supplier", lang), _t("Stall", lang),
            _t("Items to Purchase", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Private Notes", lang),
        ]
        COL_ITEMS_TO_PURCHASE = 5
        COL_SUPPLIER, COL_STALL = 3, 4
        COL_PHONE, COL_QTY = 6, 7
        COL_PRIVATE_NOTES = 8
    else:
        HDRS = [
            "#", _t("Photo", lang), _t("Floor", lang), _t("Supplier", lang),
            _t("Stall", lang), _t("Product", lang),
            _t("Items to Purchase", lang),
            _t("Case", lang), _t("Grip", lang), _t("Charm", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Order #", lang),
            _t("Etsy Shop", lang),
            _t("Private Notes", lang),
        ]
        COL_ITEMS_TO_PURCHASE = 7
        COL_CASE, COL_GRIP, COL_CHARM = 8, 9, 10
        COL_SUPPLIER, COL_STALL = 4, 5
        COL_PHONE, COL_QTY = 11, 12
        COL_PRIVATE_NOTES = 15
    COLS    = len(HDRS)
    HDR_ROW = 4
    col_end = get_column_letter(COLS)

    # Use larger row/photo sizing for the Chinese version so images are clearly visible
    _row_h    = ZH_ROW_HEIGHT  if lang == "zh" else ROW_HEIGHT
    _photo_px = ZH_PHOTO_PX    if lang == "zh" else PHOTO_PX

    # -- Title row
    ws.merge_cells(f"A1:{col_end}1")
    if lang == "zh":
        title_date = date.today().strftime("%Y年%m月%d日")
        title_text = f"购物路线  --  {title_date}"
    else:
        title_text = f"Shopping Route  --  {date.today().strftime('%B %d, %Y')}"
    ws.cell(1, 1, title_text).font = _TITLE_FONT
    ws.row_dimensions[1].height = 36

    # Three-bucket classification:
    #  • routable    – has supplier with at least shop_name OR stall filled in
    #  • needs_info  – matched a catalog entry whose shop/stall are empty AND
    #                  match score is high-confidence (>= EMPTY_ENTRY_MATCH_THRESHOLD),
    #                  meaning it IS the same product but the user hasn't filled in
    #                  the location yet.  A new amber row was already added to the
    #                  catalog so the user can fill it in.
    #  • unmatched   – no catalog entry, OR a low-confidence match against an
    #                  empty-info entry (potential false positive → also gets its
    #                  own amber row appended to the catalog via update_catalog)
    def _supplier_has_location(r: ResolvedItem) -> bool:
        return bool(r.supplier and (r.supplier.shop_name or r.supplier.stall))

    routable   = [r for r in items if _supplier_has_location(r)]
    needs_info = [r for r in items
                  if r.supplier
                  and not _supplier_has_location(r)
                  and not _needs_catalog_entry(r)]
    unmatched  = [r for r in items if not r.supplier or _needs_catalog_entry(r)]

    # Charm items: any order whose style includes a charm component.
    # These receive a dedicated CHARMS section at the bottom of the sheet,
    # completely separate from the case/grip supplier sections.
    charm_items     = [r for r in items if _style_has(r.item.style)[2]]
    total_charm_qty = sum(r.item.quantity for r in charm_items)

    supplier_stops = len({(r.supplier.shop_name, r.supplier.stall) for r in routable})
    order_count    = len({r.order.order_number for r in items})

    # -- Subtitle row
    ws.merge_cells(f"A2:{col_end}2")
    if lang == "zh":
        sub_parts = [
            f"{len(items)} 件商品",
            f"{order_count} 个订单",
            f"{supplier_stops} 个供应商",
            "按楼层从低到高排序",
        ]
        if charm_items:
            sub_parts.append(f"{total_charm_qty} 个挂件需采购（独立楼栋）")
        if needs_info:
            sub_parts.append(f"{len(needs_info)} 个待填供应商信息")
        if unmatched:
            sub_parts.append(f"{len(unmatched)} 个未匹配")
    else:
        sub_parts = [
            f"{len(items)} items",
            f"{order_count} orders",
            f"{supplier_stops} supplier stops",
            "sorted lowest to highest floor",
        ]
        if charm_items:
            sub_parts.append(
                f"{total_charm_qty} charm(s) needed \u2014 separate building"
            )
        if needs_info:
            sub_parts.append(f"{len(needs_info)} awaiting supplier info")
        if unmatched:
            sub_parts.append(f"{len(unmatched)} unmatched")
    ws.cell(2, 1, "  |  ".join(sub_parts)).font = _SUB_FONT
    ws.row_dimensions[2].height = 24

    # -- Legend row
    ws.merge_cells(f"A3:{col_end}3")
    if lang == "zh":
        legend_text = (
            "待购项列标明本单需采购的部件（手机壳 / 支架）。"
            "   |   蓝色行 = 已在目录中，请在 supplier_catalog.xlsx 填写供应商信息"
            "   |   ✦ 挂件区（紫色）= 在独立楼栋另行采购，见下方挂件区"
        )
    else:
        legend_text = (
            "Per-component status:   Pending (white)   |   Purchased (green)"
            "   |   Out of Stock (amber)   |   Out of Production (red)"
            "   |   N/A (gray) = not part of this order"
            "   |   blue rows = in catalog, fill supplier info in supplier_catalog.xlsx"
            "   |   \u2728 Charm column = N/A here; charms are purchased at a SEPARATE BUILDING \u2014 see purple section below"
        )
    ws.cell(3, 1, legend_text).font = Font("Calibri", size=9, italic=True, color="555555")
    ws.row_dimensions[3].height = 14

    # -- Header row
    for ci, h in enumerate(HDRS, 1):
        ws.cell(HDR_ROW, ci, h)
    _style_header(ws, HDR_ROW, COLS)
    ws.cell(HDR_ROW, COL_ITEMS_TO_PURCHASE).fill = PatternFill("solid", fgColor="2E7D32")
    if not _zh_route_compact:
        ws.cell(HDR_ROW, COL_CASE).fill  = PatternFill("solid", fgColor="1A6B3C")
        ws.cell(HDR_ROW, COL_GRIP).fill  = PatternFill("solid", fgColor="1A3D6B")
        ws.cell(HDR_ROW, COL_CHARM).fill = PatternFill("solid", fgColor="5B1A6B")
    ws.row_dimensions[HDR_ROW].height = 18

    # -- Group by (supplier, stall), sorted floor-ascending  (routable only)
    groups: dict[tuple[str, str], list[ResolvedItem]] = defaultdict(list)
    for r in routable:
        groups[(r.supplier.shop_name, r.supplier.stall)].append(r)
    for _gk in groups:
        groups[_gk].sort(key=_route_item_sort_key)

    sorted_keys = sorted(
        groups,
        key=lambda k: (_stall_floor(k[1]), k[1] or "\uffff", k[0]),
    )

    # Track cells to add to the component-status dropdowns
    active_case_cells:  list[str] = []
    active_grip_cells:  list[str] = []
    active_charm_cells: list[str] = []

    row = HDR_ROW + 1
    first_data_row = row
    seq = 1

    _statuses = statuses or {}

    def _write_component_cell(ws, row, col, has_component, active_cells,
                              order_num="", comp="", item_title=""):
        """Write status (or N/A) into a component cell, restoring preserved values.

        The status key is (order_num, normalized_title, comp) so that items
        from the same order number but with different product titles never
        overwrite each other's statuses.
        """
        cell = ws.cell(row, col)
        if has_component:
            norm_title     = _normalize(item_title)[:50]
            preserved      = _statuses.get((order_num, norm_title, comp))
            cell.value     = _t(preserved, lang) if preserved else _t("Pending", lang)
            cell.alignment = _CENTER
            active_cells.append(cell.coordinate)
        else:
            cell.value     = _t("N/A", lang)
            cell.fill      = _NA_FILL
            cell.font      = _NA_FONT
            cell.alignment = _CENTER

    for gidx, key in enumerate(sorted_keys):
        fill = _GROUP_FILLS[gidx % 2]
        for r in groups[key]:
            floor       = _stall_floor(r.supplier.stall)
            floor_label = f"{floor}F" if floor != 999 else "--"
            has_case, has_grip, _ = _style_has(r.item.style)
            onum = r.order.order_number
            norm_title = _normalize(r.item.title)[:50]
            case_status = _statuses.get((onum, norm_title, "case"))
            grip_status = _statuses.get((onum, norm_title, "grip"))
            items_label = _items_to_purchase(has_case, has_grip, case_status, grip_status, lang)

            ws.cell(row, 1, seq)
            # col 2 = photo
            if _zh_route_compact:
                ws.cell(row, COL_SUPPLIER, r.supplier.shop_name or "--")
                ws.cell(row, COL_STALL, r.supplier.stall or "--")
            else:
                ws.cell(row, 3, floor_label)
                ws.cell(row, 4, r.supplier.shop_name or "--")
                ws.cell(row, 5, r.supplier.stall or "--")
                ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
            itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, items_label)
            itp_cell.alignment = _CENTER
            itp_cell.font = _ITEMS_TO_PURCHASE_FONT
            if not _zh_route_compact:
                _write_component_cell(ws, row, COL_CASE,  has_case, active_case_cells, onum, "case",  r.item.title)
                _write_component_cell(ws, row, COL_GRIP,  has_grip, active_grip_cells, onum, "grip",  r.item.title)
                # Charm is ALWAYS N/A in case/grip supplier rows — tracked in the
                # dedicated Charm section (separate building) at the bottom of this sheet.
                ws.cell(row, COL_CHARM, _t("N/A", lang))
            ws.cell(row, COL_PHONE, r.item.phone_model)
            ws.cell(row, COL_QTY, r.item.quantity)
            if not _zh_route_compact:
                ws.cell(row, 13, f"#{r.order.order_number}")
                ws.cell(row, 14, r.order.etsy_shop)
            if r.order.private_notes:
                pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                pn.alignment = _WRAP

            _style_row(ws, row, COLS, fill=fill)
            if not _zh_route_compact:
                # Re-apply N/A styling after _style_row (which resets font/fill)
                for col, has in (
                    (COL_CASE, has_case), (COL_GRIP, has_grip), (COL_CHARM, False)
                ):
                    if not has:
                        c = ws.cell(row, col)
                        c.fill = _NA_FILL
                        c.font = _NA_FONT
                    else:
                        ws.cell(row, col).alignment = _CENTER
            if r.order.private_notes:
                ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP

            ws.cell(row, 1).alignment = _CENTER
            if not _zh_route_compact:
                ws.cell(row, 3).alignment = _CENTER
            ws.cell(row, COL_QTY).alignment = _CENTER
            ws.row_dimensions[row].height = _row_h
            _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
            seq += 1
            row += 1

    # -- "Needs Supplier Info" section
    #    Items whose title matched a catalog entry but shop_name and stall are
    #    both empty (amber row in catalog not yet completed by the user).
    #    Shown in pale blue so they are visually distinct from both the main
    #    table and the truly-unmatched amber section below.
    if needs_info:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if lang == "zh":
            ni_banner = (
                "(~)  目录中 \u2013 待填供应商信息  "
                "\u2192  请打开 supplier_catalog.xlsx 并填写店名（D列）和摊位（E列）"
            )
        else:
            ni_banner = (
                "(~)  In Catalog \u2013 Awaiting Supplier Info  "
                "\u2192  open supplier_catalog.xlsx and fill in Shop Name (D) + Stall (E)"
            )
        ni_hdr = ws.cell(row, 1, ni_banner)
        ni_hdr.font   = Font("Calibri", bold=True, size=11, color="1F4E79")
        ni_hdr.fill   = PatternFill("solid", fgColor="BDD7EE")
        ni_hdr.border = _BORDER
        row += 1

        for r in sorted(needs_info, key=_route_item_sort_key):
            has_case, has_grip, _ = _style_has(r.item.style)
            onum = r.order.order_number
            norm_title = _normalize(r.item.title)[:50]
            case_status = _statuses.get((onum, norm_title, "case"))
            grip_status = _statuses.get((onum, norm_title, "grip"))
            items_label = _items_to_purchase(has_case, has_grip, case_status, grip_status, lang)

            ws.cell(row, 1, seq)
            if _zh_route_compact:
                ws.cell(row, COL_SUPPLIER, r.supplier.shop_name or "\u2014")
                ws.cell(row, COL_STALL, r.supplier.stall or "\u2014")
            else:
                ws.cell(row, 3, "--")
                ws.cell(row, 4, r.supplier.shop_name or "\u2014")
                ws.cell(row, 5, r.supplier.stall or "\u2014")
                ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
            itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, items_label)
            itp_cell.alignment = _CENTER
            itp_cell.font = _ITEMS_TO_PURCHASE_FONT
            if not _zh_route_compact:
                _write_component_cell(ws, row, COL_CASE, has_case, active_case_cells, onum, "case", r.item.title)
                _write_component_cell(ws, row, COL_GRIP, has_grip, active_grip_cells, onum, "grip", r.item.title)
                ws.cell(row, COL_CHARM, _t("N/A", lang))
            ws.cell(row, COL_PHONE, r.item.phone_model)
            ws.cell(row, COL_QTY, r.item.quantity)
            if not _zh_route_compact:
                ws.cell(row, 13, f"#{r.order.order_number}")
                ws.cell(row, 14, r.order.etsy_shop)
            if r.order.private_notes:
                pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                pn.alignment = _WRAP
            _style_row(ws, row, COLS, fill=_NEEDSINFO_FILL, font=_NEEDSINFO_FONT)
            if not _zh_route_compact:
                for col, has in (
                    (COL_CASE, has_case), (COL_GRIP, has_grip), (COL_CHARM, False)
                ):
                    c = ws.cell(row, col)
                    if not has:
                        c.fill = _NA_FILL
                        c.font = _NA_FONT
                    else:
                        c.alignment = _CENTER
            if r.order.private_notes:
                ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP
            ws.cell(row, 1).alignment = _CENTER
            if not _zh_route_compact:
                ws.cell(row, 3).alignment = _CENTER
            ws.cell(row, COL_QTY).alignment = _CENTER
            ws.row_dimensions[row].height = _row_h
            _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
            seq += 1
            row += 1

    # -- Unmatched items section
    if unmatched:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        warn = ws.cell(
            row, 1,
            "(!!)  未匹配商品 -- 目录中未找到供应商" if lang == "zh"
            else "(!!)  Unmatched Items -- supplier not found in catalog",
        )
        warn.font   = Font("Calibri", bold=True, size=11, color="856404")
        warn.fill   = _WARN_FILL
        warn.border = _BORDER
        row += 1

        for r in sorted(unmatched, key=_route_item_sort_key):
            has_case, has_grip, _ = _style_has(r.item.style)
            onum = r.order.order_number
            norm_title = _normalize(r.item.title)[:50]
            case_status = _statuses.get((onum, norm_title, "case"))
            grip_status = _statuses.get((onum, norm_title, "grip"))
            items_label = _items_to_purchase(has_case, has_grip, case_status, grip_status, lang)

            ws.cell(row, 1, seq)
            if _zh_route_compact:
                ws.cell(row, COL_SUPPLIER, "???")
                ws.cell(row, COL_STALL, "???")
            else:
                ws.cell(row, 3, "--")
                ws.cell(row, 4, "???")
                ws.cell(row, 5, "???")
                ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
            itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, items_label)
            itp_cell.alignment = _CENTER
            itp_cell.font = _ITEMS_TO_PURCHASE_FONT
            if not _zh_route_compact:
                _write_component_cell(ws, row, COL_CASE, has_case, active_case_cells, onum, "case", r.item.title)
                _write_component_cell(ws, row, COL_GRIP, has_grip, active_grip_cells, onum, "grip", r.item.title)
                ws.cell(row, COL_CHARM, _t("N/A", lang))
            ws.cell(row, COL_PHONE, r.item.phone_model)
            ws.cell(row, COL_QTY, r.item.quantity)
            if not _zh_route_compact:
                ws.cell(row, 13, f"#{r.order.order_number}")
                ws.cell(row, 14, r.order.etsy_shop)
            if r.order.private_notes:
                pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                pn.alignment = _WRAP
            _style_row(ws, row, COLS, fill=_WARN_FILL, font=_WARN_FONT)
            if not _zh_route_compact:
                for col, has in (
                    (COL_CASE, has_case), (COL_GRIP, has_grip), (COL_CHARM, False)
                ):
                    c = ws.cell(row, col)
                    if not has:
                        c.fill = _NA_FILL
                        c.font = _NA_FONT
                    else:
                        c.alignment = _CENTER
            if r.order.private_notes:
                ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP
            ws.cell(row, 1).alignment = _CENTER
            if not _zh_route_compact:
                ws.cell(row, 3).alignment = _CENTER
            ws.cell(row, COL_QTY).alignment = _CENTER
            ws.row_dimensions[row].height = _row_h
            _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
            seq += 1
            row += 1

    last_data_row = row - 1

    # -- Per-component status dropdowns (only on cells with active components)
    _status_opts = ZH_STATUS_OPTIONS if lang == "zh" else STATUS_OPTIONS
    dv_formula = f'"{",".join(_status_opts)}"'
    dv_kwargs  = dict(
        type="list",
        formula1=dv_formula,
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        error="请从下拉列表中选择一个值。" if lang == "zh" else "Pick a value from the dropdown.",
        errorTitle="状态无效" if lang == "zh" else "Invalid status",
    )
    # Case and Grip dropdowns for main supplier sections (charm is in its own section below)
    if not _zh_route_compact:
        for cell_list in (active_case_cells, active_grip_cells):
            if cell_list:
                dv = DataValidation(**dv_kwargs)
                ws.add_data_validation(dv)
                for coord in cell_list:
                    dv.add(coord)

    # -- Conditional formatting: row colour = worst component status
    #
    #    Priority (first rule wins):
    #      1. Any component = Out of Production  -> RED
    #      2. Any component = Out of Stock       -> AMBER  (and none OOP)
    #      3. All included components Purchased  -> GREEN
    #      4. (no rule)                          -> white / group colour (Pending)
    if not _zh_route_compact and last_data_row >= first_data_row:
        full_range = f"A{first_data_row}:{col_end}{last_data_row}"
        r0         = first_data_row  # reference row for relative formulas
        gc = f"${get_column_letter(COL_CASE)}"
        hc = f"${get_column_letter(COL_GRIP)}"
        ic = f"${get_column_letter(COL_CHARM)}"

        oop = _t("Out of Production", lang)
        oos = _t("Out of Stock", lang)
        purchased = _t("Purchased", lang)
        na        = _t("N/A", lang)

        # Rule 1 — any component Out of Production
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[f'OR({gc}{r0}="{oop}",{hc}{r0}="{oop}",{ic}{r0}="{oop}")'],
            fill=_STATUS_FILLS["Out of Production"],
            font=_STATUS_FONTS["Out of Production"],
            stopIfTrue=True,
        ))

        # Rule 2 — any component Out of Stock (and none Out of Production)
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[
                f'AND('
                f'OR({gc}{r0}="{oos}",{hc}{r0}="{oos}",{ic}{r0}="{oos}"),'
                f'NOT(OR({gc}{r0}="{oop}",{hc}{r0}="{oop}",{ic}{r0}="{oop}"))'
                f')'
            ],
            fill=_STATUS_FILLS["Out of Stock"],
            font=_STATUS_FONTS["Out of Stock"],
            stopIfTrue=True,
        ))

        # Rule 3 — all included components Purchased (N/A counts as done)
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[
                f'AND('
                f'OR({gc}{r0}="{na}",{gc}{r0}="{purchased}"),'
                f'OR({hc}{r0}="{na}",{hc}{r0}="{purchased}"),'
                f'OR({ic}{r0}="{na}",{ic}{r0}="{purchased}")'
                f')'
            ],
            fill=_STATUS_FILLS["Purchased"],
            font=_STATUS_FONTS["Purchased"],
            stopIfTrue=True,
        ))

    # ---------------------------------------------------------------------------
    # CHARMS TO PURCHASE — dedicated section, completely separate building
    # ---------------------------------------------------------------------------
    #
    # Two sub-sections:
    #   A) Aggregated by charm code — one row per unique charm, photo + details
    #      from the Charm Library, total qty across all orders.  Status tracked
    #      per charm code via sentinel ``~C:<code>`` in the Order # column.
    #   B) Awaiting charm code — orders whose style has a charm component but
    #      no charm code assigned yet.  Shows product photo + prompt to assign.
    # ---------------------------------------------------------------------------
    if charm_items:
        row += 1   # one blank separator row

        _cshops            = charm_shops or []
        _cshops_lookup_tmp = {cs.shop_name: cs for cs in _cshops}
        total_charm_qty_c  = sum(r.item.quantity for r in charm_items)

        # Partition: items with a charm code vs items still awaiting one
        _coded_items:    list[ResolvedItem] = []
        _awaiting_items: list[ResolvedItem] = []
        for _ci in charm_items:
            _cc = (_ci.supplier.charm_code if _ci.supplier else "").strip()
            if _cc:
                _coded_items.append(_ci)
            else:
                _awaiting_items.append(_ci)

        # Aggregate coded items by charm code
        _charm_agg: dict[str, dict] = {}
        for _ci in _coded_items:
            _cc = _ci.supplier.charm_code.strip()
            if _cc not in _charm_agg:
                _lib = (charm_library or {}).get(_cc)
                _charm_agg[_cc] = {
                    "code": _cc,
                    "sku": _lib.sku if _lib else "",
                    "default_shop": _lib.default_charm_shop if _lib else "",
                    "notes": _lib.notes if _lib else "",
                    "photo_bytes": None,
                    "charm_shop": "",
                    "charm_shop_obj": None,
                    "total_qty": 0,
                    "orders": [],
                    "items": [],
                }
                _ph = charm_photo_bytes_from_folder(_cc, charm_images_dir)
                if not _ph and _lib and _lib.photo_bytes:
                    _ph = _lib.photo_bytes
                _charm_agg[_cc]["photo_bytes"] = _ph
            _charm_agg[_cc]["total_qty"] += _ci.item.quantity
            _charm_agg[_cc]["orders"].append(_ci.order.order_number)
            _charm_agg[_cc]["items"].append(_ci)
            if not _charm_agg[_cc]["charm_shop"]:
                _as = (_ci.supplier.charm_shop if _ci.supplier else "").strip()
                if _as:
                    _charm_agg[_cc]["charm_shop"] = _as
                    _charm_agg[_cc]["charm_shop_obj"] = _cshops_lookup_tmp.get(_as)

        n_unique_charms  = len(_charm_agg)
        n_missing_code   = len(_awaiting_items)
        unassigned_count = sum(
            1 for r in charm_items
            if not (r.supplier and r.supplier.charm_shop
                    and r.supplier.charm_shop in _cshops_lookup_tmp)
        )

        # --- Banner row ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if lang == "zh":
            charm_banner_text = (
                f"\u2728  \u5f85\u8d2d\u6302\u4ef6  \u2014  \u72ec\u7acb\u697c\u68cb"
                f"  \u2014  \u5171\u9700 {total_charm_qty_c} \u4e2a\u6302\u4ef6"
                f"\uff0c\u6d89\u53ca {len(charm_items)} \u4e2a\u8ba2\u5355"
            )
            if n_unique_charms:
                charm_banner_text += f"  \u2014  {n_unique_charms} \u79cd\u6302\u4ef6"
            if n_missing_code:
                charm_banner_text += (
                    f"  \u25b6  {n_missing_code} \u4e2a\u8ba2\u5355\u5f85\u5206\u914d\u6302\u4ef6\u7f16\u7801"
                    f" \u2014 \u6253\u5f00 supplier_catalog.xlsx \u2192 Product Map H\u5217"
                )
            if unassigned_count:
                charm_banner_text += (
                    f"  \u25b6  {unassigned_count} \u4e2a\u8ba2\u5355\u672a\u5206\u914d\u6302\u4ef6\u5e97"
                )
        else:
            charm_banner_text = (
                f"\u2728  CHARMS TO PURCHASE  \u2014  SEPARATE BUILDING"
                f"  \u2014  {total_charm_qty_c} charm(s) needed"
                f" across {len(charm_items)} order(s)"
            )
            if n_missing_code:
                charm_banner_text += (
                    f"  \u25b6  {n_missing_code} order(s) missing charm-code"
                    f" assignment \u2014 open supplier_catalog.xlsx"
                    f" \u2192 Product Map col H (Charm Code)"
                )
            if unassigned_count:
                charm_banner_text += (
                    f"  \u25b6  {unassigned_count} order(s) missing charm-shop"
                    f" assignment \u2014 open supplier_catalog.xlsx"
                    f" \u2192 Product Map col G (Charm Shop)"
                )
        charm_banner = ws.cell(row, 1, charm_banner_text)
        charm_banner.font      = Font("Calibri", bold=True, size=13, color="FFFFFF")
        charm_banner.fill      = _CHARM_BANNER_FILL
        charm_banner.border    = _BORDER
        charm_banner.alignment = _CENTER
        ws.row_dimensions[row].height = 26
        row += 1

        # --- Charm shops reference row ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if _cshops:
            shop_parts = [f"{s.shop_name}  ({s.stall})" for s in _cshops]
            shops_label = ("\u6302\u4ef6\u5e97\u94fa\uff1a  " if lang == "zh" else "Charm shops:   ")
            shops_text  = shops_label + "   |   ".join(shop_parts)
        else:
            shops_text = (
                "\u672a\u914d\u7f6e\u6302\u4ef6\u5e97\u94fa \u2014 \u8bf7\u5728 supplier_catalog.xlsx \u7684\u6302\u4ef6\u5e97\u94fa\u6807\u7b7e\u4e2d\u6dfb\u52a0"
                if lang == "zh"
                else "No charm shops configured \u2014 add them in the "
                     "'Charm Shops' tab of supplier_catalog.xlsx"
            )
        shops_ref = ws.cell(row, 1, shops_text)
        shops_ref.font      = Font("Calibri", bold=True, size=10, color="3D1359")
        shops_ref.fill      = _CHARM_SHOPS_FILL
        shops_ref.border    = _BORDER
        shops_ref.alignment = _CENTER
        ws.row_dimensions[row].height = 20
        row += 1

        charm_shop_lookup   = _cshops_lookup_tmp
        charm_section_cells: list[str] = []
        charm_first_row     = row

        # ===============================================================
        # SUB-SECTION A — Aggregated charm purchase list (by charm code)
        # ===============================================================
        if _charm_agg:
            if _zh_route_compact:
                _CHARM_HDRS = [
                    "#", _t("Photo", lang),
                    "\u6302\u4ef6\u7f16\u7801",
                    "SKU",
                    "\u6302\u4ef6\u5e97\u94fa",
                    "\u6446\u4f4d",
                    _t("Qty", lang),
                    "\u5907\u6ce8",
                ]
            else:
                _CHARM_HDRS = [
                    "#", _t("Photo", lang), "",
                    "Charm Code" if lang != "zh" else "\u6302\u4ef6\u7f16\u7801",
                    "SKU",
                    "Charm Shop" if lang != "zh" else "\u6302\u4ef6\u5e97\u94fa",
                    _t("Stall", lang),
                    "", "",
                    _t("Charm", lang),
                    "",
                    _t("Qty", lang),
                    "Orders" if lang != "zh" else "\u5173\u8054\u8ba2\u5355",
                    "",
                    "Notes" if lang != "zh" else "\u5907\u6ce8",
                ]
            for ci, h in enumerate(_CHARM_HDRS, 1):
                ws.cell(row, ci, h)
            _style_header(ws, row, COLS)
            if not _zh_route_compact:
                for _muted_col in (3, 8, 9, 11, 14):
                    ws.cell(row, _muted_col).fill = _NA_FILL
                    ws.cell(row, _muted_col).font = _CHARM_NA_HDR_FONT
                ws.cell(row, COL_CHARM).fill = _CHARM_HDR_FILL
            ws.row_dimensions[row].height = 18
            row += 1
            charm_first_row = row

            sorted_codes = sorted(
                _charm_agg,
                key=lambda c: (_charm_agg[c]["charm_shop"] or "\uffff", c),
            )

            for cidx, code in enumerate(sorted_codes):
                agg  = _charm_agg[code]
                fill = _CHARM_GROUP_FILLS[cidx % 2]

                _cs_name = agg["charm_shop"] or agg["default_shop"]
                _cs_obj  = charm_shop_lookup.get(_cs_name)
                if _cs_obj:
                    shop_display  = _cs_obj.shop_name
                    stall_display = _cs_obj.stall
                elif _cs_name:
                    shop_display  = f"? {_cs_name}"
                    stall_display = "?"
                else:
                    shop_display  = "--"
                    stall_display = "--"

                unique_orders  = sorted(set(agg["orders"]))
                orders_display = ", ".join(f"#{o}" for o in unique_orders)
                notes_val      = agg["notes"]

                ws.cell(row, 1, cidx + 1)
                if _zh_route_compact:
                    ws.cell(row, 3, code)
                    ws.cell(row, 4, agg["sku"])
                    ws.cell(row, 5, shop_display)
                    ws.cell(row, 6, stall_display)
                    ws.cell(row, 7, agg["total_qty"])
                    if notes_val:
                        ws.cell(row, 8, notes_val).alignment = _WRAP
                else:
                    ws.cell(row, 3, "")
                    ws.cell(row, 4, code)
                    ws.cell(row, 5, agg["sku"])
                    ws.cell(row, 6, shop_display)
                    ws.cell(row, 7, stall_display)
                    ws.cell(row, 8, "")
                    ws.cell(row, 9, "")

                    preserved = _statuses.get((code, "", "charm_agg"))
                    if not preserved:
                        _per_order = []
                        for _ri in agg["items"]:
                            _ps = _statuses.get((
                                _ri.order.order_number,
                                _normalize(_ri.item.title)[:50],
                                "charm",
                            ))
                            if _ps:
                                _per_order.append(_ps)
                        if _per_order:
                            if any(s == "Out of Production" for s in _per_order):
                                preserved = "Out of Production"
                            elif any(s == "Out of Stock" for s in _per_order):
                                preserved = "Out of Stock"
                            elif all(s == "Purchased" for s in _per_order):
                                preserved = "Purchased"

                    charm_cell = ws.cell(row, COL_CHARM)
                    charm_cell.value     = _t(preserved, lang) if preserved else _t("Pending", lang)
                    charm_cell.alignment = _CENTER
                    charm_section_cells.append(charm_cell.coordinate)

                    ws.cell(row, 11, "")
                    ws.cell(row, 12, agg["total_qty"])
                    ws.cell(row, 13, f"~C:{code}")
                    _orders_trunc = orders_display[:60] + ("\u2026" if len(orders_display) > 60 else "")
                    ws.cell(row, 14, _orders_trunc)
                    if notes_val:
                        ws.cell(row, 15, notes_val).alignment = _WRAP

                _style_row(ws, row, COLS, fill=fill)
                if not _zh_route_compact:
                    for _na_c in (3, 8, 9, 11):
                        nc = ws.cell(row, _na_c)
                        nc.fill = _NA_FILL
                        nc.font = _NA_FONT
                    ws.cell(row, COL_CHARM).alignment = _CENTER
                    _sentinel = ws.cell(row, 13)
                    _sentinel.font = Font("Calibri", size=7, color="D8D8D8")
                if notes_val:
                    ws.cell(row, 15 if not _zh_route_compact else 8).alignment = _WRAP
                ws.cell(row, 1).alignment = _CENTER
                ws.cell(row, 4 if not _zh_route_compact else 3).alignment = _CENTER
                ws.cell(row, 12 if not _zh_route_compact else 7).alignment = _CENTER
                ws.row_dimensions[row].height = _row_h
                _embed_photo(ws, agg["photo_bytes"], row, 2, _photo_px)
                row += 1

        charm_last_row = row - 1

        # DataValidation for aggregated charm status cells
        if charm_section_cells:
            dv_charm = DataValidation(**dv_kwargs)
            ws.add_data_validation(dv_charm)
            for coord in charm_section_cells:
                dv_charm.add(coord)

        # Conditional formatting for aggregated charm rows
        if not _zh_route_compact and charm_last_row >= charm_first_row:
            charm_range = f"A{charm_first_row}:{col_end}{charm_last_row}"
            cr0 = charm_first_row
            ic  = f"${get_column_letter(COL_CHARM)}"
            oop = _t("Out of Production", lang)
            oos = _t("Out of Stock", lang)
            purch = _t("Purchased", lang)

            ws.conditional_formatting.add(charm_range, FormulaRule(
                formula=[f'{ic}{cr0}="{oop}"'],
                fill=_STATUS_FILLS["Out of Production"],
                font=_STATUS_FONTS["Out of Production"],
                stopIfTrue=True,
            ))
            ws.conditional_formatting.add(charm_range, FormulaRule(
                formula=[f'{ic}{cr0}="{oos}"'],
                fill=_STATUS_FILLS["Out of Stock"],
                font=_STATUS_FONTS["Out of Stock"],
                stopIfTrue=True,
            ))
            ws.conditional_formatting.add(charm_range, FormulaRule(
                formula=[f'{ic}{cr0}="{purch}"'],
                fill=_STATUS_FILLS["Purchased"],
                font=_STATUS_FONTS["Purchased"],
                stopIfTrue=True,
            ))

        # ===============================================================
        # SUB-SECTION B — Awaiting charm code assignment
        # ===============================================================
        if _awaiting_items:
            row += 1
            _total_await_qty = sum(r.item.quantity for r in _awaiting_items)

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
            if lang == "zh":
                _await_text = (
                    f"\u23f3  \u5f85\u5206\u914d\u6302\u4ef6\u7f16\u7801  \u2014  "
                    f"{len(_awaiting_items)} \u4e2a\u8ba2\u5355"
                    f"\uff08{_total_await_qty} \u4ef6\uff09"
                    f"  \u2192  \u6253\u5f00 supplier_catalog.xlsx"
                    f" \u2192 Product Map H\u5217 (\u6302\u4ef6\u7f16\u7801)"
                )
            else:
                _await_text = (
                    f"\u23f3  AWAITING CHARM CODE ASSIGNMENT  \u2014  "
                    f"{len(_awaiting_items)} order(s)"
                    f" ({_total_await_qty} unit(s))"
                    f"  \u2192  open supplier_catalog.xlsx"
                    f" \u2192 Product Map col H (Charm Code)"
                )
            await_cell = ws.cell(row, 1, _await_text)
            await_cell.font      = Font("Calibri", bold=True, size=11, color="7D4E00")
            await_cell.fill      = PatternFill("solid", fgColor="FFF3CD")
            await_cell.border    = _BORDER
            await_cell.alignment = _CENTER
            ws.row_dimensions[row].height = 22
            row += 1

            for ci, h in enumerate(HDRS, 1):
                ws.cell(row, ci, h)
            _style_header(ws, row, COLS)
            if not _zh_route_compact:
                ws.cell(row, COL_CASE).fill  = _NA_FILL
                ws.cell(row, COL_CASE).font  = _CHARM_NA_HDR_FONT
                ws.cell(row, COL_GRIP).fill  = _NA_FILL
                ws.cell(row, COL_GRIP).font  = _CHARM_NA_HDR_FONT
                ws.cell(row, COL_CHARM).fill = PatternFill("solid", fgColor="FFF3CD")
                ws.cell(row, COL_CHARM).font = Font("Calibri", bold=True, size=11, color="7D4E00")
            ws.row_dimensions[row].height = 18
            row += 1

            _AWAIT_FILL      = PatternFill("solid", fgColor="FFFBF0")
            _AWAIT_CODE_FONT = Font("Calibri", size=9, color="7D4E00", italic=True)
            _AWAIT_CODE_FILL = PatternFill("solid", fgColor="FFF3CD")

            def _await_sort(ri: ResolvedItem) -> tuple[str, str]:
                return (_normalize(ri.item.title), ri.order.order_number)

            for aidx, r in enumerate(sorted(_awaiting_items, key=_await_sort)):
                onum = r.order.order_number
                assigned_name = (r.supplier.charm_shop if r.supplier else "") or ""
                assigned_cs   = charm_shop_lookup.get(assigned_name)
                if assigned_cs:
                    shop_display  = assigned_cs.shop_name
                    stall_display = assigned_cs.stall
                elif assigned_name:
                    shop_display  = f"? {assigned_name}"
                    stall_display = "?"
                else:
                    shop_display  = "--"
                    stall_display = "--"

                ws.cell(row, 1, aidx + 1)
                if _zh_route_compact:
                    ws.cell(row, COL_SUPPLIER, shop_display)
                    ws.cell(row, COL_STALL, stall_display)
                else:
                    ws.cell(row, 3, "--")
                    ws.cell(row, 4, shop_display)
                    ws.cell(row, 5, stall_display)
                    ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
                itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, "\u2014")
                itp_cell.alignment = _CENTER
                itp_cell.font = _ITEMS_TO_PURCHASE_FONT

                if not _zh_route_compact:
                    for na_col in (COL_CASE, COL_GRIP):
                        nc = ws.cell(row, na_col, _t("N/A", lang))
                        nc.fill      = _NA_FILL
                        nc.font      = _NA_FONT
                        nc.alignment = _CENTER
                    _await_charm = ws.cell(row, COL_CHARM)
                    _await_charm.value     = "\u23f3 Awaiting Code" if lang != "zh" else "\u23f3 \u5f85\u5206\u914d"
                    _await_charm.alignment = _CENTER
                    _await_charm.font      = _AWAIT_CODE_FONT
                    _await_charm.fill      = _AWAIT_CODE_FILL

                ws.cell(row, COL_PHONE, r.item.phone_model)
                ws.cell(row, COL_QTY, r.item.quantity)
                if not _zh_route_compact:
                    ws.cell(row, 13, f"~?#{r.order.order_number}")
                    ws.cell(row, 14, r.order.etsy_shop)
                if r.order.private_notes:
                    pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                    pn.alignment = _WRAP

                _style_row(ws, row, COLS, fill=_AWAIT_FILL)
                if not _zh_route_compact:
                    for na_col in (COL_CASE, COL_GRIP):
                        nc = ws.cell(row, na_col)
                        nc.fill = _NA_FILL
                        nc.font = _NA_FONT
                    _await_charm = ws.cell(row, COL_CHARM)
                    _await_charm.font = _AWAIT_CODE_FONT
                    _await_charm.fill = _AWAIT_CODE_FILL
                    _await_charm.alignment = _CENTER
                if r.order.private_notes:
                    ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP
                ws.cell(row, 1).alignment = _CENTER
                if not _zh_route_compact:
                    ws.cell(row, 3).alignment = _CENTER
                ws.cell(row, COL_QTY).alignment = _CENTER
                ws.row_dimensions[row].height = _row_h
                _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
                row += 1


    # -- Column widths, freeze, filter
    # Product narrowed (26) so Items to Purchase (14) is more prominent
    _photo_col_w = ZH_PHOTO_COL_W if lang == "zh" else PHOTO_COL_W
    if _zh_route_compact:
        col_widths = [4, _photo_col_w, 13, 9, 14, 18, 4, 28]
    else:
        col_widths = [4, _photo_col_w, 6, 13, 9, 26, 14, 10, 10, 10, 18, 4, 16, 15, 32]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"
    if last_data_row >= first_data_row:
        ws.auto_filter.ref = f"A{HDR_ROW}:{col_end}{last_data_row}"


# ---------------------------------------------------------------------------
# Sheet 2 -- Orders Detail
# ---------------------------------------------------------------------------

def _sheet_orders(ws, items: list[ResolvedItem], lang: str = "en", title_fn=None,
                  charm_library: dict[str, CharmLibraryEntry] | None = None,
                  charm_images_dir: Path | None = None) -> None:
    ws.sheet_properties.tabColor = "2E75B6"

    if lang == "zh":
        HDRS = [
            _t("Buyer", lang), _t("Ship To", lang), _t("Country", lang),
            _t("Order Date", lang), _t("Photo", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Supplier", lang), _t("Stall", lang), _t("Match %", lang),
            _t("Private Notes", lang),
        ]
    else:
        HDRS = [
            _t("Order #", lang), _t("Etsy Shop", lang), _t("Buyer", lang),
            _t("Ship To", lang), _t("Country", lang), _t("Order Date", lang),
            _t("Photo", lang), _t("Product", lang),
            _t("Case", lang), _t("Grip", lang), _t("Charm", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Supplier", lang), _t("Stall", lang), _t("Match %", lang),
            _t("Private Notes", lang),
        ]
    COLS = len(HDRS)

    for ci, h in enumerate(HDRS, 1):
        ws.cell(1, ci, h)
    _style_header(ws, 1, COLS)
    if lang != "zh":
        _case_col, _grip_col, _charm_col = 9, 10, 11
        ws.cell(1, _case_col).fill  = PatternFill("solid", fgColor="1A6B3C")
        ws.cell(1, _grip_col).fill  = PatternFill("solid", fgColor="1A3D6B")
        ws.cell(1, _charm_col).fill = PatternFill("solid", fgColor="5B1A6B")
    ws.row_dimensions[1].height = 18

    row = 2
    for r in sorted(items, key=lambda x: x.order.order_number):
        if lang == "zh":
            _b, _s, _c, _od, _ph, _pm, _q, _su, _st, _m, _pn = (
                1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11
            )
            ws.cell(row, _b,  r.order.buyer_name)
            ws.cell(row, _s,  r.order.ship_to_name)
            ws.cell(row, _c,  r.order.ship_to_country)
            ws.cell(row, _od, r.order.order_date)
            ws.cell(row, _pm, r.item.phone_model)
            ws.cell(row, _q,  r.item.quantity)
            ws.cell(row, _su, (r.supplier.shop_name or "--") if r.supplier else "--")
            ws.cell(row, _st, (r.supplier.stall or "--")     if r.supplier else "--")
            ws.cell(row, _m,  f"{r.match_score:.0f}%"        if r.supplier else "--")
            if r.order.private_notes:
                ws.cell(row, _pn, r.order.private_notes).alignment = _WRAP
            center_cols = (_q, _m)
        else:
            case, grip, charm = _style_flags(r.item.style)
            ws.cell(row, 1, f"#{r.order.order_number}")
            ws.cell(row, 2, r.order.etsy_shop)
            _b, _s, _c, _od, _ph, _pr, _ca, _g, _ch, _pm, _q, _su, _st, _m, _pn = (
                3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17
            )
            ws.cell(row, _b,  r.order.buyer_name)
            ws.cell(row, _s,  r.order.ship_to_name)
            ws.cell(row, _c,  r.order.ship_to_country)
            ws.cell(row, _od, r.order.order_date)
            ws.cell(row, _pr, title_fn(r.item.title) if title_fn else r.item.title)
            ws.cell(row, _ca, case)
            ws.cell(row, _g,  grip)
            ws.cell(row, _ch, charm)
            ws.cell(row, _pm, r.item.phone_model)
            ws.cell(row, _q,  r.item.quantity)
            ws.cell(row, _su, (r.supplier.shop_name or "--") if r.supplier else "--")
            ws.cell(row, _st, (r.supplier.stall or "--")     if r.supplier else "--")
            ws.cell(row, _m,  f"{r.match_score:.0f}%"        if r.supplier else "--")
            if r.order.private_notes:
                ws.cell(row, _pn, r.order.private_notes).alignment = _WRAP
            center_cols = (_ca, _g, _ch, _q, _m)

        if not r.supplier or _needs_catalog_entry(r):
            fill = _WARN_FILL                          # unmatched or false-positive – amber
        elif not r.supplier.shop_name and not r.supplier.stall:
            fill = _NEEDSINFO_FILL                     # in catalog, info pending – blue
        else:
            fill = None
        _style_row(ws, row, COLS, fill=fill)
        for cc in center_cols:
            ws.cell(row, cc).alignment = _CENTER
        if r.order.private_notes:
            ws.cell(row, _pn).alignment = _WRAP
        ws.row_dimensions[row].height = ROW_HEIGHT
        _, _, _has_charm = _style_has(r.item.style)
        _od_ph = (
            (_resolve_charm_photo_bytes(r, charm_library, charm_images_dir)
             or r.item.photo_bytes)
            if _has_charm else r.item.photo_bytes
        )
        _embed_photo(ws, _od_ph, row, _ph)
        row += 1

    if lang == "zh":
        col_widths = [18, 18, 14, 14, PHOTO_COL_W, 18, 4, 14, 10, 10, 28]
    else:
        col_widths = [16, 18, 16, 18, 14, 14, PHOTO_COL_W, 52, 6, 6, 7, 18, 4, 14, 10, 10, 32]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(COLS)}{row - 1}"


# ---------------------------------------------------------------------------
# Sheet 3 -- Summary
# ---------------------------------------------------------------------------

def _sheet_summary(ws, items: list[ResolvedItem], lang: str = "en") -> None:
    ws.sheet_properties.tabColor = "548235"

    ws.merge_cells("A1:D1")
    ws.cell(1, 1, _t("Summary", lang)).font = _TITLE_FONT
    ws.row_dimensions[1].height = 34

    row = 3
    routable_ct   = sum(1 for r in items
                        if r.supplier and (r.supplier.shop_name or r.supplier.stall))
    needs_info_ct = sum(1 for r in items
                        if r.supplier
                        and not (r.supplier.shop_name or r.supplier.stall)
                        and not _needs_catalog_entry(r))
    unmatched_ct  = sum(1 for r in items if not r.supplier or _needs_catalog_entry(r))
    stats = [
        (_t("Total orders", lang),                              len({r.order.order_number for r in items})),
        (_t("Total line items", lang),                          len(items)),
        (_t("Total quantity", lang),                            sum(r.item.quantity for r in items)),
        (_t("Ready (supplier + location)", lang),               routable_ct),
        (_t("In catalog \u2013 needs supplier info", lang),     needs_info_ct),
        (_t("Not in catalog (unmatched)", lang),                unmatched_ct),
    ]
    for label, val in stats:
        ws.cell(row, 1, label).font = _BODY_BOLD
        ws.cell(row, 2, val).font   = _BODY
        for c in (1, 2):
            ws.cell(row, c).border = _BORDER
        row += 1

    # Items per supplier (floor-sorted)
    row += 1
    ws.cell(row, 1, _t("Items per Supplier", lang)).font = _SEC_FONT
    row += 1
    for ci, h in enumerate([
        _t("Floor", lang), _t("Supplier", lang), _t("Stall", lang),
        _t("Items", lang), _t("Qty", lang),
    ], 1):
        ws.cell(row, ci, h)
    _style_header(ws, row, 5)
    row += 1

    sup_groups: dict[tuple[str, str], list[ResolvedItem]] = defaultdict(list)
    for r in items:
        if r.supplier:
            sup_groups[(r.supplier.shop_name, r.supplier.stall)].append(r)

    for (shop, stall), grp in sorted(
        sup_groups.items(),
        key=lambda x: (_stall_floor(x[0][1]), x[0][1] or "\uffff", x[0][0]),
    ):
        floor = _stall_floor(stall)
        ws.cell(row, 1, f"{floor}F" if floor != 999 else "--")
        ws.cell(row, 2, shop  or "--")
        ws.cell(row, 3, stall or "--")
        ws.cell(row, 4, len(grp))
        ws.cell(row, 5, sum(r.item.quantity for r in grp))
        _style_row(ws, row, 5)
        ws.cell(row, 1).alignment = _CENTER
        row += 1

    # Items per Etsy shop (omit for Chinese — no Etsy shop column)
    if lang != "zh":
        row += 1
        ws.cell(row, 1, _t("Items per Etsy Shop", lang)).font = _SEC_FONT
        row += 1
        for ci, h in enumerate([_t("Etsy Shop", lang), _t("Orders", lang), _t("Items", lang)], 1):
            ws.cell(row, ci, h)
        _style_header(ws, row, 3)
        row += 1

        shop_groups: dict[str, list[ResolvedItem]] = defaultdict(list)
        for r in items:
            shop_groups[r.order.etsy_shop].append(r)

        for shop, grp in sorted(shop_groups.items()):
            ws.cell(row, 1, shop)
            ws.cell(row, 2, len({r.order.order_number for r in grp}))
            ws.cell(row, 3, len(grp))
            _style_row(ws, row, 3)
            row += 1

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 20


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def generate_xlsx(items: list[ResolvedItem], output: Path,
                  statuses: dict[tuple[str, str], str] | None = None,
                  lang: str = "en",
                  title_fn=None,
                  charm_shops: list[CharmShop] | None = None,
                  charm_library: dict[str, CharmLibraryEntry] | None = None,
                  charm_images_dir: Path | None = None) -> None:
    wb = openpyxl.Workbook()
    _sheet_route(
        wb.active, items,
        statuses=statuses, lang=lang, title_fn=title_fn,
        charm_shops=charm_shops,
        charm_library=charm_library,
        charm_images_dir=charm_images_dir,
    )
    _sheet_orders(
        wb.create_sheet(_t("Orders Detail", lang)), items,
        lang=lang, title_fn=title_fn, charm_library=charm_library,
        charm_images_dir=charm_images_dir,
    )
    _sheet_summary(wb.create_sheet(_t("Summary", lang)), items, lang=lang)

    # Configure every sheet so File → Export to PDF produces readable output:
    # landscape orientation, fit all columns onto one page wide (unlimited tall),
    # and narrow margins to maximise usable area.
    from openpyxl.worksheet.page import PageMargins
    for ws in wb.worksheets:
        ws.page_setup.orientation        = "landscape"
        ws.page_setup.paperSize          = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage          = True
        ws.page_setup.fitToWidth         = 1
        ws.page_setup.fitToHeight        = 0   # unlimited pages tall
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(
            left=0.4, right=0.4, top=0.6, bottom=0.6,
            header=0.3, footer=0.3,
        )

    wb.save(output)
    log.info("Saved -> %s", output.resolve())


# ---------------------------------------------------------------------------
# HTML generation -- responsive, mobile-first, self-contained shopping route
# ---------------------------------------------------------------------------

# Status CSS class names (used in both Python-rendered HTML and JS)
_STATUS_CSS = {
    "Purchased":         "s-bought",
    "Out of Stock":      "s-oos",
    "Out of Production": "s-oop",
    "Pending":           "s-pending",
    "N/A":               "s-na",
    # ZH equivalents
    "已购买": "s-bought",
    "缺货":   "s-oos",
    "停产":   "s-oop",
    "待处理": "s-pending",
    "不适用": "s-na",
}

# Row-level status CSS class names (worst-component rule)
_ROW_STATUS_CSS = {
    "oop":      "row-oop",
    "oos":      "row-oos",
    "bought":   "row-bought",
    "pending":  "",
}


def _image_data_uri(photo_bytes: bytes | None) -> str:
    """Return a base64 data URI for PNG, WebP, or JPEG bytes, or empty string."""
    if not photo_bytes:
        return ""
    if len(photo_bytes) >= 8 and photo_bytes[:8] == b"\x89PNG\r\n\x1a\n":
        return "data:image/png;base64," + base64.b64encode(photo_bytes).decode()
    if len(photo_bytes) >= 12 and photo_bytes[:4] == b"RIFF" and photo_bytes[8:12] == b"WEBP":
        return "data:image/webp;base64," + base64.b64encode(photo_bytes).decode()
    return "data:image/jpeg;base64," + base64.b64encode(photo_bytes).decode()


def _img_b64(photo_bytes: bytes | None) -> str:
    """Return a base64 data URI (JPEG/PNG)."""
    return _image_data_uri(photo_bytes)


def _status_cls(value: str) -> str:
    return _STATUS_CSS.get(value, "s-pending")


def _html_item_card(r: ResolvedItem, seq: int, row_cls: str, lang: str,
                    is_charm: bool, title_fn=None,
                    statuses: dict | None = None,
                    charm_shop_lookup: dict | None = None,
                    etsy_shop_col: bool = True,
                    charm_library: dict[str, CharmLibraryEntry] | None = None,
                    charm_images_dir: Path | None = None) -> str:
    """Render one item as a responsive <tr>.

    Status cells include ``data-skey`` (localStorage key) and ``data-comp``
    so the JS status-picker knows exactly what to update and persist.
    N/A cells are rendered as plain non-interactive badges.
    """
    _statuses  = statuses or {}
    title_raw  = r.item.title
    title_display = title_fn(title_raw) if title_fn else title_raw
    order_num  = r.order.order_number
    norm_title = _normalize(title_raw)[:50]

    if is_charm:
        cs_lookup  = charm_shop_lookup or {}
        assigned   = (r.supplier.charm_shop if r.supplier else "") or ""
        cs_obj     = cs_lookup.get(assigned)
        shop_disp  = cs_obj.shop_name if cs_obj else (f"? {assigned}" if assigned else "--")
        stall_disp = cs_obj.stall     if cs_obj else ("?" if assigned else "--")
        floor_disp = "--"
        order_ref  = f"~#{order_num}"
    else:
        shop_disp  = r.supplier.shop_name if r.supplier else "???"
        stall_disp = r.supplier.stall     if r.supplier else "???"
        stall_val  = r.supplier.stall     if r.supplier else ""
        floor_n    = _stall_floor(stall_val)
        floor_disp = f"{floor_n}F" if floor_n != 999 else "--"
        order_ref  = f"#{order_num}"

    has_case, has_grip, _ = _style_has(r.item.style)
    na_val = _t("N/A", lang)

    def _comp_val(comp: str, has: bool) -> str:
        if not has:
            return na_val
        preserved = _statuses.get((order_num, norm_title, comp))
        return _t(preserved, lang) if preserved else _t("Pending", lang)

    if is_charm:
        pres      = _statuses.get((order_num, norm_title, "charm"))
        charm_val = _t(pres, lang) if pres else _t("Pending", lang)
        case_val  = na_val
        grip_val  = na_val
        items_label = "\u2014"
    else:
        case_val  = _comp_val("case",  has_case)
        grip_val  = _comp_val("grip",  has_grip)
        charm_val = na_val
        case_status = _statuses.get((order_num, norm_title, "case"))
        grip_status = _statuses.get((order_num, norm_title, "grip"))
        items_label = _items_to_purchase(has_case, has_grip, case_status, grip_status, lang)

    _photo_bytes = r.item.photo_bytes
    if is_charm and (charm_library or charm_images_dir):
        _lib_ph = _resolve_charm_photo_bytes(r, charm_library, charm_images_dir)
        if _lib_ph:
            _photo_bytes = _lib_ph
    img_src  = _image_data_uri(_photo_bytes)
    img_html = (
        f'<img src="{img_src}" class="item-photo" loading="lazy" alt="">'
        if img_src else
        '<div class="item-photo no-photo"></div>'
    )

    def _skey(comp: str) -> str:
        """Deterministic localStorage key — mirrors the Python (order, norm, comp) tuple."""
        return f"{order_num}|{norm_title}|{comp}"

    def _badge(val: str, comp: str, has: bool) -> str:
        """Return a badge span.  Interactive (tappable) when the component is present."""
        cls = _status_cls(val)
        if not has:
            # N/A — non-interactive, plain badge
            return f'<span class="badge {cls}">{val}</span>'
        # Interactive badge: carries storage key + component for the JS picker
        return (
            f'<span class="badge {cls} ibadge" '
            f'data-skey="{_skey(comp)}" '
            f'data-comp="{comp}" '
            f'data-val="{val}" '
            f'role="button" tabindex="0" '
            f'title="{"点击更新状态" if lang == "zh" else "Tap to update status"}">'
            f'{val}</span>'
        )

    # Escape title for use in a data attribute (strip quotes / angle brackets)
    title_safe = title_display.replace('"', '&quot;').replace('<', '&lt;').replace('>', '&gt;')

    etsy_td = (
        f'<td data-label="{_t("Etsy Shop", lang)}" class="col-shop">'
        f'{r.order.etsy_shop}</td>'
        if etsy_shop_col else ""
    )

    _zh_html_compact = lang == "zh"

    # Search index — scanned by the JS live-filter (include title/order even when those cols are hidden)
    search_text = " ".join(filter(None, [
        title_display, shop_disp, stall_disp, floor_disp,
        r.item.phone_model, order_ref,
        getattr(r.order, "etsy_shop", ""),
        r.order.private_notes,
    ])).lower().replace('"', '')

    floor_td = (
        ""
        if _zh_html_compact else
        f'  <td data-label="{_t("Floor", lang)}" class="col-floor">{floor_disp}</td>\n'
    )
    product_td = (
        ""
        if _zh_html_compact else
        f'  <td data-label="{_t("Product", lang)}" class="col-product">{title_display}</td>\n'
    )
    case_td = (
        ""
        if _zh_html_compact else
        f'  <td data-label="{_t("Case", lang)}" class="col-status">{_badge(case_val, "case", has_case and not is_charm)}</td>\n'
    )
    grip_td = (
        ""
        if _zh_html_compact else
        f'  <td data-label="{_t("Grip", lang)}" class="col-status">{_badge(grip_val, "grip", has_grip and not is_charm)}</td>\n'
    )
    charm_td = (
        ""
        if _zh_html_compact else
        f'  <td data-label="{_t("Charm", lang)}" class="col-status">{_badge(charm_val, "charm", is_charm)}</td>\n'
    )
    order_td = (
        ""
        if _zh_html_compact else
        f'  <td data-label="{_t("Order #", lang)}" class="col-order">{order_ref}</td>\n'
    )
    _pn_text = r.order.private_notes
    private_notes_td = (
        f'  <td data-label="{_t("Private Notes", lang)}" class="col-private-notes">'
        f'{_pn_text}</td>\n'
        if _pn_text else
        f'  <td data-label="{_t("Private Notes", lang)}" class="col-private-notes"></td>\n'
    )

    # Embed order + norm keys on the <tr> so JS can recompute row colour
    # after any status change without re-parsing each cell's skey string.
    return (
        f'<tr class="item-row {row_cls}" '
        f'data-search="{search_text}" '
        f'data-title="{title_safe}">\n'
        f'  <td class="col-seq">{seq}</td>\n'
        f'  <td class="col-photo">{img_html}</td>\n'
        f'{floor_td}'
        f'  <td data-label="{_t("Supplier", lang)}" class="col-supplier">{shop_disp}</td>\n'
        f'  <td data-label="{_t("Stall", lang)}" class="col-stall">{stall_disp}</td>\n'
        f'{product_td}'
        f'  <td data-label="{_t("Items to Purchase", lang)}" class="col-items{" items-done" if items_label == "\u2014" else ""}">{items_label}</td>\n'
        f'{case_td}{grip_td}{charm_td}'
        f'  <td data-label="{_t("Phone Model", lang)}" class="col-model">{r.item.phone_model}</td>\n'
        f'  <td data-label="{_t("Qty", lang)}" class="col-qty">{r.item.quantity}</td>\n'
        f'{order_td}'
        f'  {etsy_td}\n'
        f'{private_notes_td}'
        f'</tr>'
    )


def _html_charm_agg_card(
    seq: int, agg: dict, row_cls: str, lang: str,
    statuses: dict | None = None,
    charm_shop_lookup: dict | None = None,
    charm_library: dict[str, CharmLibraryEntry] | None = None,
    charm_images_dir: Path | None = None,
) -> str:
    """Render one aggregated charm row as a responsive <tr> for the HTML route."""
    _statuses = statuses or {}
    code      = agg["code"]
    sku       = agg["sku"]
    notes     = agg["notes"]

    cs_lookup = charm_shop_lookup or {}
    _cs_name  = agg.get("charm_shop") or agg.get("default_shop") or ""
    _cs_obj   = cs_lookup.get(_cs_name)
    if _cs_obj:
        shop_disp  = _cs_obj.shop_name
        stall_disp = _cs_obj.stall
    elif _cs_name:
        shop_disp  = f"? {_cs_name}"
        stall_disp = "?"
    else:
        shop_disp  = "--"
        stall_disp = "--"

    unique_orders  = sorted(set(agg["orders"]))
    orders_display = ", ".join(f"#{o}" for o in unique_orders)

    preserved = _statuses.get((code, "", "charm_agg"))
    if not preserved:
        _per_order = []
        for _ri in agg["items"]:
            _ps = _statuses.get((
                _ri.order.order_number,
                _normalize(_ri.item.title)[:50],
                "charm",
            ))
            if _ps:
                _per_order.append(_ps)
        if _per_order:
            if any(s == "Out of Production" for s in _per_order):
                preserved = "Out of Production"
            elif any(s == "Out of Stock" for s in _per_order):
                preserved = "Out of Stock"
            elif all(s == "Purchased" for s in _per_order):
                preserved = "Purchased"
    charm_val = _t(preserved, lang) if preserved else _t("Pending", lang)
    na_val    = _t("N/A", lang)

    img_src = _image_data_uri(agg.get("photo_bytes"))
    img_html = (
        f'<img src="{img_src}" class="item-photo" loading="lazy" alt="">'
        if img_src else
        '<div class="item-photo no-photo"></div>'
    )

    skey = f"~C:{code}||charm_agg"

    def _badge_agg(val: str) -> str:
        cls = _status_cls(val)
        return (
            f'<span class="badge {cls} ibadge" '
            f'data-skey="{skey}" data-comp="charm" data-val="{val}" '
            f'role="button" tabindex="0" '
            f'title="{"点击更新状态" if lang == "zh" else "Tap to update status"}">'
            f'{val}</span>'
        )

    search_text = " ".join(filter(None, [
        code, sku, shop_disp, stall_disp, orders_display, notes,
    ])).lower().replace('"', '')

    _zh = lang == "zh"
    floor_td = "" if _zh else f'  <td data-label="{_t("Floor", lang)}" class="col-floor">--</td>\n'
    product_td = (
        "" if _zh else
        f'  <td data-label="Charm Code / SKU" class="col-product">'
        f'<strong>{code}</strong>{(" — " + sku) if sku else ""}</td>\n'
    )
    case_td = "" if _zh else f'  <td class="col-status"><span class="badge s-na">{na_val}</span></td>\n'
    grip_td = "" if _zh else f'  <td class="col-status"><span class="badge s-na">{na_val}</span></td>\n'
    charm_td = "" if _zh else f'  <td data-label="{_t("Charm", lang)}" class="col-status">{_badge_agg(charm_val)}</td>\n'
    order_td = "" if _zh else f'  <td data-label="Orders" class="col-order">{orders_display}</td>\n'
    etsy_td = "" if _zh else f'<td class="col-shop">{notes}</td>'
    pn_td = f'  <td class="col-private-notes"></td>\n'

    return (
        f'<tr class="item-row {row_cls}" data-search="{search_text}" data-title="{code}">\n'
        f'  <td class="col-seq">{seq}</td>\n'
        f'  <td class="col-photo">{img_html}</td>\n'
        f'{floor_td}'
        f'  <td data-label="Charm Shop" class="col-supplier">{shop_disp}</td>\n'
        f'  <td data-label="{_t("Stall", lang)}" class="col-stall">{stall_disp}</td>\n'
        f'{product_td}'
        f'  <td class="col-items">\u2014</td>\n'
        f'{case_td}{grip_td}{charm_td}'
        f'  <td class="col-model"></td>\n'
        f'  <td data-label="{_t("Qty", lang)}" class="col-qty">{agg["total_qty"]}</td>\n'
        f'{order_td}'
        f'  {etsy_td}\n'
        f'{pn_td}'
        f'</tr>'
    )


def _html_charm_await_card(
    r: ResolvedItem, seq: int, row_cls: str, lang: str,
    charm_shop_lookup: dict | None = None,
    etsy_shop_col: bool = True,
) -> str:
    """Render an awaiting-charm-code item as a <tr> for the HTML route."""
    cs_lookup  = charm_shop_lookup or {}
    assigned   = (r.supplier.charm_shop if r.supplier else "") or ""
    cs_obj     = cs_lookup.get(assigned)
    shop_disp  = cs_obj.shop_name if cs_obj else (f"? {assigned}" if assigned else "--")
    stall_disp = cs_obj.stall     if cs_obj else ("?" if assigned else "--")

    na_val   = _t("N/A", lang)
    title_fn = None
    title    = r.item.title

    img_src  = _image_data_uri(r.item.photo_bytes)
    img_html = (
        f'<img src="{img_src}" class="item-photo" loading="lazy" alt="">'
        if img_src else
        '<div class="item-photo no-photo"></div>'
    )

    await_label = "\u23f3 Awaiting Code" if lang != "zh" else "\u23f3 \u5f85\u5206\u914d"

    search_text = " ".join(filter(None, [
        title, shop_disp, stall_disp, r.item.phone_model,
        r.order.order_number, r.order.private_notes,
    ])).lower().replace('"', '')

    _zh = lang == "zh"
    floor_td   = "" if _zh else f'  <td class="col-floor">--</td>\n'
    product_td = "" if _zh else f'  <td class="col-product">{title}</td>\n'
    case_td    = "" if _zh else f'  <td class="col-status"><span class="badge s-na">{na_val}</span></td>\n'
    grip_td    = "" if _zh else f'  <td class="col-status"><span class="badge s-na">{na_val}</span></td>\n'
    charm_td   = "" if _zh else f'  <td class="col-status"><span class="badge s-pending" style="color:#7D4E00;background:#FFF3CD;font-style:italic">{await_label}</span></td>\n'
    order_td   = "" if _zh else f'  <td class="col-order">~?#{r.order.order_number}</td>\n'
    etsy_td    = f'<td class="col-shop">{r.order.etsy_shop}</td>' if etsy_shop_col and not _zh else ""
    _pn        = r.order.private_notes
    pn_td      = f'  <td class="col-private-notes">{_pn}</td>\n' if _pn else f'  <td class="col-private-notes"></td>\n'

    return (
        f'<tr class="item-row {row_cls}" data-search="{search_text}" data-title="{title}">\n'
        f'  <td class="col-seq">{seq}</td>\n'
        f'  <td class="col-photo">{img_html}</td>\n'
        f'{floor_td}'
        f'  <td class="col-supplier">{shop_disp}</td>\n'
        f'  <td class="col-stall">{stall_disp}</td>\n'
        f'{product_td}'
        f'  <td class="col-items">\u2014</td>\n'
        f'{case_td}{grip_td}{charm_td}'
        f'  <td class="col-model">{r.item.phone_model}</td>\n'
        f'  <td class="col-qty">{r.item.quantity}</td>\n'
        f'{order_td}'
        f'  {etsy_td}\n'
        f'{pn_td}'
        f'</tr>'
    )


def _row_status_cls(case_val: str, grip_val: str, charm_val: str, lang: str) -> str:
    """Return the CSS class that colours the whole row by worst status."""
    oop = _t("Out of Production", lang)
    oos = _t("Out of Stock",      lang)
    purch = _t("Purchased",       lang)
    na    = _t("N/A",             lang)
    vals  = [case_val, grip_val, charm_val]
    if oop in vals:
        return _ROW_STATUS_CSS["oop"]
    if oos in vals:
        return _ROW_STATUS_CSS["oos"]
    if all(v in (purch, na) for v in vals):
        return _ROW_STATUS_CSS["bought"]
    return ""


def _section_banner(text: str, extra_cls: str = "") -> str:
    return f'<tr class="section-banner {extra_cls}"><td colspan="99">{text}</td></tr>'


def _group_header(label: str, group_id: str, n_items: int, lang: str) -> str:
    items_word = "件" if lang == "zh" else ("item" if n_items == 1 else "items")
    return (
        f'<tr class="group-header" data-target="{group_id}">'
        f'<td colspan="99">'
        f'<span class="chevron">▼</span> {label}'
        f'<span class="group-count">{n_items} {items_word}</span>'
        f'</td></tr>'
    )


def generate_html(items: list[ResolvedItem], output: Path,
                  statuses: dict | None = None,
                  lang: str = "en",
                  title_fn=None,
                  charm_shops: list[CharmShop] | None = None,
                  charm_library: dict[str, CharmLibraryEntry] | None = None,
                  charm_images_dir: Path | None = None) -> None:
    """Generate a self-contained responsive HTML shopping route.

    On desktop: scrollable table with sticky column headers.
    On mobile (≤768 px): each row becomes a card — photo on the left,
    all fields stacked on the right with labelled rows.
    Includes live search, collapsible supplier groups, and a progress bar.
    """
    _statuses = statuses or {}

    def _has_loc(r: ResolvedItem) -> bool:
        return bool(r.supplier and (r.supplier.shop_name or r.supplier.stall))

    routable    = [r for r in items if _has_loc(r)]
    needs_info  = [r for r in items
                   if r.supplier and not _has_loc(r) and not _needs_catalog_entry(r)]
    unmatched   = [r for r in items if not r.supplier or _needs_catalog_entry(r)]
    charm_items = [r for r in items if _style_has(r.item.style)[2]]

    groups: dict[tuple[str, str], list[ResolvedItem]] = defaultdict(list)
    for r in routable:
        groups[(r.supplier.shop_name, r.supplier.stall)].append(r)
    for _gk in groups:
        groups[_gk].sort(key=_route_item_sort_key)
    sorted_keys = sorted(
        groups,
        key=lambda k: (_stall_floor(k[1]), k[1] or "\uffff", k[0]),
    )

    supplier_stops  = len({(r.supplier.shop_name, r.supplier.stall) for r in routable})
    order_count     = len({r.order.order_number for r in items})
    total_charm_qty = sum(r.item.quantity for r in charm_items)

    # Progress stats (for the progress bar)
    na_lbl    = _t("N/A",       lang)
    purch_lbl = _t("Purchased", lang)
    def _is_done(r: ResolvedItem) -> bool:
        hc, hg, hch = _style_has(r.item.style)
        norm = _normalize(r.item.title)[:50]
        onum = r.order.order_number
        def _v(comp, has):
            if not has:
                return na_lbl
            p = _statuses.get((onum, norm, comp))
            return _t(p, lang) if p else _t("Pending", lang)
        vals = [_v("case", hc), _v("grip", hg), _v("charm", hch)]
        return all(v in (purch_lbl, na_lbl) for v in vals)
    done_count  = sum(1 for r in items if _is_done(r))
    total_count = len(items)
    pct         = round(100 * done_count / total_count) if total_count else 0

    if lang == "zh":
        title_date = date.today().strftime("%Y年%m月%d日")
        page_title = f"购物路线 — {title_date}"
        stat_chips = [
            f"{total_count} 件商品",
            f"{order_count} 个订单",
            f"{supplier_stops} 个供应商",
        ]
        if charm_items:
            stat_chips.append(f"✨ {total_charm_qty} 个挂件")
        search_placeholder = "搜索商品、供应商…"
        progress_label     = f"{done_count} / {total_count} 已完成"
        collapse_all_lbl   = "全部折叠"
        expand_all_lbl     = "全部展开"
    else:
        page_title = f"Shopping Route — {date.today().strftime('%B %d, %Y')}"
        stat_chips = [
            f"{total_count} items",
            f"{order_count} orders",
            f"{supplier_stops} stops",
        ]
        if charm_items:
            stat_chips.append(f"✨ {total_charm_qty} charms")
        search_placeholder = "Search product, supplier, order #…"
        progress_label     = f"{done_count} / {total_count} done"
        collapse_all_lbl   = "Collapse all"
        expand_all_lbl     = "Expand all"

    chips_html = "".join(f'<span class="chip">{ch}</span>' for ch in stat_chips)

    # Column headers (ZH omits floor/product/component status/order# like the XLSX route sheet)
    if lang == "zh":
        col_labels = [
            "#",
            _t("Photo", lang),
            _t("Supplier", lang),
            _t("Stall", lang),
            _t("Items to Purchase", lang),
            _t("Phone Model", lang),
            _t("Qty", lang),
            _t("Private Notes", lang),
        ]
    else:
        col_labels = [
            "#",
            _t("Photo",       lang),
            _t("Floor",       lang),
            _t("Supplier",    lang),
            _t("Stall",       lang),
            _t("Product",     lang),
            _t("Items to Purchase", lang),
            _t("Case",        lang),
            _t("Grip",        lang),
            _t("Charm",       lang),
            _t("Phone Model", lang),
            _t("Qty",         lang),
            _t("Order #",     lang),
            _t("Etsy Shop", lang),
            _t("Private Notes", lang),
        ]
    thead_cells = "".join(f"<th>{h}</th>" for h in col_labels)

    # ── Build table body ──────────────────────────────────────────────────
    body_parts: list[str] = []
    seq = 1
    show_etsy = lang != "zh"  # omit Etsy Shop for Chinese output

    for gidx, key in enumerate(sorted_keys):
        shop_name, stall = key
        floor_n   = _stall_floor(stall)
        floor_lbl = f"{floor_n}F" if floor_n != 999 else "--"
        group_id  = f"grp-{gidx}"
        label     = f"{shop_name or '—'} · {stall or '—'} ({floor_lbl})"
        n_items   = len(groups[key])
        body_parts.append(_group_header(label, group_id, n_items, lang))

        for r in groups[key]:
            hc, hg, _ = _style_has(r.item.style)
            na = _t("N/A", lang)
            onum = r.order.order_number
            norm = _normalize(r.item.title)[:50]
            def _cv(comp, has, _onum=onum, _norm=norm):
                if not has:
                    return na
                p = _statuses.get((_onum, _norm, comp))
                return _t(p, lang) if p else _t("Pending", lang)
            cv = _cv("case", hc)
            gv = _cv("grip", hg)
            row_cls = _row_status_cls(cv, gv, na, lang) + f" {group_id}"
            body_parts.append(_html_item_card(
                r, seq, row_cls, lang,
                is_charm=False, title_fn=title_fn,
                statuses=_statuses,
                etsy_shop_col=show_etsy,
                charm_library=charm_library,
                charm_images_dir=charm_images_dir,
            ))
            seq += 1

    if needs_info:
        ni_label = (
            "⚠️ 目录中 — 待填供应商信息"
            if lang == "zh" else
            "⚠️ In Catalog — Awaiting Supplier Info"
        )
        body_parts.append(_section_banner(ni_label, "banner-info"))
        for ridx, r in enumerate(sorted(needs_info, key=_route_item_sort_key)):
            body_parts.append(_html_item_card(
                r, seq, "row-info", lang,
                is_charm=False, title_fn=title_fn,
                statuses=_statuses, etsy_shop_col=show_etsy,
                charm_library=charm_library,
                charm_images_dir=charm_images_dir,
            ))
            seq += 1

    if unmatched:
        um_label = (
            "‼️ 未匹配商品 — 目录中未找到供应商"
            if lang == "zh" else
            "‼️ Unmatched Items — supplier not found in catalog"
        )
        body_parts.append(_section_banner(um_label, "banner-warn"))
        for ridx, r in enumerate(sorted(unmatched, key=_route_item_sort_key)):
            body_parts.append(_html_item_card(
                r, seq, "row-warn", lang,
                is_charm=False, title_fn=title_fn,
                statuses=_statuses, etsy_shop_col=show_etsy,
                charm_library=charm_library,
                charm_images_dir=charm_images_dir,
            ))
            seq += 1

    if charm_items:
        _cshops      = charm_shops or []
        cs_lookup    = {cs.shop_name: cs for cs in _cshops}

        # Partition into coded vs awaiting
        _h_coded:   list[ResolvedItem] = []
        _h_await:   list[ResolvedItem] = []
        for _hci in charm_items:
            _hcc = (_hci.supplier.charm_code if _hci.supplier else "").strip()
            if _hcc:
                _h_coded.append(_hci)
            else:
                _h_await.append(_hci)

        # Aggregate coded items by charm code
        _h_agg: dict[str, dict] = {}
        for _hci in _h_coded:
            _hcc = _hci.supplier.charm_code.strip()
            if _hcc not in _h_agg:
                _hlib = (charm_library or {}).get(_hcc)
                _h_agg[_hcc] = {
                    "code": _hcc,
                    "sku": _hlib.sku if _hlib else "",
                    "default_shop": _hlib.default_charm_shop if _hlib else "",
                    "notes": _hlib.notes if _hlib else "",
                    "photo_bytes": None,
                    "charm_shop": "",
                    "total_qty": 0,
                    "orders": [],
                    "items": [],
                }
                _hph = charm_photo_bytes_from_folder(_hcc, charm_images_dir)
                if not _hph and _hlib and _hlib.photo_bytes:
                    _hph = _hlib.photo_bytes
                _h_agg[_hcc]["photo_bytes"] = _hph
            _h_agg[_hcc]["total_qty"] += _hci.item.quantity
            _h_agg[_hcc]["orders"].append(_hci.order.order_number)
            _h_agg[_hcc]["items"].append(_hci)
            if not _h_agg[_hcc]["charm_shop"]:
                _has = (_hci.supplier.charm_shop if _hci.supplier else "").strip()
                if _has:
                    _h_agg[_hcc]["charm_shop"] = _has

        n_missing_code = len(_h_await)
        unassigned_c = sum(1 for r in charm_items
                           if not (r.supplier and r.supplier.charm_shop
                                   and r.supplier.charm_shop in cs_lookup))

        if lang == "zh":
            charm_banner = (
                f"\u2728 \u5f85\u8d2d\u6302\u4ef6 \u2014 \u72ec\u7acb\u697c\u68cb \u2014 \u5171\u9700 {total_charm_qty} \u4e2a\u6302\u4ef6\uff0c"
                f"\u6d89\u53ca {len(charm_items)} \u4e2a\u8ba2\u5355"
            )
            if n_missing_code:
                charm_banner += f" \u25b6 {n_missing_code} \u4e2a\u5f85\u5206\u914d\u7f16\u7801"
            if unassigned_c:
                charm_banner += f" \u25b6 {unassigned_c} \u4e2a\u672a\u5206\u914d\u5e97\u94fa"
            shops_row = (
                "\u6302\u4ef6\u5e97\u94fa\uff1a" + "  |  ".join(f"{s.shop_name} ({s.stall})" for s in _cshops)
                if _cshops else "\u672a\u914d\u7f6e\u6302\u4ef6\u5e97\u94fa"
            )
        else:
            charm_banner = (
                f"\u2728 CHARMS TO PURCHASE \u2014 SEPARATE BUILDING \u2014 "
                f"{total_charm_qty} charm(s) across {len(charm_items)} order(s)"
            )
            if n_missing_code:
                charm_banner += f" \u25b6 {n_missing_code} awaiting charm code"
            if unassigned_c:
                charm_banner += f" \u25b6 {unassigned_c} unassigned shop"
            shops_row = (
                "Charm shops:  " + "   |   ".join(f"{s.shop_name} ({s.stall})" for s in _cshops)
                if _cshops else "No charm shops configured"
            )

        body_parts.append(_section_banner(charm_banner, "banner-charm"))
        body_parts.append(
            f'<tr class="charm-shops-row"><td colspan="99">{shops_row}</td></tr>'
        )
        body_parts.append(f"<tr class='thead-repeat'>{thead_cells}</tr>")

        # Sub-section A: aggregated charm cards
        sorted_h_codes = sorted(
            _h_agg,
            key=lambda c: (_h_agg[c]["charm_shop"] or "\uffff", c),
        )
        for cidx, code in enumerate(sorted_h_codes):
            agg = _h_agg[code]
            preserved = _statuses.get((code, "", "charm_agg"))
            if not preserved:
                _hpo = []
                for _ri in agg["items"]:
                    _ps = _statuses.get((_ri.order.order_number, _normalize(_ri.item.title)[:50], "charm"))
                    if _ps:
                        _hpo.append(_ps)
                if _hpo:
                    if any(s == "Out of Production" for s in _hpo):
                        preserved = "Out of Production"
                    elif any(s == "Out of Stock" for s in _hpo):
                        preserved = "Out of Stock"
                    elif all(s == "Purchased" for s in _hpo):
                        preserved = "Purchased"
            chv = _t(preserved, lang) if preserved else _t("Pending", lang)
            na  = _t("N/A", lang)
            row_cls = _row_status_cls(na, na, chv, lang) + " charm-item"
            body_parts.append(_html_charm_agg_card(
                cidx + 1, agg, row_cls, lang,
                statuses=_statuses,
                charm_shop_lookup=cs_lookup,
                charm_library=charm_library,
                charm_images_dir=charm_images_dir,
            ))

        # Sub-section B: awaiting charm code
        if _h_await:
            await_label = (
                f"\u23f3 \u5f85\u5206\u914d\u6302\u4ef6\u7f16\u7801 \u2014 {len(_h_await)} \u4e2a\u8ba2\u5355"
                if lang == "zh" else
                f"\u23f3 AWAITING CHARM CODE \u2014 {len(_h_await)} order(s)"
            )
            body_parts.append(_section_banner(await_label, "banner-warn"))
            body_parts.append(f"<tr class='thead-repeat'>{thead_cells}</tr>")
            for aidx, r in enumerate(sorted(
                _h_await,
                key=lambda ri: (_normalize(ri.item.title), ri.order.order_number),
            )):
                body_parts.append(_html_charm_await_card(
                    r, aidx + 1, "charm-item row-warn", lang,
                    charm_shop_lookup=cs_lookup,
                    etsy_shop_col=show_etsy,
                ))

    tbody_html = "\n".join(body_parts)

    # ── JS constants (language-aware, injected into the template) ────────
    import json as _json
    js_opts        = _json.dumps([
        _t("Pending", lang), _t("Purchased", lang),
        _t("Out of Stock", lang), _t("Out of Production", lang),
    ])
    js_na          = _t("N/A", lang)
    js_purch       = _t("Purchased", lang)
    js_storage_key = f"sr_statuses_{lang}_{date.today().strftime('%Y%m%d')}"
    js_comp_labels = _json.dumps({
        "case":  _t("Case",  lang),
        "grip":  _t("Grip",  lang),
        "charm": _t("Charm", lang),
    })
    if lang == "zh":
        js_toast_saved    = "✓ 已更新"
        js_done_word      = "已完成"
        js_cancel_lbl     = "取消"
        js_update_title   = "更新状态"
        js_no_results     = "没有匹配的结果"
    else:
        js_toast_saved    = "✓ Saved"
        js_done_word      = "done"
        js_cancel_lbl     = "Cancel"
        js_update_title   = "Update Status"
        js_no_results     = "No items match your search."

    # ── Full HTML document ────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="{'zh-CN' if lang == 'zh' else 'en'}">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=5.0">
<title>{page_title}</title>
<style>
/* ── Reset ── */
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

/* ── Design tokens ── */
:root {{
  --brand:        #1F4E79;
  --brand-light:  #d0e4f5;
  --brand-mid:    #4a7ead;
  --charm-dark:   #3D1359;
  --charm-light:  #EFD9FC;
  --warn-bg:      #FFF3CD;
  --warn-fg:      #856404;
  --info-bg:      #BDD7EE;
  --info-fg:      #1F4E79;
  --oop-bg:       #FFC7CE;
  --oos-bg:       #FFEB9C;
  --ok-bg:        #C6EFCE;
  --row-a:        #EBF2FA;
  --row-b:        #ffffff;
  --charm-row-a:  #F8F0FD;
  --border:       #dde4ec;
  --radius:       10px;
  --shadow:       0 2px 12px rgba(0,0,0,.10);
  --font: -apple-system,"Segoe UI","PingFang SC","Noto Sans SC","Microsoft YaHei",Arial,sans-serif;
}}

/* ── Base ── */
body {{
  font-family: var(--font);
  background: #f0f3f7;
  color: #1a1a1a;
  font-size: 14px;
  line-height: 1.5;
  padding-bottom: 80px;
}}

/* ── Sticky header ── */
.page-header {{
  background: var(--brand);
  color: #fff;
  padding: 14px 16px 10px;
  position: sticky;
  top: 0;
  z-index: 200;
  box-shadow: 0 3px 10px rgba(0,0,0,.25);
}}
.header-top {{ display: flex; align-items: baseline; gap: 10px; flex-wrap: wrap; margin-bottom: 8px; }}
.page-header h1 {{ font-size: clamp(16px,4vw,22px); font-weight: 700; letter-spacing: -.3px; }}
.chips {{ display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 8px; }}
.chip {{
  background: rgba(255,255,255,.18);
  border: 1px solid rgba(255,255,255,.3);
  border-radius: 20px; padding: 2px 10px; font-size: 12px; white-space: nowrap;
}}

/* ── Progress bar ── */
.progress-wrap {{ display: flex; align-items: center; gap: 10px; margin-bottom: 8px; }}
.progress-bar {{ flex: 1; height: 6px; background: rgba(255,255,255,.25); border-radius: 3px; overflow: hidden; }}
.progress-fill {{ height: 100%; background: #4ade80; border-radius: 3px; transition: width .4s ease; }}
.progress-label {{ font-size: 12px; opacity: .85; white-space: nowrap; }}

/* ── Search ── */
.search-wrap {{ position: relative; }}
.search-wrap svg {{ position: absolute; left: 10px; top: 50%; transform: translateY(-50%); pointer-events: none; opacity: .5; }}
#search-input {{
  width: 100%; padding: 8px 12px 8px 34px; border: none; border-radius: 8px;
  font-size: 14px; font-family: var(--font); background: rgba(255,255,255,.15);
  color: #fff; outline: none; transition: background .2s;
}}
#search-input::placeholder {{ color: rgba(255,255,255,.6); }}
#search-input:focus {{ background: rgba(255,255,255,.25); }}

/* ── Toolbar ── */
.toolbar {{ display: flex; justify-content: flex-end; gap: 8px; padding: 8px 16px 0; }}
.toolbar button {{
  background: none; border: 1px solid var(--border); border-radius: 6px;
  padding: 4px 12px; font-size: 12px; font-family: var(--font);
  cursor: pointer; color: var(--brand); transition: background .15s;
}}
.toolbar button:hover {{ background: var(--brand-light); }}

/* ── Table wrapper ── */
.table-wrap {{
  margin: 12px 12px 0; border-radius: var(--radius);
  box-shadow: var(--shadow); overflow: hidden; background: #fff;
}}

/* ── Table ── */
table {{ border-collapse: collapse; width: 100%; }}
thead th {{
  background: var(--brand); color: #fff; padding: 10px;
  font-size: 12px; font-weight: 600; white-space: nowrap;
  text-align: left; position: sticky; top: 0; z-index: 10;
}}
.thead-repeat th {{ background: var(--charm-dark); top: 0; }}
td {{ padding: 10px; border-bottom: 1px solid var(--border); vertical-align: middle; }}
tr.item-row:last-child td {{ border-bottom: none; }}

/* ── Columns ── */
.col-seq     {{ width: 32px; text-align: center; color: #888; font-size: 12px; }}
.col-photo   {{ width: 220px; text-align: center; padding: 6px; }}
.col-floor   {{ width: 44px; text-align: center; font-weight: 700; font-size: 15px; }}
.col-supplier {{ font-weight: 600; }}
.col-stall   {{ text-align: center; font-weight: 700; font-size: 15px; color: var(--brand); }}
.col-product {{ width: 110px; max-width: 130px; font-size: 12px; line-height: 1.35; color: #555; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
.col-items   {{ font-weight: 700; font-size: 13px; text-align: center; white-space: nowrap; color: var(--brand); min-width: 100px; padding: 6px 10px; background: var(--brand-light); border-radius: 6px; }}
.col-items.items-done {{ background: #f5f7fa; color: #888; font-weight: 500; }}
.col-status  {{ text-align: center; white-space: nowrap; }}
.col-model   {{ font-size: 12px; color: #555; }}
.col-qty     {{ text-align: center; font-weight: 700; font-size: 16px; }}
.col-order   {{ text-align: center; font-size: 12px; color: #555; }}
.col-shop    {{ font-size: 12px; color: #555; }}
.col-private-notes {{ font-size: 12px; color: #8B1A1A; font-style: italic; max-width: 220px; white-space: pre-wrap; word-break: break-word; }}

/* ── Photos ── */
.item-photo {{ width: 210px; height: 210px; object-fit: contain; border-radius: 6px; display: block; margin: 0 auto; }}
.no-photo   {{ width: 210px; height: 210px; background: #eee; border-radius: 6px; margin: 0 auto; }}

/* ── Status badges ── */
.badge {{
  display: inline-block; padding: 3px 9px; border-radius: 20px;
  font-size: 12px; font-weight: 600; white-space: nowrap;
}}
.s-bought  {{ background: var(--ok-bg);  color: #276221; }}
.s-oos     {{ background: var(--oos-bg); color: #7D4E00; }}
.s-oop     {{ background: var(--oop-bg); color: #9C0006; }}
.s-pending {{ background: #F0F0F0;       color: #333; }}
.s-na      {{ background: #F5F5F5;       color: #aaa; font-style: italic; }}

/* ── Interactive (tappable) badges ── */
.ibadge {{
  cursor: pointer;
  transition: filter .12s, transform .1s, box-shadow .12s;
  outline: none;
  /* Mobile: immediate tap response, no 300ms delay (critical for WeChat/QQ Browser) */
  touch-action: manipulation;
  -webkit-tap-highlight-color: transparent;
  user-select: none;
  -webkit-user-select: none;
  /* Ensure minimum tap target on mobile (44px recommended) */
  min-height: 36px;
  min-width: 44px;
  display: inline-flex;
  align-items: center;
  justify-content: center;
}}
.ibadge:hover {{
  filter: brightness(.88);
  box-shadow: 0 2px 6px rgba(0,0,0,.18);
}}
.ibadge:active {{ transform: scale(.94); }}
.ibadge:focus-visible {{
  box-shadow: 0 0 0 3px rgba(31,78,121,.4);
}}
/* Chevron cue — pointer-events: none so taps pass through to badge (fixes X5 WebView) */
.ibadge::after {{
  content: " ▾";
  font-size: 10px;
  opacity: .55;
  letter-spacing: 0;
  pointer-events: none;
}}

/* ── Row status colouring ── */
.row-oop td    {{ background: #ffe4e6 !important; }}
.row-oos td    {{ background: #fff8e0 !important; }}
.row-bought td {{ background: #e8f8ee !important; }}
.row-info td   {{ background: #EBF5FF !important; }}
.row-warn td   {{ background: var(--warn-bg) !important; }}
tr.item-row:not(.row-oop):not(.row-oos):not(.row-bought):not(.row-info):not(.row-warn) td {{ background: var(--row-a); }}
tr.item-row:nth-child(even of .item-row):not(.row-oop):not(.row-oos):not(.row-bought):not(.row-info):not(.row-warn) td {{ background: var(--row-b); }}

/* ── Group headers ── */
.group-header td {{
  background: var(--brand-light); color: var(--brand);
  font-weight: 700; font-size: 13px; padding: 8px 14px;
  cursor: pointer; user-select: none;
}}
.group-header:hover td {{ background: #c0d9ef; }}
.group-count  {{ float: right; font-weight: 400; font-size: 12px; opacity: .7; }}
.chevron      {{ display: inline-block; transition: transform .2s; margin-right: 4px; }}
.group-header.collapsed .chevron {{ transform: rotate(-90deg); }}

/* ── Section banners ── */
.section-banner td {{ font-weight: 700; font-size: 14px; padding: 10px 16px; }}
.banner-info td  {{ background: var(--info-bg); color: var(--info-fg); }}
.banner-warn td  {{ background: var(--warn-bg); color: var(--warn-fg); }}
.banner-charm td {{ background: var(--charm-dark); color: #fff; font-size: 15px; }}
.charm-shops-row td {{
  background: var(--charm-light); color: var(--charm-dark);
  font-weight: 600; font-size: 13px; padding: 7px 16px;
}}
.charm-item:not(.row-oop):not(.row-oos):not(.row-bought) td {{ background: var(--charm-row-a); }}

/* ── Utility ── */
.hidden {{ display: none !important; }}
#no-results {{ display: none; text-align: center; padding: 40px 20px; color: #888; font-size: 15px; }}

/* ── FAB ── */
#fab-top {{
  position: fixed; bottom: 22px; right: 18px;
  width: 44px; height: 44px; border-radius: 50%;
  background: var(--brand); color: #fff; border: none;
  font-size: 20px; line-height: 44px; text-align: center;
  cursor: pointer; box-shadow: 0 4px 14px rgba(0,0,0,.3);
  display: none; z-index: 300; transition: opacity .2s;
}}
#fab-top:hover {{ background: var(--brand-mid); }}

/* ══════════════════════════
   STATUS PICKER — bottom sheet
   ══════════════════════════ */
.sheet-overlay {{
  position: fixed; inset: 0;
  background: rgba(0,0,0,.45);
  z-index: 500;
  opacity: 0; pointer-events: none;
  transition: opacity .25s;
}}
.sheet-overlay.open {{ opacity: 1; pointer-events: all; }}

.status-sheet {{
  position: fixed; bottom: 0; left: 0; right: 0;
  max-width: 500px; margin: 0 auto;
  background: #fff;
  border-radius: 20px 20px 0 0;
  padding-bottom: env(safe-area-inset-bottom, 12px);
  box-shadow: 0 -6px 40px rgba(0,0,0,.22);
  transform: translateY(100%);
  transition: transform .3s cubic-bezier(.32,.72,0,1);
  z-index: 501;
}}
.status-sheet.open {{ transform: translateY(0); }}

.sheet-handle {{
  width: 40px; height: 4px; background: #ddd;
  border-radius: 2px; margin: 12px auto 0;
}}
.sheet-header {{
  padding: 14px 20px 10px;
  border-bottom: 1px solid #eee;
  display: flex; flex-direction: column; gap: 3px;
}}
.sheet-comp-label {{
  font-size: 11px; font-weight: 700; text-transform: uppercase;
  letter-spacing: .8px; color: #888;
}}
.sheet-product-title {{
  font-size: 13px; color: #333;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  max-width: 100%;
}}
.sheet-options {{
  padding: 10px 16px 16px;
  display: flex; flex-direction: column; gap: 8px;
}}
.sheet-option {{
  display: flex; align-items: center; gap: 14px;
  padding: 15px 16px; border-radius: 12px;
  border: 2px solid transparent;
  cursor: pointer; font-size: 15px; font-weight: 500;
  transition: background .15s, border-color .15s;
  font-family: var(--font); background: #f7f8fa;
  text-align: left; width: 100%;
  touch-action: manipulation;
  -webkit-tap-highlight-color: transparent;
}}
.sheet-option:hover, .sheet-option:focus {{ background: #eef2f7; outline: none; }}
.sheet-option.active {{
  border-color: var(--brand); background: #e8f0f9;
}}
.opt-radio {{
  width: 18px; height: 18px; border-radius: 50%;
  border: 2px solid #ccc; flex-shrink: 0;
  display: flex; align-items: center; justify-content: center;
  transition: border-color .15s, background .15s;
}}
.sheet-option.active .opt-radio {{
  border-color: var(--brand); background: var(--brand);
}}
.sheet-option.active .opt-radio::after {{
  content: "";
  width: 7px; height: 7px;
  background: #fff; border-radius: 50%;
}}
/* Colour hint strip on each option */
.opt-swatch {{
  width: 10px; height: 32px; border-radius: 4px; flex-shrink: 0;
}}
.swatch-pending  {{ background: #ddd; }}
.swatch-bought   {{ background: var(--ok-bg); }}
.swatch-oos      {{ background: var(--oos-bg); }}
.swatch-oop      {{ background: var(--oop-bg); }}
.sheet-cancel {{
  display: block; width: calc(100% - 32px); margin: 0 16px 12px;
  padding: 13px; border: 1px solid #ddd; border-radius: 12px;
  background: none; font-size: 15px; font-family: var(--font);
  cursor: pointer; color: #555; transition: background .15s;
}}
.sheet-cancel:hover {{ background: #f5f5f5; }}

/* ── Toast notification ── */
#toast {{
  position: fixed; bottom: 90px; left: 50%;
  transform: translateX(-50%) translateY(12px);
  background: #1a1a1a; color: #fff;
  padding: 10px 22px; border-radius: 22px;
  font-size: 13px; font-weight: 500;
  opacity: 0; pointer-events: none;
  transition: opacity .2s, transform .2s;
  z-index: 600; white-space: nowrap;
  box-shadow: 0 4px 16px rgba(0,0,0,.25);
}}
#toast.show {{ opacity: 1; transform: translateX(-50%) translateY(0); }}

/* ══════════════════════════════════════════
   MOBILE  ≤ 768 px  →  CARD LAYOUT
   ══════════════════════════════════════════ */
@media (max-width: 768px) {{
  .table-wrap {{ margin: 8px 8px 0; border-radius: 8px; }}
  thead {{ display: none; }}
  .thead-repeat {{ display: none; }}
  table, tbody, tr, td {{ display: block; width: 100%; }}

  tr.item-row {{
    position: relative;
    margin-bottom: 2px;
    padding: 10px 12px 10px 235px;
    min-height: 225px;
    border-bottom: 2px solid var(--border);
  }}
  tr.item-row:last-child {{ border-bottom: none; }}

  .col-photo {{
    position: absolute; top: 10px; left: 8px;
    width: 210px; padding: 0; text-align: left;
  }}
  .item-photo, .no-photo {{ width: 210px; height: 210px; }}
  .col-seq {{ display: none; }}

  td:not(.col-photo):not(.col-seq) {{
    padding: 3px 0; border: none; background: transparent !important;
    font-size: 13px; display: flex; align-items: center; gap: 6px;
  }}
  td::before {{
    content: attr(data-label);
    flex-shrink: 0; font-size: 11px; font-weight: 600;
    color: #888; min-width: 54px;
    text-transform: uppercase; letter-spacing: .4px;
  }}

  .col-product {{
    font-size: 12px; color: #555; max-width: 100%;
  }}
  .col-product::before {{ font-size: 10px; }}
  .col-items {{
    font-weight: 700; background: var(--brand-light); padding: 8px 12px;
    border-radius: 6px; display: block; text-align: center;
  }}
  .col-items::before {{ font-size: 11px; }}

  .col-floor, .col-stall {{ display: inline-flex !important; width: auto !important; }}
  .col-floor {{ margin-right: 12px; }}
  .col-qty {{ font-size: 18px; }}
  .col-status {{ flex-wrap: wrap; gap: 4px; }}

  /* Make badges larger on mobile for easier tapping */
  .badge {{ padding: 6px 14px; font-size: 13px; }}

  .group-header td {{ font-size: 14px; padding: 10px 14px; }}

  /* Row colouring on mobile: apply to <tr> not <td> */
  .row-oop   {{ background: #ffe4e6; }}
  .row-oos   {{ background: #fff8e0; }}
  .row-bought {{ background: #e8f8ee; }}
  .row-info  {{ background: #EBF5FF; }}
  .row-warn  {{ background: var(--warn-bg); }}
  .row-oop td, .row-oos td, .row-bought td,
  .row-info td, .row-warn td {{ background: transparent !important; }}
  tr.item-row:not(.row-oop):not(.row-oos):not(.row-bought):not(.row-info):not(.row-warn) {{ background: #fff; }}
}}
</style>
</head>
<body>

<!-- ── Sticky header ── -->
<div class="page-header">
  <div class="header-top"><h1>{page_title}</h1></div>
  <div class="chips">{chips_html}</div>
  <div class="progress-wrap">
    <div class="progress-bar">
      <div class="progress-fill" id="prog-fill" style="width:{pct}%"></div>
    </div>
    <span class="progress-label" id="prog-label">{progress_label}</span>
  </div>
  <div class="search-wrap">
    <svg width="16" height="16" fill="none" stroke="white" stroke-width="2"
         viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
    <input id="search-input" type="search" autocomplete="off"
           placeholder="{search_placeholder}" oninput="filterRows()">
  </div>
</div>

<!-- ── Toolbar ── -->
<div class="toolbar">
  <button onclick="collapseAll()">{collapse_all_lbl}</button>
  <button onclick="expandAll()">{expand_all_lbl}</button>
</div>

<!-- ── Table ── -->
<div class="table-wrap">
  <table>
    <thead><tr>{thead_cells}</tr></thead>
    <tbody id="main-tbody">
{tbody_html}
    </tbody>
  </table>
  <div id="no-results">{js_no_results}</div>
</div>

<!-- ── Status picker bottom sheet ── -->
<div id="sheet-overlay" class="sheet-overlay" onclick="closePicker()" ontouchend="closePicker()"></div>
<div id="status-sheet" class="status-sheet" role="dialog" aria-modal="true">
  <div class="sheet-handle"></div>
  <div class="sheet-header">
    <span class="sheet-comp-label" id="sheet-comp-label"></span>
    <span class="sheet-product-title" id="sheet-product-title"></span>
  </div>
  <div class="sheet-options" id="sheet-options"></div>
  <button class="sheet-cancel" onclick="closePicker()">{js_cancel_lbl}</button>
</div>

<!-- ── Toast ── -->
<div id="toast"></div>

<!-- ── Scroll-to-top FAB ── -->
<button id="fab-top" onclick="window.scrollTo({{top:0,behavior:'smooth'}})" title="Back to top">↑</button>

<script>
// ── Configuration (injected by Python) ─────────────────────────────────────
var STATUS_OPTIONS  = {js_opts};          // [Pending, Purchased, OOS, OOP]
var STATUS_NA       = "{js_na}";
var STATUS_PURCH    = "{js_purch}";
var STORAGE_KEY     = "{js_storage_key}";
var COMP_LABELS     = {js_comp_labels};
var DONE_WORD       = "{js_done_word}";
var TOAST_SAVED     = "{js_toast_saved}";
var UPDATE_TITLE    = "{js_update_title}";

// Map status value → CSS class name
var STATUS_CLS = {{}};
STATUS_CLS[STATUS_OPTIONS[0]] = 's-pending';
STATUS_CLS[STATUS_OPTIONS[1]] = 's-bought';
STATUS_CLS[STATUS_OPTIONS[2]] = 's-oos';
STATUS_CLS[STATUS_OPTIONS[3]] = 's-oop';
STATUS_CLS[STATUS_NA]         = 's-na';

// Map status value → swatch class
var STATUS_SWATCH = {{}};
STATUS_SWATCH[STATUS_OPTIONS[0]] = 'swatch-pending';
STATUS_SWATCH[STATUS_OPTIONS[1]] = 'swatch-bought';
STATUS_SWATCH[STATUS_OPTIONS[2]] = 'swatch-oos';
STATUS_SWATCH[STATUS_OPTIONS[3]] = 'swatch-oop';

// ── localStorage helpers ────────────────────────────────────────────────────
function _loadAll() {{
  try {{ return JSON.parse(localStorage.getItem(STORAGE_KEY) || '{{}}'); }} catch(e) {{ return {{}}; }}
}}
function _saveAll(data) {{
  try {{ localStorage.setItem(STORAGE_KEY, JSON.stringify(data)); }} catch(e) {{}}
}}
function saveStatus(skey, val) {{
  var d = _loadAll(); d[skey] = val; _saveAll(d);
}}

// ── Apply a value to a badge DOM element ───────────────────────────────────
function applyBadge(badge, val) {{
  badge.textContent = val + ' ▾';   // keep the caret
  badge.dataset.val = val;
  badge.className = badge.className.replace(/\\bs-\\S+/g, '').trim();
  badge.classList.add(STATUS_CLS[val] || 's-pending');
  // Strip extra trailing space before ▾ inserted by the ::after pseudo
  badge.textContent = val;   // real text; ::after adds the ▾ via CSS
}}

// ── Row colour update (worst-component logic) ──────────────────────────────
function updateRowStatus(tr) {{
  var badges = tr.querySelectorAll('.col-status .badge');
  var vals = Array.from(badges).map(function(b) {{ return b.dataset.val || b.textContent.trim(); }});
  var OOP   = STATUS_OPTIONS[3];
  var OOS   = STATUS_OPTIONS[2];
  tr.classList.remove('row-oop', 'row-oos', 'row-bought');
  if (!vals.length) return;
  if (vals.indexOf(OOP) !== -1) {{
    tr.classList.add('row-oop');
  }} else if (vals.indexOf(OOS) !== -1) {{
    tr.classList.add('row-oos');
  }} else if (vals.every(function(v) {{ return v === STATUS_PURCH || v === STATUS_NA; }})) {{
    tr.classList.add('row-bought');
  }}
}}

// ── Progress bar ───────────────────────────────────────────────────────────
function updateProgress() {{
  var rows  = document.querySelectorAll('.item-row');
  var total = rows.length;
  var done  = 0;
  rows.forEach(function(tr) {{
    var badges = tr.querySelectorAll('.col-status .badge');
    var vals   = Array.from(badges).map(function(b) {{ return b.dataset.val || b.textContent.trim(); }});
    if (!vals.length) return;
    if (vals.every(function(v) {{ return v === STATUS_PURCH || v === STATUS_NA; }})) done++;
  }});
  var pct = total ? Math.round(100 * done / total) : 0;
  var fill  = document.getElementById('prog-fill');
  var label = document.getElementById('prog-label');
  if (fill)  fill.style.width  = pct + '%';
  if (label) label.textContent = done + ' / ' + total + ' ' + DONE_WORD;
}}

// ── Load saved statuses on page open ──────────────────────────────────────
function loadSavedStatuses() {{
  var saved = _loadAll();
  document.querySelectorAll('.ibadge').forEach(function(badge) {{
    var skey = badge.dataset.skey;
    if (skey && saved[skey] !== undefined) {{
      applyBadge(badge, saved[skey]);
    }}
  }});
  document.querySelectorAll('.item-row').forEach(updateRowStatus);
  updateProgress();
}}

// ── Status picker (bottom sheet) ───────────────────────────────────────────
var _activeBadge = null;

function openPicker(badge) {{
  _activeBadge = badge;
  var comp        = badge.dataset.comp || '';
  var currentVal  = badge.dataset.val  || badge.textContent.trim();
  var row         = badge.closest('.item-row');
  var title       = row ? (row.dataset.title || '') : '';

  document.getElementById('sheet-comp-label').textContent    = COMP_LABELS[comp] || comp;
  document.getElementById('sheet-product-title').textContent = title;

  var container = document.getElementById('sheet-options');
  container.innerHTML = '';
  STATUS_OPTIONS.forEach(function(opt) {{
    var btn = document.createElement('button');
    btn.className = 'sheet-option' + (opt === currentVal ? ' active' : '');
    btn.setAttribute('type', 'button');
    // Radio circle
    var radio = document.createElement('span');
    radio.className = 'opt-radio';
    // Colour swatch
    var swatch = document.createElement('span');
    swatch.className = 'opt-swatch ' + (STATUS_SWATCH[opt] || '');
    // Label
    var lbl = document.createElement('span');
    lbl.textContent = opt;
    btn.appendChild(radio);
    btn.appendChild(swatch);
    btn.appendChild(lbl);
    function doPick() {{ pickStatus(opt); }}
    btn.addEventListener('touchend', function(e) {{ e.preventDefault(); doPick(); }}, {{ passive: false }});
    btn.addEventListener('click', doPick);
    container.appendChild(btn);
  }});

  var overlay = document.getElementById('sheet-overlay');
  var sheet   = document.getElementById('status-sheet');
  overlay.classList.add('open');
  sheet.classList.add('open');
  // Focus first option for keyboard accessibility
  var first = container.querySelector('.sheet-option');
  if (first) setTimeout(function() {{ first.focus(); }}, 60);
}}

function closePicker() {{
  document.getElementById('sheet-overlay').classList.remove('open');
  document.getElementById('status-sheet').classList.remove('open');
  if (_activeBadge) {{ _activeBadge.focus(); }}
  _activeBadge = null;
}}

function pickStatus(val) {{
  if (!_activeBadge) return;
  var skey = _activeBadge.dataset.skey;
  applyBadge(_activeBadge, val);
  saveStatus(skey, val);
  var row = _activeBadge.closest('.item-row');
  if (row) updateRowStatus(row);
  updateProgress();
  closePicker();
  showToast(TOAST_SAVED + ': ' + val);
}}

// Close sheet on Escape key
document.addEventListener('keydown', function(e) {{
  if (e.key === 'Escape') closePicker();
}});

// ── Toast ──────────────────────────────────────────────────────────────────
var _toastTimer = null;
function showToast(msg) {{
  var t = document.getElementById('toast');
  t.textContent = msg;
  t.classList.add('show');
  clearTimeout(_toastTimer);
  _toastTimer = setTimeout(function() {{ t.classList.remove('show'); }}, 2200);
}}

// ── Attach tap/click + keyboard handlers to interactive badges ─────────────
// Mobile (WeChat, QQ Browser, X5 WebView): click often does NOT fire on touch.
// We MUST use touchend to get reliable tap response. preventDefault stops the
// delayed click, so we handle both touch and mouse in one path.
function handleBadgeOpen(badge, e) {{
  if (e) {{ e.preventDefault(); e.stopPropagation(); }}
  var now = Date.now();
  if (typeof handleBadgeOpen._last === 'number' && now - handleBadgeOpen._last < 350) return;
  handleBadgeOpen._last = now;
  openPicker(badge);
}}
document.querySelectorAll('.ibadge').forEach(function(badge) {{
  badge.addEventListener('touchend', function(e) {{
    handleBadgeOpen(badge, e);
  }}, {{ passive: false }});
  badge.addEventListener('click', function(e) {{
    handleBadgeOpen(badge, e);
  }});
  badge.addEventListener('keydown', function(e) {{
    if (e.key === 'Enter' || e.key === ' ') {{ e.preventDefault(); handleBadgeOpen(badge); }}
  }});
}});

// ── Collapsible supplier groups ────────────────────────────────────────────
document.querySelectorAll('.group-header').forEach(function(hdr) {{
  hdr.addEventListener('click', function() {{
    var gid       = hdr.dataset.target;
    var collapsed = hdr.classList.toggle('collapsed');
    document.querySelectorAll('tr.' + gid).forEach(function(row) {{
      row.classList.toggle('hidden', collapsed);
    }});
  }});
}});

function collapseAll() {{
  document.querySelectorAll('.group-header').forEach(function(hdr) {{
    var gid = hdr.dataset.target;
    hdr.classList.add('collapsed');
    document.querySelectorAll('tr.' + gid).forEach(function(r) {{ r.classList.add('hidden'); }});
  }});
}}
function expandAll() {{
  document.querySelectorAll('.group-header').forEach(function(hdr) {{
    var gid = hdr.dataset.target;
    hdr.classList.remove('collapsed');
    document.querySelectorAll('tr.' + gid).forEach(function(r) {{ r.classList.remove('hidden'); }});
  }});
}}

// ── Live search ────────────────────────────────────────────────────────────
function filterRows() {{
  var q    = document.getElementById('search-input').value.trim().toLowerCase();
  var rows = document.querySelectorAll('tr.item-row');
  var vis  = 0;
  rows.forEach(function(row) {{
    var match = !q || (row.dataset.search || '').indexOf(q) !== -1;
    row.classList.toggle('hidden', !match);
    if (match) vis++;
  }});
  document.getElementById('no-results').style.display = (!vis && q) ? 'block' : 'none';
  if (q) expandAll();
}}

// ── Scroll-to-top FAB ──────────────────────────────────────────────────────
window.addEventListener('scroll', function() {{
  document.getElementById('fab-top').style.display = window.scrollY > 300 ? 'block' : 'none';
}});

// ── Init ───────────────────────────────────────────────────────────────────
loadSavedStatuses();
</script>
</body>
</html>"""

    output.write_text(html, encoding="utf-8")
    log.info("HTML saved -> %s", output.resolve())


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Generate a shopping route from Etsy order PDFs + supplier catalog."
    )
    ap.add_argument("pdfs", nargs="*", help="PDF files (default: all *.pdf in cwd or input/)")
    ap.add_argument(
        "--project-dir",
        default="",
        metavar="DIR",
        help=(
            "Project root for organized layout. When set, uses data/, input/, "
            "output/, cache/ subdirs. Run from project root; omit for legacy flat layout."
        ),
    )
    ap.add_argument("--catalog",  default=CATALOG_FILE, help="Supplier catalog .xlsx")
    ap.add_argument("--output",   default=OUTPUT_FILE,  help="Output .xlsx path")
    ap.add_argument("--cache",    default=CACHE_FILE,   help="Order cache .json path")
    ap.add_argument(
        "--threshold", type=int, default=MATCH_THRESHOLD,
        help="Fuzzy-match score cutoff 0-100 (default %(default)s)",
    )
    ap.add_argument(
        "--no-catalog-update", action="store_true",
        help="Skip writing new products back to the supplier catalog",
    )
    ap.add_argument(
        "--reset", action="store_true",
        help="Ignore existing cache and rebuild from the supplied PDFs only",
    )
    ap.add_argument(
        "--purge-purchased", action="store_true",
        help=(
            "Read the statuses already entered in shopping_route.xlsx and drop "
            "every item whose ALL active components are 'Purchased' or 'Out of Production'. "
            "Out-of-Production items are recorded to out_of_production_log.csv. "
            "Items with any component still 'Out of Stock' or 'Pending' "
            "are kept. The cache is trimmed to match so purged orders "
            "never reappear. Run this after returning from a shopping trip."
        ),
    )
    ap.add_argument(
        "--new-batch", action="store_true",
        help=(
            "Add a new batch of today's order PDFs on top of existing orders. "
            "Only PDFs whose filenames have NOT been seen in a previous run are "
            "parsed; all others are skipped (their orders are already in the cache). "
            "After processing, every newly-ingested PDF name is recorded in the "
            "cache so it is skipped automatically on future runs. "
            "Use this each time you download a fresh set of order PDFs and want "
            "to merge them into the existing shopping route without re-scanning "
            "old files. Pass specific PDFs as arguments or let the script "
            "auto-discover all *.pdf files in the current directory. "
            "If shopping_route.xlsx is missing from the output folder, every "
            "PDF in input/ is processed (same as a clean rebuild from PDFs only)."
        ),
    )
    ap.add_argument(
        "--refresh-catalog", action="store_true",
        help=(
            "Re-read supplier_catalog.xlsx and regenerate shopping_route.xlsx "
            "without processing any PDF files. Use this after you have filled in "
            "Shop Name / Stall for amber (new-product) rows in supplier_catalog.xlsx. "
            "Every order in the cache is re-matched against the updated catalog so "
            "items that previously lacked supplier info are promoted to the correct "
            "floor in the shopping route automatically."
        ),
    )
    ap.add_argument(
        "--rebuild-catalog", action="store_true",
        help=(
            "Re-sort supplier_catalog.xlsx in-place so that every row with Shop "
            "Name / Stall filled in is grouped by shop → stall → title, "
            "and amber rows still awaiting info are placed at the end.  All product "
            "photos are re-attached at their new positions.  A timestamped backup is "
            "saved under data/supplier_catalog_backups/ before changes.  Run this after filling in supplier "
            "details for previously-unknown products.  Also refreshes the Suppliers "
            "sheet from unique Product Map shop+stall pairs."
        ),
    )
    ap.add_argument(
        "--sync-suppliers", action="store_true",
        help=(
            "Update the Suppliers sheet from Product Map: append one row per distinct "
            "non-empty Shop Name (col D) + Stall (col E) pair.  Never duplicates an "
            "existing pair; leaves Mall, Floor, Address, etc. unchanged.  Saves the "
            "workbook in place (no PDF processing)."
        ),
    )
    ap.add_argument(
        "--clear-product-map-charm-codes",
        action="store_true",
        help=(
            "Clear Product Map column G (Charm Code) on every product row and save "
            "supplier_catalog.xlsx. Use when H contains wrong text (e.g. shop names); "
            "re-select only valid codes from the Charm Library list (dropdown)."
        ),
    )
    ap.add_argument(
        "--renumber-charms",
        action="store_true",
        help=(
            "Renumber Charm Library codes to match their current top-to-bottom row order "
            "in the sheet.  Typical use: after reordering rows in Excel (e.g. inserting a "
            "new charm between two existing ones), run this to reassign clean sequential "
            "codes (CH-00001, CH-00002 ...), update every Product Map Charm Code reference, "
            "and rename on-disk charm image files.  "
            "Pair with --renumber-charms-dry-run to preview the mapping first."
        ),
    )
    ap.add_argument(
        "--renumber-charms-dry-run",
        action="store_true",
        help=(
            "With --renumber-charms: print the planned old -> new code mapping without "
            "writing any changes to the workbook or disk."
        ),
    )
    ap.add_argument(
        "--mark-product-discontinued",
        default="",
        metavar="TITLE",
        help=(
            "Move the matching product row from Product Map to the Discontinued Products "
            "sheet (with timestamp and photo). Match by exact title after normalization, "
            "or a unique case-insensitive substring. "
            "Then run --refresh-catalog. See also --mark-product-discontinued-row."
        ),
    )
    ap.add_argument(
        "--mark-product-discontinued-row",
        type=int,
        default=0,
        metavar="N",
        help=(
            "Move Excel row N from Product Map to Discontinued Products (data rows start at 2). "
            "Alternative to --mark-product-discontinued when you know the row number."
        ),
    )
    ap.add_argument(
        "--init-charm-shops", action="store_true",
        help=(
            "Add the 'Charm Shops' reference sheet to supplier_catalog.xlsx if it "
            "does not already exist, pre-populated with the default charm shops. "
            "Run this once after first installing this update, or any time the sheet "
            "is accidentally deleted. Existing data is never overwritten."
        ),
    )
    ap.add_argument(
        "--chinese", action="store_true",
        help=(
            "Also generate a Simplified Chinese version of the shopping route. "
            "The Chinese file is saved alongside the English one with '_zh' appended "
            "to the filename (e.g. shopping_route_zh.xlsx). Column layout, sheet "
            "structure (Shopping Route, Orders Detail, Summary), and all tabs match "
            "the English version; only the language differs (Simplified Chinese). "
            "Product titles are auto-translated via Google Translate (requires "
            "internet; translations are cached locally)."
        ),
    )
    ap.add_argument(
        "--chinese-exclude-shops",
        metavar="SHOPS",
        default="",
        help=(
            "Comma-separated list of Etsy shop names to exclude from the Chinese "
            "shopping route (case-insensitive). Useful when certain shops are "
            "handled directly by you and do not need to be passed to the Chinese "
            "employee.  Example: --chinese-exclude-shops Y2KASEofficial,MyShop"
        ),
    )
    ap.add_argument(
        "--html", action="store_true",
        help=(
            "Also generate a self-contained HTML file alongside the Excel file. "
            "The HTML file embeds all product photos at full quality and is "
            "optimised for mobile browsers — ideal for sharing with employees "
            "who shop in person. The file is saved with the same base name as "
            "the Excel output (e.g. shopping_route.html). "
            "Combine with --chinese to also produce shopping_route_zh.html."
        ),
    )
    ap.add_argument(
        "--charm-images-dir",
        default="",
        metavar="DIR",
        help=(
            "Directory of charm photos named <Charm Code>.png|.jpg|.jpeg|.webp "
            "(overrides Excel embeds). With --project-dir, defaults to "
            "data/charm_images/ under the project root."
        ),
    )
    ap.add_argument(
        "--export-charm-manifest",
        action="store_true",
        help=(
            "Only write charm_manifest.json (Charm Library + charm_images/) and exit. "
            "On normal runs the same file is updated automatically unless --no-charm-manifest."
        ),
    )
    ap.add_argument(
        "--charm-manifest-output",
        default="",
        metavar="FILE",
        help=(
            "Output path for --export-charm-manifest "
            "(default: data/charm_manifest.json with --project-dir, else next to the catalog)."
        ),
    )
    ap.add_argument(
        "--no-charm-manifest",
        action="store_true",
        help=(
            "On a normal run, skip writing charm_manifest.json (still written when using "
            "--export-charm-manifest unless you pass a custom path there)."
        ),
    )
    ap.add_argument(
        "--import-charm-images",
        action="store_true",
        help=(
            "Rename charm image files matching --import-charm-pattern (default Screenshot*.png; "
            "repeat the flag for multiple globs) "
            "to the next free Charm Codes under --import-charm-prefix, append Charm Library rows "
            "with embedded photos, and save supplier_catalog.xlsx.  Scans subfolders.  "
            "Optional: --import-charm-vision-sku + vision API (CHARM_VISION_API_KEY / OPENAI_API_KEY "
            "or local Ollama) to suggest SKU (column C).  "
            "Pair with --import-charm-dry-run to preview renames only."
        ),
    )
    ap.add_argument(
        "--import-charm-dry-run",
        action="store_true",
        help="With --import-charm-images: list planned renames only; no disk or catalog changes.",
    )
    ap.add_argument(
        "--import-charm-pattern",
        action="append",
        dest="import_charm_patterns",
        metavar="GLOB",
        default=None,
        help=(
            "Basename glob for --import-charm-images (repeat for multiple patterns). "
            "If omitted, defaults to Screenshot*.png. "
            "Matched recursively under the charm images folder."
        ),
    )
    ap.add_argument(
        "--import-charm-prefix",
        default="CH-",
        metavar="PREFIX",
        help=(
            "Charm code prefix for new files (default %(default)s).  Numeric suffix is zero-padded: "
            "matches width of existing PREFIX+digits rows/files, else at least 5 digits (CH-00001), "
            "and widens automatically past 9999."
        ),
    )
    ap.add_argument(
        "--import-charm-vision-sku", "--import-charm-vision-names",
        dest="import_charm_vision_sku",
        action="store_true",
        help=(
            "With --import-charm-images: fill Charm Library **SKU** (C) using a vision-capable "
            "OpenAI-compatible chat model (default gpt-4o-mini on api.openai.com).  "
            "Set CHARM_VISION_API_KEY or OPENAI_API_KEY, or point --openai-base-url / "
            "CHARM_VISION_BASE_URL at OpenRouter, Ollama (e.g. http://127.0.0.1:11434/v1), LM Studio, etc. "
            "Review every label.  Not used during --import-charm-dry-run."
        ),
    )
    ap.add_argument(
        "--openai-api-key",
        default="",
        metavar="KEY",
        help=(
            "API key for charm vision SKU (Bearer).  If unset, uses CHARM_VISION_API_KEY, "
            "then OPENAI_API_KEY, then OPENROUTER_API_KEY.  Omit for some local servers (localhost)."
        ),
    )
    ap.add_argument(
        "--openai-model",
        default="gpt-4o-mini",
        metavar="MODEL",
        help="Vision model id for OpenAI-compatible /chat/completions (default %(default)s).",
    )
    ap.add_argument(
        "--openai-base-url",
        default="",
        metavar="URL",
        help=(
            "OpenAI-compatible API base URL ending in /v1, e.g. https://api.openai.com/v1, "
            "https://openrouter.ai/api/v1, http://127.0.0.1:11434/v1 (Ollama).  "
            "If unset, uses CHARM_VISION_BASE_URL, OPENAI_BASE_URL, else api.openai.com."
        ),
    )
    ap.add_argument(
        "--fill-charm-sku", "--fill-charm-display-names",
        dest="fill_charm_sku",
        action="store_true",
        help=(
            "Fill empty Charm Library **SKU** (C) via OpenAI-compatible vision API, using "
            "data/charm_images/<Charm Code>.ext when present else the embedded row photo. "
            "Requires CHARM_VISION_API_KEY / OPENAI_API_KEY or --openai-api-key unless the "
            "base URL is local (Ollama/LM Studio).  Not required with --fill-charm-sku-dry-run."
        ),
    )
    ap.add_argument(
        "--fill-charm-sku-dry-run", "--fill-charm-display-names-dry-run",
        dest="fill_charm_sku_dry_run",
        action="store_true",
        help="With --fill-charm-sku: list target rows only; no API calls or save.",
    )
    ap.add_argument(
        "--fill-charm-sku-overwrite", "--fill-charm-display-names-overwrite",
        dest="fill_charm_sku_overwrite",
        action="store_true",
        help="With --fill-charm-sku: replace SKU even when column C is not empty.",
    )
    ap.add_argument(
        "--list-catalog-backups",
        action="store_true",
        help=(
            "List timestamped supplier_catalog.xlsx backups (newest first) under "
            "data/supplier_catalog_backups/ and exit."
        ),
    )
    ap.add_argument(
        "--restore-catalog-backup",
        default="",
        metavar="PATH",
        help=(
            "Replace supplier_catalog.xlsx with this backup. The current file is copied "
            "into the backups folder first (reason before_restore). PATH may be absolute "
            "or relative to supplier_catalog_backups/."
        ),
    )
    ap.add_argument(
        "--restore-latest-catalog-backup",
        action="store_true",
        help=(
            "Replace supplier_catalog.xlsx with the newest backup in "
            "supplier_catalog_backups/ (current file is backed up first)."
        ),
    )
    args = ap.parse_args()

    if not args.import_charm_patterns:
        args.import_charm_patterns = ["Screenshot*.png"]

    if args.import_charm_dry_run and not args.import_charm_images:
        ap.error("--import-charm-dry-run requires --import-charm-images.")

    if args.import_charm_vision_sku and not args.import_charm_images:
        ap.error("--import-charm-vision-sku requires --import-charm-images.")

    if args.fill_charm_sku_dry_run and not args.fill_charm_sku:
        ap.error("--fill-charm-sku-dry-run requires --fill-charm-sku.")

    if args.fill_charm_sku_overwrite and not args.fill_charm_sku:
        ap.error("--fill-charm-sku-overwrite requires --fill-charm-sku.")

    if args.fill_charm_sku and args.import_charm_images:
        ap.error("Run --fill-charm-sku and --import-charm-images in separate invocations.")

    if args.renumber_charms_dry_run and not args.renumber_charms:
        ap.error("--renumber-charms-dry-run requires --renumber-charms.")

    if (args.restore_catalog_backup or "").strip() and args.restore_latest_catalog_backup:
        ap.error(
            "Use only one of --restore-catalog-backup and --restore-latest-catalog-backup."
        )

    if args.refresh_catalog and args.new_batch:
        ap.error("--refresh-catalog and --new-batch are mutually exclusive.")

    if args.rebuild_catalog and args.new_batch:
        ap.error("--rebuild-catalog and --new-batch are mutually exclusive.")

    if args.mark_product_discontinued_row and args.mark_product_discontinued.strip():
        ap.error(
            "Use only one of --mark-product-discontinued and --mark-product-discontinued-row."
        )

    # Resolve paths: --project-dir enables organized layout (data/, input/, output/, cache/)
    if args.project_dir:
        proj = Path(args.project_dir).resolve()
        catalog_path = proj / "data" / "supplier_catalog.xlsx"
        output_path  = proj / "output" / "shopping_route.xlsx"
        cache_path   = proj / "cache" / "orders_cache.json"
        input_dir    = proj / "input"
        oop_log_path = proj / "cache" / "out_of_production_log.csv"
        trans_cache_path = proj / "cache" / "translations_zh_cache.json"
        charm_images_dir = (
            Path(args.charm_images_dir).resolve()
            if args.charm_images_dir.strip()
            else proj / "data" / CHARM_IMAGES_DIR_NAME
        )
    else:
        catalog_path = Path(args.catalog)
        output_path  = Path(args.output)
        cache_path   = Path(args.cache)
        input_dir    = Path(".")
        oop_log_path = Path(OOP_LOG_FILE)
        trans_cache_path = Path(ZH_TRANS_CACHE)
        charm_images_dir = (
            Path(args.charm_images_dir).resolve()
            if args.charm_images_dir.strip()
            else (Path("data") / CHARM_IMAGES_DIR_NAME).resolve()
        )

    charm_manifest_default_path = (
        (Path(args.project_dir).resolve() / "data" / CHARM_MANIFEST_FILE)
        if args.project_dir.strip()
        else (catalog_path.parent / CHARM_MANIFEST_FILE)
    ).resolve()

    try:
        charm_images_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        log.warning("Could not create charm images directory %s: %s", charm_images_dir, exc)

    # ------------------------------------------------------------------ #
    # --list-catalog-backups / --restore-catalog-backup                  #
    # ------------------------------------------------------------------ #
    if args.list_catalog_backups:
        bdir = catalog_backup_dir(catalog_path)
        backups = list_supplier_catalog_backups(catalog_path)
        print(f"\nSupplier catalog backups under:\n  {bdir.resolve()}\n")
        if not backups:
            print(
                "  (none yet - backups are created automatically before the tool "
                "writes the catalog)\n"
            )
        else:
            for p in backups:
                print(f"  {p.name}")
            print()
        sys.exit(0)

    if args.restore_latest_catalog_backup or (args.restore_catalog_backup or "").strip():
        if args.restore_latest_catalog_backup:
            backups = list_supplier_catalog_backups(catalog_path)
            if not backups:
                log.error(
                    "No backups found in %s",
                    catalog_backup_dir(catalog_path),
                )
                sys.exit(1)
            chosen = backups[0]
        else:
            raw = (args.restore_catalog_backup or "").strip()
            chosen = Path(raw)
            if not chosen.is_absolute():
                chosen = catalog_backup_dir(catalog_path) / raw
            chosen = chosen.resolve()
            if not chosen.is_file():
                log.error("Backup file not found: %s", chosen)
                sys.exit(1)
        try:
            restore_supplier_catalog(catalog_path, chosen)
        except OSError as exc:
            log.error("%s", exc)
            sys.exit(1)
        print(f"\n{'=' * 60}")
        print(f"  [OK]  Restored supplier catalog from:\n        {chosen}")
        print(f"  [>>]  Active file:  {catalog_path.resolve()}")
        print(f"{'=' * 60}\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --mark-product-discontinued / --mark-product-discontinued-row        #
    # ------------------------------------------------------------------ #
    if args.mark_product_discontinued_row or args.mark_product_discontinued.strip():
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        try:
            if args.mark_product_discontinued_row:
                if args.mark_product_discontinued_row < 2:
                    ap.error("--mark-product-discontinued-row must be >= 2 (row 1 is the header).")
                title = mark_product_map_discontinued_by_row(
                    catalog_path, args.mark_product_discontinued_row
                )
                print(
                    f"\n  [OK]  Product moved to '{DISCONTINUED_SHEET}' sheet.\n"
                    f"        {title[:120]}\n"
                )
            else:
                row_num, title = mark_product_map_discontinued(
                    catalog_path, args.mark_product_discontinued.strip()
                )
                print(
                    f"\n  [OK]  Product moved to '{DISCONTINUED_SHEET}' sheet.\n"
                    f"        {title[:120]}\n"
                )
        except (ValueError, OSError) as e:
            log.error("%s", e)
            sys.exit(1)
        print("  [>>]  Run with --refresh-catalog to update the shopping route.\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --fill-charm-sku: vision labels for existing library rows           #
    # ------------------------------------------------------------------ #
    if args.fill_charm_sku:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        oa_key_fill = _resolve_charm_vision_api_key(args.openai_api_key)
        oa_base_fill = _resolve_charm_vision_base_url(
            args.openai_base_url.strip() or None
        )
        if (
            not args.fill_charm_sku_dry_run
            and not oa_key_fill
            and not _charm_vision_base_allows_empty_key(oa_base_fill)
        ):
            log.error(
                "Set CHARM_VISION_API_KEY or OPENAI_API_KEY (or --openai-api-key) for "
                "--fill-charm-sku, or use --openai-base-url / CHARM_VISION_BASE_URL for a "
                "local OpenAI-compatible server."
            )
            sys.exit(1)
        init_charm_shops_sheet(catalog_path)
        init_charm_library_sheet(catalog_path)
        n_fill, fill_lines = fill_charm_library_vision_sku(
            catalog_path,
            charm_images_dir,
            openai_api_key=oa_key_fill,
            openai_model=args.openai_model.strip() or "gpt-4o-mini",
            openai_base_url=oa_base_fill,
            overwrite=args.fill_charm_sku_overwrite,
            dry_run=args.fill_charm_sku_dry_run,
        )
        print(f"\n{'=' * 60}")
        print("  Charm Library — SKU (vision)")
        for ln in fill_lines:
            print(f"  {ln}")
        print(f"  ({n_fill} row(s) {'would be ' if args.fill_charm_sku_dry_run else ''}updated)")
        print(f"{'=' * 60}\n")
        if any("[error]" in x for x in fill_lines):
            sys.exit(1)
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --import-charm-images: bulk rename + Charm Library rows               #
    # ------------------------------------------------------------------ #
    if args.import_charm_images:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        oa_key = _resolve_charm_vision_api_key(args.openai_api_key)
        oa_base = _resolve_charm_vision_base_url(
            args.openai_base_url.strip() or None
        )
        if (
            args.import_charm_vision_sku
            and not args.import_charm_dry_run
            and not oa_key
            and not _charm_vision_base_allows_empty_key(oa_base)
        ):
            log.error(
                "Set CHARM_VISION_API_KEY or OPENAI_API_KEY (or --openai-api-key) when using "
                "--import-charm-vision-sku, or set CHARM_VISION_BASE_URL to a local server "
                "(e.g. Ollama http://127.0.0.1:11434/v1)."
            )
            sys.exit(1)
        init_charm_shops_sheet(catalog_path)
        init_charm_library_sheet(catalog_path)
        n_imp, imp_lines = import_charm_screenshot_assets(
            charm_images_dir,
            catalog_path,
            patterns=args.import_charm_patterns,
            dry_run=args.import_charm_dry_run,
            prefix=args.import_charm_prefix,
            vision_sku=args.import_charm_vision_sku,
            openai_api_key=oa_key or None,
            openai_model=args.openai_model.strip() or "gpt-4o-mini",
            openai_base_url=oa_base,
        )
        print(f"\n{'=' * 60}")
        print("  Charm image import")
        for ln in imp_lines:
            print(f"  {ln}")
        print(f"  ({n_imp} file(s) {'would be ' if args.import_charm_dry_run else ''}imported)")
        print(f"{'=' * 60}\n")
        imp_failed = any("[error]" in x for x in imp_lines)
        if imp_failed:
            sys.exit(1)
        if args.export_charm_manifest:
            m_out = (
                Path(args.charm_manifest_output).resolve()
                if args.charm_manifest_output.strip()
                else charm_manifest_default_path
            )
            n_manifest = export_charm_manifest(catalog_path, charm_images_dir, m_out)
            print(f"  [OK]  Charm manifest  ->  {m_out}")
            print(f"        {n_manifest} charm(s) indexed (library and/or disk)\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --export-charm-manifest: JSON index for websites / tooling          #
    # ------------------------------------------------------------------ #
    if args.export_charm_manifest:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        init_charm_shops_sheet(catalog_path)
        init_charm_library_sheet(catalog_path)
        m_out = (
            Path(args.charm_manifest_output).resolve()
            if args.charm_manifest_output.strip()
            else charm_manifest_default_path
        )
        n_manifest = export_charm_manifest(catalog_path, charm_images_dir, m_out)
        print(f"\n{'=' * 60}")
        print(f"  [OK]  Charm manifest  ->  {m_out}")
        print(f"        {n_manifest} charm(s) indexed (library and/or disk)")
        print(f"{'=' * 60}\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --init-charm-shops: one-time setup of the Charm Shops sheet.        #
    # ------------------------------------------------------------------ #
    if args.init_charm_shops:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        init_charm_shops_sheet(catalog_path)
        init_charm_library_sheet(catalog_path)
        print(f"\n{'=' * 60}")
        print(f"  [OK]  'Charm Shops' + 'Charm Library' ready in  ->  {catalog_path.resolve()}")
        print(f"\n  Next steps:")
        print(f"    • Charm Shops  — edit shop + stall rows; save, then run the generator.")
        print(f"    • Charm Library — one row per charm; see the guide row on that sheet.")
        print(f"    • Product Map  — column G = same charm code as Library column B.")
        print(f"    • Images       — optional: data/charm_images/<Code>.png (overrides embed).")
        print(f"    • Manifest     — .\\run.ps1 --export-charm-manifest")
        print(f"{'=' * 60}\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --rebuild-catalog: re-sort the catalog in-place and exit early.     #
    # No PDFs, no cache, no shopping route regeneration needed.           #
    # ------------------------------------------------------------------ #
    if args.rebuild_catalog:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        init_charm_shops_sheet(catalog_path)
        init_charm_library_sheet(catalog_path)
        rebuild_catalog(catalog_path)
        print(f"\n{'=' * 60}")
        print(f"  [OK]  Catalog rebuilt and sorted  ->  {catalog_path.resolve()}")
        print(
            f"  [>>]  Catalog backups folder      ->  {catalog_backup_dir(catalog_path).resolve()}"
        )
        print(f"\n  Tip: run --refresh-catalog next to update shopping_route.xlsx")
        print(f"{'=' * 60}\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --sync-suppliers: Suppliers sheet from Product Map only             #
    # (After --rebuild-catalog so rebuild alone does not exit early here.) #
    # ------------------------------------------------------------------ #
    if args.sync_suppliers:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        wb = openpyxl.load_workbook(catalog_path)
        ensure_catalog_column_layout(wb)
        n = sync_suppliers_from_product_map(wb)
        if CATALOG_SHEET in wb.sheetnames:
            _refresh_all_product_map_validations(wb, wb[CATALOG_SHEET])
        set_supplier_catalog_active_to_product_map(wb)
        backup_supplier_catalog_before_write(catalog_path, "sync_suppliers")
        wb.save(catalog_path)
        print(f"\n{'=' * 60}")
        print(f"  [OK]  Suppliers synced  ->  {catalog_path.resolve()}")
        print(f"        {n} new supplier row(s) appended (if any); existing rows unchanged.")
        print(f"{'=' * 60}\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --clear-product-map-charm-codes                                     #
    # ------------------------------------------------------------------ #
    if args.clear_product_map_charm_codes:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        n = clear_product_map_charm_codes(catalog_path)
        print(f"\n{'=' * 60}")
        print(f"  [OK]  Product Map Charm Code column (H) cleared  ->  {catalog_path.resolve()}")
        print(f"        {n} cell(s) cleared. Re-pick codes from Charm Library (column B) via the dropdown.")
        print(f"{'=' * 60}\n")
        sys.exit(0)

    # ------------------------------------------------------------------ #
    # --renumber-charms: reassign sequential codes to match row order     #
    # ------------------------------------------------------------------ #
    if args.renumber_charms:
        if not catalog_path.exists():
            log.error("Catalog file not found: %s", catalog_path)
            sys.exit(1)
        init_charm_shops_sheet(catalog_path)
        init_charm_library_sheet(catalog_path)
        n_ren, ren_lines = renumber_charm_library(
            catalog_path,
            charm_images_dir=charm_images_dir,
            dry_run=args.renumber_charms_dry_run,
        )
        print(f"\n{'=' * 60}")
        print(
            "  Charm Library — renumber codes"
            + (" (dry run)" if args.renumber_charms_dry_run else "")
        )
        for ln in ren_lines:
            print(f"  {ln}")
        if args.renumber_charms_dry_run:
            print(f"  ({n_ren} code(s) would be renumbered — run without --renumber-charms-dry-run to apply)")
        else:
            print(f"  ({n_ren} code(s) renumbered)")
        print(f"{'=' * 60}\n")
        if any("[error]" in x for x in ren_lines):
            sys.exit(1)
        sys.exit(0)

    # If the shopping route file is gone, treat this as a fresh run from the
    # current PDFs only — do not merge in orders_cache or recover from a
    # deleted Excel.  Exception: --refresh-catalog rebuilds from cache alone
    # and must still load the cache when the xlsx is missing.
    output_missing = not output_path.exists()
    ignore_prior_orders = args.reset or (
        output_missing and not args.refresh_catalog
    )

    # ------------------------------------------------------------------ #
    # Step 1 -- Preserve any statuses the user already updated in Excel   #
    # ------------------------------------------------------------------ #
    existing_statuses = load_existing_statuses(output_path)

    # ------------------------------------------------------------------ #
    # Step 2 -- Load previously cached orders (from earlier runs)         #
    #           ALWAYS also read the current Excel as a safety net so    #
    #           that no data in the file can ever be silently deleted.    #
    # ------------------------------------------------------------------ #
    if ignore_prior_orders:
        cached_items:   list[ResolvedItem] = []
        excel_items:    list[ResolvedItem] = []
        processed_pdfs: set[str]           = set()
        if output_missing and not args.reset and not args.refresh_catalog:
            log.info(
                "Output file %s not found — rebuilding from input PDFs only "
                "(ignoring %s and any prior shopping route data).",
                output_path.name,
                cache_path.name,
            )
    else:
        cached_items, processed_pdfs = load_cache(cache_path)
        excel_items                  = load_items_from_xlsx(output_path)

    # Merge: cache (has photos) takes priority; Excel fills in anything missing
    _seen_prior: set[tuple[str, str]] = set()
    prior_items: list[ResolvedItem]   = []
    for r in cached_items + excel_items:
        key = (r.order.order_number, _normalize(r.item.title)[:50])
        if key not in _seen_prior:
            _seen_prior.add(key)
            prior_items.append(r)

    if len(prior_items) > len(cached_items):
        log.info(
            "Safety-net: recovered %d item(s) from existing Excel that were absent from cache",
            len(prior_items) - len(cached_items),
        )

    cached_items      = prior_items
    cached_order_nums = {r.order.order_number for r in cached_items}

    # ------------------------------------------------------------------ #
    # Step 3 -- Discover and parse NEW PDFs                               #
    # --refresh-catalog skips this step entirely: no PDFs are touched.   #
    # ------------------------------------------------------------------ #
    if args.refresh_catalog:
        # Snapshot supplier info BEFORE re-matching so we can report changes.
        _before: dict[str, tuple[str, str]] = {
            _normalize(r.item.title): (
                r.supplier.shop_name if r.supplier else "",
                r.supplier.stall     if r.supplier else "",
            )
            for r in cached_items
        }
        if not cached_items:
            log.error(
                "--refresh-catalog requires an existing cache (%s) or shopping route "
                "(%s). Run a normal pass first to ingest your order PDFs.",
                cache_path, output_path,
            )
            sys.exit(1)
        log.info(
            "Refresh-catalog mode: re-matching %d cached item(s) against updated %s "
            "-- no PDFs will be processed.",
            len(cached_items), catalog_path.name,
        )
        pdf_paths = []
    else:
        pdf_paths = (
            [Path(p) for p in args.pdfs] if args.pdfs
            else sorted(input_dir.glob("*.pdf"))
        )

        # --new-batch: skip any PDF whose filename was already recorded in the
        # cache from a previous run so only genuinely new files are re-parsed.
        if args.new_batch and pdf_paths:
            skipped = [p for p in pdf_paths if p.name in processed_pdfs]
            pdf_paths = [p for p in pdf_paths if p.name not in processed_pdfs]
            if skipped:
                log.info(
                    "New-batch: skipping %d already-processed PDF(s): %s",
                    len(skipped), ", ".join(p.name for p in skipped),
                )
            if not pdf_paths:
                log.info("New-batch: no new PDF files found -- re-generating from cache only.")

        if not pdf_paths and not cached_items:
            log.error("No PDF files found in %s and cache is empty.", input_dir.resolve())
            sys.exit(1)

        if pdf_paths:
            log.info("Found %d PDF(s) to process: %s",
                     len(pdf_paths), ", ".join(p.name for p in pdf_paths))
        else:
            log.info("No PDFs to process -- re-generating from cache only.")

    if not catalog_path.exists():
        log.error("Catalog not found: %s", catalog_path)
        sys.exit(1)

    # Ensure charm infrastructure + Product Map columns exist, then load catalog
    # so columns G/H/I are visible on the first run after an upgrade.
    init_charm_shops_sheet(catalog_path)
    init_charm_library_sheet(catalog_path)
    catalog = load_catalog(catalog_path)
    charm_shops = load_charm_shops(catalog_path)
    charm_library = load_charm_library(catalog_path)
    if charm_images_dir.is_dir():
        _n_ch = len(_disk_charm_files_index(charm_images_dir))
        log.info("Charm image folder: %s (%d charm image(s))", charm_images_dir, _n_ch)

    all_new_orders: list[Order] = []
    for pdf in pdf_paths:
        orders     = parse_pdf(pdf)
        item_count = sum(len(o.items) for o in orders)
        photos     = sum(1 for o in orders for it in o.items if it.photo_bytes)
        log.info("  %-30s  %d orders, %d items (%d with photos)",
                 pdf.name, len(orders), item_count, photos)
        all_new_orders.extend(orders)
        processed_pdfs.add(pdf.name)   # mark this PDF as ingested

    # ------------------------------------------------------------------ #
    # Step 3b -- Re-match ALL prior items against the current catalog   #
    #            Picks up any supplier name / location changes the user  #
    #            made to supplier_catalog.xlsx since the last run.       #
    # ------------------------------------------------------------------ #
    if cached_items:
        log.info("Re-matching %d prior item(s) against current catalog "
                 "(picks up catalog edits) ...", len(cached_items))
        rematched_all = match_items(
            [r.order for r in cached_items], catalog, args.threshold
        )
        # Swap in updated supplier info; keep original item data (photos, etc.)
        updated_cached: list[ResolvedItem] = []
        for orig, new_match in zip(cached_items, rematched_all):
            if new_match.supplier:
                updated_cached.append(ResolvedItem(
                    order       = orig.order,
                    item        = orig.item,
                    supplier    = new_match.supplier,
                    match_score = new_match.match_score,
                ))
            else:
                updated_cached.append(orig)   # keep whatever we had before
        cached_items = updated_cached

    # --refresh-catalog: diff old vs new supplier assignments and report clearly
    if args.refresh_catalog and cached_items:
        promoted:     list[str] = []   # had no location → now has one
        info_updated: list[str] = []   # location changed (user corrected it)
        still_empty:  list[str] = []   # still no supplier info after re-match

        for r in cached_items:
            key         = _normalize(r.item.title)
            old_shop, old_stall = _before.get(key, ("", ""))
            new_shop = r.supplier.shop_name if r.supplier else ""
            new_stall = r.supplier.stall    if r.supplier else ""

            if not r.supplier or (not new_shop and not new_stall):
                still_empty.append(r.item.title[:70])
            elif not old_shop and not old_stall and (new_shop or new_stall):
                promoted.append(
                    f"  {r.item.title[:55]:<55}  ->  {new_shop or '?'} / {new_stall or '?'}"
                )
            elif (old_shop, old_stall) != (new_shop, new_stall):
                info_updated.append(
                    f"  {r.item.title[:55]:<55}  "
                    f"{old_shop or '?'}/{old_stall or '?'}  ->  "
                    f"{new_shop or '?'}/{new_stall or '?'}"
                )

        if promoted:
            log.info("Promoted (no info → routable):  %d item(s)", len(promoted))
            for line in promoted:
                log.info(line)
        if info_updated:
            log.info("Supplier info updated:  %d item(s)", len(info_updated))
            for line in info_updated:
                log.info(line)
        if still_empty:
            log.info(
                "Still awaiting supplier info:  %d item(s)  "
                "(fill them in supplier_catalog.xlsx and re-run --refresh-catalog)",
                len(still_empty),
            )

    # ------------------------------------------------------------------ #
    # Step 4 -- Match new orders and deduplicate against cache            #
    # ------------------------------------------------------------------ #
    new_resolved: list[ResolvedItem] = []
    if all_new_orders:
        new_resolved = match_items(all_new_orders, catalog, args.threshold)

    truly_new = [r for r in new_resolved
                 if r.order.order_number not in cached_order_nums]
    duplicates = len(new_resolved) - len(truly_new)
    if duplicates:
        log.info("Skipped %d item(s) already in cache (duplicate PDF re-scan)", duplicates)

    # ------------------------------------------------------------------ #
    # Step 5 -- Merge: cached (old) + truly new                           #
    # ------------------------------------------------------------------ #
    all_resolved = cached_items + truly_new
    log.info(
        "Total: %d item(s)  [%d from cache + %d new]",
        len(all_resolved), len(cached_items), len(truly_new),
    )
    if not all_resolved:
        log.error("No order items to write.")
        sys.exit(1)

    # ------------------------------------------------------------------ #
    # Step 5b -- (optional) Purge purchased sections                      #
    #                                                                      #
    # Reads the statuses already entered in the current Excel file and    #
    # removes or trims any item whose procurement sections are complete.  #
    #                                                                      #
    # Two sections are evaluated independently:                           #
    #   • Case / Grip section  (supplier floors)                          #
    #   • Charm section        (separate building)                        #
    #                                                                      #
    # If the Case/Grip section is fully purchased, those components are   #
    # stripped from the item's style and the item no longer appears in    #
    # the supplier-floor rows on the next run.  Likewise for the Charm    #
    # section.  An item is removed entirely only when both sections are   #
    # complete.  "Out of Production" is treated as complete (same as     #
    # Purchased); those items are purged and logged to OOP log.           #
    # Out-of-Stock / Pending components are kept for follow-up.           #
    # ------------------------------------------------------------------ #
    purged_count        = 0   # items removed entirely (all sections done)
    partial_purge_count = 0   # items whose style was trimmed (one section done)
    oop_records: list[tuple[str, str, str, str]] = []
    if args.purge_purchased:
        # Bridge aggregated charm statuses to per-order keys so the purge
        # logic (which checks per-order (order_num, norm_title, "charm"))
        # works with the new ~C:<code> aggregated format.
        _agg_charm_statuses = {
            k: v for k, v in existing_statuses.items()
            if len(k) == 3 and k[2] == "charm_agg"
        }
        if _agg_charm_statuses:
            for r in all_resolved:
                if not _style_has(r.item.style)[2]:
                    continue
                _cc = (r.supplier.charm_code if r.supplier else "").strip()
                if not _cc:
                    continue
                _agg_val = _agg_charm_statuses.get((_cc, "", "charm_agg"))
                if _agg_val:
                    _nt = _normalize(r.item.title)[:50]
                    _per_key = (r.order.order_number, _nt, "charm")
                    if _per_key not in existing_statuses:
                        existing_statuses[_per_key] = _agg_val

        new_resolved: list[ResolvedItem] = []
        for r in all_resolved:
            remaining_style = _compute_remaining_style(
                r.order.order_number, r.item.title, r.item.style, existing_statuses
            )
            if remaining_style is None:
                # Every section complete → drop this item entirely
                oop_entries = _get_oop_components_being_purged(
                    r.order.order_number, r.item.title, r.item.style, existing_statuses
                )
                for order_num, product_title, component in oop_entries:
                    oop_records.append((order_num, product_title, component, r.order.etsy_shop))
                purged_count += 1
            else:
                if remaining_style != r.item.style:
                    # One section was completed; strip it so only the pending
                    # section appears in the regenerated shopping route.
                    oop_entries = _get_oop_components_being_purged(
                        r.order.order_number, r.item.title, r.item.style, existing_statuses
                    )
                    for order_num, product_title, component in oop_entries:
                        oop_records.append((order_num, product_title, component, r.order.etsy_shop))
                    log.debug(
                        "Partial purge order %s: style '%s' -> '%s'",
                        r.order.order_number, r.item.style, remaining_style,
                    )
                    r.item.style = remaining_style
                    partial_purge_count += 1
                new_resolved.append(r)
        if oop_records:
            _append_to_oop_log(oop_records, oop_log_path)
        all_resolved = new_resolved

        if purged_count or partial_purge_count:
            log.info(
                "Purge: removed %d fully-complete item(s); "
                "%d item(s) partially trimmed (one section done); "
                "%d item(s) remaining",
                purged_count, partial_purge_count, len(all_resolved),
            )
        else:
            log.info("Purge: no purchased sections found -- nothing removed")
        if not all_resolved:
            log.info("All items have been purchased! Shopping route is now empty.")
            # Still write an empty (header-only) file so old data is not left stale

        # Retain only the statuses for items that are still in the route.
        # Use (order_num, norm_title) pairs so we never accidentally carry over
        # statuses from a purged item that shares an order number with a kept one.
        # Aggregated charm statuses (charm_agg) are kept if any remaining item
        # references that charm code.
        kept_keys = {
            (r.order.order_number, _normalize(r.item.title)[:50])
            for r in all_resolved
        }
        _kept_charm_codes = {
            (r.supplier.charm_code.strip() if r.supplier else "")
            for r in all_resolved
            if _style_has(r.item.style)[2]
        } - {""}
        existing_statuses = {
            k: v for k, v in existing_statuses.items()
            if (
                (len(k) == 3 and k[2] == "charm_agg" and k[0] in _kept_charm_codes)
                or (k[0], k[1]) in kept_keys
            )
        }

    # ------------------------------------------------------------------ #
    # Step 6 -- Update supplier catalog for ALL unmatched items           #
    #                                                                      #
    # We pass all_resolved (not just truly_new) so that:                  #
    #  • Products from new PDFs that have no catalog entry are added.      #
    #  • Old cached items that are still unmatched (e.g. the amber row     #
    #    was deleted, or --no-catalog-update was used on a prior run)      #
    #    are also caught and appended.                                      #
    # The duplicate guard inside update_catalog (exact-normalised title    #
    # check) ensures nothing is added twice even across multiple runs.     #
    # ------------------------------------------------------------------ #
    new_added = 0
    if not args.no_catalog_update:
        new_added = update_catalog(catalog_path, all_resolved)

    # ------------------------------------------------------------------ #
    # Step 7 -- Save updated cache (all orders, inc. newly added)         #
    # ------------------------------------------------------------------ #
    save_cache(cache_path, all_resolved, processed_pdfs)

    # ------------------------------------------------------------------ #
    # Step 8 -- Generate Excel (with preserved statuses overlaid)         #
    # ------------------------------------------------------------------ #
    generate_xlsx(all_resolved, output_path, statuses=existing_statuses,
                  charm_shops=charm_shops, charm_library=charm_library,
                  charm_images_dir=charm_images_dir)

    # Step 8b -- Optionally generate a Simplified Chinese version
    zh_item_count   = 0   # used later in the summary print
    zh_excluded_ct  = 0
    zh_excluded_shops_str = ""
    title_fn        = None   # set below if --chinese; reused for HTML
    zh_items: list[ResolvedItem] = []
    if args.chinese:
        zh_path = output_path.with_stem(output_path.stem + "_zh")

        # Filter out any Etsy shops the user wants excluded from the Chinese file
        excluded_shops: set[str] = set()
        if args.chinese_exclude_shops:
            excluded_shops = {s.strip().lower() for s in args.chinese_exclude_shops.split(",") if s.strip()}
        zh_items = [
            r for r in all_resolved
            if r.order.etsy_shop.lower() not in excluded_shops
        ]
        zh_item_count  = len(zh_items)
        zh_excluded_ct = len(all_resolved) - zh_item_count
        zh_excluded_shops_str = args.chinese_exclude_shops or ""

        if excluded_shops:
            log.warning(
                "Chinese file: EXCLUDING %d of %d item(s) from shop(s): %s  "
                "-- Chinese file will have %d item(s)",
                zh_excluded_ct, len(all_resolved),
                ", ".join(sorted(excluded_shops)),
                zh_item_count,
            )

        # Build product-title translator (cached on disk to avoid repeat API calls)
        trans_cache = _load_trans_cache(trans_cache_path)

        def _make_title_fn(cache: dict, cpath: Path):
            def _fn(title: str) -> str:
                return _translate_title(title, cache, cpath)
            return _fn

        title_fn = _make_title_fn(trans_cache, trans_cache_path)

        generate_xlsx(zh_items, zh_path, statuses=existing_statuses, lang="zh",
                      title_fn=title_fn, charm_shops=charm_shops,
                      charm_library=charm_library,
                      charm_images_dir=charm_images_dir)
        log.info("Chinese version saved -> %s", zh_path.resolve())

    # ------------------------------------------------------------------ #
    # Step 8c -- Optionally generate HTML file(s)                         #
    # ------------------------------------------------------------------ #
    if args.html:
        html_path = output_path.with_suffix(".html")
        generate_html(all_resolved, html_path, statuses=existing_statuses,
                      lang="en", charm_shops=charm_shops,
                      charm_library=charm_library,
                      charm_images_dir=charm_images_dir)
        if args.chinese and zh_items:
            zh_html_path = output_path.with_stem(output_path.stem + "_zh").with_suffix(".html")
            generate_html(zh_items, zh_html_path, statuses=existing_statuses,
                          lang="zh", title_fn=title_fn, charm_shops=charm_shops,
                          charm_library=charm_library,
                          charm_images_dir=charm_images_dir)
            log.info("Chinese HTML saved -> %s", zh_html_path.resolve())

    # ------------------------------------------------------------------ #
    # Step 8d -- Charm manifest (auto): library + disk + route snapshot     #
    # ------------------------------------------------------------------ #
    manifest_written_path: Path | None = None
    manifest_charm_count = 0
    if not args.no_charm_manifest:
        try:
            _route_snap = {
                "order_line_items": len(all_resolved),
                "lines_with_charm_in_order_style": sum(
                    1 for r in all_resolved if _style_has(r.item.style)[2]
                ),
                "matched_lines_with_charm_code": sum(
                    1 for r in all_resolved
                    if r.supplier and (r.supplier.charm_code or "").strip()
                ),
            }
            manifest_charm_count = export_charm_manifest(
                catalog_path,
                charm_images_dir,
                charm_manifest_default_path,
                route_snapshot=_route_snap,
            )
            manifest_written_path = charm_manifest_default_path
        except Exception as exc:
            log.warning("Charm manifest (auto) failed: %s", exc)

    routable_ct   = sum(1 for r in all_resolved
                        if r.supplier and (r.supplier.shop_name or r.supplier.stall))
    needs_info_ct = sum(1 for r in all_resolved
                        if r.supplier
                        and not (r.supplier.shop_name or r.supplier.stall)
                        and not _needs_catalog_entry(r))
    unmatched_ct  = sum(1 for r in all_resolved
                        if not r.supplier or _needs_catalog_entry(r))

    print(f"\n{'=' * 60}")
    if args.refresh_catalog:
        print(f"  [>>]  Catalog refresh complete  ({len(all_resolved)} order item(s) re-matched)")
    if purged_count:
        print(f"  [X]   {purged_count} fully-complete item(s) removed from route")
    if oop_records:
        print(f"  [OOP] {len(oop_records)} Out-of-Production record(s) logged to {OOP_LOG_FILE}")
    if partial_purge_count:
        print(f"  [~X]  {partial_purge_count} item(s) partially trimmed "
              f"(one section done; other section still pending)")
    print(f"  [OK]  {routable_ct} item(s) ready  (supplier + location known)")
    if needs_info_ct:
        print(f"  [~]   {needs_info_ct} item(s) in catalog – "
              f"fill Shop Name / Stall in {catalog_path.name}, then re-run --refresh-catalog")
    if unmatched_ct:
        print(f"  [!!]  {unmatched_ct} item(s) not in catalog  "
              f"(check {catalog_path.name} for new amber rows to fill in)")
    if new_added:
        print(f"  [+]   {new_added} new product(s) appended to {catalog_path.name}  "
              f"– open it and fill in Shop Name / Stall / Price")
    if cached_items and not purged_count and not partial_purge_count and not args.refresh_catalog:
        print(f"  [>>]  {len(cached_items)} prior order(s) carried over from cache")
    remaining_ct = len(all_resolved)
    if purged_count or partial_purge_count:
        print(f"  [>>]  {remaining_ct} item(s) still need attention")
    print(f"  --->  {output_path.resolve()}  ({len(all_resolved)} items)")
    if manifest_written_path is not None:
        print(
            f"  [JSON] {manifest_written_path.resolve()}  "
            f"(charm index, {manifest_charm_count} charm(s))"
        )
    if args.html:
        html_path = output_path.with_suffix(".html")
        print(f"  [HTML] {html_path.resolve()}  (mobile-friendly, share this file)")
    if args.chinese:
        zh_path = output_path.with_stem(output_path.stem + "_zh")
        print(f"  [ZH]  {zh_path.resolve()}  ({zh_item_count} items  |  Simplified Chinese)")
        if args.html and zh_items:
            zh_html_path = zh_path.with_suffix(".html")
            print(f"  [ZH HTML] {zh_html_path.resolve()}  (mobile-friendly Chinese)")
        if zh_excluded_ct:
            sep = "!" * 60
            print(f"\n  {sep}")
            print(f"  !!  WARNING: CHINESE FILE HAS FEWER ITEMS THAN ENGLISH  !!")
            print(f"  {sep}")
            print(f"  !!  English  : {len(all_resolved)} items")
            print(f"  !!  Chinese  : {zh_item_count} items  ({zh_excluded_ct} item(s) intentionally excluded)")
            print(f"  !!  Excluded : {zh_excluded_shops_str}")
            print(f"  !!")
            print(f"  !!  This is correct ONLY if you used --chinese-exclude-shops on purpose.")
            print(f"  !!  To include ALL items in the Chinese file, re-run WITHOUT that flag.")
            print(f"  {sep}\n")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()
