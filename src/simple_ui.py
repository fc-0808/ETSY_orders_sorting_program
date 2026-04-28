#!/usr/bin/env python3
"""
Simple desktop UI for the shopping-route script.

Plain language for anyone new — no Etsy jargon required.
Uses tkinter (included with Python on Windows). Double-click run_ui.ps1 or run from project folder.
UI language: defaults to 中文; use the language button for English.
Layout: title + quick-open row, then the gray step strip, then tabbed panels (PDF / Charms drop zones mirror each other).
"""

from __future__ import annotations

import os
import queue
import re
import secrets
import shutil
import subprocess
import sys
import threading
from datetime import date
from collections.abc import Callable
from io import BytesIO
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

PROJECT_ROOT = Path(__file__).resolve().parent.parent
GENERATOR = PROJECT_ROOT / "src" / "generate_shopping_route.py"

try:
    from generate_shopping_route import (
        CATALOG_SHEET,
        CHARM_LIBRARY_SHEET,
        CHARM_SHOPS_SHEET,
        ProductMapPickerRow,
        extract_photos_from_xlsx,
        get_catalog_photo_map,
        normalize_catalog_charm_shops,
        import_charm_screenshot_assets,
        list_product_map_rows_for_picker,
        sort_suppliers_sheet,
        write_charm_library_skus,
        load_cache,
        load_catalog,
        load_charm_library,
        load_charm_shops,
        mark_product_map_discontinued_by_row,
        reorder_charm_library_rows,
        save_cache,
        update_product_map_cells,
        update_product_map_photo,
        Order,
        OrderItem,
        ResolvedItem,
        _style_has,
        _normalize,
    )
except ImportError:
    CATALOG_SHEET = "Product Map"  # type: ignore[assignment, misc]
    CHARM_LIBRARY_SHEET = "Charm Library"  # type: ignore[assignment, misc]
    CHARM_SHOPS_SHEET = "Charm Shops"  # type: ignore[assignment, misc]
    ProductMapPickerRow = None  # type: ignore[assignment, misc]
    extract_photos_from_xlsx = None  # type: ignore[assignment, misc]
    get_catalog_photo_map = None  # type: ignore[assignment, misc]
    normalize_catalog_charm_shops = None  # type: ignore[assignment, misc]
    import_charm_screenshot_assets = None  # type: ignore[assignment, misc]
    list_product_map_rows_for_picker = None  # type: ignore[assignment, misc]
    sort_suppliers_sheet = None  # type: ignore[assignment, misc]
    write_charm_library_skus = None  # type: ignore[assignment, misc]
    load_cache = None  # type: ignore[assignment, misc]
    load_catalog = None  # type: ignore[assignment, misc]
    load_charm_library = None  # type: ignore[assignment, misc]
    load_charm_shops = None  # type: ignore[assignment, misc]
    mark_product_map_discontinued_by_row = None  # type: ignore[assignment, misc]
    reorder_charm_library_rows = None  # type: ignore[assignment, misc]
    save_cache = None  # type: ignore[assignment, misc]
    update_product_map_cells = None  # type: ignore[assignment, misc]
    update_product_map_photo = None  # type: ignore[assignment, misc]
    Order = None  # type: ignore[assignment, misc]
    OrderItem = None  # type: ignore[assignment, misc]
    ResolvedItem = None  # type: ignore[assignment, misc]
    _style_has = None  # type: ignore[assignment, misc]
    _normalize = None  # type: ignore[assignment, misc]

try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None  # type: ignore[assignment, misc]
    ImageTk = None  # type: ignore[assignment, misc]

try:
    from supplier_catalog_backup import (
        backup_supplier_catalog_before_write,
        catalog_backup_dir,
        list_supplier_catalog_backups,
        restore_supplier_catalog,
    )
except ImportError:
    backup_supplier_catalog_before_write = None  # type: ignore[assignment, misc]
    catalog_backup_dir = None  # type: ignore[assignment, misc]
    list_supplier_catalog_backups = None  # type: ignore[assignment, misc]
    restore_supplier_catalog = None  # type: ignore[assignment, misc]

# Same layout as generate_shopping_route.py with --project-dir .
DATA_DIR = PROJECT_ROOT / "data"
OUTPUT_DIR = PROJECT_ROOT / "output"
CHARM_IMAGES_DIR_NAME = "charm_images"
DEFAULT_CHARM_IMAGES_DIR = DATA_DIR / CHARM_IMAGES_DIR_NAME
# Staged copies use this prefix; import matches them via charm_import_pattern_argv() (must not match CH-*.png).
CHARM_INCOMING_PREFIX = "__incoming__"
CHARM_INCOMING_PATTERN = f"{CHARM_INCOMING_PREFIX}*"
CHARM_IMAGE_EXTS = frozenset({".png", ".jpg", ".jpeg", ".webp"})


def charm_import_pattern_argv() -> list[str]:
    """Globs for --import-charm-images: UI-staged copies and classic Screenshot*.png files."""
    return [
        "--import-charm-pattern",
        CHARM_INCOMING_PATTERN,
        "--import-charm-pattern",
        "Screenshot*.png",
    ]


def _auto_sku_from_stem(stem: str) -> str:
    """Derive a CHM-WORD1-WORD2-... SKU suggestion from a staged file's stem.

    Strips the ``__incoming__<8hex>_`` prefix, splits the remaining text
    on separators, removes noise tokens (numbers, generic words), uppercases
    each part, and joins with dashes — e.g.:
      ``__incoming__a1b2c3d4_pink_pearl_necklace``  →  ``CHM-PINK-PEARL-NECKL``
      ``__incoming__a1b2c3d4_Screenshot_20240115``  →  ``""``  (all noise)
    """
    clean = re.sub(r"^__incoming__[0-9a-fA-F]+_", "", stem)
    raw_parts = re.split(r"[_\-\s\.]+", clean)
    _NOISE = {
        "screenshot", "img", "image", "photo", "copy", "new", "file",
        "jpeg", "jpg", "png", "webp", "pic", "pics",
    }
    parts = [
        w.upper() for w in raw_parts
        if w and not w.isdigit() and w.lower() not in _NOISE
    ]
    if not parts:
        return ""
    # Up to 5 parts, each capped at 5 characters for a compact CHM-style code
    return "CHM-" + "-".join(p[:5] for p in parts[:5])
ORDER_PDF_EXT = ".pdf"
DEFAULT_ORDER_INPUT_DIR = PROJECT_ROOT / "input"
DEFAULT_BACKUP_DIR = PROJECT_ROOT / "backup"
_ORDER_PDF_MMdd_SUFFIX = re.compile(r"_(\d{4})\.pdf$", re.IGNORECASE)
# Match Charms tab drop target height for a consistent look.
DROP_ZONE_H = 92

FILE_SUPPLIER_CATALOG = DATA_DIR / "supplier_catalog.xlsx"
FILE_ORDERS_CACHE = PROJECT_ROOT / "cache" / "orders_cache.json"
FILE_DELETED_ORDERS = PROJECT_ROOT / "cache" / "deleted_orders.json"
FILE_SHOPPING_ROUTE = OUTPUT_DIR / "shopping_route.xlsx"
FILE_SHOPPING_ROUTE_ZH = OUTPUT_DIR / "shopping_route_zh.xlsx"
FILE_SHOPPING_ROUTE_SIMPLE = OUTPUT_DIR / "shopping_route_simple.xlsx"
FILE_SHOPPING_HTML = OUTPUT_DIR / "shopping_route.html"
FILE_SHOPPING_HTML_ZH = OUTPUT_DIR / "shopping_route_zh.html"

windnd = None
if sys.platform == "win32":
    try:
        import windnd  # type: ignore[import-untyped, unused-ignore]
    except ImportError:
        windnd = None

# Readable, calm palette (works with ttk “clam” on Windows).
COLORS = {
    "app": "#eef2f7",
    "card": "#ffffff",
    "hero": "#1d4ed8",
    "accent": "#1e40af",
    "accent_soft": "#dbeafe",
    "run": "#047857",
    "run_hover": "#065f46",
    "text": "#0f172a",
    "muted": "#64748b",
    "border": "#cbd5e1",
    "strip": "#dbeafe",
    "strip_text": "#1e3a8a",
    "strip_accent": "#2563eb",
    "log_bg": "#f8fafc",
    "drop_zone": "#eff6ff",
    "drop_border": "#93c5fd",
    "drop_hint": "#3b82f6",
    "separator": "#e2e8f0",
}

# Window layout (logical pixels; OS display scaling still applies).
UI_DEFAULT_W = 1240
UI_DEFAULT_H = 1020
UI_MIN_W = 1020
UI_MIN_H = 760
UI_MARGIN_X = 16
UI_LOG_LINES = 10

Lang = str  # "en" | "zh"

# Main tasks: id -> (title, blurb) per language
JOB_TEXT: dict[str, dict[Lang, tuple[str, str]]] = {
    "new_batch": {
        "en": (
            "Process new order PDFs",
            "Use when you have new Etsy order PDFs. Add them with the drop zone above or the input folder — "
            "only files never processed before are read. Then run this job (or use «Run: process new order PDFs» "
            "in that panel). This updates your shopping list files.",
        ),
        "zh": (
            "处理新的订单 PDF",
            "有新的 Etsy 订单 PDF 时使用。可用上方拖放区或 input 文件夹加入文件——只会处理从未处理过的 PDF。"
            "然后选本项并点绿色「运行」，或点拖放区旁的「运行：处理新订单 PDF」。会更新采购清单相关文件。",
        ),
    },
    "refresh_catalog": {
        "en": (
            "Rebuild the shopping list (no new PDFs)",
            "Use after editing the supplier product spreadsheet (supplier_catalog.xlsx). "
            "Uses saved order data — you do not need new PDFs.",
        ),
        "zh": (
            "重新生成采购清单（不需要新 PDF）",
            "在修改供应商商品表（supplier_catalog.xlsx）后使用。沿用已保存的订单数据，不需要新的 PDF。",
        ),
    },
    "purge_purchased": {
        "en": (
            "Clean up after shopping",
            "Use after you opened the shopping list spreadsheet and marked what you already bought. "
            "This removes finished items from the list and updates saved data.",
        ),
        "zh": (
            "购物后清理清单",
            "在采购清单表格里标记「已买」等项目后使用。会从清单中移除已完成项并更新保存的数据。",
        ),
    },
    "rebuild_catalog": {
        "en": (
            "Re-sort the supplier spreadsheet only",
            "Reorders rows in supplier_catalog.xlsx by floor/stall. "
            "If you also need new shopping list files, run “Rebuild the shopping list” afterward.",
        ),
        "zh": (
            "仅重新排序供应商表格",
            "按楼层/摊位重排 supplier_catalog.xlsx。若还需要新的采购清单文件，请之后再选「重新生成采购清单」。",
        ),
    },
    "reset": {
        "en": (
            "Advanced: start over from PDFs only",
            "Ignores saved progress and rebuilds from whatever PDFs are in the input folder now. "
            "Only use this if your manager asked — it is stronger than a normal refresh.",
        ),
        "zh": (
            "高级：仅从 PDF 重新开始",
            "忽略已保存进度，只根据 input 文件夹里现有的 PDF 重建。除非主管要求，否则不要随意使用。",
        ),
    },
}

JOB_ORDER = list(JOB_TEXT.keys())

# Charm: one clear step at a time (avoids mixing incompatible flags). Internal id -> argv flags.
CHARM_MODE_FLAGS: dict[str, list[str]] = {
    "skip": [],
    "setup_excel": ["--init-charm-shops"],
    "try_add_photos": ["--import-charm-images", "--import-charm-dry-run"],
    "add_photos": ["--import-charm-images"],
    "export_list": ["--export-charm-manifest"],
    "try_fill_codes": ["--fill-charm-sku-dry-run"],
    "fill_codes": ["--fill-charm-sku"],
    "try_renumber": ["--renumber-charms", "--renumber-charms-dry-run"],
    "renumber": ["--renumber-charms"],
}

# Order shown in the UI
CHARM_MODE_ORDER = [
    "skip",
    "setup_excel",
    "try_add_photos",
    "add_photos",
    "export_list",
    "try_fill_codes",
    "fill_codes",
    "try_renumber",
    "renumber",
]

# id -> (radio title, one-line hint) per language — keep titles short; details live in charm_run_note / section A.
CHARM_MODE_TEXT: dict[str, dict[Lang, tuple[str, str]]] = {
    "skip": {
        "en": (
            "None (default)",
            "Green Run on tab 1 runs only the main job you selected there.",
        ),
        "zh": (
            "不用（默认）",
            "绿色「运行」只执行第 1 页已选的主要任务。",
        ),
    },
    "setup_excel": {
        "en": (
            "One-time: add Charm sheets to the workbook",
            "Creates Charm Shops + Charm Library if missing. Does not import images.",
        ),
        "zh": (
            "一次性：在商品表里添加挂饰工作表",
            "若无则创建 Charm Shops、Charm Library；不导入图片。",
        ),
    },
    "try_add_photos": {
        "en": (
            "Preview charm import (dry run)",
            "Prints planned renames only — no changes to disk or Excel.",
        ),
        "zh": (
            "预览挂饰导入（试运行）",
            "仅在输出中列出将重命名的文件，不写磁盘、不改 Excel。",
        ),
    },
    "add_photos": {
        "en": (
            "Import charm images (via Run)",
            "Same result as Import into workbook above; this path uses green Run instead.",
        ),
        "zh": (
            "导入挂饰图（通过「运行」）",
            "与上方「导入到工作簿」效果相同，只是改用绿色「运行」执行。",
        ),
    },
    "export_list": {
        "en": (
            "Export charm index (JSON)",
            "Sidecar file for web or scripts — not the shopping list.",
        ),
        "zh": (
            "导出挂饰索引（JSON）",
            "供网页或脚本用的辅助文件，不是采购清单。",
        ),
    },
    "try_fill_codes": {
        "en": (
            "Preview AI SKU fill (dry run)",
            "Lists target rows; no API calls or saves.",
        ),
        "zh": (
            "预览 AI 填写 SKU（试运行）",
            "列出将处理的行；不调接口、不保存。",
        ),
    },
    "fill_codes": {
        "en": (
            "Fill empty SKUs in Charm Library (AI)",
            "Needs CHARM_VISION_* / OPENAI_* (or local base URL) set on this PC.",
        ),
        "zh": (
            "用 AI 补全 Charm Library 的空 SKU",
            "需本机已配置 CHARM_VISION_* / OPENAI_* 或本机视觉接口地址。",
        ),
    },
    "try_renumber": {
        "en": (
            "Preview renumber charm codes (dry run)",
            "Shows the planned old → new code mapping without making any changes.",
        ),
        "zh": (
            "预览重新编号挂饰代码（试运行）",
            "仅显示旧代码→新代码的对应关系，不写入任何修改。",
        ),
    },
    "renumber": {
        "en": (
            "Renumber charm codes (apply)",
            "Reassigns codes by row order, updates Product Map references and renames image files.",
        ),
        "zh": (
            "重新编号挂饰代码（应用）",
            "按行顺序重新分配代码，自动更新商品表引用并重命名图片文件。",
        ),
    },
}

# Fixed chrome strings
CHROME: dict[Lang, dict[str, str]] = {
    "en": {
        "win_title": "Shopping list builder",
        "lang_btn": "中文",
        "header_title": "Turn order PDFs into shopping lists and store maps.",
        "tab_orders": "1 — Orders & outputs",
        "tab_charms": "2 — Charms (optional)",
        "step_strip": (
            "  ①  Add PDFs (drop zone / Browse / «Open input folder» on this tab)   →   ②  Pick ONE main job (left)   →   "
            "③  Green Run below (or «Run: process new order PDFs» after adding PDFs)"
        ),
        "footer_run_hint": "Read the messages in the output box when it finishes.",
        "quick_hint": (
            "Quick open — opens in Excel or your browser when the file exists; run the tool first if not. "
            "Top group: supplier catalog and data folder. "
            "Bottom group: shopping list outputs separated by format (Excel / Web)."
        ),
        "quick_group_supplier": "Supplier — product data (catalog lives here)",
        "quick_group_route": "Shopping list — generated outputs",
        "quick_group_excel": "\U0001F4CA  Excel",
        "quick_group_web":   "\U0001F310  Web",
        "quick_catalog": "Open supplier workbook",
        "btn_catalog_backups": "Catalog backups…",
        "cat_backup_title": "Supplier catalog backups",
        "cat_backup_hint": (
            "Timestamped copies are saved automatically before this program changes "
            "supplier_catalog.xlsx. Close Excel before restoring."
        ),
        "cat_backup_open_folder": "Open backup folder",
        "cat_backup_snapshot": "Snapshot now",
        "cat_backup_snapshot_ok": "Snapshot saved:\n{path}",
        "cat_backup_snapshot_no_file": "supplier_catalog.xlsx was not found — nothing to copy.",
        "cat_backup_restore": "Restore selected",
        "cat_backup_none": "(No backups yet — run any catalog action to create one.)",
        "cat_backup_restore_confirm": (
            "Replace the current supplier_catalog.xlsx with this backup?\n\n{path}\n\n"
            "Close Excel first. The current file will be copied to the backup folder first."
        ),
        "cat_backup_restore_ok": "Catalog restored from backup.",
        "cat_backup_restore_pick": "Select a backup in the list first.",
        "cat_backup_close": "Close",
        "btn_edit_products": "Edit products…",
        "btn_orders_dashboard": "Orders Dashboard…",
        "edit_title": "Edit Product Map",
        "edit_heading": "Assign shop, stall, charm code — or mark as discontinued",
        "edit_intro": (
            "Select a product row, then use the fields on the right to assign or change "
            "the Shop Name, Stall, Charm Shop, and Charm Code — or mark the product as "
            "discontinued to remove it from future shopping lists."
        ),
        "edit_search_label": "Search",
        "edit_search_tip": "Matches title, shop, or stall",
        "edit_col_photo": "Photo",
        "edit_col_row": "Row",
        "edit_col_title": "Title (short)",
        "edit_col_shop": "Shop",
        "edit_col_stall": "Stall",
        "edit_col_charm_shop": "Charm Shop",
        "edit_col_charm_code": "Charm Code",
        "edit_lbl_shop": "Shop Name",
        "edit_lbl_stall": "Stall",
        "edit_lbl_charm_shop": "Charm Shop",
        "edit_lbl_charm_code": "Charm Code",
        "edit_btn_save": "Save",
        "edit_btn_close": "Close",
        "edit_saved": "Saved — {title}",
        "edit_no_selection": "Select a product row in the table first.",
        "edit_no_import": "Catalog helpers are unavailable (check installation).",
        "edit_empty": "No product rows on the Product Map sheet.",
        "edit_preview_title": "Edit — selected product",
        "edit_preview_placeholder": "Click a row to edit its fields.",
        "edit_hover_tip": "Tip: rest the pointer on a thumbnail in the Photo column — a larger photo pops up next to the cursor.",
        "edit_btn_upload_photo": "Upload new photo\u2026",
        "edit_upload_photo_title": "Choose product photo",
        "edit_upload_photo_pending": "\u2605 New photo ready — click Save to apply",
        "edit_photo_saved": "Photo updated — {title}",
        "edit_photo_save_fail": "Could not save photo: {err}",
        "edit_charm_pick": "Pick…",
        "edit_charm_unknown": "Code not found in Charm Library.",
        "charm_picker_title": "Choose a charm",
        "charm_picker_heading": "Click a charm to select it",
        "charm_picker_search": "Filter",
        "charm_picker_search_tip": "Matches code or SKU",
        "charm_picker_col_photo": "Photo",
        "charm_picker_col_code": "Code",
        "charm_picker_col_sku": "SKU",
        "charm_picker_col_shop": "Default Shop",
        "charm_picker_none": "(none — clear charm code)",
        "charm_picker_hover_tip": "Tip: hover over the Photo column — a larger image pops up next to the cursor.",
        "edit_danger_zone": "Danger zone",
        "edit_danger_note": "This action moves the product permanently out of the active catalog.",
        "edit_btn_discontinue": "Mark as discontinued\u2026",
        "edit_discontinue_confirm_title": "Mark as discontinued",
        "edit_discontinue_confirm": (
            "Move this product to the Discontinued Products sheet?\n\n"
            "{title}\n\n"
            "It will no longer appear on future shopping lists.\n"
            "The record is kept for reference — it is not deleted."
        ),
        "edit_discontinue_done": (
            "Moved to Discontinued Products sheet.\n"
            "Run \u00abRebuild the shopping list\u00bb next to update the route."
        ),
        "discontinue_title": "Mark product as discontinued",
        "discontinue_heading": "Stop matching a catalog product",
        "discontinue_intro": (
            "When a supplier no longer sells an item, mark it here. The product row is moved "
            "from Product Map to the Discontinued Products sheet — it stays for your records "
            "but new orders will not match it."
        ),
        "discontinue_step1": "1. Use the photo plus Shop and Stall to confirm the correct row.",
        "discontinue_step2": "2. Click a row, check the preview, then «Mark selected» (or double-click). The product is moved to a Discontinued Products sheet.",
        "discontinue_step3": "3. Run «Rebuild the shopping list» on the main tab so the route updates.",
        "discontinue_search_label": "Search",
        "discontinue_search_tip": "Matches title, shop, or stall",
        "discontinue_hover_tip": "Tip: rest the pointer on a thumbnail in the Image column — a larger photo pops up next to the cursor.",
        "discontinue_preview_title": "Preview — selected product",
        "discontinue_preview_placeholder": "Click a row to see the full title and a larger photo.",
        "discontinue_col_thumb": "Photo",
        "discontinue_col_row": "Row",
        "discontinue_col_shop": "Shop",
        "discontinue_col_stall": "Stall",
        "discontinue_col_title": "Title (short)",
        "discontinue_apply": "Mark selected as discontinued",
        "discontinue_done": "Product moved to Discontinued Products sheet. Regenerate the shopping list next.",
        "discontinue_empty": "No product rows on the Product Map sheet.",
        "discontinue_close": "Close",
        "discontinue_no_import": "Catalog helpers are unavailable (check installation).",
        "discontinue_no_pillow": (
            "Install Pillow for product thumbnails:  pip install Pillow  (then restart this app)."
        ),
        "discontinue_no_selection": "Select a product row in the table first.",
        "quick_data_folder": "Open data folder",
        "quick_route": "English — detailed",
        "quick_route_simple": "English — simple",
        "quick_route_zh": "Chinese — simple",
        "quick_html": "Web page",
        "quick_html_zh": "Web — Chinese",
        "open_route_btn": "\U0001F4C2  Open Route \u25be",
        "open_route_menu_title": "Open Shopping Route",
        "open_route_en_detail": "\U0001F4CA  English — Detailed  (shopping_route.xlsx)",
        "open_route_en_simple": "\U0001F4CB  English — Simple  (shopping_route_simple.xlsx)",
        "open_route_zh": "\U0001F1E8\U0001F1F3  Chinese — Simple  (shopping_route_zh.xlsx)",
        "open_route_html_en": "\U0001F310  Web Page — English  (shopping_route.html)",
        "open_route_html_zh": "\U0001F310  Web Page — Chinese  (shopping_route_zh.html)",
        "open_route_missing": "not generated yet",
        "file_missing_title": "File not found",
        "file_missing_body": "This file is not there yet:\n{path}\nRun the tool above first, or ask a supervisor. For Chinese or web files, tick those options before running.",
        "file_open_fail_title": "Could not open",
        "file_open_fail_body": "{err}",
        "main_frame": "Main job (choose one)",
        "also_frame": "Also create",
        "chk_excel_zh": "Chinese shopping list (Excel)",
        "chk_html": "Phone-friendly web page of the list",
        "chk_no_catalog": "Do not add new products to the supplier spreadsheet this run",
        "charm_drop_section": "A — Add a new charm picture to the catalog",
        "charm_how_body": (
            "What actually happens\n"
            "• Drag or Browse copies your file into the charm images folder (Open charm folder) under a temporary name.\n"
            "• Import into workbook renames it on disk to the next free code (e.g. CH-00012.png) and adds one new row on the "
            "Charm Library sheet in supplier_catalog.xlsx, with the picture embedded. That is the master list of charms.\n"
            "• It does not add rows to Charm Shops — that sheet is only your list of physical market stalls (setup once).\n"
            "• First-time shop setup: run One-time: add Charm sheets… in section B once (supervisor), then imports work.\n"
            "Formats: PNG, JPG/JPEG, WebP."
        ),
        "charm_drop_blurb": "Drop or browse, then click Import into workbook — that is the normal path for new photos.",
        "charm_run_section": "B — Optional: charm task for green Run",
        "charm_run_note": (
            "Choose one option below, or leave None. If you pick any charm task, the next green Run performs only that task "
            "(tab 1’s main job is skipped for that run). To add new charm photos in the usual way, use section A and "
            "Import into workbook — you rarely need section B for that."
        ),
        "charm_smart": "Also use AI for SKU column (C) when Run does a charm import — PC must be configured by a supervisor",
        "charm_drop_hint_dnd": "Drop image files here",
        "charm_drop_hint_no_dnd": "Drag-and-drop: on Windows, install “windnd” (pip install windnd) and restart this app. You can always use Browse.",
        "charm_drop_browse": "Browse for photos…",
        "charm_drop_browse_title": "Choose charm photos",
        "charm_drop_import": "Import into workbook (rename + Excel)",
        "charm_drop_open_folder": "Open charm images folder",
        "charm_drop_vision": (
            "Suggest SKU with AI when importing (env CHARM_VISION_* / OPENAI_*; "
            "or CHARM_VISION_BASE_URL for Ollama, OpenRouter, LM Studio, …)"
        ),
        "charm_msg_title": "Charm photos",
        "charm_drop_no_valid": "None of those files are supported images (PNG, JPG/JPEG, WebP).",
        "charm_drop_nothing_to_import": "No staged photos to import. Drop files or use Browse first.",
        "charm_drop_staged": "Staged {n} photo(s). Click «Import into workbook» to rename and update Excel.\n",
        "charm_import_start": "\n--- Charm image import ---\n",
        "charm_reorder_section": "C — Reorder charms",
        "charm_reorder_body": (
            "Drag charms up or down in the panel to change their order — "
            "no need to touch the Excel file.  "
            "After reordering, codes are reassigned CH-00001, CH-00002 … by position "
            "and all Product Map references update automatically."
        ),
        "charm_reorder_btn_open": "Open charm reorder panel\u2026",
        "charm_reorder_start_preview": "\n--- Renumber charm codes (preview) ---\n",
        "charm_reorder_start_apply": "\n--- Renumber charm codes ---\n",
        "charm_reorder_nothing": "Charm Library not found or empty — nothing to reorder.",
        "charm_reorder_no_import": "Catalog helpers are unavailable (check installation).",
        # Reorder dialog strings
        "reorder_title": "Reorder Charm Library",
        "reorder_heading": "Drag and drop to reorder",
        "reorder_intro": (
            "Drag rows up or down, or select a row and use the \u2191 \u2193 buttons.  "
            "The \"New Code\" column shows the code each charm will receive after applying.  "
            "Click Apply to save — the Excel file and Product Map are updated automatically."
        ),
        "reorder_col_photo": "Photo",
        "reorder_col_code": "Current Code",
        "reorder_col_new": "New Code",
        "reorder_col_sku": "SKU",
        "reorder_col_shop": "Default Shop",
        "reorder_btn_up": "\u2191  Move up",
        "reorder_btn_down": "\u2193  Move down",
        "reorder_btn_top": "\u21d1  To top",
        "reorder_btn_bottom": "\u21d3  To bottom",
        "reorder_btn_apply": "Apply reorder & renumber",
        "reorder_btn_close": "Close",
        "reorder_confirm_title": "Apply reorder",
        "reorder_confirm_body": (
            "This will rewrite the Charm Library rows in the new order,\n"
            "reassign codes CH-00001, CH-00002 \u2026 by position,\n"
            "and update all Product Map references.\n\n"
            "Make sure supplier_catalog.xlsx is closed in Excel first.\n\n"
            "Continue?"
        ),
        "reorder_busy": "Applying\u2026 please wait.",
        "reorder_done": "Charm Library reordered and saved.\nRun \u00abRebuild the shopping list\u00bb to update the route.",
        "reorder_empty": "No charms found in the Charm Library.",
        "reorder_no_catalog": "supplier_catalog.xlsx not found.\nRun \u00abOne-time: add Charm sheets\u00bb first.",
        "pdf_drop_frame": "Order PDFs",
        "pdf_drop_blurb": (
            "Add Etsy order PDFs to the project input folder — same idea as the Charms photo box on the other tab. "
            "You can drop files or whole folders (PDFs are found inside). Then choose «Process new order PDFs» "
            "and the green Run, or click «Run: process new order PDFs» here to run that job immediately "
            "(uses the checkboxes on the right for Chinese / HTML / catalog options)."
        ),
        "pdf_drop_hint_dnd": "Drop PDF files or folders here",
        "pdf_drop_hint_no_dnd": (
            "Drag-and-drop: on Windows, install “windnd” (pip install windnd) and restart this app. "
            "You can always use Browse."
        ),
        "pdf_drop_browse": "Browse for PDFs…",
        "pdf_drop_browse_title": "Choose order PDFs",
        "pdf_drop_run": "Run: process new order PDFs",
        "pdf_drop_open_folder": "Open input folder",
        "pdf_drop_move_backup": "Move PDFs to backup…",
        "pdf_drop_no_valid": "None of those paths contained a .pdf file.",
        "pdf_drop_copied": "Copied {n} PDF file(s) to the input folder.\n",
        "pdf_new_batch_start": "\n--- Process new order PDFs ---\n",
        "pdf_backup_empty": "There are no PDF files in the input folder.",
        "pdf_backup_confirm_title": "Move to backup",
        "pdf_backup_confirm_body": (
            "Move {n} PDF file(s) from the input folder into the backup folder?\n\n"
            "Each file goes under backup\\MMDD\\ (month+day). If the name ends with _MMDD.pdf "
            "and that date is valid, that MMDD is used; otherwise today’s date is used.\n\n"
            "{path}"
        ),
        "pdf_backup_log_start": "\n--- Move order PDFs to backup ---\n",
        "pdf_backup_log_line": "  {src}  →  backup\\{mmdd}\\{dest}\n",
        "pdf_backup_done": "Moved {n} PDF file(s) to backup.\n",
        "pdf_backup_errors": "Could not move {n} file(s) (see log).\n",
        "opts_frame": "Optional fields (usually leave empty)",
        "threshold_hint": "Match strictness (0–100, higher = stricter). Empty = default.",
        "zh_exclude_hint": "Exclude shops from Chinese file (comma-separated). Only if Chinese Excel is checked.",
        "charm_dir_hint": "Custom charm photos folder (full path). Empty = default folder.",
        "run": "Run",
        "output_frame": "Output",
        "log_ready": (
            "Ready. Tab 1: order PDFs + main job + green Run. Tab 2: section A adds charm photos to the catalog; "
            "section B is only for charm steps triggered by Run.\n"
        ),
        "msg_busy_title": "Busy",
        "msg_busy": "Already running — wait for it to finish.",
        "msg_missing_title": "Missing file",
        "msg_missing": "Cannot find:\n",
        "log_start": "\n--- Starting ---\n",
        "log_messages": "\n--- Messages ---\n",
        "log_finished": "\n--- Finished (exit code {code}) ---\n",
        "log_error": "\nError: {e}\n",
    },
    "zh": {
        "win_title": "采购清单生成工具",
        "lang_btn": "English",
        "header_title": "把订单 PDF 转成采购清单和卖场路线图。",
        "tab_orders": "1 — 订单与输出",
        "tab_charms": "2 — 挂饰（可选）",
        "step_strip": (
            "  ①  加入 PDF（本页拖放 / 浏览 /「打开输入文件夹」）   →   ②  左侧选「一项」主要任务   →   "
            "③  下方绿色「运行」（或加完 PDF 后点「运行：处理新订单 PDF」）"
        ),
        "footer_run_hint": "完成后请看下方输出框里的提示信息。",
        "quick_hint": (
            "快速打开 — 文件已存在时用 Excel 或浏览器打开；没有时请先在下方「运行」生成。"
            "上面一组是供应商商品表及其所在文件夹；下面一组是采购清单生成结果（按格式分为 Excel 与网页两组）。"
        ),
        "quick_group_supplier": "供应商 — 商品数据（商品表在此文件夹中）",
        "quick_group_route": "采购清单 — 生成结果",
        "quick_group_excel": "\U0001F4CA  Excel",
        "quick_group_web":   "\U0001F310  网页版",
        "quick_catalog": "打开商品表（Excel）",
        "btn_catalog_backups": "商品表备份…",
        "cat_backup_title": "供应商商品表备份",
        "cat_backup_hint": (
            "在程序修改 supplier_catalog.xlsx 之前会自动保存带时间戳的副本。"
            "恢复前请先关闭 Excel。"
        ),
        "cat_backup_open_folder": "打开备份文件夹",
        "cat_backup_snapshot": "立即快照",
        "cat_backup_snapshot_ok": "快照已保存：\n{path}",
        "cat_backup_snapshot_no_file": "未找到 supplier_catalog.xlsx，无法复制。",
        "cat_backup_restore": "恢复所选备份",
        "cat_backup_none": "（尚无备份——执行任意会写入商品表的操作后即可生成。）",
        "cat_backup_restore_confirm": (
            "用此备份替换当前的 supplier_catalog.xlsx？\n\n{path}\n\n"
            "请先关闭 Excel。当前文件会先复制到备份文件夹后再替换。"
        ),
        "cat_backup_restore_ok": "已从备份恢复商品表。",
        "cat_backup_restore_pick": "请先在列表中选择一项备份。",
        "cat_backup_close": "关闭",
        "btn_edit_products": "编辑商品…",
        "btn_orders_dashboard": "订单总览…",
        "edit_title": "编辑商品表",
        "edit_heading": "分配店铺、摊位与挂饰代码，或标记为停售",
        "edit_intro": (
            "选择一行商品，在右侧面板中分配或更改店铺、摊位、挂饰店和挂饰代码，"
            "或将其标记为停售以将其从后续采购清单中移除。"
        ),
        "edit_search_label": "搜索",
        "edit_search_tip": "匹配标题、店铺或摊位",
        "edit_col_photo": "照片",
        "edit_col_row": "行",
        "edit_col_title": "标题（简短）",
        "edit_col_shop": "店铺",
        "edit_col_stall": "摊位",
        "edit_col_charm_shop": "挂饰店",
        "edit_col_charm_code": "挂饰代码",
        "edit_lbl_shop": "店铺名称",
        "edit_lbl_stall": "摊位",
        "edit_lbl_charm_shop": "挂饰店",
        "edit_lbl_charm_code": "挂饰代码",
        "edit_btn_save": "保存",
        "edit_btn_close": "关闭",
        "edit_saved": "已保存 — {title}",
        "edit_no_selection": "请先在表格中选择一个商品行。",
        "edit_no_import": "商品表辅助功能不可用（请检查安装）。",
        "edit_empty": "Product Map 工作表中没有商品行。",
        "edit_preview_title": "编辑 — 已选商品",
        "edit_preview_placeholder": "点击一行以编辑其字段。",
        "edit_hover_tip": "提示：将鼠标悬停在照片列的缩略图上——光标旁边会弹出更大的照片。",
        "edit_btn_upload_photo": "上传新照片\u2026",
        "edit_upload_photo_title": "选择商品照片",
        "edit_upload_photo_pending": "\u2605 新照片已就绪——点击「保存」以应用",
        "edit_photo_saved": "照片已更新 — {title}",
        "edit_photo_save_fail": "无法保存照片：{err}",
        "edit_charm_pick": "选择…",
        "edit_charm_unknown": "挂饰库中未找到此代码。",
        "charm_picker_title": "选择挂饰",
        "charm_picker_heading": "点击一个挂饰以选择",
        "charm_picker_search": "筛选",
        "charm_picker_search_tip": "匹配代码或 SKU",
        "charm_picker_col_photo": "照片",
        "charm_picker_col_code": "代码",
        "charm_picker_col_sku": "SKU",
        "charm_picker_col_shop": "默认挂饰店",
        "charm_picker_none": "（无——清除挂饰代码）",
        "charm_picker_hover_tip": "提示：将鼠标悬停在照片栏上，旁边会弹出较大的图片。",
        "edit_danger_zone": "危险操作",
        "edit_danger_note": "此操作将商品永久移出活跃商品目录。",
        "edit_btn_discontinue": "标记为停售\u2026",
        "edit_discontinue_confirm_title": "标记为停售",
        "edit_discontinue_confirm": (
            "将此商品移至「Discontinued Products」工作表？\n\n"
            "{title}\n\n"
            "它将不再出现在后续采购清单中。\n"
            "记录将保留以供参考——不会删除。"
        ),
        "edit_discontinue_done": (
            "已移至「Discontinued Products」工作表。\n"
            "请接着在主界面运行「重新生成采购清单」以更新路线。"
        ),
        "discontinue_title": "标记商品为停售",
        "discontinue_heading": "从目录中排除某款商品",
        "discontinue_intro": (
            "供应商不再供货时在此标记。商品行会从「Product Map」移至「Discontinued Products」工作表——"
            "记录仍会保留，但新订单不会再自动匹配到它。"
        ),
        "discontinue_step1": "1. 对照左侧缩略图，并结合「店名 / 摊位」确认是哪一款。",
        "discontinue_step2": "2. 单击一行查看详情，再点「标记所选」（或双击）。商品会被移到「Discontinued Products」工作表。",
        "discontinue_step3": "3. 回到主界面执行「重新生成采购清单」，路线图才会更新。",
        "discontinue_search_label": "搜索",
        "discontinue_search_tip": "匹配标题、店名或摊位",
        "discontinue_hover_tip": "提示：将鼠标停留在左侧「图片」列的缩略图上片刻，光标旁会弹出放大商品图。",
        "discontinue_preview_title": "预览 — 当前选中",
        "discontinue_preview_placeholder": "点击表格中的一行，可查看完整标题和放大图片。",
        "discontinue_col_thumb": "图片",
        "discontinue_col_row": "行",
        "discontinue_col_shop": "店名",
        "discontinue_col_stall": "摊位",
        "discontinue_col_title": "标题（简）",
        "discontinue_apply": "将所选标记为停售",
        "discontinue_done": "商品已移至「Discontinued Products」工作表。请接着在主界面重新生成采购清单。",
        "discontinue_empty": "Product Map 工作表中没有商品行。",
        "discontinue_close": "关闭",
        "discontinue_no_import": "无法加载商品表功能（请检查安装）。",
        "discontinue_no_pillow": (
            "安装 Pillow 后可显示商品缩略图：  pip install Pillow  （装好后重启本程序）"
        ),
        "discontinue_no_selection": "请先在表格中选择一行商品。",
        "quick_data_folder": "打开商品数据文件夹",
        "quick_route": "英文 — 详细版",
        "quick_route_simple": "英文 — 简版",
        "quick_route_zh": "中文 — 简版",
        "quick_html": "网页版",
        "quick_html_zh": "网页 — 中文版",
        "open_route_btn": "\U0001F4C2  打开路线表 \u25be",
        "open_route_menu_title": "打开采购路线文件",
        "open_route_en_detail": "\U0001F4CA  英文 — 详细版  (shopping_route.xlsx)",
        "open_route_en_simple": "\U0001F4CB  英文 — 简版  (shopping_route_simple.xlsx)",
        "open_route_zh": "\U0001F1E8\U0001F1F3  中文 — 简版  (shopping_route_zh.xlsx)",
        "open_route_html_en": "\U0001F310  网页版 — 英文  (shopping_route.html)",
        "open_route_html_zh": "\U0001F310  网页版 — 中文  (shopping_route_zh.html)",
        "open_route_missing": "尚未生成",
        "file_missing_title": "还没有这个文件",
        "file_missing_body": "目前找不到：\n{path}\n请先在上方运行一次，或询问主管。若要中文版或网页，运行前请勾选对应选项。",
        "file_open_fail_title": "无法打开",
        "file_open_fail_body": "{err}",
        "main_frame": "主要任务（选一项）",
        "also_frame": "同时生成",
        "chk_excel_zh": "中文版采购清单（Excel）",
        "chk_html": "手机友好的网页版清单",
        "chk_no_catalog": "本次运行不向供应商表添加新商品行",
        "charm_drop_section": "A — 把新挂饰照片加入商品表",
        "charm_how_body": (
            "具体会做什么\n"
            "• 拖放或浏览会把文件复制到挂饰图片文件夹（见「打开挂饰照片文件夹」），文件名暂时带前缀。\n"
            "• 点「导入到工作簿」后：磁盘上的文件会重命名为下一个可用编码（如 CH-00012.png），并在 supplier_catalog.xlsx 的 Charm Library "
            "工作表新增一行，图片嵌在表里。这里是「每个挂饰一条」的主表。\n"
            "• 不会往 Charm Shops 表加行——那个表只是卖场摊位清单，一般在 setup 时填好即可。\n"
            "• 第一次使用：请主管在 B 区执行一次「一次性：在商品表里添加挂饰工作表」，之后才能正常导入。\n"
            "支持格式：PNG、JPG/JPEG、WebP。"
        ),
        "charm_drop_blurb": "拖放或浏览选图后，点「导入到工作簿」——这是添加新挂饰图的常规做法。",
        "charm_run_section": "B — 可选：挂饰任务（配合绿色「运行」）",
        "charm_run_note": (
            "以下任选一项，或保持「不用」。若选了某项挂饰任务，下一次绿色「运行」将只执行该任务，"
            "不会同时跑第 1 页的主要任务。日常添加新挂饰图请用上方 A 区并点「导入到工作簿」；多数情况不必动 B 区。"
        ),
        "charm_smart": "在「运行」执行挂饰导入时，同时用 AI 填 SKU（C 列）；需主管在本机配置接口",
        "charm_drop_hint_dnd": "将图片文件拖放到此处",
        "charm_drop_hint_no_dnd": "拖放：在 Windows 上请安装 windnd（pip install windnd）并重启本程序。也可始终使用「浏览」按钮。",
        "charm_drop_browse": "浏览照片…",
        "charm_drop_browse_title": "选择挂饰照片",
        "charm_drop_import": "导入到工作簿（重命名并写 Excel）",
        "charm_drop_open_folder": "打开挂饰照片文件夹",
        "charm_drop_vision": (
            "导入时用 AI 建议 SKU（环境变量 CHARM_VISION_* / OPENAI_*；"
            "本机或其他服务可设 CHARM_VISION_BASE_URL，如 Ollama、OpenRouter 等）"
        ),
        "charm_msg_title": "挂饰照片",
        "charm_drop_no_valid": "所选文件里没有支持的图片格式（PNG、JPG/JPEG、WebP）。",
        "charm_drop_nothing_to_import": "没有待导入的暂存照片。请先拖放文件或使用浏览。",
        "charm_drop_staged": "已暂存 {n} 张照片。请点击「导入到工作簿」以重命名并更新 Excel。\n",
        "charm_import_start": "\n--- 挂饰图片导入 ---\n",
        "charm_reorder_section": "C — 重排挂饰顺序",
        "charm_reorder_body": (
            "在面板中上下拖动挂饰行即可更改顺序，无需打开 Excel 文件。"
            "应用后，代码将按位置重新分配为 CH-00001、CH-00002\u2026，"
            "商品表中所有引用也自动更新。"
        ),
        "charm_reorder_btn_open": "打开挂饰排序面板\u2026",
        "charm_reorder_start_preview": "\n--- 重新编号挂饰代码（预览）---\n",
        "charm_reorder_start_apply": "\n--- 重新编号挂饰代码 ---\n",
        "charm_reorder_nothing": "未找到 Charm Library 或为空——无需重排。",
        "charm_reorder_no_import": "商品表辅助功能不可用（请检查安装）。",
        # Reorder dialog strings
        "reorder_title": "挂饰排序",
        "reorder_heading": "拖放调整顺序",
        "reorder_intro": (
            "上下拖动行，或选中行后点击 \u2191 \u2193 按钮。"
            "「新代码」列显示应用后各挂饰将获得的代码。"
            "点击「应用」保存——Excel 文件与商品表均自动更新。"
        ),
        "reorder_col_photo": "照片",
        "reorder_col_code": "当前代码",
        "reorder_col_new": "新代码",
        "reorder_col_sku": "SKU",
        "reorder_col_shop": "默认挂饰店",
        "reorder_btn_up": "\u2191 上移",
        "reorder_btn_down": "\u2193 下移",
        "reorder_btn_top": "\u21d1 移到顶部",
        "reorder_btn_bottom": "\u21d3 移到底部",
        "reorder_btn_apply": "应用排序并重新编号",
        "reorder_btn_close": "关闭",
        "reorder_confirm_title": "应用排序",
        "reorder_confirm_body": (
            "将按新顺序重写 Charm Library 行，\n"
            "按位置重新分配 CH-00001、CH-00002\u2026，\n"
            "并更新所有商品表引用。\n\n"
            "请先确保 supplier_catalog.xlsx 在 Excel 中已关闭。\n\n"
            "继续？"
        ),
        "reorder_busy": "应用中\u2026请稍候。",
        "reorder_done": "Charm Library 已重排并保存。\n请运行「重新生成采购清单」以更新路线。",
        "reorder_empty": "Charm Library 中未找到挂饰。",
        "reorder_no_catalog": "未找到 supplier_catalog.xlsx。\n请先运行「一次性：在商品表里添加挂饰工作表」。",
        "pdf_drop_frame": "订单 PDF",
        "pdf_drop_blurb": (
            "把 Etsy 订单 PDF 放进项目的 input 文件夹——用法与「挂饰」标签页的照片拖放区一致。"
            "可拖文件或整个文件夹（会自动查找其中的 PDF）。然后左侧选「处理新的订单 PDF」再点绿色「运行」，"
            "或在本区点击「运行：处理新订单 PDF」立刻执行该任务（右侧勾选的中文版 / 网页 / 商品表选项会一并生效）。"
        ),
        "pdf_drop_hint_dnd": "将 PDF 文件或文件夹拖放到此处",
        "pdf_drop_hint_no_dnd": (
            "拖放：在 Windows 上请安装 windnd（pip install windnd）并重启本程序。也可始终使用「浏览」按钮。"
        ),
        "pdf_drop_browse": "浏览 PDF…",
        "pdf_drop_browse_title": "选择订单 PDF",
        "pdf_drop_run": "运行：处理新订单 PDF",
        "pdf_drop_open_folder": "打开输入文件夹",
        "pdf_drop_move_backup": "将 PDF 移至备份…",
        "pdf_drop_no_valid": "这些路径里没有找到 .pdf 文件。",
        "pdf_drop_copied": "已将 {n} 个 PDF 复制到 input 文件夹。\n",
        "pdf_new_batch_start": "\n--- 处理新订单 PDF ---\n",
        "pdf_backup_empty": "input 文件夹中没有 PDF 文件。",
        "pdf_backup_confirm_title": "移至备份",
        "pdf_backup_confirm_body": (
            "是否将 input 文件夹中的 {n} 个 PDF 移到 backup 文件夹？\n\n"
            "每个文件会放入 backup\\MMDD\\（月+日，共四位）。若文件名以 _MMDD.pdf 结尾且日期有效，"
            "则使用该 MMDD；否则使用今天的日期。\n\n"
            "{path}"
        ),
        "pdf_backup_log_start": "\n--- 将订单 PDF 移至备份 ---\n",
        "pdf_backup_log_line": "  {src}  →  backup\\{mmdd}\\{dest}\n",
        "pdf_backup_done": "已将 {n} 个 PDF 移到备份。\n",
        "pdf_backup_errors": "有 {n} 个文件未能移动（见日志）。\n",
        "opts_frame": "可选项（通常留空）",
        "threshold_hint": "匹配严格程度（0–100，越大越严）。留空则用默认。",
        "zh_exclude_hint": "中文版要排除的店铺（英文逗号分隔）。仅在选择中文版 Excel 时有效。",
        "charm_dir_hint": "自定义挂饰照片文件夹（完整路径）。留空则用默认目录。",
        "run": "运行",
        "output_frame": "输出",
        "log_ready": (
            "就绪。第 1 页：订单 PDF、主要任务与绿色「运行」。第 2 页：A 区把挂饰照片写入商品表；"
            "B 区仅在使用「运行」做挂饰专项时需要。\n"
        ),
        "msg_busy_title": "请稍候",
        "msg_busy": "正在运行中，请等待完成。",
        "msg_missing_title": "缺少文件",
        "msg_missing": "找不到文件：\n",
        "log_start": "\n--- 开始 ---\n",
        "log_messages": "\n--- 提示信息 ---\n",
        "log_finished": "\n--- 结束（退出码 {code}）---\n",
        "log_error": "\n错误：{e}\n",
    },
}


def flag_for_job(job_id: str) -> str | None:
    return {
        "new_batch": "--new-batch",
        "refresh_catalog": "--refresh-catalog",
        "purge_purchased": "--purge-purchased",
        "rebuild_catalog": "--rebuild-catalog",
        "reset": "--reset",
    }.get(job_id)


def _decode_windnd_paths(files: object) -> list[Path]:
    out: list[Path] = []
    seq = list(files) if files is not None else []
    for f in seq:
        if isinstance(f, str):
            if f:
                out.append(Path(f))
            continue
        if isinstance(f, bytes):
            text: str | None = None
            for enc in ("utf-8", "mbcs"):
                try:
                    text = f.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                text = f.decode("utf-8", errors="replace")
        else:
            text = str(f)
        if text:
            out.append(Path(text))
    return out


def _expand_pdf_paths(paths: list[Path]) -> list[Path]:
    """Collect .pdf files from dropped files and from directories (recursive)."""
    out: list[Path] = []
    for p in paths:
        try:
            rp = p if isinstance(p, Path) else Path(p)
            rp = rp.expanduser()
        except (TypeError, ValueError):
            continue
        try:
            if rp.is_file():
                if rp.suffix.lower() == ORDER_PDF_EXT:
                    out.append(rp.resolve())
            elif rp.is_dir():
                for child in rp.rglob("*"):
                    if child.is_file() and child.suffix.lower() == ORDER_PDF_EXT:
                        out.append(child.resolve())
        except OSError:
            continue
    return out


def _unique_pdf_dest(input_dir: Path, src: Path) -> Path:
    """Pick input_dir / name that does not overwrite an existing file."""
    input_dir.mkdir(parents=True, exist_ok=True)
    stem = re.sub(r"[^\w\-.]+", "_", src.stem, flags=re.UNICODE).strip("._") or "order"
    stem = stem[:120]
    ext = ORDER_PDF_EXT
    candidate = input_dir / f"{stem}{ext}"
    if not candidate.exists():
        return candidate
    for i in range(2, 10_000):
        alt = input_dir / f"{stem} ({i}){ext}"
        if not alt.exists():
            return alt
    return input_dir / f"{stem}_{secrets.token_hex(4)}{ext}"


def _valid_mmdd(token: str) -> bool:
    """True if token is MMDD for a real calendar day (uses a leap year so Feb 29 is allowed)."""
    if len(token) != 4 or not token.isdigit():
        return False
    month = int(token[:2])
    day = int(token[2:])
    try:
        date(2000, month, day)
    except ValueError:
        return False
    return True


def _mmdd_folder_for_order_pdf(path: Path) -> str:
    """Pick backup subfolder name MMDD: from filename *_MMDD.pdf if valid, else today."""
    m = _ORDER_PDF_MMdd_SUFFIX.search(path.name)
    if m:
        tok = m.group(1)
        if _valid_mmdd(tok):
            return tok
    today = date.today()
    return f"{today.month:02d}{today.day:02d}"


def _unique_path_in_dir(directory: Path, filename: str) -> Path:
    """Pick directory / filename that does not overwrite an existing file."""
    directory.mkdir(parents=True, exist_ok=True)
    candidate = directory / filename
    if not candidate.exists():
        return candidate
    stem = Path(filename).stem
    ext = Path(filename).suffix
    for i in range(2, 10_000):
        alt = directory / f"{stem} ({i}){ext}"
        if not alt.exists():
            return alt
    return directory / f"{stem}_{secrets.token_hex(4)}{ext}"


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self._lang: Lang = "zh"
        self.minsize(UI_MIN_W, UI_MIN_H)
        self._wrap_sync_id: int | None = None

        self._job_var = tk.StringVar(value=JOB_ORDER[0])
        self._chinese = tk.BooleanVar(value=True)
        self._html = tk.BooleanVar(value=True)
        self._no_catalog = tk.BooleanVar(value=False)
        self._charm_mode_var = tk.StringVar(value="skip")
        self._charm_smart_import = tk.BooleanVar(value=False)
        self._charm_drop_vision = tk.BooleanVar(value=False)

        self._log_q: queue.Queue[str] = queue.Queue()
        self._run_busy = False

        # Widgets updated on language change
        self._w_title: ttk.Label | None = None
        self._w_quick_hint: ttk.Label | None = None
        self._w_quick_grp_sup: ttk.Label | None = None
        self._w_quick_grp_route: ttk.Label | None = None
        self._w_quick_grp_excel: ttk.Label | None = None
        self._w_quick_grp_web:   ttk.Label | None = None
        self._quick_file_btns: list[tuple[ttk.Button, str]] = []
        self._btn_edit_products: ttk.Button | None = None
        self._btn_catalog_backups: ttk.Button | None = None
        self._w_btn_data: ttk.Button | None = None
        self._w_lang: ttk.Button | None = None
        self._main_frame: ttk.LabelFrame | None = None
        self._job_rbs: list[tuple[ttk.Radiobutton, str]] = []
        self._job_blurbs: list[tuple[ttk.Label, str]] = []
        self._also_frame: ttk.LabelFrame | None = None
        self._chk_zh: ttk.Checkbutton | None = None
        self._chk_html: ttk.Checkbutton | None = None
        self._chk_nc: ttk.Checkbutton | None = None
        self._charm_run_frame: ttk.LabelFrame | None = None
        self._w_charm_how: ttk.Label | None = None
        self._w_charm_run_note: ttk.Label | None = None
        self._charm_mode_rbs: list[tuple[ttk.Radiobutton, str]] = []
        self._charm_mode_blurbs: list[tuple[ttk.Label, str]] = []
        self._charm_smart_cb: ttk.Checkbutton | None = None
        self._opts_frame: ttk.LabelFrame | None = None
        self._w_th_lbl: ttk.Label | None = None
        self._w_zx_lbl: ttk.Label | None = None
        self._w_cd_lbl: ttk.Label | None = None
        self._run_btn: tk.Button | None = None
        self._w_run_hint: tk.Label | None = None
        self._log_frame: ttk.LabelFrame | None = None
        self._notebook: ttk.Notebook | None = None
        self._tab_main: ttk.Frame | None = None
        self._tab_charms: ttk.Frame | None = None
        self._main_inner: ttk.Frame | None = None
        self._charm_inner: ttk.Frame | None = None
        self._w_step_strip: tk.Label | None = None
        self._charm_canvas: tk.Canvas | None = None
        self._charm_vsb: ttk.Scrollbar | None = None
        self._main_canvas: tk.Canvas | None = None
        self._main_vsb: ttk.Scrollbar | None = None
        self._charm_drop_frame: ttk.LabelFrame | None = None
        self._w_charm_drop_blurb: ttk.Label | None = None
        self._w_charm_drop_hint: tk.Label | None = None
        self._charm_drop_zone: tk.Frame | None = None
        self._btn_charm_browse: ttk.Button | None = None
        self._btn_charm_import: ttk.Button | None = None
        self._btn_charm_open_folder: ttk.Button | None = None
        self._charm_drop_vision_cb: ttk.Checkbutton | None = None
        self._charm_reorder_frame: ttk.LabelFrame | None = None
        self._w_charm_reorder_body: ttk.Label | None = None
        self._btn_charm_reorder_open: ttk.Button | None = None
        self._pdf_drop_frame: ttk.LabelFrame | None = None
        self._w_pdf_drop_blurb: ttk.Label | None = None
        self._pdf_drop_zone: tk.Frame | None = None
        self._w_pdf_drop_hint: tk.Label | None = None
        self._btn_pdf_browse: ttk.Button | None = None
        self._btn_pdf_run_new: ttk.Button | None = None
        self._btn_pdf_open_folder: ttk.Button | None = None
        self._btn_pdf_move_backup: ttk.Button | None = None

        self._build()
        self._apply_language()
        self._place_initial_window()
        self.after(200, self._drain_log)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _place_initial_window(self) -> None:
        self.update_idletasks()
        w, h = UI_DEFAULT_W, UI_DEFAULT_H
        sw = max(self.winfo_screenwidth(), w + 1)
        sh = max(self.winfo_screenheight(), h + 1)
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2 - 24)
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _schedule_wrap_sync(self, event: tk.Event) -> None:
        if event.widget is not self:
            return
        if self._wrap_sync_id is not None:
            try:
                self.after_cancel(self._wrap_sync_id)
            except tk.TclError:
                pass
        self._wrap_sync_id = self.after(120, self._sync_text_wraps)

    def _sync_text_wraps(self) -> None:
        self._wrap_sync_id = None
        self.update_idletasks()
        w = self.winfo_width()
        if w < 80:
            return
        edge = UI_MARGIN_X * 2 + 32
        wrap_full = max(360, w - edge)
        wrap_strip = max(300, w - edge)
        half = max(260, (w - edge) // 2 - 56)
        run_hint = max(260, w - 300)
        charm = max(400, w - edge - 24)
        if self._w_quick_hint:
            self._w_quick_hint.config(wraplength=wrap_full)
        if self._w_quick_grp_sup:
            self._w_quick_grp_sup.config(wraplength=wrap_full)
        if self._w_quick_grp_route:
            self._w_quick_grp_route.config(wraplength=wrap_full)
        if self._w_step_strip:
            self._w_step_strip.config(wraplength=wrap_strip)
        if self._w_run_hint:
            self._w_run_hint.config(wraplength=run_hint)
        for lbl, _jid in self._job_blurbs:
            lbl.config(wraplength=half)
        for lbl, _mid in self._charm_mode_blurbs:
            lbl.config(wraplength=max(260, charm - 24))
        if self._w_charm_how:
            self._w_charm_how.config(wraplength=max(320, charm - 8))
        if self._w_charm_run_note:
            self._w_charm_run_note.config(wraplength=max(320, charm - 8))
        if self._w_charm_reorder_body:
            self._w_charm_reorder_body.config(wraplength=max(320, charm - 8))
        if self._w_charm_drop_blurb:
            self._w_charm_drop_blurb.config(wraplength=max(280, charm - 24))
        if self._w_charm_drop_hint:
            self._w_charm_drop_hint.config(wraplength=max(240, w - 200))
        if self._w_pdf_drop_blurb:
            self._w_pdf_drop_blurb.config(wraplength=max(280, wrap_full - 40))
        if self._w_pdf_drop_hint:
            self._w_pdf_drop_hint.config(wraplength=max(240, w - 200))
        self.after_idle(self._sync_main_scroll)
        self.after_idle(self._sync_charm_scroll)

    def _on_close(self) -> None:
        if self._wrap_sync_id is not None:
            try:
                self.after_cancel(self._wrap_sync_id)
            except tk.TclError:
                pass
        try:
            self.unbind_all("<MouseWheel>")
            self.unbind_all("<Button-4>")
            self.unbind_all("<Button-5>")
        except tk.TclError:
            pass
        self.destroy()

    def _t(self, key: str, **fmt: object) -> str:
        s = CHROME[self._lang][key]
        return s.format(**fmt) if fmt else s

    def _hook_windnd_drop(self, widget: tk.Misc | None, handler: Callable[[list[Path]], None]) -> None:
        """Register *handler* for file/folder drops on *widget* (Windows + windnd only)."""
        if windnd is None or widget is None:
            return

        def _on_drop(files: object, _obj: object | None = None) -> None:
            paths = _decode_windnd_paths(files)
            self.after(0, lambda p=list(paths): handler(p))

        try:
            windnd.hook_dropfiles(widget, _on_drop, force_unicode=True)
        except (tk.TclError, OSError, AttributeError):
            pass

    def _reveal_directory(self, path: Path) -> None:
        """Open a directory in the OS file manager (cross-platform)."""
        path.mkdir(parents=True, exist_ok=True)
        target = str(path.resolve())
        try:
            if sys.platform == "win32":
                os.startfile(target)
            elif sys.platform == "darwin":
                subprocess.run(["open", target], check=False)
            else:
                subprocess.run(["xdg-open", target], check=False)
        except OSError as e:
            messagebox.showerror(self._t("file_open_fail_title"), self._t("file_open_fail_body", err=e))

    def _toggle_language(self) -> None:
        self._lang = "zh" if self._lang == "en" else "en"
        self._apply_language()

    def _apply_language(self) -> None:
        self.title(CHROME[self._lang]["win_title"])
        if self._w_lang:
            # Button label = the language you switch TO
            self._w_lang.config(text=CHROME[self._lang]["lang_btn"])
        if self._w_title:
            self._w_title.config(text=CHROME[self._lang]["header_title"])
        if self._w_quick_hint:
            self._w_quick_hint.config(text=CHROME[self._lang]["quick_hint"])
        if self._w_quick_grp_sup:
            self._w_quick_grp_sup.config(text=CHROME[self._lang]["quick_group_supplier"])
        if self._w_quick_grp_route:
            self._w_quick_grp_route.config(text=CHROME[self._lang]["quick_group_route"])
        if self._w_quick_grp_excel:
            self._w_quick_grp_excel.config(text=CHROME[self._lang]["quick_group_excel"])
        if self._w_quick_grp_web:
            self._w_quick_grp_web.config(text=CHROME[self._lang]["quick_group_web"])
        for btn, key in self._quick_file_btns:
            btn.config(text=CHROME[self._lang][key])
        if self._btn_edit_products:
            self._btn_edit_products.config(text=CHROME[self._lang]["btn_edit_products"])
        if hasattr(self, "_btn_orders_dashboard") and self._btn_orders_dashboard:
            self._btn_orders_dashboard.config(text=CHROME[self._lang]["btn_orders_dashboard"])
        if self._btn_catalog_backups:
            self._btn_catalog_backups.config(text=CHROME[self._lang]["btn_catalog_backups"])
        if self._w_btn_data:
            self._w_btn_data.config(text=CHROME[self._lang]["quick_data_folder"])
        if self._main_frame:
            self._main_frame.config(text=CHROME[self._lang]["main_frame"])
        for rb, job_id in self._job_rbs:
            title, _ = JOB_TEXT[job_id][self._lang]
            rb.config(text=title)
        for lbl, job_id in self._job_blurbs:
            _, blurb = JOB_TEXT[job_id][self._lang]
            lbl.config(text=blurb)
        if self._also_frame:
            self._also_frame.config(text=CHROME[self._lang]["also_frame"])
        if self._chk_zh:
            self._chk_zh.config(text=CHROME[self._lang]["chk_excel_zh"])
        if self._chk_html:
            self._chk_html.config(text=CHROME[self._lang]["chk_html"])
        if self._chk_nc:
            self._chk_nc.config(text=CHROME[self._lang]["chk_no_catalog"])
        if self._charm_drop_frame:
            self._charm_drop_frame.config(text=CHROME[self._lang]["charm_drop_section"])
        if self._w_charm_how:
            self._w_charm_how.config(text=CHROME[self._lang]["charm_how_body"])
        if self._charm_run_frame:
            self._charm_run_frame.config(text=CHROME[self._lang]["charm_run_section"])
        if self._w_charm_run_note:
            self._w_charm_run_note.config(text=CHROME[self._lang]["charm_run_note"])
        for rb, mode_id in self._charm_mode_rbs:
            title, _ = CHARM_MODE_TEXT[mode_id][self._lang]
            rb.config(text=title)
        for lbl, mode_id in self._charm_mode_blurbs:
            _, blurb = CHARM_MODE_TEXT[mode_id][self._lang]
            lbl.config(text=blurb)
        if self._charm_smart_cb:
            self._charm_smart_cb.config(text=CHROME[self._lang]["charm_smart"])
        if self._w_charm_drop_blurb:
            self._w_charm_drop_blurb.config(text=CHROME[self._lang]["charm_drop_blurb"])
        if self._w_charm_drop_hint:
            hk = "charm_drop_hint_dnd" if windnd is not None else "charm_drop_hint_no_dnd"
            self._w_charm_drop_hint.config(text=CHROME[self._lang][hk])
        if self._btn_charm_browse:
            self._btn_charm_browse.config(text=CHROME[self._lang]["charm_drop_browse"])
        if self._btn_charm_import:
            self._btn_charm_import.config(text=CHROME[self._lang]["charm_drop_import"])
        if self._btn_charm_open_folder:
            self._btn_charm_open_folder.config(text=CHROME[self._lang]["charm_drop_open_folder"])
        if self._charm_drop_vision_cb:
            self._charm_drop_vision_cb.config(text=CHROME[self._lang]["charm_drop_vision"])
        if self._charm_reorder_frame:
            self._charm_reorder_frame.config(text=CHROME[self._lang]["charm_reorder_section"])
        if self._w_charm_reorder_body:
            self._w_charm_reorder_body.config(text=CHROME[self._lang]["charm_reorder_body"])
        if self._btn_charm_reorder_open:
            self._btn_charm_reorder_open.config(
                text=CHROME[self._lang]["charm_reorder_btn_open"]
            )
        if self._pdf_drop_frame:
            self._pdf_drop_frame.config(text=CHROME[self._lang]["pdf_drop_frame"])
        if self._w_pdf_drop_blurb:
            self._w_pdf_drop_blurb.config(text=CHROME[self._lang]["pdf_drop_blurb"])
        if self._w_pdf_drop_hint:
            hk = "pdf_drop_hint_dnd" if windnd is not None else "pdf_drop_hint_no_dnd"
            self._w_pdf_drop_hint.config(text=CHROME[self._lang][hk])
        if self._btn_pdf_browse:
            self._btn_pdf_browse.config(text=CHROME[self._lang]["pdf_drop_browse"])
        if self._btn_pdf_run_new:
            self._btn_pdf_run_new.config(text=CHROME[self._lang]["pdf_drop_run"])
        if self._btn_pdf_open_folder:
            self._btn_pdf_open_folder.config(text=CHROME[self._lang]["pdf_drop_open_folder"])
        if self._btn_pdf_move_backup:
            self._btn_pdf_move_backup.config(text=CHROME[self._lang]["pdf_drop_move_backup"])
        if self._opts_frame:
            self._opts_frame.config(text=CHROME[self._lang]["opts_frame"])
        if self._w_th_lbl:
            self._w_th_lbl.config(text=CHROME[self._lang]["threshold_hint"])
        if self._w_zx_lbl:
            self._w_zx_lbl.config(text=CHROME[self._lang]["zh_exclude_hint"])
        if self._w_cd_lbl:
            self._w_cd_lbl.config(text=CHROME[self._lang]["charm_dir_hint"])
        if self._run_btn:
            self._run_btn.config(text=CHROME[self._lang]["run"])
        if self._w_run_hint:
            self._w_run_hint.config(text=CHROME[self._lang]["footer_run_hint"])
        if self._log_frame:
            self._log_frame.config(text=CHROME[self._lang]["output_frame"])
        if self._notebook and self._tab_main and self._tab_charms:
            self._notebook.tab(self._tab_main, text=CHROME[self._lang]["tab_orders"])
            self._notebook.tab(self._tab_charms, text=CHROME[self._lang]["tab_charms"])
        if self._w_step_strip:
            self._w_step_strip.config(text=CHROME[self._lang]["step_strip"])
        self.after_idle(self._sync_charm_scroll)
        self.after_idle(self._sync_main_scroll)
        self.after_idle(self._sync_text_wraps)

    def _sync_main_scroll(self) -> None:
        if self._main_canvas is None:
            return
        self.update_idletasks()
        box = self._main_canvas.bbox("all")
        if box:
            self._main_canvas.configure(scrollregion=box)
        self._maybe_toggle_main_vsb()

    def _maybe_toggle_main_vsb(self) -> None:
        c = self._main_canvas
        v = self._main_vsb
        if c is None or v is None:
            return
        try:
            c.update_idletasks()
            bbox = c.bbox("all")
            if not bbox:
                return
            content_h = bbox[3] - bbox[1]
            view_h = c.winfo_height()
        except tk.TclError:
            return
        if content_h <= view_h + 12:
            v.pack_forget()
        elif not v.winfo_ismapped():
            v.pack(side=tk.RIGHT, fill=tk.Y)

    def _sync_charm_scroll(self) -> None:
        if self._charm_canvas is None:
            return
        self.update_idletasks()
        box = self._charm_canvas.bbox("all")
        if box:
            self._charm_canvas.configure(scrollregion=box)
        self._maybe_toggle_charm_vsb()

    def _maybe_toggle_charm_vsb(self) -> None:
        c = self._charm_canvas
        v = self._charm_vsb
        if c is None or v is None:
            return
        try:
            c.update_idletasks()
            bbox = c.bbox("all")
            if not bbox:
                return
            content_h = bbox[3] - bbox[1]
            view_h = c.winfo_height()
        except tk.TclError:
            return
        if content_h <= view_h + 12:
            v.pack_forget()
        elif not v.winfo_ismapped():
            v.pack(side=tk.RIGHT, fill=tk.Y)

    def _setup_styles(self) -> None:
        s = ttk.Style()
        try:
            s.theme_use("clam")
        except tk.TclError:
            pass
        C = COLORS
        s.configure(".", background=C["app"])
        s.configure("TFrame", background=C["app"])
        s.configure("App.TFrame", background=C["app"])
        s.configure("Card.TFrame", background=C["card"])
        s.configure("TNotebook", background=C["app"], borderwidth=0)
        s.configure("TNotebook.Tab", padding=[20, 11], font=("Segoe UI", 10, "bold"))
        s.map("TNotebook.Tab",
              background=[("selected", C["card"]), ("!selected", C["app"])],
              foreground=[("selected", C["accent"]), ("!selected", C["muted"])])
        s.configure("Card.TLabelframe", background=C["card"], relief="solid", borderwidth=1,
                    bordercolor=C["border"])
        s.configure("Card.TLabelframe.Label", background=C["card"], foreground=C["accent"],
                    font=("Segoe UI", 11, "bold"))
        s.configure("TLabel", background=C["app"], foreground=C["text"], font=("Segoe UI", 10))
        s.configure("Card.TLabel", background=C["card"], foreground=C["text"], font=("Segoe UI", 10))
        s.configure("Muted.TLabel", background=C["card"], foreground=C["muted"], font=("Segoe UI", 10))
        s.configure("Title.TLabel", background=C["app"], foreground=C["text"], font=("Segoe UI", 16, "bold"))
        s.configure("Sub.TLabel", background=C["app"], foreground=C["muted"], font=("Segoe UI", 10))
        s.configure("TRadiobutton", background=C["card"], foreground=C["text"], font=("Segoe UI", 10))
        s.configure("TCheckbutton", background=C["card"], foreground=C["text"], font=("Segoe UI", 10))
        s.configure("Tool.TButton", font=("Segoe UI", 10), padding=(11, 7))
        s.map("Tool.TButton",
              background=[("active", C["accent_soft"])],
              foreground=[("active", C["accent"])])
        s.configure("Run.TButton", background=C["run"], foreground="#ffffff",
                    font=("Segoe UI", 12, "bold"), padding=(28, 12))
        s.map("Run.TButton", background=[("active", C["run_hover"]), ("disabled", "#94a3b8")])

    def _build(self) -> None:
        self.configure(bg=COLORS["app"])
        self._setup_styles()
        # Root layout must use grid (not pack) so the notebook cannot swallow vertical space
        # reserved for Run + log. On Windows, ttk.Notebook + pack(expand=True) often leaves
        # siblings clipped or painted under the native tab control; grid + explicit row weights fixes it.
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        hero = tk.Frame(self, bg=COLORS["hero"], height=5, highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        self.grid_rowconfigure(0, minsize=5)

        pad = {"padx": UI_MARGIN_X, "pady": 6}
        # Header + strip only (no expand): keeps the notebook + Run bar from fighting for height inside one frame.
        top_block = ttk.Frame(self, style="App.TFrame")
        top_block.grid(row=1, column=0, sticky="ew")

        header = ttk.Frame(top_block, style="App.TFrame")
        header.pack(fill=tk.X, **pad)

        top_row = ttk.Frame(header, style="App.TFrame")
        top_row.pack(fill=tk.X)
        self._w_lang = ttk.Button(top_row, text="", command=self._toggle_language, width=10, style="Tool.TButton")
        self._w_lang.pack(side=tk.RIGHT)

        self._w_title = ttk.Label(top_row, text="", style="Title.TLabel")
        self._w_title.pack(anchor=tk.W, fill=tk.X, expand=True)

        self._w_quick_hint = ttk.Label(header, text="", style="Sub.TLabel", wraplength=1080)
        self._w_quick_hint.pack(anchor=tk.W, pady=(6, 4))

        quick_rows = ttk.Frame(header, style="App.TFrame")
        quick_rows.pack(anchor=tk.W, fill=tk.X)

        def _mk_open_btn(parent: ttk.Frame, path: Path, chrome_key: str) -> None:
            b = ttk.Button(parent, text="", command=lambda p=path: self._open_file_path(p), style="Tool.TButton")
            b.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 4))
            self._quick_file_btns.append((b, chrome_key))

        grp_sup = ttk.Frame(quick_rows, style="App.TFrame")
        grp_sup.pack(anchor=tk.W, fill=tk.X, pady=(0, 2))
        self._w_quick_grp_sup = ttk.Label(grp_sup, text="", style="Sub.TLabel", wraplength=1060)
        self._w_quick_grp_sup.pack(anchor=tk.W, pady=(0, 4))
        row_sup = ttk.Frame(grp_sup, style="App.TFrame")
        row_sup.pack(anchor=tk.W, fill=tk.X)
        # Left cluster: file openers
        sup_left = ttk.Frame(row_sup, style="App.TFrame")
        sup_left.pack(side=tk.LEFT)
        _mk_open_btn(sup_left, FILE_SUPPLIER_CATALOG, "quick_catalog")
        self._btn_catalog_backups = ttk.Button(
            sup_left, text="", command=self._open_catalog_backups_dialog, style="Tool.TButton"
        )
        self._btn_catalog_backups.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 4))
        self._w_btn_data = ttk.Button(sup_left, text="", command=self._open_data_folder, style="Tool.TButton")
        self._w_btn_data.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 4))
        # Divider
        ttk.Separator(row_sup, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=(8, 8), pady=(0, 4))
        # Right cluster: editor / dashboard tools
        sup_right = ttk.Frame(row_sup, style="App.TFrame")
        sup_right.pack(side=tk.LEFT)
        self._btn_edit_products = ttk.Button(
            sup_right, text="", command=self._open_edit_products_dialog, style="Tool.TButton"
        )
        self._btn_edit_products.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 4))
        self._btn_orders_dashboard = ttk.Button(
            sup_right, text="", command=self._open_orders_dashboard, style="Tool.TButton"
        )
        self._btn_orders_dashboard.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 4))

        ttk.Separator(quick_rows, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(6, 8))

        grp_route = ttk.Frame(quick_rows, style="App.TFrame")
        grp_route.pack(anchor=tk.W, fill=tk.X, pady=(0, 0))
        self._w_quick_grp_route = ttk.Label(grp_route, text="", style="Sub.TLabel", wraplength=1060)
        self._w_quick_grp_route.pack(anchor=tk.W, pady=(0, 6))

        # ── Excel sub-row ─────────────────────────────────────────────────────
        row_excel = ttk.Frame(grp_route, style="App.TFrame")
        row_excel.pack(anchor=tk.W, fill=tk.X, pady=(0, 2))
        self._w_quick_grp_excel = ttk.Label(
            row_excel, text="", style="Muted.TLabel",
            font=("Segoe UI", 8, "bold"), foreground="#6b7280", width=8,
        )
        self._w_quick_grp_excel.pack(side=tk.LEFT, padx=(0, 4), pady=(0, 0))
        _mk_open_btn(row_excel, FILE_SHOPPING_ROUTE, "quick_route")
        _mk_open_btn(row_excel, FILE_SHOPPING_ROUTE_SIMPLE, "quick_route_simple")
        _mk_open_btn(row_excel, FILE_SHOPPING_ROUTE_ZH, "quick_route_zh")

        # ── Web sub-row (no horizontal rule — spacing + labels keep it clean) ─
        row_web = ttk.Frame(grp_route, style="App.TFrame")
        row_web.pack(anchor=tk.W, fill=tk.X, pady=(2, 0))
        self._w_quick_grp_web = ttk.Label(
            row_web, text="", style="Muted.TLabel",
            font=("Segoe UI", 8, "bold"), foreground="#6b7280", width=8,
        )
        self._w_quick_grp_web.pack(side=tk.LEFT, padx=(0, 4), pady=(0, 0))
        _mk_open_btn(row_web, FILE_SHOPPING_HTML, "quick_html")
        _mk_open_btn(row_web, FILE_SHOPPING_HTML_ZH, "quick_html_zh")

        strip = tk.Frame(top_block, bg=COLORS["strip"], highlightthickness=0)
        strip.pack(fill=tk.X, pady=(8, 0))
        strip_bar = tk.Frame(strip, bg=COLORS["strip_accent"], width=5, highlightthickness=0)
        strip_bar.pack(side=tk.LEFT, fill=tk.Y)
        self._w_step_strip = tk.Label(
            strip,
            text="",
            bg=COLORS["strip"],
            fg=COLORS["strip_text"],
            font=("Segoe UI", 10, "bold"),
            wraplength=1060,
            justify=tk.LEFT,
            padx=14,
            pady=11,
        )
        self._w_step_strip.pack(side=tk.LEFT, anchor=tk.W, fill=tk.X, expand=True)

        self._notebook = ttk.Notebook(self)
        self._notebook.grid(row=2, column=0, sticky="nsew", padx=(UI_MARGIN_X, UI_MARGIN_X), pady=(10, 0))

        self._tab_main = ttk.Frame(self._notebook, style="Card.TFrame")
        self._tab_charms = ttk.Frame(self._notebook, style="Card.TFrame")
        self._notebook.add(self._tab_main, text=" ")
        self._notebook.add(self._tab_charms, text=" ")
        self._notebook.select(0)

        # Notebook panes clip children; tall content needs an inner canvas (same idea as Charms tab).
        main_wrap = ttk.Frame(self._tab_main, style="Card.TFrame")
        main_wrap.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        self._main_canvas = tk.Canvas(main_wrap, bg=COLORS["card"], highlightthickness=0, borderwidth=0)
        self._main_vsb = ttk.Scrollbar(main_wrap, orient=tk.VERTICAL, command=self._main_canvas.yview)
        self._main_canvas.configure(yscrollcommand=self._main_vsb.set)
        main_inner = ttk.Frame(self._main_canvas, style="Card.TFrame")
        self._main_inner = main_inner
        main_win = self._main_canvas.create_window((0, 0), window=main_inner, anchor=tk.NW)

        def _main_cw(event: tk.Event) -> None:
            self._main_canvas.itemconfigure(main_win, width=event.width)
            self.after_idle(self._maybe_toggle_main_vsb)

        def _main_configure(_: tk.Event | None = None) -> None:
            self._main_canvas.configure(scrollregion=self._main_canvas.bbox("all"))
            self.after_idle(self._maybe_toggle_main_vsb)

        self._main_canvas.bind("<Configure>", _main_cw)
        main_inner.bind("<Configure>", _main_configure)
        self._main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._main_vsb.pack(side=tk.RIGHT, fill=tk.Y)

        wrap_main = max(360, 1020)
        self._pdf_drop_frame = ttk.LabelFrame(main_inner, text="", style="Card.TLabelframe", padding=14)
        self._pdf_drop_frame.pack(fill=tk.X, padx=10, pady=(10, 10))
        self._w_pdf_drop_blurb = ttk.Label(
            self._pdf_drop_frame,
            text="",
            style="Muted.TLabel",
            wraplength=wrap_main - 24,
            justify=tk.LEFT,
        )
        self._w_pdf_drop_blurb.pack(anchor=tk.W)
        self._pdf_drop_zone = tk.Frame(
            self._pdf_drop_frame,
            bg=COLORS["drop_zone"],
            highlightthickness=2,
            highlightbackground=COLORS["drop_border"],
            height=DROP_ZONE_H,
            highlightcolor=COLORS["accent"],
        )
        self._pdf_drop_zone.pack(fill=tk.X, pady=(8, 10))
        self._pdf_drop_zone.pack_propagate(False)
        self._w_pdf_drop_hint = tk.Label(
            self._pdf_drop_zone,
            text="",
            bg=COLORS["drop_zone"],
            fg=COLORS["drop_hint"],
            font=("Segoe UI", 10),
            wraplength=760,
            justify=tk.CENTER,
        )
        self._w_pdf_drop_hint.pack(expand=True)
        pdf_btns = ttk.Frame(self._pdf_drop_frame, style="Card.TFrame")
        pdf_btns.pack(fill=tk.X)
        self._btn_pdf_browse = ttk.Button(
            pdf_btns,
            text="",
            command=self._pdf_browse,
            style="Tool.TButton",
        )
        self._btn_pdf_browse.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_pdf_run_new = ttk.Button(
            pdf_btns,
            text="",
            command=self._pdf_run_new_batch,
            style="Tool.TButton",
        )
        self._btn_pdf_run_new.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_pdf_open_folder = ttk.Button(
            pdf_btns,
            text="",
            command=self._open_input,
            style="Tool.TButton",
        )
        self._btn_pdf_open_folder.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_pdf_move_backup = ttk.Button(
            pdf_btns,
            text="",
            command=self._pdf_move_to_backup,
            style="Tool.TButton",
        )
        self._btn_pdf_move_backup.pack(side=tk.LEFT, padx=(0, 8))

        self._hook_windnd_drop(self._pdf_drop_zone, self._pdf_on_paths_dropped)

        cols = ttk.Frame(main_inner, style="Card.TFrame")
        cols.pack(fill=tk.X, padx=10, pady=(0, 10))
        left_col = ttk.Frame(cols, style="Card.TFrame")
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
        right_col = ttk.Frame(cols, style="Card.TFrame")
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))

        self._main_frame = ttk.LabelFrame(left_col, text="", style="Card.TLabelframe", padding=14)
        self._main_frame.pack(fill=tk.X, pady=(0, 10))

        n_jobs = len(JOB_ORDER)
        for i, job_id in enumerate(JOB_ORDER):
            title, blurb = JOB_TEXT[job_id]["en"]
            rb = ttk.Radiobutton(self._main_frame, text=title, value=job_id, variable=self._job_var)
            rb.pack(anchor=tk.W, pady=(0, 2))
            self._job_rbs.append((rb, job_id))
            lb = ttk.Label(self._main_frame, text=blurb, style="Muted.TLabel", wraplength=480)
            lb.pack(anchor=tk.W, padx=(22, 0), pady=(0, 10 if i < n_jobs - 1 else 0))
            self._job_blurbs.append((lb, job_id))

        self._also_frame = ttk.LabelFrame(left_col, text="", style="Card.TLabelframe", padding=14)
        self._also_frame.pack(fill=tk.X, pady=(0, 0))
        self._chk_zh = ttk.Checkbutton(self._also_frame, text="", variable=self._chinese)
        self._chk_zh.pack(anchor=tk.W)
        self._chk_html = ttk.Checkbutton(self._also_frame, text="", variable=self._html)
        self._chk_html.pack(anchor=tk.W)
        self._chk_nc = ttk.Checkbutton(self._also_frame, text="", variable=self._no_catalog)
        self._chk_nc.pack(anchor=tk.W)

        self._opts_frame = ttk.LabelFrame(right_col, text="", style="Card.TLabelframe", padding=14)
        self._opts_frame.pack(fill=tk.X)
        self._w_th_lbl = ttk.Label(self._opts_frame, text="", style="Card.TLabel")
        self._w_th_lbl.pack(anchor=tk.W)
        self._threshold = ttk.Entry(self._opts_frame, width=14)
        self._threshold.pack(anchor=tk.W, pady=(2, 8))
        self._w_zx_lbl = ttk.Label(self._opts_frame, text="", style="Card.TLabel")
        self._w_zx_lbl.pack(anchor=tk.W)
        self._zh_exclude = ttk.Entry(self._opts_frame, width=52)
        self._zh_exclude.pack(anchor=tk.W, pady=(2, 8))
        self._w_cd_lbl = ttk.Label(self._opts_frame, text="", style="Card.TLabel")
        self._w_cd_lbl.pack(anchor=tk.W)
        self._charm_dir = ttk.Entry(self._opts_frame, width=52)
        self._charm_dir.pack(anchor=tk.W, pady=(2, 0))

        self.after(120, _main_configure)

        ch_wrap = ttk.Frame(self._tab_charms, style="Card.TFrame")
        ch_wrap.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        self._charm_canvas = tk.Canvas(ch_wrap, bg=COLORS["card"], highlightthickness=0, borderwidth=0)
        self._charm_vsb = ttk.Scrollbar(ch_wrap, orient=tk.VERTICAL, command=self._charm_canvas.yview)
        self._charm_canvas.configure(yscrollcommand=self._charm_vsb.set)
        ch_inner = ttk.Frame(self._charm_canvas, style="Card.TFrame")
        self._charm_inner = ch_inner
        ch_win = self._charm_canvas.create_window((0, 0), window=ch_inner, anchor=tk.NW)

        def _charm_cw(event: tk.Event) -> None:
            self._charm_canvas.itemconfigure(ch_win, width=event.width)
            self.after_idle(self._maybe_toggle_charm_vsb)

        def _charm_configure(_: tk.Event | None = None) -> None:
            self._charm_canvas.configure(scrollregion=self._charm_canvas.bbox("all"))
            self.after_idle(self._maybe_toggle_charm_vsb)

        self._charm_canvas.bind("<Configure>", _charm_cw)
        ch_inner.bind("<Configure>", _charm_configure)
        self._charm_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._charm_vsb.pack(side=tk.RIGHT, fill=tk.Y)

        def _pointer_in_canvas(c: tk.Canvas | None, event: tk.Event) -> bool:
            if c is None:
                return False
            try:
                rx, ry = c.winfo_rootx(), c.winfo_rooty()
                rw, rh = c.winfo_width(), c.winfo_height()
            except tk.TclError:
                return False
            return rw > 2 and rh > 2 and rx <= event.x_root < rx + rw and ry <= event.y_root < ry + rh

        def _canvas_for_wheel_widget(widget: tk.Misc | None) -> tk.Canvas | None:
            """Map the widget under the pointer to the notebook tab canvas that should scroll."""
            if widget is None:
                return None
            mi, ci = self._main_inner, self._charm_inner
            w: tk.Misc | None = widget
            for _ in range(64):
                if w is None or w is self:
                    break
                if ci is not None and w is ci:
                    return self._charm_canvas
                if mi is not None and w is mi:
                    return self._main_canvas
                try:
                    parent = w.winfo_parent()
                    w = self.nametowidget(parent) if parent else None
                except tk.TclError:
                    break
            return None

        def _canvas_for_wheel_event(event: tk.Event) -> tk.Canvas | None:
            c = _canvas_for_wheel_widget(event.widget)
            if c is not None:
                return c
            try:
                sel = self._notebook.select() if self._notebook else ""
            except tk.TclError:
                sel = ""
            if sel == str(self._tab_charms):
                order = (self._charm_canvas, self._main_canvas)
            elif sel == str(self._tab_main):
                order = (self._main_canvas, self._charm_canvas)
            else:
                order = (self._main_canvas, self._charm_canvas)
            for cand in order:
                if _pointer_in_canvas(cand, event):
                    return cand
            return None

        def _on_mousewheel(event: tk.Event) -> str | None:
            delta = int(-1 * (event.delta / 120)) if getattr(event, "delta", 0) else 0
            if not delta:
                return None
            target = _canvas_for_wheel_event(event)
            if target is None:
                return None
            target.yview_scroll(delta, "units")
            return "break"

        def _on_linux_scroll(event: tk.Event) -> str | None:
            delta = -1 if event.num == 4 else 1
            target = _canvas_for_wheel_event(event)
            if target is None:
                return None
            target.yview_scroll(delta, "units")
            return "break"

        self.bind_all("<MouseWheel>", _on_mousewheel)
        self.bind_all("<Button-4>", _on_linux_scroll)
        self.bind_all("<Button-5>", _on_linux_scroll)

        def _nb_refresh_scroll(_: tk.Event) -> None:
            self.after_idle(self._sync_main_scroll)
            self.after_idle(self._sync_charm_scroll)

        self._notebook.bind("<<NotebookTabChanged>>", _nb_refresh_scroll)

        wrap_ch = 1020
        self._charm_drop_frame = ttk.LabelFrame(ch_inner, text="", style="Card.TLabelframe", padding=14)
        self._charm_drop_frame.pack(fill=tk.X)
        self._w_charm_how = ttk.Label(
            self._charm_drop_frame,
            text="",
            style="Muted.TLabel",
            wraplength=wrap_ch - 16,
            justify=tk.LEFT,
        )
        self._w_charm_how.pack(anchor=tk.W, pady=(0, 10))
        self._w_charm_drop_blurb = ttk.Label(
            self._charm_drop_frame,
            text="",
            style="Muted.TLabel",
            wraplength=wrap_ch - 24,
            justify=tk.LEFT,
        )
        self._w_charm_drop_blurb.pack(anchor=tk.W)

        self._charm_drop_zone = tk.Frame(
            self._charm_drop_frame,
            bg=COLORS["drop_zone"],
            highlightthickness=2,
            highlightbackground=COLORS["drop_border"],
            height=DROP_ZONE_H,
            highlightcolor=COLORS["accent"],
        )
        self._charm_drop_zone.pack(fill=tk.X, pady=(8, 10))
        self._charm_drop_zone.pack_propagate(False)
        self._w_charm_drop_hint = tk.Label(
            self._charm_drop_zone,
            text="",
            bg=COLORS["drop_zone"],
            fg=COLORS["drop_hint"],
            font=("Segoe UI", 10),
            wraplength=760,
            justify=tk.CENTER,
        )
        self._w_charm_drop_hint.pack(expand=True)

        drop_btns = ttk.Frame(self._charm_drop_frame, style="Card.TFrame")
        drop_btns.pack(fill=tk.X)
        self._btn_charm_browse = ttk.Button(
            drop_btns,
            text="",
            command=self._charm_browse_photos,
            style="Tool.TButton",
        )
        self._btn_charm_browse.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_charm_import = ttk.Button(
            drop_btns,
            text="",
            command=self._charm_import_workbook,
            style="Tool.TButton",
        )
        self._btn_charm_import.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_charm_open_folder = ttk.Button(
            drop_btns,
            text="",
            command=self._open_charm_images_folder,
            style="Tool.TButton",
        )
        self._btn_charm_open_folder.pack(side=tk.LEFT, padx=(0, 8))

        self._charm_drop_vision_cb = ttk.Checkbutton(
            self._charm_drop_frame,
            text="",
            variable=self._charm_drop_vision,
        )
        self._charm_drop_vision_cb.pack(anchor=tk.W, pady=(10, 0))

        self._hook_windnd_drop(self._charm_drop_zone, self._charm_on_paths_dropped)

        self._charm_run_frame = ttk.LabelFrame(ch_inner, text="", style="Card.TLabelframe", padding=14)
        self._charm_run_frame.pack(fill=tk.X, pady=(14, 0))
        self._w_charm_run_note = ttk.Label(
            self._charm_run_frame,
            text="",
            style="Muted.TLabel",
            wraplength=wrap_ch - 16,
            justify=tk.LEFT,
        )
        self._w_charm_run_note.pack(anchor=tk.W, pady=(0, 12))

        for mode_id in CHARM_MODE_ORDER:
            title, blurb = CHARM_MODE_TEXT[mode_id]["en"]
            rb = ttk.Radiobutton(
                self._charm_run_frame,
                text=title,
                value=mode_id,
                variable=self._charm_mode_var,
            )
            rb.pack(anchor=tk.W)
            self._charm_mode_rbs.append((rb, mode_id))
            lb = ttk.Label(self._charm_run_frame, text=blurb, style="Muted.TLabel", wraplength=wrap_ch - 24)
            lb.pack(anchor=tk.W, padx=(22, 0), pady=(0, 4))
            self._charm_mode_blurbs.append((lb, mode_id))

        self._charm_smart_cb = ttk.Checkbutton(
            self._charm_run_frame,
            text="",
            variable=self._charm_smart_import,
        )
        self._charm_smart_cb.pack(anchor=tk.W, pady=(8, 0))

        def _sync_charm_smart_state(*_: object) -> None:
            m = self._charm_mode_var.get()
            ok = m in ("try_add_photos", "add_photos")
            if ok:
                self._charm_smart_cb.state(["!disabled"])
            else:
                self._charm_smart_cb.state(["disabled"])
                self._charm_smart_import.set(False)

        self._charm_mode_var.trace_add("write", lambda *_: _sync_charm_smart_state())
        _sync_charm_smart_state()

        # ---- Section C: Reorder / Renumber charm codes ----
        self._charm_reorder_frame = ttk.LabelFrame(
            ch_inner, text="", style="Card.TLabelframe", padding=14
        )
        self._charm_reorder_frame.pack(fill=tk.X, pady=(14, 0))
        self._w_charm_reorder_body = ttk.Label(
            self._charm_reorder_frame,
            text="",
            style="Muted.TLabel",
            wraplength=wrap_ch - 16,
            justify=tk.LEFT,
        )
        self._w_charm_reorder_body.pack(anchor=tk.W, pady=(0, 10))
        reorder_btns = ttk.Frame(self._charm_reorder_frame, style="Card.TFrame")
        reorder_btns.pack(fill=tk.X)
        self._btn_charm_reorder_open = ttk.Button(
            reorder_btns,
            text="",
            command=self._open_charm_reorder_dialog,
            style="Tool.TButton",
        )
        self._btn_charm_reorder_open.pack(side=tk.LEFT, padx=(0, 8))

        self.after(120, _charm_configure)

        foot_sep = tk.Frame(self, bg=COLORS["separator"], height=1, highlightthickness=0)
        foot_sep.grid(row=3, column=0, sticky="ew", pady=(10, 0))
        foot = tk.Frame(self, bg=COLORS["accent_soft"], highlightthickness=0)
        foot.grid(row=4, column=0, sticky="ew")
        fin = tk.Frame(foot, bg=COLORS["accent_soft"])
        fin.pack(fill=tk.X, padx=UI_MARGIN_X, pady=14)
        self._run_btn = tk.Button(
            fin,
            text="",
            command=self._run,
            bg=COLORS["run"],
            fg="#ffffff",
            activebackground=COLORS["run_hover"],
            activeforeground="#ffffff",
            font=("Segoe UI", 12, "bold"),
            padx=28,
            pady=10,
            cursor="hand2",
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=0,
        )
        self._run_btn.pack(side=tk.LEFT)

        self._w_run_hint = tk.Label(
            fin,
            text="",
            bg=COLORS["accent_soft"],
            fg=COLORS["muted"],
            font=("Segoe UI", 10),
            wraplength=760,
            justify=tk.LEFT,
        )
        self._w_run_hint.pack(side=tk.LEFT, padx=(18, 0), anchor=tk.NW)

        self._log_frame = ttk.LabelFrame(self, text="", style="Card.TLabelframe", padding=10)
        self._log_frame.grid(row=5, column=0, sticky="ew", padx=UI_MARGIN_X, pady=(4, 14))
        self._log = scrolledtext.ScrolledText(
            self._log_frame,
            height=UI_LOG_LINES,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg=COLORS["log_bg"],
            fg=COLORS["text"],
            insertbackground=COLORS["text"],
            selectbackground=COLORS["accent_soft"],
            selectforeground=COLORS["text"],
            relief=tk.FLAT,
            borderwidth=0,
        )
        self._log.pack(fill=tk.BOTH, expand=True)

        self._append_log(CHROME[self._lang]["log_ready"])

        def _stack_run_and_log_above_notebook() -> None:
            # Native ttk.Notebook can paint above later siblings; keep Run + output on top.
            try:
                foot_sep.lift(self._notebook)
                foot.lift(foot_sep)
                self._log_frame.lift(foot)
            except tk.TclError:
                pass

        self.after_idle(_stack_run_and_log_above_notebook)

        self.bind("<Configure>", self._schedule_wrap_sync, add="+")

    def _append_log(self, text: str) -> None:
        self._log.insert(tk.END, text)
        self._log.see(tk.END)

    def _drain_log(self) -> None:
        try:
            while True:
                self._append_log(self._log_q.get_nowait())
        except queue.Empty:
            pass
        self.after(200, self._drain_log)

    def _open_input(self) -> None:
        self._reveal_directory(DEFAULT_ORDER_INPUT_DIR)

    def _list_input_order_pdfs(self) -> list[Path]:
        """Top-level .pdf files in the order input folder (non-recursive)."""
        d = DEFAULT_ORDER_INPUT_DIR
        if not d.is_dir():
            return []
        out: list[Path] = []
        try:
            for child in d.iterdir():
                if child.is_file() and child.suffix.lower() == ORDER_PDF_EXT:
                    out.append(child)
        except OSError:
            return []
        out.sort(key=lambda p: p.name.lower())
        return out

    def _pdf_move_to_backup(self) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        pdfs = self._list_input_order_pdfs()
        if not pdfs:
            messagebox.showinfo(self._t("pdf_drop_frame"), self._t("pdf_backup_empty"))
            return
        in_path = str(DEFAULT_ORDER_INPUT_DIR.resolve())
        if not messagebox.askyesno(
            self._t("pdf_backup_confirm_title"),
            self._t("pdf_backup_confirm_body", n=len(pdfs), path=in_path),
            parent=self,
        ):
            return
        self._append_log(self._t("pdf_backup_log_start"))
        ok = 0
        err = 0
        for src in pdfs:
            mmdd = _mmdd_folder_for_order_pdf(src)
            dest_dir = DEFAULT_BACKUP_DIR / mmdd
            try:
                dest = _unique_path_in_dir(dest_dir, src.name)
                shutil.move(str(src), str(dest))
                ok += 1
                self._append_log(
                    self._t("pdf_backup_log_line", src=src.name, mmdd=mmdd, dest=dest.name),
                )
            except OSError:
                err += 1
                self._append_log(f"  ! failed: {src.name}\n")
        if ok:
            self._append_log(self._t("pdf_backup_done", n=ok))
        if err:
            self._append_log(self._t("pdf_backup_errors", n=err))

    def _collect_generator_args(self, job_id: str, *, include_charm_steps: bool = True) -> list[str]:
        args: list[str] = ["--project-dir", str(PROJECT_ROOT)]
        main_flag = flag_for_job(job_id)
        if main_flag:
            args.append(main_flag)
        if self._no_catalog.get():
            args.append("--no-catalog-update")
        if self._chinese.get():
            args.append("--chinese")
        if self._html.get():
            args.append("--html")
        th = self._threshold.get().strip()
        if th:
            args.extend(["--threshold", th])
        zx = self._zh_exclude.get().strip()
        if zx:
            args.extend(["--chinese-exclude-shops", zx])
        cd = self._charm_dir.get().strip()
        if cd:
            args.extend(["--charm-images-dir", cd])
        if include_charm_steps:
            mode = self._charm_mode_var.get()
            args.extend(CHARM_MODE_FLAGS.get(mode, []))
            if mode in ("try_add_photos", "add_photos"):
                args.extend(charm_import_pattern_argv())
            if self._charm_smart_import.get() and mode in ("try_add_photos", "add_photos"):
                args.append("--import-charm-vision-sku")
        return args

    def _spawn_generator(self, gen_args: list[str], *, log_intro: str) -> None:
        self._run_busy = True
        self._set_chrome_busy(True)
        self._append_log(log_intro)
        cmd = [sys.executable, str(GENERATOR), *gen_args]
        self._append_log(" ".join(cmd) + "\n\n")

        def work() -> None:
            try:
                proc = subprocess.run(
                    cmd,
                    cwd=str(PROJECT_ROOT),
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                )
                out = proc.stdout or ""
                err = proc.stderr or ""
                self._log_q.put(out)
                if err:
                    self._log_q.put(self._t("log_messages") + err)
                self._log_q.put(self._t("log_finished", code=proc.returncode))
            except Exception as e:
                self._log_q.put(self._t("log_error", e=e))
            finally:

                def done() -> None:
                    self._run_busy = False
                    self._set_chrome_busy(False)

                self.after(0, done)

        threading.Thread(target=work, daemon=True).start()

    def _copy_pdfs_to_input(self, paths: list[Path]) -> int:
        input_dir = DEFAULT_ORDER_INPUT_DIR
        pdfs = _expand_pdf_paths(paths)
        n = 0
        for src in pdfs:
            try:
                dest = _unique_pdf_dest(input_dir, src)
                shutil.copy2(src, dest)
                n += 1
            except OSError:
                continue
        return n

    def _pdf_on_paths_dropped(self, paths: list[Path]) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        if not paths:
            return
        n = self._copy_pdfs_to_input(paths)
        if n == 0:
            messagebox.showinfo(self._t("pdf_drop_frame"), self._t("pdf_drop_no_valid"))
            return
        self._append_log(self._t("pdf_drop_copied", n=n))

    def _pdf_browse(self) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        picked = filedialog.askopenfilenames(
            parent=self,
            filetypes=[("PDF", "*.pdf"), ("All", "*.*")],
            title=self._t("pdf_drop_browse_title"),
        )
        if not picked:
            return
        self._pdf_on_paths_dropped([Path(s) for s in picked])

    def _pdf_run_new_batch(self) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        if not GENERATOR.is_file():
            messagebox.showerror(self._t("msg_missing_title"), self._t("msg_missing") + str(GENERATOR))
            return
        self._spawn_generator(
            self._collect_generator_args("new_batch", include_charm_steps=False),
            log_intro=self._t("pdf_new_batch_start"),
        )

    def _open_orders_dashboard(self) -> None:
        if load_cache is None or _style_has is None or _normalize is None:
            messagebox.showerror("Missing imports", "Cannot open dashboard: core imports unavailable.")
            return
        if not FILE_ORDERS_CACHE.exists():
            messagebox.showinfo("No data", "No order cache found.\nRun a job first to populate orders.")
            return

        items, _ = load_cache(FILE_ORDERS_CACHE)
        if not items:
            messagebox.showinfo("No data", "Order cache is empty.\nRun a job first.")
            return

        title_to_row: dict[str, int] = {}
        if list_product_map_rows_for_picker is not None and FILE_SUPPLIER_CATALOG.exists():
            try:
                for pr in list_product_map_rows_for_picker(FILE_SUPPLIER_CATALOG):
                    full_key = _normalize(pr.title)
                    # Register both the full key and the 50-char-truncated variant.
                    # Items cached before this fix stored norm_title with [:50];
                    # registering both ensures either lookup hits the right row.
                    title_to_row[full_key] = pr.row_num
                    title_to_row[full_key[:50]] = pr.row_num
            except Exception:
                pass

        sup_shops: list[str] = []
        sup_stalls: list[str] = []
        sup_shop_stalls: dict[str, str] = {}   # shop_name → stall
        sup_stall_shops: dict[str, str] = {}   # stall → shop_name
        try:
            import openpyxl as _xl
            _wb = _xl.load_workbook(FILE_SUPPLIER_CATALOG, read_only=True, data_only=True)
            if "Suppliers" in _wb.sheetnames:
                _ws = _wb["Suppliers"]
                _shop_ci = _stall_ci = None
                for ci in range(1, 20):
                    h = str(_ws.cell(1, ci).value or "").strip().lower()
                    if h == "shop name":
                        _shop_ci = ci
                    elif h == "stall":
                        _stall_ci = ci
                for r in _ws.iter_rows(min_row=2, values_only=False):
                    _sv = str(r[_shop_ci - 1].value or "").strip() if _shop_ci else ""
                    _stv = str(r[_stall_ci - 1].value or "").strip() if _stall_ci else ""
                    if _sv and _sv not in sup_shops:
                        sup_shops.append(_sv)
                    if _stv and _stv not in sup_stalls:
                        sup_stalls.append(_stv)
                    if _sv and _stv:
                        sup_shop_stalls.setdefault(_sv, _stv)
                        sup_stall_shops.setdefault(_stv, _sv)
            _wb.close()
        except Exception:
            pass

        charm_library: dict = {}
        charm_codes: list[str] = []
        charm_shop_names: list[str] = []
        charm_shop_stalls: dict[str, str] = {}   # shop_name → stall from Charm Shops tab
        try:
            if load_charm_library is not None and FILE_SUPPLIER_CATALOG.exists():
                charm_library = load_charm_library(FILE_SUPPLIER_CATALOG)
                charm_codes = list(charm_library.keys())
            if load_charm_shops is not None and FILE_SUPPLIER_CATALOG.exists():
                for cs in load_charm_shops(FILE_SUPPLIER_CATALOG):
                    if cs.shop_name:
                        charm_shop_names.append(cs.shop_name)
                        if cs.stall:
                            charm_shop_stalls[cs.shop_name] = cs.stall
        except Exception:
            pass

        # Load canonical product photos from the catalog (norm_title → jpeg bytes).
        # Every order for the same product always shows the same photo this way,
        # regardless of which PDF batch it was originally extracted from.
        catalog_photos: dict[str, bytes] = {}
        if get_catalog_photo_map is not None and FILE_SUPPLIER_CATALOG.exists():
            try:
                catalog_photos = get_catalog_photo_map(FILE_SUPPLIER_CATALOG)
            except Exception:
                pass

        _OrdersDashboardDialog(
            self, items,
            title_to_row=title_to_row,
            supplier_shops=sup_shops,
            supplier_stalls=sup_stalls,
            supplier_shop_stalls=sup_shop_stalls,
            supplier_stall_shops=sup_stall_shops,
            charm_codes=charm_codes,
            charm_shop_names=charm_shop_names,
            charm_library=charm_library,
            charm_shop_stalls=charm_shop_stalls,
            catalog_photos=catalog_photos,
        )

    def _open_edit_products_dialog(self) -> None:
        if update_product_map_cells is None or list_product_map_rows_for_picker is None:
            messagebox.showerror(
                self._t("msg_missing_title"), self._t("edit_no_import"),
            )
            return
        if not FILE_SUPPLIER_CATALOG.exists():
            messagebox.showinfo(
                self._t("msg_missing_title"),
                self._t("msg_missing") + str(FILE_SUPPLIER_CATALOG),
            )
            return
        try:
            all_rows = list_product_map_rows_for_picker(FILE_SUPPLIER_CATALOG)
        except OSError as e:
            messagebox.showerror(self._t("file_open_fail_title"), str(e))
            return
        if not all_rows:
            messagebox.showinfo(self._t("edit_title"), self._t("edit_empty"))
            return

        row_photos: dict[int, bytes] = {}
        try:
            if extract_photos_from_xlsx is not None:
                row_photos = extract_photos_from_xlsx(
                    FILE_SUPPLIER_CATALOG, sheet_name=CATALOG_SHEET, photo_col_idx=0,
                )
        except Exception:
            pass

        supplier_shops: list[str] = []
        supplier_stalls: list[str] = []
        try:
            import openpyxl as _xl
            _wb = _xl.load_workbook(FILE_SUPPLIER_CATALOG, read_only=True, data_only=True)
            if "Suppliers" in _wb.sheetnames:
                _ws_s = _wb["Suppliers"]
                _shop_ci = _stall_ci = None
                for ci in range(1, 20):
                    h = str(_ws_s.cell(1, ci).value or "").strip().lower()
                    if h == "shop name":
                        _shop_ci = ci
                    elif h == "stall":
                        _stall_ci = ci
                for r in _ws_s.iter_rows(min_row=2, values_only=False):
                    if _shop_ci is not None:
                        v = str(r[_shop_ci - 1].value or "").strip()
                        if v and v not in supplier_shops:
                            supplier_shops.append(v)
                    if _stall_ci is not None:
                        v = str(r[_stall_ci - 1].value or "").strip()
                        if v and v not in supplier_stalls:
                            supplier_stalls.append(v)
            _wb.close()
        except Exception:
            pass

        charm_shop_names: list[str] = []
        charm_entries: dict = {}
        try:
            if load_charm_shops is not None:
                for cs in load_charm_shops(FILE_SUPPLIER_CATALOG):
                    if cs.shop_name and cs.shop_name not in charm_shop_names:
                        charm_shop_names.append(cs.shop_name)
            if load_charm_library is not None:
                charm_entries = load_charm_library(FILE_SUPPLIER_CATALOG)
        except Exception:
            pass

        _ProductMapEditorDialog(
            self,
            all_rows,
            row_photos,
            supplier_shops=supplier_shops,
            supplier_stalls=supplier_stalls,
            charm_shop_names=charm_shop_names,
            charm_entries=charm_entries,
        )

    def _open_discontinue_dialog(self) -> None:
        if (
            list_product_map_rows_for_picker is None
            or mark_product_map_discontinued_by_row is None
            or extract_photos_from_xlsx is None
        ):
            messagebox.showerror(self._t("msg_missing_title"), self._t("discontinue_no_import"))
            return
        if not FILE_SUPPLIER_CATALOG.exists():
            messagebox.showinfo(self._t("msg_missing_title"), self._t("msg_missing") + str(FILE_SUPPLIER_CATALOG))
            return
        try:
            all_rows = list_product_map_rows_for_picker(FILE_SUPPLIER_CATALOG)
        except OSError as e:
            messagebox.showerror(self._t("file_open_fail_title"), str(e))
            return
        if not all_rows:
            messagebox.showinfo(self._t("msg_missing_title"), self._t("discontinue_empty"))
            return

        pil_ok = Image is not None and ImageTk is not None
        try:
            row_photos: dict[int, bytes] = extract_photos_from_xlsx(
                FILE_SUPPLIER_CATALOG,
                sheet_name=CATALOG_SHEET,
                photo_col_idx=0,
            )
        except Exception:
            row_photos = {}

        d = tk.Toplevel(self)
        d.title(self._t("discontinue_title"))
        d.transient(self)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.minsize(1020, 620)
        d.geometry("1320x820")

        tk_img_refs: list[object] = []
        row_by_iid: dict[str, ProductMapPickerRow] = {}
        preview_photo_ref: list[object] = []

        def bytes_to_thumb(b: bytes | None, px: int) -> object | None:
            if not pil_ok or Image is None or ImageTk is None:
                return None
            if not b:
                im = Image.new("RGB", (px, px), (226, 232, 240))
            else:
                try:
                    im = Image.open(BytesIO(b))
                except Exception:
                    im = Image.new("RGB", (px, px), (226, 232, 240))
                else:
                    if im.mode not in ("RGB", "RGBA"):
                        im = im.convert("RGBA") if "A" in im.mode else im.convert("RGB")
                    if im.mode == "RGBA":
                        bg = Image.new("RGB", im.size, (255, 255, 255))
                        bg.paste(im, mask=im.split()[3])
                        im = bg
                    im.thumbnail((px, px), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            tk_img_refs.append(ph)
            return ph

        # ---- root layout: grid (hero, search, hover hint, body, buttons) ----
        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(3, weight=1)  # main content row expands

        # Row 0 — hero banner (fixed height)
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero, text=self._t("discontinue_heading"),
            font=("Segoe UI", 12, "bold"), fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(10, 4))
        for key in ("discontinue_intro", "discontinue_step1", "discontinue_step2", "discontinue_step3"):
            tk.Label(
                hero, text=self._t(key), wraplength=1240, justify=tk.LEFT,
                font=("Segoe UI", 9), fg="#e2e8f0", bg=COLORS["hero"],
            ).pack(anchor=tk.W, padx=14, pady=(0, 2))
        tk.Label(hero, text="", bg=COLORS["hero"], height=0).pack(pady=(0, 6))

        # Row 1 — search bar (fixed height)
        search_frame = tk.Frame(d, bg=COLORS["app"])
        search_frame.grid(row=1, column=0, sticky="ew", padx=14, pady=(8, 4))
        tk.Label(
            search_frame, text=self._t("discontinue_search_label"),
            font=("Segoe UI", 10, "bold"), bg=COLORS["app"], fg=COLORS["text"],
        ).pack(side=tk.LEFT)
        filt = tk.StringVar()
        ef = tk.Entry(search_frame, textvariable=filt, font=("Segoe UI", 10))
        ef.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 8))
        tk.Label(
            search_frame, text=self._t("discontinue_search_tip"),
            font=("Segoe UI", 8), bg=COLORS["app"], fg=COLORS["muted"],
        ).pack(side=tk.LEFT)
        tk.Label(
            d,
            text=self._t("discontinue_hover_tip"),
            font=("Segoe UI", 8),
            bg=COLORS["app"],
            fg=COLORS["muted"],
            wraplength=1220,
            justify=tk.LEFT,
        ).grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 2))

        # Row 3 — main content (tree + preview, expands)
        body = tk.Frame(d, bg=COLORS["app"])
        body.grid(row=3, column=0, sticky="nsew", padx=14, pady=(4, 0))
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=1, minsize=360)
        body.grid_rowconfigure(0, weight=1)

        # ---- Left: Treeview ----
        tree_frame = tk.Frame(body, bg=COLORS["app"])
        tree_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        sty = ttk.Style(d)
        row_h = 80 if pil_ok else 24
        try:
            sty.configure("Disc.Treeview", rowheight=row_h, font=("Segoe UI", 9))
            sty.configure("Disc.Treeview.Heading", font=("Segoe UI", 9, "bold"))
            sty.map("Disc.Treeview", background=[("selected", "#dbeafe")])
        except tk.TclError:
            pass

        cols = ("row", "shop", "stall", "title")
        tree = ttk.Treeview(
            tree_frame, columns=cols, show="tree headings",
            selectmode="browse", style="Disc.Treeview",
        )
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        tree.heading("#0", text=self._t("discontinue_col_thumb"))
        tree.column("#0", width=88 if pil_ok else 20, stretch=False, anchor=tk.CENTER)
        tree.heading("row", text=self._t("discontinue_col_row"))
        tree.column("row", width=40, stretch=False, anchor=tk.CENTER)
        tree.heading("shop", text=self._t("discontinue_col_shop"))
        tree.column("shop", width=100, stretch=False)
        tree.heading("stall", text=self._t("discontinue_col_stall"))
        tree.column("stall", width=62, stretch=False, anchor=tk.CENTER)
        tree.heading("title", text=self._t("discontinue_col_title"))
        tree.column("title", width=380, stretch=True)

        # ---- Right: Preview panel ----
        preview = tk.Frame(body, bg=COLORS["card"], highlightthickness=1, highlightbackground=COLORS["border"])
        preview.grid(row=0, column=1, sticky="nsew")
        preview.grid_rowconfigure(3, weight=1)
        preview.grid_columnconfigure(0, weight=1)

        tk.Label(
            preview, text=self._t("discontinue_preview_title"),
            font=("Segoe UI", 10, "bold"), bg=COLORS["card"], fg=COLORS["text"],
            anchor=tk.W,
        ).grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 2))

        preview_meta = tk.Label(
            preview, text="", bg=COLORS["card"], fg=COLORS["muted"],
            font=("Segoe UI", 9), anchor=tk.W, justify=tk.LEFT, wraplength=340,
        )
        preview_meta.grid(row=1, column=0, sticky="ew", padx=10)

        preview_img = tk.Label(preview, bg="#f1f5f9", borderwidth=0, highlightthickness=0)
        preview_img.grid(row=2, column=0, padx=10, pady=(8, 4))

        preview_title = tk.Text(
            preview, wrap=tk.WORD, font=("Segoe UI", 10),
            bg=COLORS["card"], fg=COLORS["text"], relief=tk.FLAT,
            highlightthickness=0, padx=6, pady=6,
        )
        preview_title.grid(row=3, column=0, sticky="nsew", padx=10, pady=(4, 10))
        preview_title.configure(state=tk.DISABLED)

        def set_preview(r: ProductMapPickerRow | None) -> None:
            preview_photo_ref.clear()
            preview_img.configure(image="", text="")
            preview_title.configure(state=tk.NORMAL)
            preview_title.delete("1.0", tk.END)
            if r is None:
                preview_meta.configure(text=self._t("discontinue_preview_placeholder"))
                preview_title.configure(state=tk.DISABLED)
                return
            shop_str = r.shop_name or "—"
            stall_str = r.stall or "—"
            lines = [
                f"Row {r.row_num}",
                f"{self._t('discontinue_col_shop')}: {shop_str}",
                f"{self._t('discontinue_col_stall')}: {stall_str}",
            ]
            preview_meta.configure(text="\n".join(lines), fg=COLORS["text"])
            preview_title.insert(tk.END, r.title)
            preview_title.configure(state=tk.DISABLED)
            big = bytes_to_thumb(row_photos.get(r.row_num), 240)
            if big is not None:
                preview_img.configure(image=big, text="")
                preview_photo_ref.append(big)
            else:
                no_photo = "(no photo)" if self._lang == "en" else "（无图）"
                preview_img.configure(image="", text=no_photo, fg=COLORS["muted"], font=("Segoe UI", 9))

        set_preview(None)

        # Hover zoom popup (large image near cursor; debounced per row)
        _hover_after_id: list[object | None] = [None]
        _hover_tip_win: list[tk.Toplevel | None] = [None]
        _hover_photo: list[object] = []
        _hover_active_iid: list[str | None] = [None]
        HOVER_DELAY_MS = 220
        # Display cap: upscale/downscale to fit; shrink on small displays (≤ half the shorter screen side, max 720)
        try:
            _sh_scr = int(d.winfo_screenheight())
            _sw_scr = int(d.winfo_screenwidth())
        except tk.TclError:
            _sh_scr, _sw_scr = 1080, 1920
        HOVER_MAX_PX = min(720, max(300, min(_sh_scr, _sw_scr) // 2))

        def _hide_product_hover() -> None:
            if _hover_after_id[0] is not None:
                try:
                    d.after_cancel(_hover_after_id[0])
                except tk.TclError:
                    pass
                _hover_after_id[0] = None
            _hover_active_iid[0] = None
            if _hover_tip_win[0] is not None:
                try:
                    _hover_tip_win[0].destroy()
                except tk.TclError:
                    pass
                _hover_tip_win[0] = None
            _hover_photo.clear()

        def _photo_for_hover(raw: bytes) -> object | None:
            if not pil_ok or Image is None or ImageTk is None:
                return None
            try:
                im = Image.open(BytesIO(raw))
            except Exception:
                return None
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA") if "A" in im.mode else im.convert("RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            w, h = im.size
            if w >= 1 and h >= 1:
                scale = min(HOVER_MAX_PX / w, HOVER_MAX_PX / h)
                nw = max(1, int(round(w * scale)))
                nh = max(1, int(round(h * scale)))
                if (nw, nh) != (w, h):
                    im = im.resize((nw, nh), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            _hover_photo.append(ph)
            return ph

        def _show_hover_for_iid(iid: str) -> None:
            _hover_after_id[0] = None
            px, py = tree.winfo_pointerx(), tree.winfo_pointery()
            lx = px - tree.winfo_rootx()
            ly = py - tree.winfo_rooty()
            if lx < 0 or ly < 0 or lx >= tree.winfo_width() or ly >= tree.winfo_height():
                _hide_product_hover()
                return
            if tree.identify_region(lx, ly) != "tree":
                _hover_active_iid[0] = None
                return
            if tree.identify_row(ly) != iid:
                _hover_active_iid[0] = None
                return
            r = row_by_iid.get(iid)
            if r is None:
                _hover_active_iid[0] = None
                return
            raw = row_photos.get(r.row_num)
            if not raw:
                _hover_active_iid[0] = None
                return
            ph = _photo_for_hover(raw)
            if ph is None:
                _hover_active_iid[0] = None
                return
            if _hover_tip_win[0] is not None:
                try:
                    _hover_tip_win[0].destroy()
                except tk.TclError:
                    pass
                _hover_tip_win[0] = None
            tip = tk.Toplevel(d)
            tip.overrideredirect(True)
            try:
                tip.attributes("-topmost", True)
            except tk.TclError:
                pass
            tip.configure(bg="#ffffff", bd=0, highlightthickness=0)
            tk.Label(
                tip,
                image=ph,
                bg="#ffffff",
                bd=0,
                highlightthickness=0,
                relief=tk.FLAT,
            ).pack()
            tip.update_idletasks()
            tw, th = tip.winfo_reqwidth(), tip.winfo_reqheight()
            sw, sh = tip.winfo_screenwidth(), tip.winfo_screenheight()
            x = min(max(12, px + 20), sw - tw - 12)
            y = min(max(12, py + 20), sh - th - 12)
            tip.geometry(f"+{x}+{y}")
            _hover_tip_win[0] = tip

        def on_tree_motion(_e: tk.Event | None = None) -> None:
            if not pil_ok:
                return
            px, py = tree.winfo_pointerx(), tree.winfo_pointery()
            lx = px - tree.winfo_rootx()
            ly = py - tree.winfo_rooty()
            if lx < 0 or ly < 0 or lx >= tree.winfo_width() or ly >= tree.winfo_height():
                _hide_product_hover()
                return
            if tree.identify_region(lx, ly) != "tree":
                _hide_product_hover()
                return
            iid = tree.identify_row(ly)
            if not iid:
                _hide_product_hover()
                return
            r = row_by_iid.get(iid)
            if r is None or not row_photos.get(r.row_num):
                _hide_product_hover()
                return
            if iid == _hover_active_iid[0]:
                return
            _hide_product_hover()
            _hover_active_iid[0] = iid
            _hover_after_id[0] = d.after(HOVER_DELAY_MS, lambda i=iid: _show_hover_for_iid(i))

        def on_tree_leave(_e: tk.Event | None = None) -> None:
            _hide_product_hover()

        tree.bind("<Motion>", on_tree_motion)
        tree.bind("<Leave>", on_tree_leave)

        def row_matches_query(r: ProductMapPickerRow, q: str) -> bool:
            if not q:
                return True
            blob = " ".join((r.title, r.shop_name, r.stall)).lower()
            return q in blob

        def refill_tree() -> None:
            _hide_product_hover()
            tree.delete(*tree.get_children())
            row_by_iid.clear()
            tk_img_refs.clear()
            q = filt.get().strip().lower()
            for r in all_rows:
                if not row_matches_query(r, q):
                    continue
                iid = f"r{r.row_num}"
                row_by_iid[iid] = r
                t_short = r.title if len(r.title) <= 80 else r.title[:77] + "…"
                dash = "—"
                raw = row_photos.get(r.row_num)
                thumb = bytes_to_thumb(raw, 72) if pil_ok else None
                kw: dict = dict(
                    values=(r.row_num, r.shop_name or dash, r.stall or dash, t_short),
                )
                if thumb is not None:
                    kw["image"] = thumb
                    kw["text"] = ""
                else:
                    kw["text"] = dash
                tree.insert("", tk.END, iid=iid, **kw)

        def on_tree_select(_evt: object | None = None) -> None:
            sel = tree.selection()
            if not sel:
                set_preview(None)
                return
            set_preview(row_by_iid.get(sel[0]))

        filt.trace_add("write", lambda *_: refill_tree())
        refill_tree()
        tree.bind("<<TreeviewSelect>>", on_tree_select)

        def do_mark() -> None:
            sel = tree.selection()
            if not sel:
                messagebox.showinfo(
                    self._t("discontinue_title"), self._t("discontinue_no_selection"), parent=d,
                )
                return
            r = row_by_iid.get(sel[0])
            if r is None:
                return
            mark_btn.config(state=tk.DISABLED)
            close_btn.config(state=tk.DISABLED)
            prev_cursor = d.cget("cursor")
            d.config(cursor="watch")
            d.update_idletasks()
            try:
                mark_product_map_discontinued_by_row(FILE_SUPPLIER_CATALOG, r.row_num)
            except Exception as e:
                messagebox.showerror(self._t("file_open_fail_title"), str(e), parent=d)
                return
            finally:
                # Always restore cursor and re-enable buttons regardless of success or failure,
                # so the dialog is never left in a locked state on unexpected exceptions.
                d.config(cursor=prev_cursor)
                mark_btn.config(state=tk.NORMAL)
                close_btn.config(state=tk.NORMAL)
            nonlocal all_rows
            try:
                row_photos.clear()
                row_photos.update(
                    extract_photos_from_xlsx(FILE_SUPPLIER_CATALOG, sheet_name=CATALOG_SHEET, photo_col_idx=0)
                )
            except Exception:
                pass
            all_rows = list_product_map_rows_for_picker(FILE_SUPPLIER_CATALOG)
            refill_tree()
            set_preview(None)
            # Redraw list before the modal so the removed row is not visible behind the message
            d.update_idletasks()
            d.update()
            messagebox.showinfo(
                self._t("discontinue_title"), self._t("discontinue_done"), parent=d,
            )

        def _on_tree_double_click(evt: tk.Event) -> None:
            # Only trigger on actual data rows/thumbnails, not column headings or separators.
            region = tree.identify_region(evt.x, evt.y)
            if region not in ("tree", "cell"):
                return
            do_mark()

        tree.bind("<Double-Button-1>", _on_tree_double_click)

        # Row 4 — buttons (fixed height)
        btn_bar = tk.Frame(d, bg=COLORS["app"])
        btn_bar.grid(row=4, column=0, sticky="ew", padx=14, pady=(8, 12))
        mark_btn = tk.Button(
            btn_bar, text=self._t("discontinue_apply"), font=("Segoe UI", 10, "bold"),
            bg="#dc2626", fg="#ffffff", activebackground="#b91c1c", activeforeground="#ffffff",
            relief=tk.FLAT, padx=18, pady=6, cursor="hand2", command=do_mark,
        )
        mark_btn.pack(side=tk.LEFT)
        def _close_discontinue_dialog() -> None:
            _hide_product_hover()
            d.destroy()

        close_btn = tk.Button(
            btn_bar, text=self._t("discontinue_close"), font=("Segoe UI", 10),
            bg=COLORS["strip"], fg=COLORS["text"], relief=tk.FLAT, padx=14, pady=6,
            cursor="hand2", command=_close_discontinue_dialog,
        )
        close_btn.pack(side=tk.RIGHT)
        d.protocol("WM_DELETE_WINDOW", _close_discontinue_dialog)

        ef.focus_set()

    def _open_catalog_backups_dialog(self) -> None:
        if (
            list_supplier_catalog_backups is None
            or restore_supplier_catalog is None
            or catalog_backup_dir is None
            or backup_supplier_catalog_before_write is None
        ):
            messagebox.showinfo(self._t("msg_missing_title"), self._t("edit_no_import"))
            return

        dlg = tk.Toplevel(self)
        dlg.title(self._t("cat_backup_title"))
        dlg.minsize(420, 320)
        dlg.geometry("520x380")
        dlg.transient(self)
        dlg.configure(bg=COLORS["app"])

        outer = ttk.Frame(dlg, style="App.TFrame", padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text=self._t("cat_backup_hint"), wraplength=480).pack(
            anchor=tk.W, pady=(0, 8)
        )

        lb_frame = ttk.Frame(outer)
        lb_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        scroll = ttk.Scrollbar(lb_frame)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        listbox = tk.Listbox(
            lb_frame,
            height=12,
            font=("Segoe UI", 10),
            yscrollcommand=scroll.set,
            selectmode=tk.SINGLE,
        )
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.config(command=listbox.yview)

        paths: list[Path] = []

        def refresh() -> None:
            nonlocal paths
            listbox.delete(0, tk.END)
            paths = list_supplier_catalog_backups(FILE_SUPPLIER_CATALOG)
            if not paths:
                listbox.insert(tk.END, self._t("cat_backup_none"))
                return
            for p in paths:
                listbox.insert(tk.END, p.name)

        refresh()

        btn_row = ttk.Frame(outer)
        btn_row.pack(fill=tk.X, pady=(4, 0))

        def open_folder() -> None:
            bdir = catalog_backup_dir(FILE_SUPPLIER_CATALOG)
            bdir.mkdir(parents=True, exist_ok=True)
            self._open_file_path(bdir)

        def snapshot() -> None:
            if not FILE_SUPPLIER_CATALOG.is_file():
                messagebox.showinfo(
                    self._t("msg_missing_title"),
                    self._t("cat_backup_snapshot_no_file"),
                )
                return
            dest = backup_supplier_catalog_before_write(
                FILE_SUPPLIER_CATALOG, "manual_snapshot"
            )
            if dest:
                messagebox.showinfo(
                    self._t("cat_backup_title"),
                    self._t("cat_backup_snapshot_ok", path=str(dest)),
                )
                refresh()

        def restore() -> None:
            if not paths:
                return
            sel = listbox.curselection()
            if not sel:
                messagebox.showinfo(
                    self._t("cat_backup_title"),
                    self._t("cat_backup_restore_pick"),
                )
                return
            idx = int(sel[0])
            if idx < 0 or idx >= len(paths):
                return
            chosen = paths[idx]
            if not messagebox.askyesno(
                self._t("cat_backup_title"),
                self._t("cat_backup_restore_confirm", path=chosen.name),
            ):
                return
            try:
                restore_supplier_catalog(FILE_SUPPLIER_CATALOG, chosen)
            except OSError as exc:
                messagebox.showerror(self._t("cat_backup_title"), str(exc))
                return
            messagebox.showinfo(
                self._t("cat_backup_title"), self._t("cat_backup_restore_ok")
            )
            refresh()

        ttk.Button(
            btn_row, text=self._t("cat_backup_open_folder"), command=open_folder
        ).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(
            btn_row, text=self._t("cat_backup_snapshot"), command=snapshot
        ).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(
            btn_row, text=self._t("cat_backup_restore"), command=restore
        ).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(
            btn_row, text=self._t("cat_backup_close"), command=dlg.destroy
        ).pack(side=tk.RIGHT)

        dlg.focus_set()

    def _open_data_folder(self) -> None:
        self._reveal_directory(DATA_DIR)

    def _open_file_path(self, path: Path) -> None:
        if not path.exists():
            messagebox.showinfo(self._t("file_missing_title"), self._t("file_missing_body", path=str(path)))
            return
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", str(path)], check=False)
            else:
                subprocess.run(["xdg-open", str(path)], check=False)
        except OSError as e:
            messagebox.showerror(self._t("file_open_fail_title"), self._t("file_open_fail_body", err=e))

    def _resolve_charm_images_dir(self) -> Path:
        raw = self._charm_dir.get().strip() if self._charm_dir else ""
        if raw:
            return Path(raw).expanduser().resolve()
        return DEFAULT_CHARM_IMAGES_DIR.resolve()

    def _set_chrome_busy(self, busy: bool) -> None:
        if busy:
            ttk_state = ["disabled"]
            if self._run_btn:
                self._run_btn.config(state=tk.DISABLED, bg="#94a3b8", fg="#f1f5f9")
        else:
            ttk_state = ["!disabled"]
            if self._run_btn:
                self._run_btn.config(state=tk.NORMAL, bg=COLORS["run"], fg="#ffffff")
        for b in (
            self._btn_edit_products,
            self._btn_catalog_backups,
            self._btn_charm_browse,
            self._btn_charm_import,
            self._btn_charm_open_folder,
            self._btn_charm_reorder_open,
            self._btn_pdf_browse,
            self._btn_pdf_run_new,
            self._btn_pdf_open_folder,
            self._btn_pdf_move_backup,
        ):
            if b is not None:
                b.state(ttk_state)
        if self._charm_drop_vision_cb is not None:
            self._charm_drop_vision_cb.state(ttk_state)

    def _charm_stage_files(self, paths: list[Path]) -> int:
        dest = self._resolve_charm_images_dir()
        dest.mkdir(parents=True, exist_ok=True)
        n = 0
        for src in paths:
            try:
                p = src if isinstance(src, Path) else Path(src)
                if not p.is_file():
                    continue
                ext = p.suffix.lower()
                if ext not in CHARM_IMAGE_EXTS:
                    continue
                stem_clean = re.sub(r"[^\w\-.]+", "_", p.stem, flags=re.UNICODE)
                stem_clean = stem_clean.strip("._") or "photo"
                stem_clean = stem_clean[:72]
                name = f"{CHARM_INCOMING_PREFIX}{secrets.token_hex(4)}_{stem_clean}{ext}"
                shutil.copy2(p, dest / name)
                n += 1
            except OSError:
                continue
        return n

    def _charm_on_paths_dropped(self, paths: list[Path]) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        if not paths:
            return
        n = self._charm_stage_files(paths)
        if n == 0:
            messagebox.showinfo(self._t("charm_msg_title"), self._t("charm_drop_no_valid"))
            return
        self._append_log(self._t("charm_drop_staged", n=n))

    def _charm_browse_photos(self) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        picked = filedialog.askopenfilenames(
            parent=self,
            filetypes=[
                ("Images", "*.png *.jpg *.jpeg *.webp"),
                ("PNG", "*.png"),
                ("JPEG", "*.jpg *.jpeg"),
                ("WebP", "*.webp"),
                ("All", "*.*"),
            ],
            title=self._t("charm_drop_browse_title"),
        )
        if not picked:
            return
        self._charm_on_paths_dropped([Path(s) for s in picked])

    def _open_charm_images_folder(self) -> None:
        self._reveal_directory(self._resolve_charm_images_dir())

    def _charm_import_workbook(self) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        if not GENERATOR.is_file():
            messagebox.showerror(self._t("msg_missing_title"), self._t("msg_missing") + str(GENERATOR))
            return
        dest = self._resolve_charm_images_dir()
        dest.mkdir(parents=True, exist_ok=True)
        staged = [
            p
            for p in dest.iterdir()
            if p.is_file()
            and p.name.startswith(CHARM_INCOMING_PREFIX)
            and p.suffix.lower() in CHARM_IMAGE_EXTS
        ]
        if not staged:
            messagebox.showinfo(self._t("charm_msg_title"), self._t("charm_drop_nothing_to_import"))
            return

        args: list[str] = [
            sys.executable,
            str(GENERATOR),
            "--project-dir",
            str(PROJECT_ROOT),
            "--import-charm-images",
            *charm_import_pattern_argv(),
        ]
        cd = self._charm_dir.get().strip()
        if cd:
            args.extend(["--charm-images-dir", cd])
        if self._charm_drop_vision.get():
            args.append("--import-charm-vision-sku")

        self._run_busy = True
        self._set_chrome_busy(True)
        self._append_log(self._t("charm_import_start"))
        self._append_log(" ".join(args) + "\n\n")

        def work() -> None:
            try:
                proc = subprocess.run(
                    args,
                    cwd=str(PROJECT_ROOT),
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                )
                out = proc.stdout or ""
                err = proc.stderr or ""
                self._log_q.put(out)
                if err:
                    self._log_q.put(self._t("log_messages") + err)
                self._log_q.put(self._t("log_finished", code=proc.returncode))
            except Exception as e:
                self._log_q.put(self._t("log_error", e=e))
            finally:

                def done() -> None:
                    self._run_busy = False
                    self._set_chrome_busy(False)

                self.after(0, done)

        threading.Thread(target=work, daemon=True).start()

    def _open_charm_reorder_dialog(self) -> None:
        if reorder_charm_library_rows is None or load_charm_library is None:
            messagebox.showerror(
                self._t("msg_missing_title"),
                self._t("charm_reorder_no_import"),
            )
            return
        if not FILE_SUPPLIER_CATALOG.exists():
            messagebox.showinfo(
                self._t("msg_missing_title"),
                self._t("reorder_no_catalog"),
            )
            return
        try:
            entries = load_charm_library(FILE_SUPPLIER_CATALOG)
        except Exception as e:
            messagebox.showerror(self._t("file_open_fail_title"), str(e))
            return
        if not entries:
            messagebox.showinfo(
                self._t("reorder_title"), self._t("reorder_empty")
            )
            return
        charm_images_dir = self._resolve_charm_images_dir()
        _CharmReorderDialog(self, entries, FILE_SUPPLIER_CATALOG, charm_images_dir)

    def _run(self) -> None:
        if self._run_busy:
            messagebox.showinfo(self._t("msg_busy_title"), self._t("msg_busy"))
            return
        if not GENERATOR.is_file():
            messagebox.showerror(self._t("msg_missing_title"), self._t("msg_missing") + str(GENERATOR))
            return
        job = self._job_var.get()
        self._spawn_generator(
            self._collect_generator_args(job),
            log_intro=self._t("log_start"),
        )


# ======================================================================
# Orders Dashboard Dialog — standalone window with photo thumbnails,
# filter buttons, mode toggle (Case/Grip vs Charms), and inline editing.
# ======================================================================

class _OrdersDashboardDialog:
    _THUMB = 64
    _ROW_H = 74

    _CG_FILTERS = ("all", "ready", "needs_info", "unmatched")
    _CH_FILTERS = ("all", "assigned", "needs_code", "needs_shop")

    _FILTER_LABELS = {
        "en": {"all": "All", "ready": "Ready", "needs_info": "Needs Info",
               "unmatched": "Unmatched", "assigned": "Assigned",
               "needs_code": "Needs Code", "needs_shop": "Needs Shop"},
        "zh": {"all": "\u5168\u90e8", "ready": "\u5c31\u7eea",
               "needs_info": "\u5f85\u586b\u4fe1\u606f",
               "unmatched": "\u672a\u5339\u914d", "assigned": "\u5df2\u5206\u914d",
               "needs_code": "\u5f85\u5206\u914d\u7f16\u7801",
               "needs_shop": "\u5f85\u5206\u914d\u5e97\u94fa"},
    }

    _FILTER_COLORS = {
        "all": ("#1F4E79", "#ffffff"), "ready": ("#047857", "#ffffff"),
        "needs_info": ("#1e40af", "#ffffff"), "unmatched": ("#92400e", "#ffffff"),
        "assigned": ("#5b21b6", "#ffffff"), "needs_code": ("#92400e", "#ffffff"),
        "needs_shop": ("#1e40af", "#ffffff"),
    }

    def __init__(
        self, parent: App, items: list,
        *, title_to_row: dict[str, int],
        supplier_shops: list[str], supplier_stalls: list[str],
        supplier_shop_stalls: dict[str, str] | None = None,
        supplier_stall_shops: dict[str, str] | None = None,
        charm_codes: list[str], charm_shop_names: list[str],
        charm_library: dict | None = None,
        charm_shop_stalls: dict[str, str] | None = None,
        catalog_photos: dict[str, bytes] | None = None,
    ) -> None:
        self._parent = parent
        self._lang = parent._lang
        self._items = items
        self._title_to_row = title_to_row
        self._sup_shops = supplier_shops
        self._sup_stalls = supplier_stalls
        # Bidirectional mappings: shop ↔ stall (built from Suppliers sheet)
        self._sup_shop_stalls: dict[str, str] = supplier_shop_stalls or {}
        self._sup_stall_shops: dict[str, str] = supplier_stall_shops or {}
        self._charm_codes = charm_codes
        self._charm_shops = charm_shop_names
        self._charm_library: dict = charm_library or {}
        # shop_name → stall mapping from the Charm Shops tab
        self._charm_shop_stalls: dict[str, str] = charm_shop_stalls or {}
        # Canonical product photos from Product Map (norm_title → jpeg bytes).
        # Kept mutable so photo-upload updates propagate live without a restart.
        self._catalog_photos: dict[str, bytes] = dict(catalog_photos or {})
        self._tk_img_refs: list[object] = []
        self._hover_photo: list[object] = []
        self._hover_tip: tk.Toplevel | None = None
        self._hover_after_id: list[object | None] = [None]
        self._hover_active_iid: list[str | None] = [None]
        self._preview_photo_ref: list[object] = []
        self._charm_lib_photo_ref: list[object] = []
        # Photo preview in the detail panel (kept to prevent GC of PhotoImage)
        self._detail_photo_ref: list[object] = []
        # Charm purchase summary popup (kept to bring-to-front instead of re-creating)
        self._summary_win: tk.Toplevel | None = None
        self._summary_photo_refs: list[object] = []
        self._charm_tile_frames: dict[str, tk.Frame] = {}
        self._charm_tile_info: dict[str, dict] = {}
        self._selected_charm_code: str = ""
        self._selected_indices: list[int] = []
        self._gallery_hint: tk.Label | None = None
        self._recently_used_frame: tk.Frame | None = None
        self._ru_photo_refs: list[object] = []
        # Column sort state (default Stall asc — applied again in _configure_columns)
        self._sort_col: str = "stall"
        self._sort_dir: str = "asc"   # "asc" | "desc"
        # Base heading labels (without sort indicators) — populated in _configure_columns
        self._col_base_labels: dict[str, str] = {}
        self._tile_hover_after: list[object | None] = [None]
        self._tile_hover_tip: tk.Toplevel | None = None
        self._tile_hover_photo_ref: list[object] = []
        self._pil_ok = Image is not None and ImageTk is not None
        # Purchase-status dict: {(order_num, norm_title, component): status_str}
        # component = "case" | "grip" | "charm"
        # Only non-Pending values are stored (Pending is the default/omitted key).
        self._pstatuses: dict[tuple[str, str, str], str] = {}
        # Purchase-list button + its separator — both live inside _filter_frame
        # and must be destroyed/recreated on every mode switch to avoid duplicates.
        self._purchase_list_btn: tk.Button | None = None
        self._purchase_list_sep: tk.Frame | None = None

        self._mode = "casegrip"
        self._active_filter = "all"

        self._cg_items: list[dict] = []
        self._ch_items: list[dict] = []
        self._selected: dict | None = None
        self._build_item_lists()

        d = tk.Toplevel(parent)
        self._d = d
        d.title("Orders Dashboard" if self._lang == "en" else "\u8ba2\u5355\u603b\u89c8")
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        # Width accommodates Case/Grip/Buy-Status purchase-status columns
        d.geometry("1640x880")
        d.minsize(1260, 680)
        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(2, weight=1)

        # ── Row 0: Hero ───────────────────────────────────────────────
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(hero, text="Orders Dashboard" if self._lang == "en" else "\u8ba2\u5355\u603b\u89c8",
                 font=("Segoe UI", 14, "bold"), fg="#ffffff",
                 bg=COLORS["hero"]).pack(anchor=tk.W, padx=16, pady=(12, 2))
        tk.Label(hero,
                 text=("View and edit all current orders — product photos, supplier info, charm assignments."
                       if self._lang == "en" else
                       "\u67e5\u770b\u5e76\u7f16\u8f91\u6240\u6709\u5f53\u524d\u8ba2\u5355 \u2014 "
                       "\u4ea7\u54c1\u56fe\u7247\u3001\u4f9b\u5e94\u5546\u4fe1\u606f\u3001\u6302\u4ef6\u5206\u914d\u3002"),
                 font=("Segoe UI", 10), fg="#dbeafe",
                 bg=COLORS["hero"]).pack(anchor=tk.W, padx=16, pady=(0, 10))

        # ── Row 1: Mode toggle + filter bar + search ──────────────────
        toolbar = tk.Frame(d, bg=COLORS["strip"], highlightthickness=0)
        toolbar.grid(row=1, column=0, sticky="ew")

        mode_frame = tk.Frame(toolbar, bg=COLORS["strip"])
        mode_frame.pack(side=tk.LEFT, padx=(12, 0), pady=6)
        self._btn_mode_cg = tk.Button(
            mode_frame, text="\U0001F4F1 Case / Grip", font=("Segoe UI", 10, "bold"),
            relief=tk.FLAT, bd=0, padx=14, pady=5, cursor="hand2",
            command=lambda: self._set_mode("casegrip"))
        self._btn_mode_cg.pack(side=tk.LEFT, padx=(0, 2))
        self._btn_mode_ch = tk.Button(
            mode_frame, text="\u2728 Charms", font=("Segoe UI", 10, "bold"),
            relief=tk.FLAT, bd=0, padx=14, pady=5, cursor="hand2",
            command=lambda: self._set_mode("charms"))
        self._btn_mode_ch.pack(side=tk.LEFT)

        sep = tk.Frame(toolbar, bg=COLORS["border"], width=1, highlightthickness=0)
        sep.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=4)

        self._filter_frame = tk.Frame(toolbar, bg=COLORS["strip"])
        self._filter_frame.pack(side=tk.LEFT, pady=6)
        self._filter_btns: list[tk.Button] = []

        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._populate_tree())
        se = ttk.Entry(toolbar, textvariable=self._search_var,
                       font=("Segoe UI", 10), width=30)
        se.pack(side=tk.RIGHT, padx=(0, 14), pady=6)
        tk.Label(toolbar, text="\U0001F50D", font=("Segoe UI", 11),
                 bg=COLORS["strip"], fg=COLORS["muted"]).pack(side=tk.RIGHT)

        # "+ Add Order" button — opens the manual-order dialog
        self._btn_add_order = tk.Button(
            toolbar,
            text="+ Add Order" if self._lang == "en" else "+ \u6dfb\u52a0\u8ba2\u5355",
            font=("Segoe UI", 10, "bold"),
            relief=tk.FLAT, bd=0, padx=14, pady=5, cursor="hand2",
            bg="#047857", fg="#ffffff",
            activebackground="#065f46", activeforeground="#ffffff",
            command=self._open_add_manual_order,
        )
        self._btn_add_order.pack(side=tk.RIGHT, padx=(0, 14), pady=6)

        # ── Row 2: Body (tree left + detail panel right) ──────────────
        body = tk.Frame(d, bg=COLORS["app"])
        body.grid(row=2, column=0, sticky="nsew", padx=12, pady=(8, 0))
        # Tree takes all available horizontal space; right panel is a FIXED 330px
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=0)   # never resizes
        body.grid_rowconfigure(0, weight=1)

        # Tree frame
        tf = tk.Frame(body, bg=COLORS["app"])
        tf.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        sty = ttk.Style()
        row_h = self._ROW_H if self._pil_ok else 24
        sty.configure("Dash.Treeview", rowheight=row_h, font=("Segoe UI", 10),
                       background="#ffffff", fieldbackground="#ffffff")
        sty.configure("Dash.Treeview.Heading", font=("Segoe UI", 10, "bold"),
                       background=COLORS["accent"], foreground="#ffffff",
                       relief="flat", padding=(6, 6))
        sty.map("Dash.Treeview.Heading", background=[("active", COLORS["hero"])])
        sty.map("Dash.Treeview", background=[("selected", COLORS["accent"])],
                foreground=[("selected", "#ffffff")])

        self._tree = ttk.Treeview(
            tf, show="tree headings", selectmode="extended", style="Dash.Treeview",
        )
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=self._tree.yview)
        hsb = ttk.Scrollbar(tf, orient=tk.HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self._tree.bind("<<TreeviewSelect>>", self._on_select)
        self._tree.bind("<Motion>", self._on_tree_motion)
        self._tree.bind("<Leave>", lambda _: self._hide_hover())
        # Inline status picker — fires after row selection is already resolved
        self._tree.bind("<ButtonRelease-1>", self._on_tree_cell_click)

        self._tree.heading("#0", text="Photo", anchor=tk.CENTER)
        self._tree.column("#0", width=68 if self._pil_ok else 20, stretch=False, anchor=tk.CENTER)

        # Tag colors
        for tag, bg, fg in [
            ("ready",              "#d1fae5", "#065f46"),
            ("needs_info",         "#dbeafe", "#1e3a8a"),
            ("unmatched",          "#fef3c7", "#78350f"),
            ("assigned",           "#ede9fe", "#4c1d95"),
            ("needs_code",         "#fef3c7", "#78350f"),
            ("needs_shop",         "#dbeafe", "#1e3a8a"),
            ("alt_ready",          "#ecfdf5", "#065f46"),
            ("alt_needs_info",     "#eff6ff", "#1e3a8a"),
            ("alt_unmatched",      "#fffbeb", "#78350f"),
            ("alt_assigned",       "#f5f3ff", "#4c1d95"),
            ("alt_needs_code",     "#fffbeb", "#78350f"),
            ("alt_needs_shop",     "#eff6ff", "#1e3a8a"),
            # Purchase-status override tags (same palette as the route Excel)
            ("ps_purchased",       "#c6efce", "#276221"),
            ("ps_oos",             "#ffeb9c", "#7d4e00"),
            ("ps_oop",             "#ffc7ce", "#9c0006"),
            ("alt_ps_purchased",   "#dff5e4", "#276221"),
            ("alt_ps_oos",         "#fff8cc", "#7d4e00"),
            ("alt_ps_oop",         "#ffe4e8", "#9c0006"),
        ]:
            self._tree.tag_configure(tag, background=bg, foreground=fg)

        # ── Detail panel (right) — FIXED 330px, never resizes ────────
        panel_outer = tk.Frame(body, width=330, bg=COLORS["card"],
                               highlightthickness=1, highlightbackground=COLORS["border"])
        panel_outer.grid(row=0, column=1, sticky="ns")
        panel_outer.pack_propagate(False)   # ← enforces the fixed width
        panel_outer.grid_propagate(False)
        self._panel = panel_outer

        # ── Top: compact order info strip (hidden when nothing selected) ─
        # No product photo — it is already visible in the tree column.
        self._order_info_frame = tk.Frame(panel_outer, bg="#f8faff",
                                          highlightthickness=0)
        # (packed dynamically in _on_select / _clear_detail)

        self._detail_title = tk.Label(
            self._order_info_frame, text="", font=("Segoe UI", 11, "bold"),
            fg=COLORS["text"], bg="#f8faff",
            wraplength=296, justify=tk.LEFT, anchor=tk.NW,
            cursor="hand2",
        )
        self._detail_title.pack(fill=tk.X, padx=10, pady=(10, 2))

        def _copy_title(_event=None):
            title = self._detail_title.cget("text")
            if not title:
                return
            self._d.clipboard_clear()
            self._d.clipboard_append(title)
            # Brief colour flash to confirm the copy
            orig_fg = self._detail_title.cget("fg")
            self._detail_title.config(fg="#047857")
            self._d.after(400, lambda: self._detail_title.config(fg=orig_fg))

        self._detail_title.bind("<Button-1>", _copy_title)

        self._detail_meta = tk.Label(
            self._order_info_frame, text="", font=("Segoe UI", 8),
            fg=COLORS["muted"], bg="#f8faff", anchor=tk.W,
            wraplength=296, justify=tk.LEFT,
        )
        self._detail_meta.pack(fill=tk.X, padx=10, pady=(0, 2))

        self._detail_notes = tk.Label(
            self._order_info_frame, text="", font=("Segoe UI", 8, "italic"),
            fg="#7c3aed", bg="#f8faff", anchor=tk.W,
            wraplength=296, justify=tk.LEFT,
        )
        self._detail_notes.pack(fill=tk.X, padx=10, pady=(0, 8))

        # ── Photo preview + upload (shown only for single-selection) ──────
        self._detail_photo_frame = tk.Frame(self._order_info_frame, bg="#f8faff")
        # (packed dynamically in _on_select; hidden via pack_forget in _clear_detail)

        self._detail_photo_lbl = tk.Label(
            self._detail_photo_frame, bg="#eef2ff",
            bd=1, relief=tk.SOLID, highlightthickness=0,
            cursor="hand2",
        )
        self._detail_photo_lbl.pack(pady=(4, 3))

        _upload_text = (
            ("\U0001F4F7  Upload Photo" if self._lang == "en"
             else "\U0001F4F7  上传图片")
        )
        self._upload_photo_btn = tk.Button(
            self._detail_photo_frame,
            text=_upload_text,
            font=("Segoe UI", 9), relief=tk.FLAT, bd=0,
            padx=8, pady=4, cursor="hand2",
            bg="#e0e7ff", fg="#3730a3",
            activebackground="#c7d2fe", activeforeground="#1e1b4b",
            command=self._upload_photo_for_order,
        )
        self._upload_photo_btn.pack(fill=tk.X, padx=10, pady=(0, 2))

        self._upload_photo_status = tk.Label(
            self._detail_photo_frame, text="",
            font=("Segoe UI", 8, "bold"), fg="#047857", bg="#f8faff",
            wraplength=290, justify=tk.CENTER,
        )
        self._upload_photo_status.pack(fill=tk.X, padx=10, pady=(0, 6))

        # Named separator so we can use pack(before=) when needed
        self._order_info_sep = tk.Frame(
            self._order_info_frame, bg=COLORS["border"], height=1,
            highlightthickness=0,
        )
        self._order_info_sep.pack(fill=tk.X)

        # ── Middle/main area: filled dynamically per mode ─────────────
        # For Case/Grip: Supplier/Stall comboboxes (always visible, grayed until selection)
        # For Charms:    Scrollable charm library tile gallery (always visible)
        self._mode_panel = tk.Frame(panel_outer, bg=COLORS["card"])
        self._mode_panel.pack(fill=tk.BOTH, expand=True)

        # ── Bottom: save controls (hidden when nothing selected) ───────
        self._save_controls_frame = tk.Frame(panel_outer, bg=COLORS["card"])
        # (packed dynamically)
        self._save_status_lbl = tk.Label(
            self._save_controls_frame, text="", font=("Segoe UI", 9, "bold"),
            fg="#047857", bg=COLORS["card"])

        # Edit fields storage
        self._edit_widgets: dict[str, object] = {}

        # ── Row 3: Footer ─────────────────────────────────────────────
        foot = tk.Frame(d, bg=COLORS["app"])
        foot.grid(row=3, column=0, sticky="ew", padx=14, pady=(8, 12))
        ttk.Button(foot, text="Close" if self._lang == "en" else "\u5173\u95ed",
                   command=d.destroy, style="Tool.TButton").pack(side=tk.RIGHT)
        self._regen_btn = ttk.Button(
            foot, text="Regenerate Shopping Route" if self._lang == "en" else "\u91cd\u65b0\u751f\u6210\u91c7\u8d2d\u6e05\u5355",
            command=self._regen, style="Tool.TButton")
        self._regen_btn.pack(side=tk.LEFT)
        # Open Route dropdown — lists all generated file variants
        self._btn_open_route = tk.Button(
            foot,
            text=self._parent._t("open_route_btn"),
            font=("Segoe UI", 10, "bold"),
            relief=tk.FLAT, bd=0, padx=14, pady=5, cursor="hand2",
            bg=COLORS["accent"], fg="#ffffff",
            activebackground=COLORS["hero"], activeforeground="#ffffff",
            command=self._open_route,
        )
        self._btn_open_route.pack(side=tk.LEFT, padx=(10, 0))
        # Delete selected orders button — styled red as a destructive action
        self._btn_delete = tk.Button(
            foot,
            text="\U0001F5D1  Delete Selected" if self._lang == "en" else "\U0001F5D1  删除所选",
            font=("Segoe UI", 10, "bold"),
            relief=tk.FLAT, bd=0, padx=14, pady=5, cursor="hand2",
            bg="#dc2626", fg="#ffffff",
            activebackground="#b91c1c", activeforeground="#ffffff",
            command=self._delete_selected_orders,
        )
        self._btn_delete.pack(side=tk.LEFT, padx=(10, 0))
        self._save_status = tk.Label(foot, text="", font=("Segoe UI", 10, "bold"),
                                     fg="#047857", bg=COLORS["app"])
        self._save_status.pack(side=tk.LEFT, padx=(16, 0))

        # Load purchase statuses BEFORE _set_mode so the first tree render
        # already reflects any persisted statuses.
        self._load_pstatuses_cache()

        self._set_mode("casegrip")
        d.protocol("WM_DELETE_WINDOW", d.destroy)

        # ── Keyboard shortcuts ─────────────────────────────────────────
        # Ctrl+A — select all visible rows in the current tree
        def _select_all(_event=None):
            self._tree.selection_set(self._tree.get_children())
            self._on_select()
        d.bind("<Control-a>", _select_all)
        d.bind("<Control-A>", _select_all)
        # Escape — clear selection
        def _deselect(_event=None):
            self._tree.selection_remove(self._tree.selection())
            self._clear_detail()
        d.bind("<Escape>", _deselect)
        # Enter / Return — save if something is selected
        d.bind("<Return>", lambda _: self._save())

    # ── Item classification ───────────────────────────────────────────

    def _build_item_lists(self) -> None:
        cg: list[dict] = []
        ch: list[dict] = []
        for r in self._items:
            has_case, has_grip, has_charm = _style_has(r.item.style)
            sup = r.supplier
            shop = sup.shop_name if sup else ""
            stall = sup.stall if sup else ""
            c_code = sup.charm_code if sup else ""
            # Charm shop resolution — Charm Library's default_charm_shop is the
            # canonical 1:1 source of truth (enforced by normalize_catalog_charm_
            # shops).  Falls back to the per-product entry for codes not yet
            # registered in the library.  Exact same priority as the shopping
            # route's aggregation, guaranteeing dashboard ↔ Excel consistency.
            c_shop = ""
            if c_code:
                _lib_entry = self._charm_library.get(c_code)
                if _lib_entry and getattr(_lib_entry, "default_charm_shop", ""):
                    c_shop = _lib_entry.default_charm_shop
            if not c_shop:
                c_shop = sup.charm_shop if sup else ""
            has_loc = bool(sup and (shop or stall))
            norm = _normalize(r.item.title)   # full key — no truncation

            # Canonical photo: prefer the Product Map photo over the PDF photo.
            # This guarantees every order for the same product uses the same image.
            canonical_photo = (
                self._catalog_photos.get(norm)
                or self._catalog_photos.get(norm[:50])
                or r.item.photo_bytes
            )

            base = {
                "order": r.order.order_number, "title": r.item.title,
                "phone": r.item.phone_model, "qty": r.item.quantity,
                "photo_bytes": canonical_photo,
                "norm_title": norm,
                "private_notes": r.order.private_notes,
                "buyer": r.order.buyer_name, "date": r.order.order_date,
                "etsy_shop": r.order.etsy_shop,
            }

            if has_case or has_grip:
                if has_loc:
                    st = "ready"
                elif sup is not None:
                    # Matched a catalog entry but supplier/stall not filled yet —
                    # open supplier_catalog.xlsx and fill in Shop Name + Stall.
                    st = "needs_info"
                else:
                    st = "unmatched"
                cg.append({**base, "supplier": shop, "stall": stall,
                           "case": "\u2713" if has_case else "",
                           "grip": "\u2713" if has_grip else "", "status": st})

            if has_charm:
                if c_code and c_shop:
                    cst = "assigned"
                elif not c_code:
                    # No charm code yet — this is always the first gap to fill:
                    # you must know which specific charm before assigning a shop.
                    cst = "needs_code"
                else:
                    # Code is known; still needs a shop to buy it from.
                    cst = "needs_shop"
                # Stall comes from the Charm Shops tab, not the product supplier
                charm_stall = self._charm_shop_stalls.get(c_shop, "") if c_shop else ""
                ch.append({**base, "charm_code": c_code, "charm_shop": c_shop,
                           "stall": charm_stall, "status": cst})

        self._cg_items = cg
        self._ch_items = ch

    # ── Thumbnails ────────────────────────────────────────────────────

    def _thumb(self, raw: bytes | None) -> object | None:
        if not self._pil_ok or not raw or Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            im.thumbnail((self._THUMB, self._THUMB), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._tk_img_refs.append(ph)
            return ph
        except Exception:
            return None

    def _make_preview(self, raw: bytes | None, max_dim: int = 180) -> object | None:
        if not self._pil_ok or not raw or Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            w, h = im.size
            if w > max_dim or h > max_dim:
                ratio = min(max_dim / w, max_dim / h)
                im = im.resize((int(w * ratio), int(h * ratio)), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._preview_photo_ref.clear()
            self._preview_photo_ref.append(ph)
            return ph
        except Exception:
            return None

    # ── Mode toggle ───────────────────────────────────────────────────

    def _set_mode(self, mode: str) -> None:
        self._cancel_tile_hover()
        self._mode = mode
        self._active_filter = "all"
        self._selected = None

        if mode == "casegrip":
            self._btn_mode_cg.config(bg=COLORS["accent"], fg="#ffffff")
            self._btn_mode_ch.config(bg=COLORS["strip"], fg=COLORS["text"])
            filters = self._CG_FILTERS
        else:
            self._btn_mode_cg.config(bg=COLORS["strip"], fg=COLORS["text"])
            self._btn_mode_ch.config(bg="#5b21b6", fg="#ffffff")
            filters = self._CH_FILTERS

        # Destroy the purchase-list button and its separator from any previous
        # charms-mode visit before rebuilding the filter bar. Without this,
        # every charms → casegrip → charms round-trip appends a second copy.
        if self._purchase_list_sep is not None:
            self._purchase_list_sep.destroy()
            self._purchase_list_sep = None
        if self._purchase_list_btn is not None:
            self._purchase_list_btn.destroy()
            self._purchase_list_btn = None

        for btn in self._filter_btns:
            btn.destroy()
        self._filter_btns.clear()

        labels = self._FILTER_LABELS.get(self._lang, self._FILTER_LABELS["en"])
        for fid in filters:
            bg, fg = self._FILTER_COLORS[fid]
            btn = tk.Button(
                self._filter_frame, text=f" {labels[fid]} ",
                font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0,
                padx=10, pady=3, cursor="hand2",
                command=lambda f=fid: self._set_filter(f),
            )
            btn.pack(side=tk.LEFT, padx=(0, 4))
            self._filter_btns.append(btn)

        # Charms mode: show the 🛒 Purchase List button; close it on mode switch
        if mode == "charms":
            self._purchase_list_sep = tk.Frame(self._filter_frame, bg=COLORS["border"],
                                               width=1, highlightthickness=0)
            self._purchase_list_sep.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 8), pady=4)
            _pl_text = (
                "\U0001F6D2  Purchase List" if self._lang == "en"
                else "\U0001F6D2  采购清单"
            )
            self._purchase_list_btn = tk.Button(
                self._filter_frame, text=_pl_text,
                font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0,
                padx=10, pady=3, cursor="hand2",
                bg="#1e1b4b", fg="#e0e7ff",
                activebackground="#312e81", activeforeground="#ffffff",
                command=self._show_charm_purchase_summary,
            )
            self._purchase_list_btn.pack(side=tk.LEFT)
        else:
            # Switching away from charms — close the summary popup if open
            if self._summary_win and self._summary_win.winfo_exists():
                self._summary_win.destroy()
            self._summary_win = None

        self._configure_columns()
        self._build_mode_panel()   # builds the always-visible charm gallery or CG form
        self._update_filter_styles()
        self._populate_tree()
        self._clear_detail()

    def _set_filter(self, fid: str) -> None:
        self._active_filter = fid
        self._update_filter_styles()
        self._populate_tree()

    def _update_filter_styles(self) -> None:
        labels = self._FILTER_LABELS.get(self._lang, self._FILTER_LABELS["en"])
        filters = self._CG_FILTERS if self._mode == "casegrip" else self._CH_FILTERS
        for btn, fid in zip(self._filter_btns, filters):
            bg, fg = self._FILTER_COLORS[fid]
            if fid == self._active_filter:
                btn.config(bg=bg, fg=fg)
            else:
                btn.config(bg="#e2e8f0", fg="#475569")

    # ── Column configuration (with sort binding) ─────────────────────
    #
    # Sortable columns get a click command that calls _sort_by_column.
    # Non-sortable columns (#, Photo, Case checkmark, Grip checkmark, Qty)
    # are left without a command so the cursor stays default.
    # Sort indicators (▲ / ▼) are appended to the active heading label.

    # Maps column key → dict key in each item dict.
    # Keys starting with "_ps_" are sentinel values handled specially in the
    # sort engine (they rank by _PSTATUS_SORT_ORDER, not alphabetically).
    _CG_SORT_MAP: dict[str, str] = {
        "order": "order", "product": "title", "etsy_shop": "etsy_shop",
        "supplier": "supplier",
        "stall": "stall",  "phone": "phone",   "status": "status",
        "notes": "private_notes",
        "case":  "_ps_case",   # purchase-status sort
        "grip":  "_ps_grip",   # purchase-status sort
    }
    _CH_SORT_MAP: dict[str, str] = {
        "order": "order", "product": "title", "etsy_shop": "etsy_shop",
        "charm_code": "charm_code",
        "charm_shop": "charm_shop", "stall": "stall", "status": "status",
        "notes": "private_notes",
        "buy_status": "_ps_charm",  # purchase-status sort
    }

    # Purchase-status options — mirrors STATUS_OPTIONS in generate_shopping_route.py
    _PURCHASE_STATUSES: list[str] = [
        "Pending", "Purchased", "Out of Stock", "Out of Production",
    ]
    # Priority for determining the "worst" status (highest-priority first)
    _PSTATUS_PRIORITY: list[str] = [
        "Out of Production", "Out of Stock", "Pending", "Purchased",
    ]
    # Ascending sort order: Pending first (most action needed) → Purchased last
    _PSTATUS_SORT_ORDER: dict[str, int] = {
        "Pending": 0, "Out of Stock": 1, "Out of Production": 2,
        "Purchased": 3,
    }
    # Display text shown in the tree cells (abbreviated for column width)
    _PSTATUS_DISPLAY: dict[str, str] = {
        "Pending":           "Pending",
        "Purchased":         "Purchased",
        "Out of Stock":      "Out of Stock",
        "Out of Production": "Out of Prod.",
    }
    # Maps English status value → tree row tag name
    _PSTATUS_TAG: dict[str, str] = {
        "Purchased":         "ps_purchased",
        "Out of Stock":      "ps_oos",
        "Out of Production": "ps_oop",
    }

    def _configure_columns(self) -> None:
        tree = self._tree
        self._col_base_labels.clear()

        if tree["columns"]:
            tree["columns"] = ()

        def _hdr(col: str, label: str, *, anchor=tk.W, sortable: bool = True) -> None:
            """Register heading with optional sort command and store base label."""
            self._col_base_labels[col] = label
            cmd = (lambda c=col: self._sort_by_column(c)) if sortable else None
            if cmd:
                tree.heading(col, text=label, anchor=anchor, command=cmd)
            else:
                tree.heading(col, text=label, anchor=anchor)

        if self._mode == "casegrip":
            tree["columns"] = ("seq", "order", "product", "etsy_shop", "supplier", "stall",
                               "case", "grip", "phone", "qty", "status", "notes")
            _hdr("seq",       "#",             anchor=tk.CENTER, sortable=False)
            _hdr("order",     "Order #",       anchor=tk.CENTER, sortable=True)
            _hdr("product",   "Product",       sortable=True)
            _hdr("etsy_shop", "Etsy Shop",     sortable=True)
            _hdr("supplier",  "Supplier",      sortable=True)
            _hdr("stall",     "Stall",         anchor=tk.CENTER, sortable=True)
            # Case / Grip columns now show the per-component purchase status
            _hdr("case",      "Case",          anchor=tk.CENTER, sortable=True)
            _hdr("grip",      "Grip",          anchor=tk.CENTER, sortable=True)
            _hdr("phone",     "Phone Model",   sortable=True)
            _hdr("qty",       "Qty",           anchor=tk.CENTER, sortable=False)
            _hdr("status",    "Match",         anchor=tk.CENTER, sortable=True)
            _hdr("notes",     "Private Notes", sortable=True)
            tree.column("seq",       width=32,  minwidth=26,  anchor=tk.CENTER, stretch=False)
            tree.column("order",     width=95,  minwidth=80,  anchor=tk.CENTER, stretch=False)
            tree.column("product",   width=200, minwidth=120, stretch=True)
            tree.column("etsy_shop", width=110, minwidth=70,  stretch=False)
            tree.column("supplier",  width=90,  minwidth=60,  stretch=False)
            tree.column("stall",     width=58,  minwidth=44,  anchor=tk.CENTER, stretch=False)
            tree.column("case",      width=108, minwidth=72,  anchor=tk.CENTER, stretch=False)
            tree.column("grip",      width=108, minwidth=72,  anchor=tk.CENTER, stretch=False)
            tree.column("phone",     width=118, minwidth=80,  stretch=False)
            tree.column("qty",       width=32,  minwidth=26,  anchor=tk.CENTER, stretch=False)
            tree.column("status",    width=72,  minwidth=50,  anchor=tk.CENTER, stretch=False)
            tree.column("notes",     width=160, minwidth=80,  stretch=False)
        else:
            tree["columns"] = ("seq", "order", "product", "etsy_shop", "charm_code", "charm_shop",
                               "stall", "qty", "status", "buy_status", "notes")
            _hdr("seq",        "#",             anchor=tk.CENTER, sortable=False)
            _hdr("order",      "Order #",       anchor=tk.CENTER, sortable=True)
            _hdr("product",    "Product",       sortable=True)
            _hdr("etsy_shop",  "Etsy Shop",     sortable=True)
            _hdr("charm_code", "Charm Code",    anchor=tk.CENTER, sortable=True)
            _hdr("charm_shop", "Charm Shop",    sortable=True)
            _hdr("stall",      "Stall",         anchor=tk.CENTER, sortable=True)
            _hdr("qty",        "Qty",           anchor=tk.CENTER, sortable=False)
            _hdr("status",     "Match",         anchor=tk.CENTER, sortable=True)
            # Charm purchase status column
            _hdr("buy_status", "Buy Status",    anchor=tk.CENTER, sortable=True)
            _hdr("notes",      "Private Notes", sortable=True)
            tree.column("seq",        width=32,  minwidth=26,  anchor=tk.CENTER, stretch=False)
            tree.column("order",      width=95,  minwidth=80,  anchor=tk.CENTER, stretch=False)
            tree.column("product",    width=200, minwidth=120, stretch=True)
            tree.column("etsy_shop",  width=110, minwidth=70,  stretch=False)
            tree.column("charm_code", width=90,  minwidth=60,  anchor=tk.CENTER, stretch=False)
            tree.column("charm_shop", width=120, minwidth=70,  stretch=False)
            tree.column("stall",      width=58,  minwidth=44,  anchor=tk.CENTER, stretch=False)
            tree.column("qty",        width=32,  minwidth=26,  anchor=tk.CENTER, stretch=False)
            tree.column("status",     width=90,  minwidth=55,  anchor=tk.CENTER, stretch=False)
            tree.column("buy_status", width=108, minwidth=72,  anchor=tk.CENTER, stretch=False)
            tree.column("notes",      width=160, minwidth=80,  stretch=False)

        # Always open / return to a mode with Stall ascending (matches shopping-route order).
        self._sort_col = "stall"
        self._sort_dir = "asc"

    # ── Sort engine ───────────────────────────────────────────────────

    def _sort_by_column(self, col: str) -> None:
        """Toggle sort direction if same column, reset to ascending if new column."""
        if self._sort_col == col:
            self._sort_dir = "desc" if self._sort_dir == "asc" else "asc"
        else:
            self._sort_col = col
            self._sort_dir = "asc"
        self._populate_tree()

    def _refresh_sort_indicators(self) -> None:
        """Append ▲ / ▼ to the active column heading; clear all others."""
        sort_map = self._CH_SORT_MAP if self._mode == "charms" else self._CG_SORT_MAP
        UP, DN = " ▲", " ▼"
        for col, base in self._col_base_labels.items():
            if col not in (list(self._tree["columns"])):
                continue
            if col == self._sort_col and col in sort_map:
                indicator = DN if self._sort_dir == "desc" else UP
                self._tree.heading(col, text=base + indicator)
            else:
                self._tree.heading(col, text=base)

    # ── Edit fields ───────────────────────────────────────────────────

    def _build_mode_panel(self) -> None:
        """Rebuild the middle area of the right panel for the current mode.

        Case/Grip mode: Supplier + Stall comboboxes (always visible).
        Charms mode:    Scrollable charm-library tile gallery (always visible).
        The order-info strip (top) and save controls (bottom) are managed
        separately by _on_select / _clear_detail.
        """
        for w in self._mode_panel.winfo_children():
            w.destroy()
        self._edit_widgets.clear()
        self._charm_lib_photo_ref.clear()
        self._charm_tile_frames.clear()
        self._selected_charm_code = ""

        # Clear dynamically-added children from the save-controls frame so
        # repeated mode switches don't accumulate multiple Save buttons.
        # _save_status_lbl lives here too but must be preserved.
        for w in list(self._save_controls_frame.winfo_children()):
            if w is not self._save_status_lbl:
                w.destroy()

        if self._mode == "casegrip":
            self._build_casegrip_form()
        else:
            self._build_charm_gallery()

    # ── Case / Grip mode: simple supplier form ────────────────────────

    def _build_casegrip_form(self) -> None:
        f = tk.Frame(self._mode_panel, bg=COLORS["card"])
        f.pack(fill=tk.X, padx=12, pady=12)
        # Two equal columns: Supplier (0) and Stall (1)
        f.columnconfigure(0, weight=1)
        f.columnconfigure(1, weight=1)

        # ── Row 0: column headers ──────────────────────────────────────
        tk.Label(f,
                 text="SUPPLIER" if self._lang == "en" else "\u4f9b\u5e94\u5546",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).grid(row=0, column=0, sticky=tk.W, pady=(0, 3))
        tk.Label(f,
                 text="STALL" if self._lang == "en" else "\u6444\u4f4d",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).grid(row=0, column=1, sticky=tk.W,
                                         padx=(8, 0), pady=(0, 3))

        # ── Row 1: comboboxes side by side ─────────────────────────────
        sv = tk.StringVar()
        cb = ttk.Combobox(f, textvariable=sv, values=self._sup_shops,
                          font=("Segoe UI", 10))
        cb.grid(row=1, column=0, sticky=tk.EW, pady=(0, 0))
        self._edit_widgets["sup_var"] = sv
        self._edit_widgets["sup_cb"] = cb

        stv = tk.StringVar()
        scb = ttk.Combobox(f, textvariable=stv, values=self._sup_stalls,
                           font=("Segoe UI", 10))
        scb.grid(row=1, column=1, sticky=tk.EW, padx=(8, 0), pady=(0, 0))
        self._edit_widgets["stall_var"] = stv
        self._edit_widgets["stall_cb"] = scb

        # ── Row 2: autofill hints (shared row, muted small text) ───────
        hint_row = tk.Frame(f, bg=COLORS["card"])
        hint_row.grid(row=2, column=0, columnspan=2, sticky=tk.EW, pady=(3, 0))
        self._sup_autofill_lbl = tk.Label(hint_row, text="",
                                          font=("Segoe UI", 8), fg="#047857",
                                          bg=COLORS["card"])
        self._sup_autofill_lbl.pack(side=tk.LEFT)
        self._stall_autofill_lbl = tk.Label(hint_row, text="",
                                            font=("Segoe UI", 8), fg="#047857",
                                            bg=COLORS["card"])
        self._stall_autofill_lbl.pack(side=tk.LEFT, padx=(12, 0))

        # ── Row 3: Save button spanning both columns ────────────────────
        save_btn = ttk.Button(f,
                              text="Save Supplier & Stall" if self._lang == "en"
                              else "\u4fdd\u5b58\u4f9b\u5e94\u5546\u4e0e\u6444\u4f4d",
                              command=self._save, style="Tool.TButton")
        save_btn.grid(row=3, column=0, columnspan=2, sticky=tk.EW, pady=(10, 0))
        self._edit_widgets["save_btn"] = save_btn

        # ── Bidirectional auto-fill ────────────────────────────────────
        def _autofill_stall(_event=None) -> None:
            """When a shop is chosen, auto-fill stall if the mapping is known."""
            shop = sv.get().strip()
            if not shop:
                self._sup_autofill_lbl.config(text="")
                return
            stall = self._sup_shop_stalls.get(shop, "")
            if stall and not stv.get().strip():
                stv.set(stall)
                self._stall_autofill_lbl.config(
                    text="\u2713 Stall auto-filled" if self._lang == "en"
                    else "\u2713 \u6444\u4f4d\u5df2\u81ea\u52a8\u586b\u5165"
                )
            else:
                self._stall_autofill_lbl.config(text="")
            self._sup_autofill_lbl.config(text="")

        def _autofill_shop(_event=None) -> None:
            """When a stall is chosen, auto-fill shop if the mapping is known."""
            stall = stv.get().strip()
            if not stall:
                self._stall_autofill_lbl.config(text="")
                return
            shop = self._sup_stall_shops.get(stall, "")
            if shop and not sv.get().strip():
                sv.set(shop)
                self._sup_autofill_lbl.config(
                    text="\u2713 Supplier auto-filled" if self._lang == "en"
                    else "\u2713 \u4f9b\u5e94\u5546\u5df2\u81ea\u52a8\u586b\u5165"
                )
            else:
                self._sup_autofill_lbl.config(text="")
            self._stall_autofill_lbl.config(text="")

        def _clear_autofill_hints(_event=None) -> None:
            self._sup_autofill_lbl.config(text="")
            self._stall_autofill_lbl.config(text="")

        cb.bind("<<ComboboxSelected>>",  _autofill_stall)
        scb.bind("<<ComboboxSelected>>", _autofill_shop)
        sv.trace_add("write",  lambda *_: _autofill_stall())
        stv.trace_add("write", lambda *_: _autofill_shop())
        cb.bind("<Key>",  _clear_autofill_hints)
        scb.bind("<Key>", _clear_autofill_hints)

        # ── Save controls (hidden until a row is selected) ────────────
        # Only feedback label + Manage Suppliers remain here — the Save
        # button now lives directly in the form above.
        scf = self._save_controls_frame

        tk.Frame(scf, bg=COLORS["border"], height=1).pack(fill=tk.X, pady=(0, 0))
        self._save_status_lbl.pack(anchor=tk.W, padx=14, pady=(6, 4))

        tk.Frame(scf, bg=COLORS["border"], height=1).pack(fill=tk.X)
        ttk.Button(
            scf,
            text="Manage Suppliers\u2026" if self._lang == "en"
            else "\u7ba1\u7406\u4f9b\u5e94\u5546\u2026",
            command=lambda: _SuppliersManagerDialog(
                self._d, FILE_SUPPLIER_CATALOG, self._lang,
                on_close=self._reload_supplier_dropdowns,
            ),
            style="Tool.TButton",
        ).pack(fill=tk.X, padx=14, pady=(10, 10))

    def _reload_supplier_dropdowns(self) -> None:
        """Re-read the Suppliers sheet and refresh the Supplier/Stall comboboxes."""
        shops: list[str] = []
        stalls: list[str] = []
        shop_stalls: dict[str, str] = {}
        stall_shops: dict[str, str] = {}
        try:
            import openpyxl as _xl
            _wb = _xl.load_workbook(
                FILE_SUPPLIER_CATALOG, read_only=True, data_only=True
            )
            if "Suppliers" in _wb.sheetnames:
                _ws = _wb["Suppliers"]
                _shop_ci = _stall_ci = None
                for ci in range(1, 15):
                    h = str(_ws.cell(1, ci).value or "").strip().lower()
                    if h == "shop name":
                        _shop_ci = ci
                    elif h == "stall":
                        _stall_ci = ci
                for r in _ws.iter_rows(min_row=2, values_only=False):
                    sv  = str(r[_shop_ci  - 1].value or "").strip() if _shop_ci  else ""
                    stv = str(r[_stall_ci - 1].value or "").strip() if _stall_ci else ""
                    if sv and sv not in shops:
                        shops.append(sv)
                    if stv and stv not in stalls:
                        stalls.append(stv)
                    if sv and stv:
                        shop_stalls.setdefault(sv, stv)
                        stall_shops.setdefault(stv, sv)
            _wb.close()
        except Exception:
            pass

        # Persist updated data on self
        self._sup_shops       = shops
        self._sup_stalls      = stalls
        self._sup_shop_stalls = shop_stalls
        self._sup_stall_shops = stall_shops

        # Refresh live combobox widgets if the form is currently open
        sup_cb   = self._edit_widgets.get("sup_cb")
        stall_cb = self._edit_widgets.get("stall_cb")
        if sup_cb is not None:
            try:
                sup_cb["values"] = shops
            except Exception:
                pass
        if stall_cb is not None:
            try:
                stall_cb["values"] = stalls
            except Exception:
                pass

    # ── Charms mode: permanent charm tile gallery ─────────────────────

    def _build_charm_gallery(self) -> None:
        """
        Always-visible charm library panel (Charms mode only).

        Layout uses grid so that the canvas row reliably fills
        all vertical space between the header/filter and the save controls.
        Every tile widget gets a <MouseWheel> binding (fixing the scroll bug),
        plus <Enter>/<Leave> for hover-zoom.
        """
        TILE_W = 168
        TILE_H = 138
        TILE_IMG_SZ = 92
        COLS = 2
        BG = COLORS["card"]

        outer = tk.Frame(self._mode_panel, bg=BG)
        outer.pack(fill=tk.BOTH, expand=True)
        # row 0=header label, row 1=action toolbar, row 2=filter,
        # row 3=canvas (expands), row 4=save controls
        outer.rowconfigure(3, weight=1)
        outer.columnconfigure(0, weight=1)

        # ── Row 0: section header (label only) ────────────────────────
        lbl_row = tk.Frame(outer, bg=COLORS["strip"])
        lbl_row.grid(row=0, column=0, sticky="ew")
        tk.Label(lbl_row,
                 text="\u2728  Charm Library \u2014 click a charm to assign it" if self._lang == "en"
                 else "\u2728  \u6302\u4ef6\u5e93 \u2014 \u70b9\u51fb\u5373\u53ef\u5206\u914d",
                 font=("Segoe UI", 9, "bold"), fg="#5b21b6",
                 bg=COLORS["strip"]).pack(anchor=tk.W, padx=10, pady=4)

        # ── Row 1: action toolbar ──────────────────────────────────────
        toolbar_row = tk.Frame(outer, bg=BG)
        toolbar_row.grid(row=1, column=0, sticky="ew", padx=8, pady=(4, 2))
        toolbar_row.columnconfigure(0, weight=1)
        toolbar_row.columnconfigure(1, weight=1)
        ttk.Button(toolbar_row,
                   text="+ Import Photos" if self._lang == "en" else "+ \u5bfc\u5165\u6302\u4ef6\u7167\u7247",
                   command=self._charm_import_photos,
                   style="Tool.TButton").grid(row=0, column=0, sticky="ew", padx=(0, 3))
        ttk.Button(toolbar_row,
                   text="Sort / Reorder" if self._lang == "en" else "\u91cd\u6392\u7f16\u7801",
                   command=self._charm_reorder,
                   style="Tool.TButton").grid(row=0, column=1, sticky="ew", padx=(3, 0))

        # ── Row 2: SKU / code filter ───────────────────────────────────
        filter_row = tk.Frame(outer, bg=BG)
        filter_row.grid(row=2, column=0, sticky="ew", padx=8, pady=(4, 2))
        filter_row.columnconfigure(0, weight=1)

        self._charm_filter_var = tk.StringVar()
        self._charm_filter_var.trace_add("write", lambda *_: self._layout_charm_tiles())
        filter_entry = ttk.Entry(filter_row, textvariable=self._charm_filter_var,
                                 font=("Segoe UI", 9))
        filter_entry.grid(row=0, column=0, sticky="ew")
        filter_entry.insert(0, "\U0001F50D  Filter by code or SKU\u2026" if self._lang == "en"
                            else "\U0001F50D  \u6309\u7f16\u7801\u6216SKU\u641c\u7d22\u2026")
        filter_entry.config(foreground=COLORS["muted"])

        def _filter_focus_in(_):
            if self._charm_filter_var.get().startswith("\U0001F50D"):
                filter_entry.delete(0, tk.END)
                filter_entry.config(foreground=COLORS["text"])
        def _filter_focus_out(_):
            if not filter_entry.get().strip():
                filter_entry.delete(0, tk.END)
                filter_entry.insert(0, "\U0001F50D  Filter by code or SKU\u2026" if self._lang == "en"
                                    else "\U0001F50D  \u6309\u7f16\u7801\u6216SKU\u641c\u7d22\u2026")
                filter_entry.config(foreground=COLORS["muted"])
                self._charm_filter_var.set("")
        filter_entry.bind("<FocusIn>", _filter_focus_in)
        filter_entry.bind("<FocusOut>", _filter_focus_out)

        clr_btn = tk.Button(filter_row, text="\u2715", font=("Segoe UI", 8),
                            bg=BG, fg=COLORS["muted"], relief=tk.FLAT,
                            bd=0, cursor="hand2", padx=4,
                            command=lambda: [self._charm_filter_var.set(""),
                                            filter_entry.delete(0, tk.END),
                                            filter_entry.insert(0,
                                                "\U0001F50D  Filter by code or SKU\u2026" if self._lang == "en"
                                                else "\U0001F50D  \u6309\u7f16\u7801\u6216SKU\u641c\u7d22\u2026"),
                                            filter_entry.config(foreground=COLORS["muted"])])
        clr_btn.grid(row=0, column=1, padx=(4, 0))

        # ── Row 3: scrollable canvas (fills remaining height) ──────────
        canvas_wrap = tk.Frame(outer, bg=BG)
        canvas_wrap.grid(row=3, column=0, sticky="nsew")
        canvas_wrap.rowconfigure(0, weight=1)
        canvas_wrap.columnconfigure(0, weight=1)

        gallery_canvas = tk.Canvas(canvas_wrap, bg=BG,
                                   highlightthickness=0, borderwidth=0)
        gallery_vsb = ttk.Scrollbar(canvas_wrap, orient=tk.VERTICAL,
                                    command=gallery_canvas.yview)
        gallery_canvas.configure(yscrollcommand=gallery_vsb.set)
        gallery_canvas.grid(row=0, column=0, sticky="nsew")
        gallery_vsb.grid(row=0, column=1, sticky="ns")
        self._gallery_canvas = gallery_canvas

        grid_frame = tk.Frame(gallery_canvas, bg=BG)
        self._charm_grid_frame = grid_frame
        grid_win = gallery_canvas.create_window((0, 0), window=grid_frame, anchor=tk.NW)
        grid_frame.grid_columnconfigure(0, weight=1)
        grid_frame.grid_columnconfigure(1, weight=1)

        def _sync_scroll(_e=None):
            gallery_canvas.configure(scrollregion=gallery_canvas.bbox("all") or (0, 0, 0, 0))
        grid_frame.bind("<Configure>", _sync_scroll)

        def _fit_grid_width(e):
            gallery_canvas.itemconfigure(grid_win, width=e.width)
        gallery_canvas.bind("<Configure>", _fit_grid_width)

        # Mousewheel helper — bound to EVERY tile widget to fix scroll dead zones
        def _scroll(e):
            gallery_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        self._gallery_scroll_fn = _scroll
        gallery_canvas.bind("<MouseWheel>", _scroll)
        grid_frame.bind("<MouseWheel>", _scroll)

        # ── Hover-zoom state ───────────────────────────────────────────
        self._tile_hover_after: list[object | None] = [None]
        self._tile_hover_tip: tk.Toplevel | None = None
        self._tile_hover_photo_ref: list[object] = []

        # ── Build charm usage counts ───────────────────────────────────
        usage: dict[str, int] = {}
        for item in self._ch_items:
            cc = item.get("charm_code", "")
            if cc:
                usage[cc] = usage.get(cc, 0) + 1

        # ── Build tile widgets ─────────────────────────────────────────
        self._charm_tile_info: dict[str, dict] = {}   # {code: {'tile':…, 'sku':…, 'photo':…}}

        for code in self._charm_codes:
            entry = self._charm_library.get(code)
            sku_text = (entry.sku if entry and entry.sku else "")

            tile = tk.Frame(grid_frame, bg="#f5f3ff", width=TILE_W, height=TILE_H,
                            highlightthickness=2, highlightbackground="#e9d5ff",
                            cursor="hand2")
            tile.grid_propagate(False)

            # Photo
            photo_bytes = entry.photo_bytes if (entry and entry.photo_bytes) else None
            ph = self._make_charm_card_photo(photo_bytes, max_dim=TILE_IMG_SZ) if photo_bytes else None
            img_lbl = tk.Label(tile, bg="#f5f3ff", cursor="hand2")
            if ph:
                self._charm_lib_photo_ref.append(ph)
                img_lbl.config(image=ph)
            else:
                img_lbl.config(text="\U0001F48E", font=("Segoe UI", 18), fg="#c4b5fd")
            img_lbl.place(relx=0.5, rely=0.34, anchor=tk.CENTER)

            # Code + SKU labels
            code_lbl = tk.Label(tile, text=code, font=("Segoe UI", 8, "bold"),
                                fg="#5b21b6", bg="#f5f3ff", cursor="hand2")
            sku_short = sku_text[:18] if sku_text else ""
            sku_lbl = tk.Label(tile, text=sku_short, font=("Segoe UI", 7),
                               fg="#9333ea", bg="#f5f3ff", cursor="hand2")
            code_lbl.place(relx=0.5, rely=0.72, anchor=tk.CENTER)
            sku_lbl.place(relx=0.5, rely=0.87, anchor=tk.CENTER)

            # Usage badge (top-right corner) — how many current orders use this charm
            n_orders = usage.get(code, 0)
            if n_orders:
                badge = tk.Label(tile, text=f"\u00d7{n_orders}",
                                 font=("Segoe UI", 7, "bold"),
                                 fg="#ffffff", bg="#7c3aed", cursor="hand2",
                                 padx=3, pady=1)
                badge.place(relx=1.0, rely=0.0, anchor=tk.NE, x=-3, y=3)
            else:
                badge = None

            all_widgets = [w for w in (tile, img_lbl, code_lbl, sku_lbl, badge) if w]

            # Single click: select charm
            def _click(_, c=code):
                self._on_charm_tile_click(c)
            # Double click: select charm AND immediately save
            def _dblclick(_, c=code):
                self._on_charm_tile_double_click(c)
            for w in all_widgets:
                w.bind("<Button-1>", _click)
                w.bind("<Double-Button-1>", _dblclick)

            # Mousewheel binding on every tile child (the critical scroll fix)
            for w in all_widgets:
                w.bind("<MouseWheel>", _scroll)

            # Hover zoom — bind ONLY to the photo label, not the tile background or text
            def _enter(_, c=code):
                self._schedule_tile_hover(c)
            def _leave(_):
                self._cancel_tile_hover()
            img_lbl.bind("<Enter>", _enter)
            img_lbl.bind("<Leave>", _leave)

            self._charm_tile_frames[code] = tile
            self._charm_tile_info[code] = {
                "tile": tile, "sku": sku_text, "photo_bytes": photo_bytes,
            }

        # Initial tile layout (all visible)
        self._layout_charm_tiles()
        # Guarantee scrollregion is set after rendering
        self._d.after(120, _sync_scroll)

        # ── Row 4: Save controls ───────────────────────────────────────
        bottom = tk.Frame(outer, bg=BG)
        bottom.grid(row=4, column=0, sticky="ew")

        tk.Frame(bottom, bg=COLORS["border"], height=1).pack(fill=tk.X, pady=(6, 8))

        shop_row = tk.Frame(bottom, bg=BG)
        shop_row.pack(fill=tk.X, padx=8, pady=(0, 4))
        tk.Label(shop_row,
                 text="Charm Shop" if self._lang == "en" else "\u6302\u4ef6\u5e97\u94fa",
                 font=("Segoe UI", 9, "bold"), fg=COLORS["muted"],
                 bg=BG).pack(anchor=tk.W, pady=(0, 2))
        shop_var = tk.StringVar()
        shop_cb = ttk.Combobox(shop_row, textvariable=shop_var,
                               values=[""] + self._charm_shops,
                               font=("Segoe UI", 10))
        shop_cb.pack(fill=tk.X)
        self._edit_widgets["shop_var"] = shop_var
        self._edit_widgets["shop_cb"] = shop_cb

        btn_row = tk.Frame(bottom, bg=BG)
        btn_row.pack(fill=tk.X, padx=8, pady=(6, 4))
        save_btn = ttk.Button(btn_row,
                              text="\U0001F4BE  Save" if self._lang == "en" else "\U0001F4BE  \u4fdd\u5b58",
                              command=self._save, style="Tool.TButton")
        save_btn.pack(side=tk.LEFT)
        # "Clear charm" shortcut button
        clear_btn = ttk.Button(btn_row,
                               text="\u2715  Clear charm" if self._lang == "en" else "\u2715  \u6e05\u9664\u6302\u4ef6",
                               command=self._clear_charm_assignment, style="Tool.TButton")
        clear_btn.pack(side=tk.LEFT, padx=(8, 0))

        self._save_status_lbl.config(bg=BG)
        self._save_status_lbl.pack(anchor=tk.W, padx=8, pady=(0, 2))
        self._edit_widgets["save_btn"] = save_btn

        # No-selection hint
        self._gallery_hint = tk.Label(
            outer,
            text="\u2190  Select an order on the left to assign a charm" if self._lang == "en"
            else "\u2190  \u5728\u5de6\u4fa7\u9009\u62e9\u8ba2\u5355\u4ee5\u5206\u914d\u6302\u4ef6",
            font=("Segoe UI", 9), fg=COLORS["muted"], bg=BG,
            wraplength=290, justify=tk.CENTER,
        )
        self._gallery_hint.place(relx=0.5, rely=0.62, anchor=tk.CENTER)

    # ── Recently-used pinned strip ────────────────────────────────────

    def _rebuild_recently_used_row(self) -> None:
        """No-op — recently-used strip has been removed."""

    # ── Charm tile layout (filter-aware) ─────────────────────────────

    def _layout_charm_tiles(self) -> None:
        """Re-grid tiles that match the current filter query."""
        if not hasattr(self, "_charm_tile_info"):
            return
        q = ""
        if hasattr(self, "_charm_filter_var"):
            raw = self._charm_filter_var.get().strip()
            # Ignore the placeholder text
            if not raw.startswith("\U0001F50D"):
                q = raw.lower()

        COLS = 2
        seq = 0
        for code, info in self._charm_tile_info.items():
            tile = info["tile"]
            matches = (not q) or (q in code.lower()) or (q in info["sku"].lower())
            if matches:
                tile.grid(row=seq // COLS, column=seq % COLS,
                          padx=4, pady=4, sticky="nsew")
                seq += 1
            else:
                tile.grid_remove()

        if hasattr(self, "_gallery_canvas"):
            self._d.after(50, lambda: self._gallery_canvas.configure(
                scrollregion=self._gallery_canvas.bbox("all") or (0, 0, 0, 0)
            ))

    # ── Charm tile hover zoom ─────────────────────────────────────────

    def _schedule_tile_hover(self, code: str) -> None:
        self._cancel_tile_hover()
        self._tile_hover_after[0] = self._d.after(250, lambda: self._show_tile_hover(code))

    def _cancel_tile_hover(self) -> None:
        if self._tile_hover_after[0] is not None:
            self._d.after_cancel(self._tile_hover_after[0])
            self._tile_hover_after[0] = None
        if self._tile_hover_tip is not None:
            try:
                self._tile_hover_tip.destroy()
            except Exception:
                pass
            self._tile_hover_tip = None
        self._tile_hover_photo_ref.clear()

    def _show_tile_hover(self, code: str) -> None:
        if not self._pil_ok or Image is None or ImageTk is None:
            return
        info = self._charm_tile_info.get(code) if hasattr(self, "_charm_tile_info") else None
        raw = info.get("photo_bytes") if info else None
        if not raw:
            return
        entry = self._charm_library.get(code)
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg_im = Image.new("RGB", im.size, (255, 255, 255))
                bg_im.paste(im, mask=im.split()[3])
                im = bg_im
            w, h = im.size
            max_dim = 320
            if w > max_dim or h > max_dim:
                ratio = min(max_dim / w, max_dim / h)
                im = im.resize((int(w * ratio), int(h * ratio)), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._tile_hover_photo_ref.clear()
            self._tile_hover_photo_ref.append(ph)
        except Exception:
            return

        tip = tk.Toplevel(self._d)
        tip.overrideredirect(True)
        tip.attributes("-topmost", True)
        tip.configure(bg="#ffffff", highlightthickness=2,
                      highlightbackground="#7c3aed")
        # Photo
        tk.Label(tip, image=ph, bg="#ffffff", bd=0).pack()
        # Code + SKU caption
        caption = code
        if entry and entry.sku:
            caption += f"  ·  {entry.sku}"
        tk.Label(tip, text=caption, font=("Segoe UI", 9, "bold"),
                 fg="#5b21b6", bg="#f5f3ff", pady=4).pack(fill=tk.X)
        tip.update_idletasks()

        px = self._d.winfo_pointerx()
        py = self._d.winfo_pointery()
        sw = self._d.winfo_screenwidth()
        sh = self._d.winfo_screenheight()
        tw = tip.winfo_reqwidth()
        th = tip.winfo_reqheight()
        x = min(max(8, px + 16), sw - tw - 8)
        y = min(max(8, py - th - 8), sh - th - 8)
        tip.geometry(f"+{x}+{y}")
        self._tile_hover_tip = tip

    # ── Clear charm assignment ────────────────────────────────────────

    def _row_num_for_item(self, d: dict) -> int | None:
        """Resolve the Product Map row number for an item dict.

        Tries the full normalized title first, then a 50-char prefix (backward
        compat for items cached before the [:50] truncation was removed), then
        a starts-with scan as a last resort for very long catalog titles.
        """
        # Primary key (full normalized order title)
        row = self._title_to_row.get(d["norm_title"])
        if row is not None:
            return row
        # Backward compat: title stored with [:50] truncation
        short = d["norm_title"][:50]
        row = self._title_to_row.get(short)
        if row is not None:
            return row
        # Last resort: find any catalog key that starts with the item's norm_title
        for key, r in self._title_to_row.items():
            if key.startswith(d["norm_title"]) or d["norm_title"].startswith(key):
                return r
        return None

    def _upload_photo_for_order(self) -> None:
        """Open a file dialog to replace the product photo for the selected order.

        The new image is:
          1. Re-encoded as a JPEG and written to the Product Map sheet in
             supplier_catalog.xlsx for the matched catalog row.
          2. Applied immediately to EVERY in-memory order that shares the same
             product title, so the tree thumbnails refresh live without a restart.
          3. Stored in self._catalog_photos so subsequent _build_item_lists calls
             (e.g. after re-generate) also pick up the new canonical image.
        """
        if not self._pil_ok or Image is None or ImageTk is None:
            from tkinter import messagebox
            messagebox.showerror(
                "PIL not available" if self._lang == "en" else "\u65e0\u6cd5\u5904\u7406\u56fe\u7247",
                "Pillow library is required to upload photos."
                if self._lang == "en" else
                "\u9700\u8981 Pillow \u5e93\u624d\u80fd\u4e0a\u4f20\u56fe\u7247\u3002",
            )
            return

        if update_product_map_photo is None:
            from tkinter import messagebox
            messagebox.showerror(
                "Unavailable",
                "Photo update function not imported — check installation."
                if self._lang == "en" else
                "\u56fe\u7247\u66f4\u65b0\u529f\u80fd\u672a\u5bfc\u5165\uff0c\u8bf7\u68c0\u67e5\u5b89\u88c5\u3002",
            )
            return

        d = self._selected
        if not d:
            return

        self._refresh_title_to_row()
        row_num = self._row_num_for_item(d)
        if row_num is None:
            from tkinter import messagebox
            messagebox.showwarning(
                "Product not in catalog" if self._lang == "en"
                else "\u4e0d\u5728\u76ee\u5f55\u4e2d",
                ("This product has no catalog entry yet.\n"
                 "Run a job first to add it to supplier_catalog.xlsx.")
                if self._lang == "en" else
                "\u6b64\u5546\u54c1\u5c1a\u672a\u5efa\u7acb\u76ee\u5f55\u6761\u76ee\uff0c\u8bf7\u5148\u8fd0\u884c\u4e00\u6b21\u4efb\u52a1\u5c06\u5176\u6dfb\u52a0\u5230 supplier_catalog.xlsx\u3002",
            )
            return

        from tkinter import filedialog
        path_str = filedialog.askopenfilename(
            parent=self._d,
            title="Choose product photo" if self._lang == "en" else "\u9009\u62e9\u5546\u54c1\u56fe\u7247",
            filetypes=[
                ("Image files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif"),
                ("All files", "*.*"),
            ],
        )
        if not path_str:
            return

        # ── Re-encode as JPEG ─────────────────────────────────────────────
        try:
            from io import BytesIO
            im = Image.open(path_str)
            if im.mode not in ("RGB",):
                if im.mode == "RGBA":
                    bg = Image.new("RGB", im.size, (255, 255, 255))
                    bg.paste(im, mask=im.split()[3])
                    im = bg
                else:
                    im = im.convert("RGB")
            buf = BytesIO()
            im.save(buf, format="JPEG", quality=90)
            new_bytes = buf.getvalue()
        except Exception as exc:
            from tkinter import messagebox
            messagebox.showerror(
                "Image error" if self._lang == "en" else "\u56fe\u7247\u9519\u8bef",
                f"Could not open image:\n{exc}",
            )
            return

        # ── Write to supplier_catalog.xlsx ────────────────────────────────
        try:
            update_product_map_photo(FILE_SUPPLIER_CATALOG, row_num, new_bytes)
        except Exception as exc:
            from tkinter import messagebox
            messagebox.showerror(
                "Save error" if self._lang == "en" else "\u4fdd\u5b58\u9519\u8bef",
                f"Failed to write photo to catalog:\n{exc}",
            )
            return

        # ── Update canonical photo cache ──────────────────────────────────
        norm = d["norm_title"]
        self._catalog_photos[norm] = new_bytes
        if norm[:50] not in self._catalog_photos or self._catalog_photos[norm[:50]] is not new_bytes:
            self._catalog_photos[norm[:50]] = new_bytes

        # ── Propagate to every in-memory order with the same product title ─
        updated_count = 0
        for item_list in (self._cg_items, self._ch_items):
            for item_d in item_list:
                if (item_d["norm_title"] == norm
                        or item_d["norm_title"][:50] == norm[:50]):
                    item_d["photo_bytes"] = new_bytes
                    updated_count += 1

        # ── Refresh the photo preview in the detail panel ─────────────────
        ph = self._make_preview(new_bytes, max_dim=115)
        if ph:
            self._detail_photo_ref.clear()
            self._detail_photo_ref.append(ph)
            self._detail_photo_lbl.config(image=ph, text="", width=115, height=115)

        # ── Refresh tree thumbnails (old refs cleared to avoid memory bloat) ─
        self._tk_img_refs.clear()
        self._populate_tree()

        # ── Feedback ─────────────────────────────────────────────────────
        if self._lang == "en":
            msg = (f"\u2713 Photo saved  \u2014  updated {updated_count} "
                   f"order{'s' if updated_count != 1 else ''}")
        else:
            msg = f"\u2713 \u56fe\u7247\u5df2\u4fdd\u5b58\uff0c\u5171\u66f4\u65b0 {updated_count} \u4e2a\u8ba2\u5355"
        self._upload_photo_status.config(text=msg)

    def _clear_charm_assignment(self) -> None:
        """Remove the charm code (and optionally shop) for the selected order."""
        d = self._selected
        if not d or update_product_map_cells is None:
            return
        self._refresh_title_to_row()
        row_num = self._row_num_for_item(d)
        if row_num is None:
            return
        try:
            update_product_map_cells(FILE_SUPPLIER_CATALOG, row_num,
                                     charm_code="", charm_shop="")
            d["charm_code"] = ""
            d["charm_shop"] = ""
            d["status"] = "needs_code"
            self._selected_charm_code = ""
            self._highlight_charm_tile("")
            if "shop_var" in self._edit_widgets:
                self._edit_widgets["shop_var"].set("")
            self._save_status_lbl.config(
                text="\u2713 Charm cleared" if self._lang == "en" else "\u2713 \u5df2\u6e05\u9664\u6302\u4ef6"
            )
            self._populate_tree()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", str(e))

    # ── Charm tile click ──────────────────────────────────────────────

    # ── Import charm photos ─────────────────────────────────────────

    def _charm_import_photos(self) -> None:
        """Browse for charm photos, show the SKU dialog, then import into the workbook."""
        from tkinter import filedialog as _fd

        filetypes = [
            ("Image files", "*.png *.jpg *.jpeg *.webp"),
            ("All files", "*.*"),
        ]
        paths = _fd.askopenfilenames(
            parent=self._d,
            title="Select charm photo(s) to import" if self._lang == "en"
                  else "\u9009\u62e9\u8981\u5bfc\u5165\u7684\u6302\u4ef6\u7167\u7247",
            filetypes=filetypes,
        )
        if not paths:
            return

        # Stage every selected file as __incoming__<token>_<stem>.<ext>
        charm_dir = DEFAULT_CHARM_IMAGES_DIR
        charm_dir.mkdir(parents=True, exist_ok=True)
        staged_files: list[Path] = []
        for p in paths:
            try:
                src = Path(p)
                if not src.is_file():
                    continue
                ext = src.suffix.lower()
                if ext not in CHARM_IMAGE_EXTS:
                    continue
                stem = (
                    re.sub(r"[^\w\-.]+", "_", src.stem, flags=re.UNICODE)
                    .strip("._") or "photo"
                )
                name = f"{CHARM_INCOMING_PREFIX}{secrets.token_hex(4)}_{stem[:72]}{ext}"
                dest = charm_dir / name
                shutil.copy2(src, dest)
                staged_files.append(dest)
            except OSError:
                continue

        if not staged_files:
            messagebox.showinfo(
                "No files" if self._lang == "en" else "\u65e0\u6587\u4ef6",
                "No valid image files were selected." if self._lang == "en"
                else "\u672a\u9009\u62e9\u6709\u6548\u7684\u56fe\u7247\u6587\u4ef6\u3002",
                parent=self._d,
            )
            return

        # Show the SKU dialog — user can review/edit auto-generated SKUs
        dlg = _CharmImportSkuDialog(self._d, staged_files, self._lang)

        if not dlg.confirmed:
            # User cancelled — remove the staged files so nothing is left over
            for f in staged_files:
                try:
                    f.unlink()
                except OSError:
                    pass
            return

        sku_overrides: dict[str, str] = dlg.result   # {staged_stem: sku_text}
        n_staged = len(staged_files)

        self._save_status_lbl.config(
            text=f"Importing {n_staged} photo(s)\u2026" if self._lang == "en"
            else f"\u6b63\u5728\u5bfc\u5165 {n_staged} \u5f20\u7167\u7247\u2026",
        )
        self._d.update_idletasks()

        def _work() -> None:
            n_imported = 0
            try:
                if import_charm_screenshot_assets is not None:
                    n_imported, imp_lines = import_charm_screenshot_assets(
                        charm_dir,
                        FILE_SUPPLIER_CATALOG,
                        patterns=[CHARM_INCOMING_PATTERN],
                        sku_overrides=sku_overrides,
                    )
                    for line in imp_lines:
                        self._parent._log_q.put(line + "\n")
                else:
                    # Fallback to subprocess when direct import is unavailable
                    import subprocess as _sp
                    proc = _sp.run(
                        [
                            sys.executable, str(GENERATOR),
                            "--project-dir", str(PROJECT_ROOT),
                            "--import-charm-images",
                            *charm_import_pattern_argv(),
                        ],
                        cwd=str(PROJECT_ROOT),
                        capture_output=True, text=True,
                        encoding="utf-8", errors="replace",
                    )
                    if proc.stdout:
                        self._parent._log_q.put(proc.stdout)
                    if proc.stderr:
                        self._parent._log_q.put(proc.stderr)
                    n_imported = n_staged
            except Exception as e:
                self._parent._log_q.put(f"Import error: {e}\n")
            finally:
                _n = n_imported or n_staged

                def _done() -> None:
                    self._rebuild_gallery()
                    self._save_status_lbl.config(
                        text=(
                            f"\u2713 Imported {_n} charm(s)" if self._lang == "en"
                            else f"\u2713 \u5df2\u5bfc\u5165 {_n} \u4e2a\u6302\u4ef6"
                        ),
                    )

                self._d.after(0, _done)

        threading.Thread(target=_work, daemon=True).start()

    # ── Reorder charm codes ───────────────────────────────────────────

    def _charm_reorder(self) -> None:
        """Open the existing _CharmReorderDialog; rebuild gallery on close."""
        if reorder_charm_library_rows is None or load_charm_library is None:
            from tkinter import messagebox
            messagebox.showerror(
                "Unavailable",
                "Charm reorder is not available (missing imports).",
            )
            return
        if not FILE_SUPPLIER_CATALOG.exists():
            from tkinter import messagebox
            messagebox.showinfo("No catalog", str(FILE_SUPPLIER_CATALOG))
            return
        try:
            entries = load_charm_library(FILE_SUPPLIER_CATALOG)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", str(e))
            return
        if not entries:
            from tkinter import messagebox
            messagebox.showinfo(
                "Empty" if self._lang == "en" else "\u7a7a",
                "No charms in the library yet." if self._lang == "en"
                else "\u6302\u4ef6\u5e93\u4e2d\u6682\u65e0\u6302\u4ef6\u3002",
            )
            return

        charm_dir = DEFAULT_CHARM_IMAGES_DIR
        # The reorder dialog is modal (grab_set) — execution blocks here until it closes
        _CharmReorderDialog(self._parent, entries, FILE_SUPPLIER_CATALOG, charm_dir)
        # After close: rebuild everything with the new code order
        self._rebuild_gallery()
        self._save_status_lbl.config(
            text="\u2713 Charm codes reordered" if self._lang == "en"
            else "\u2713 \u6302\u4ef6\u7f16\u7801\u5df2\u91cd\u6392",
        )

    # ── Rebuild gallery after import / reorder ────────────────────────

    def _rebuild_gallery(self) -> None:
        """Reload charm library from disk and rebuild the entire gallery + tree."""
        try:
            if load_charm_library is not None:
                self._charm_library = load_charm_library(FILE_SUPPLIER_CATALOG)
                self._charm_codes = list(self._charm_library.keys())
            if load_charm_shops is not None:
                shops = load_charm_shops(FILE_SUPPLIER_CATALOG)
                self._charm_shops = [cs.shop_name for cs in shops if cs.shop_name]
                self._charm_shop_stalls = {cs.shop_name: cs.stall for cs in shops if cs.shop_name and cs.stall}
        except Exception:
            pass
        self._build_item_lists()
        if self._mode == "charms":
            self._build_mode_panel()
        self._populate_tree()

    def _on_charm_tile_click(self, code: str) -> None:
        """Single click: select the charm tile and always enforce the canonical
        shop — a charm code must always map to the same supplier."""
        if "code_var" in self._edit_widgets:
            self._edit_widgets["code_var"].set(code)
        self._selected_charm_code = code
        self._highlight_charm_tile(code)
        canonical = self._canonical_charm_shop(code)
        if canonical:
            sv = self._edit_widgets.get("shop_var")
            if sv:
                sv.set(canonical)   # always override — never allow a different shop

    def _on_charm_tile_double_click(self, code: str) -> None:
        """Double click: select the charm and immediately save all selected orders."""
        self._on_charm_tile_click(code)
        if self._selected_indices:
            self._save()

    def _highlight_charm_tile(self, code: str) -> None:
        """Give the selected tile a vivid purple border; reset all others."""
        for c, tile in self._charm_tile_frames.items():
            if c == code:
                tile.config(highlightbackground="#7c3aed", highlightthickness=3,
                            bg="#ede9fe")
                for child in tile.winfo_children():
                    child.config(bg="#ede9fe")
            else:
                tile.config(highlightbackground="#e9d5ff", highlightthickness=2,
                            bg="#f5f3ff")
                for child in tile.winfo_children():
                    try:
                        child.config(bg="#f5f3ff")
                    except Exception:
                        pass

    # ── Compatibility shims for old code paths ────────────────────────

    def _build_edit_fields(self) -> None:
        """Legacy entry point — delegates to _build_mode_panel."""
        self._build_mode_panel()

    # ── Charm code photo preview ──────────────────────────────────────

    def _update_charm_code_preview(self, code: str) -> None:
        """Update the live charm photo preview card whenever the code changes."""
        if not hasattr(self, "_charm_prev_img_lbl"):
            return
        entry = self._charm_library.get(code) if code else None
        if entry is not None and entry.photo_bytes and self._pil_ok:
            ph = self._make_charm_card_photo(entry.photo_bytes, max_dim=84)
            if ph:
                self._charm_lib_photo_ref.clear()
                self._charm_lib_photo_ref.append(ph)
                self._charm_prev_img_lbl.config(image=ph, text="")
                self._charm_prev_code_lbl.config(text=code)
                sku_parts = []
                if entry.sku:
                    sku_parts.append(entry.sku)
                if entry.default_charm_shop:
                    sku_parts.append(f"Default: {entry.default_charm_shop}")
                if entry.notes:
                    sku_parts.append(entry.notes)
                self._charm_prev_sku_lbl.config(
                    text="  |  ".join(sku_parts) if sku_parts else code,
                )
                return
        # No photo or empty code — show placeholder
        self._charm_lib_photo_ref.clear()
        self._charm_prev_img_lbl.config(image="", text="")
        if code and entry is not None:
            self._charm_prev_code_lbl.config(text=code)
            self._charm_prev_sku_lbl.config(
                text=(entry.sku or "(no photo)") if entry else "(not in library)")
        elif code:
            self._charm_prev_code_lbl.config(text=code)
            self._charm_prev_sku_lbl.config(text="Code not found in Charm Library")
        else:
            self._charm_prev_code_lbl.config(text="")
            self._charm_prev_sku_lbl.config(text="Select a charm code above")

    def _make_charm_card_photo(self, raw: bytes, max_dim: int = 84) -> object | None:
        """Create a square-cropped thumbnail for the charm preview card."""
        if not self._pil_ok or not raw or Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            w, h = im.size
            if w > max_dim or h > max_dim:
                ratio = min(max_dim / w, max_dim / h)
                im = im.resize((int(w * ratio), int(h * ratio)), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            return ph
        except Exception:
            return None

    # ── Tree population ───────────────────────────────────────────────

    def _populate_tree(self) -> None:
        tree = self._tree
        tree.delete(*tree.get_children())
        self._tk_img_refs.clear()

        items = self._cg_items if self._mode == "casegrip" else self._ch_items
        filt   = self._active_filter
        q      = self._search_var.get().strip().lower()

        # ── 1. Filter ──────────────────────────────────────────────────
        visible: list[tuple[int, dict]] = []
        for orig_idx, d in enumerate(items):
            if filt != "all" and d["status"] != filt:
                continue
            if q:
                if q not in " ".join(str(v) for v in d.values()).lower():
                    continue
            visible.append((orig_idx, d))

        # ── 2. Sort (stable) ───────────────────────────────────────────
        sort_map = self._CH_SORT_MAP if self._mode == "charms" else self._CG_SORT_MAP
        item_key = sort_map.get(self._sort_col, "")
        if item_key:
            reverse = (self._sort_dir == "desc")
            if item_key == "stall":
                # Rank by the exact row order of the Stall column in the
                # Suppliers sheet of supplier_catalog.xlsx.
                _stall_rank: dict[str, int] = {
                    s: i for i, s in enumerate(self._sup_stalls)
                }
                _fallback = len(self._sup_stalls)
                def _sort_key(pair: tuple):  # type: ignore[misc]
                    s = str(pair[1].get("stall") or "").strip()
                    return _stall_rank.get(s, _fallback)
            elif item_key.startswith("_ps_"):
                # Purchase-status sort: rank by _PSTATUS_SORT_ORDER priority.
                # _ps_case / _ps_grip / _ps_charm
                comp = item_key[4:]   # "case", "grip", or "charm"
                _so = self._PSTATUS_SORT_ORDER
                def _sort_key(pair: tuple):  # type: ignore[misc]
                    d_ = pair[1]
                    s = self._pstatuses.get(
                        (d_.get("order", ""), (d_.get("norm_title") or "")[:50], comp),
                        "Pending",
                    )
                    return _so.get(s, 99)
            else:
                def _sort_key(pair: tuple) -> str:  # type: ignore[misc]
                    v = pair[1].get(item_key, "") or ""
                    s = str(v).strip()
                    return s.lower() if s else "\uffff"
            visible.sort(key=_sort_key, reverse=reverse)

        # ── 3. Render ──────────────────────────────────────────────────
        labels = self._FILTER_LABELS.get(self._lang, self._FILTER_LABELS["en"])
        for seq, (orig_idx, d) in enumerate(visible, start=1):
            # Purchase status takes visual priority over supplier-match status
            ps = self._worst_pstatus(d)
            if ps != "Pending":
                base_tag = self._PSTATUS_TAG.get(ps, d["status"])
            else:
                base_tag = d["status"]
            tag = base_tag if seq % 2 == 1 else f"alt_{base_tag}"
            thumb = self._thumb(d.get("photo_bytes"))
            kw: dict = {}
            if thumb is not None:
                kw["image"] = thumb
                kw["text"] = ""
            else:
                kw["text"] = "\u2014"
            st_text = labels.get(d["status"], d["status"])
            title_short = d["title"]
            if len(title_short) > 52:
                title_short = title_short[:50] + "\u2026"
            raw_notes = (d.get("private_notes") or "").strip()
            notes_short = (raw_notes[:38] + "\u2026") if len(raw_notes) > 40 else raw_notes

            # Compute per-component purchase status display text.
            # Use [:50] to match the key format used by generate_shopping_route.py.
            _order  = d.get("order", "")
            _norm   = (d.get("norm_title") or "")[:50]
            _pd     = self._PSTATUS_DISPLAY

            if self._mode == "casegrip":
                # Show purchase status in Case / Grip cells; "—" when N/A
                if d["case"]:
                    _cs = self._pstatuses.get((_order, _norm, "case"), "Pending")
                    case_cell = _pd.get(_cs, _cs)
                else:
                    case_cell = "\u2014"
                if d["grip"]:
                    _gs = self._pstatuses.get((_order, _norm, "grip"), "Pending")
                    grip_cell = _pd.get(_gs, _gs)
                else:
                    grip_cell = "\u2014"
                vals = (seq, f"#{d['order']}", title_short,
                        d.get("etsy_shop") or "\u2014",
                        d["supplier"] or "\u2014", d["stall"] or "\u2014",
                        case_cell, grip_cell, d["phone"],
                        d["qty"], st_text, notes_short)
            else:
                _chps = self._pstatuses.get((_order, _norm, "charm"), "Pending")
                charm_buy_cell = _pd.get(_chps, _chps)
                vals = (seq, f"#{d['order']}", title_short,
                        d.get("etsy_shop") or "\u2014",
                        d.get("charm_code") or "\u2014",
                        d.get("charm_shop") or "\u2014",
                        d.get("stall") or "\u2014",
                        d["qty"], st_text, charm_buy_cell, notes_short)
            tree.insert("", tk.END, iid=str(orig_idx), values=vals, tags=(tag,), **kw)

        # Update column heading sort indicators
        self._refresh_sort_indicators()

    # ── Selection → detail panel ──────────────────────────────────────

    def _on_select(self, _event=None) -> None:
        sel = self._tree.selection()
        if not sel:
            self._clear_detail()
            return

        items = self._cg_items if self._mode == "casegrip" else self._ch_items

        # Collect all selected valid indices
        indices = []
        for iid in sel:
            try:
                idx = int(iid)
                if 0 <= idx < len(items):
                    indices.append(idx)
            except ValueError:
                pass
        if not indices:
            return

        self._selected_indices = indices
        # Primary item = the most recently clicked (last in selection)
        primary = items[indices[-1]]
        self._selected = primary

        n = len(indices)
        is_multi = n > 1

        # ── Show compact order-info strip ──────────────────────────────
        self._order_info_frame.pack(fill=tk.X, before=self._mode_panel)

        if is_multi:
            # Multi-selection summary
            total_qty = sum(items[i]["qty"] for i in indices)
            if self._lang == "en":
                self._detail_title.config(
                    text=f"{n} orders selected  \u2014  batch assign charm"
                )
                self._detail_meta.config(
                    text=f"Total qty: {total_qty}  \u2022  Ctrl+click or Shift+click to add/remove"
                )
            else:
                self._detail_title.config(
                    text=f"\u5df2\u9009 {n} \u4e2a\u8ba2\u5355 \u2014 \u6279\u91cf\u5206\u914d\u6302\u4ef6"
                )
                self._detail_meta.config(
                    text=f"\u603b\u6570\u91cf: {total_qty}  \u2022  Ctrl/Shift \u70b9\u51fb\u6dfb\u52a0\u9009\u62e9"
                )
            self._detail_notes.config(text="")
            # Hide photo frame for multi-select
            self._detail_photo_frame.pack_forget()
        else:
            d = primary
            self._detail_title.config(text=d["title"])
            parts = [f"Order #{d['order']}"]
            if d.get("phone"):
                parts.append(d["phone"])
            parts.append(f"Qty: {d['qty']}")
            if d.get("buyer"):
                parts.append(d["buyer"])
            if d.get("etsy_shop"):
                parts.append(d["etsy_shop"])
            self._detail_meta.config(text="  \u2022  ".join(parts))
            notes = d.get("private_notes", "")
            self._detail_notes.config(text=f"\U0001F4CB {notes}" if notes else "")

            # ── Photo preview ──────────────────────────────────────────
            self._upload_photo_status.config(text="")
            raw = d.get("photo_bytes")
            if raw and self._pil_ok:
                ph = self._make_preview(raw, max_dim=115)
                if ph:
                    self._detail_photo_ref.clear()
                    self._detail_photo_ref.append(ph)
                    self._detail_photo_lbl.config(image=ph, text="", width=115, height=115)
                else:
                    self._detail_photo_lbl.config(
                        image="",
                        text="No photo" if self._lang == "en" else "\u65e0\u56fe\u7247",
                        width=16, height=4,
                    )
            else:
                self._detail_photo_ref.clear()
                self._detail_photo_lbl.config(
                    image="",
                    text="No photo" if self._lang == "en" else "\u65e0\u56fe\u7247",
                    width=16, height=4,
                )
            # Photo preview + upload are only shown for Case/Grip.
            # In Charms mode the product photo is irrelevant — the charm gallery
            # on the right already shows the canonical charm images.
            if self._mode == "casegrip":
                self._detail_photo_frame.pack(
                    fill=tk.X, before=self._order_info_sep
                )
                self._upload_photo_btn.pack(fill=tk.X, padx=10, pady=(0, 2))
                self._upload_photo_status.pack(fill=tk.X, padx=10, pady=(0, 6))
            else:
                self._detail_photo_frame.pack_forget()
                self._upload_photo_btn.pack_forget()
                self._upload_photo_status.pack_forget()

        # ── Show save controls ─────────────────────────────────────────
        if self._mode == "casegrip":
            self._save_controls_frame.pack(fill=tk.X, padx=12, pady=(8, 0))

        self._save_status_lbl.config(text="")

        # ── Populate edit fields ───────────────────────────────────────
        if self._mode == "casegrip":
            # For multi-select in CG mode, pre-fill from primary only if uniform
            sups  = {items[i].get("supplier", "") for i in indices}
            stalls = {items[i].get("stall", "") for i in indices}
            self._edit_widgets["sup_var"].set(next(iter(sups)) if len(sups) == 1 else "")
            self._edit_widgets["stall_var"].set(next(iter(stalls)) if len(stalls) == 1 else "")
        else:
            # For charms: show current code only if all selected have the same one
            codes = {items[i].get("charm_code", "") for i in indices}
            charm_code = next(iter(codes)) if len(codes) == 1 else ""
            shops = {items[i].get("charm_shop", "") for i in indices}
            if "shop_var" in self._edit_widgets:
                self._edit_widgets["shop_var"].set(next(iter(shops)) if len(shops) == 1 else "")
            # Highlight the shared charm tile (or deselect if mixed)
            self._selected_charm_code = charm_code
            self._highlight_charm_tile(charm_code)
            if self._gallery_hint is not None:
                self._gallery_hint.place_forget()


    def _clear_detail(self) -> None:
        """Reset the panel to its default no-selection state."""
        self._selected = None
        self._selected_indices = []
        self._order_info_frame.pack_forget()
        self._save_controls_frame.pack_forget()
        self._detail_title.config(text="")
        self._detail_meta.config(text="")
        self._detail_notes.config(text="")
        self._save_status_lbl.config(text="")
        # Clear photo preview
        self._detail_photo_frame.pack_forget()
        self._detail_photo_lbl.config(image="", text="")
        self._detail_photo_ref.clear()
        self._upload_photo_status.config(text="")
        self._highlight_charm_tile("")
        self._selected_charm_code = ""
        if self._mode == "charms" and self._gallery_hint is not None:
            self._gallery_hint.place(relx=0.5, rely=0.55, anchor=tk.CENTER)

    # ── Purchase status helpers ───────────────────────────────────────

    def _worst_pstatus(self, d: dict) -> str:
        """Return the worst purchase status for a tree row.

        For CG items computes the worst across all present components (case, grip).
        For charm items returns the single charm status.
        "Pending" is the default when no explicit status is stored.
        Priority (worst-first): Out of Production > Out of Stock > Pending > Purchased.
        """
        order = d.get("order", "")
        norm  = (d.get("norm_title") or "")[:50]   # must match xlsx [:50] key format
        if self._mode == "casegrip":
            present: list[str] = []
            if d.get("case"):
                present.append(self._pstatuses.get((order, norm, "case"), "Pending"))
            if d.get("grip"):
                present.append(self._pstatuses.get((order, norm, "grip"), "Pending"))
            if not present:
                return "Pending"
            for p in self._PSTATUS_PRIORITY:
                if p in present:
                    return p
            return present[0]
        else:
            return self._pstatuses.get((order, norm, "charm"), "Pending")


    def _load_pstatuses_cache(self) -> None:
        """Load persisted purchase statuses from the JSON cache file."""
        import json as _json
        cache_path = OUTPUT_DIR / "route_statuses_cache.json"
        if not cache_path.exists():
            return
        try:
            with open(cache_path, encoding="utf-8") as fh:
                raw: dict = _json.load(fh)
            for k_str, val in raw.items():
                parts = k_str.split("\x00", 2)
                if len(parts) == 3 and val in self._PURCHASE_STATUSES:
                    order_num, norm_title, comp = parts
                    # Enforce [:50] so keys always match the xlsx round-trip format
                    self._pstatuses[(order_num, norm_title[:50], comp)] = val
        except Exception:
            pass

    def _save_pstatuses_cache(self) -> None:
        """Write all current purchase statuses to the JSON cache file."""
        import json as _json
        cache_path = OUTPUT_DIR / "route_statuses_cache.json"
        try:
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            raw = {"\x00".join(k): v for k, v in self._pstatuses.items()}
            with open(cache_path, "w", encoding="utf-8") as fh:
                _json.dump(raw, fh, indent=2, ensure_ascii=False)
        except Exception:
            pass

    # ── Canonical charm-shop lookup ───────────────────────────────────

    def _canonical_charm_shop(self, charm_code: str) -> str:
        """Return the single authoritative charm shop for *charm_code*, or ''
        if none has been established yet.

        Priority order (mirrors normalize_catalog_charm_shops):
          1. Charm Library's ``default_charm_shop`` — set by the user or the
             pipeline normaliser; this is the ground-truth.
          2. Any existing ``charm_shop`` already assigned to an order with this
             code in the current session (first-found wins).

        Callers should use this to enforce the rule: one code → one shop.
        """
        if not charm_code:
            return ""
        lib_entry = self._charm_library.get(charm_code)
        if lib_entry and getattr(lib_entry, "default_charm_shop", ""):
            return lib_entry.default_charm_shop
        for item_d in self._ch_items:
            if item_d.get("charm_code") == charm_code:
                shop = (item_d.get("charm_shop") or "").strip()
                if shop:
                    return shop
        return ""

    # ── Charm Purchase Summary ────────────────────────────────────────

    def _build_charm_purchase_data(self) -> list[dict]:
        """Aggregate ch_items by charm_code and return sorted purchase rows.

        Each row dict:
            code, sku, shop, stall, total_qty, photo_bytes, notes,
            status_by_qty  — {status_string: cumulative_qty}  e.g.
                             {"Purchased": 2, "Pending": 1}
        Sorted (shop, code) — same order as the shopping-route Excel charm section.
        """
        agg: dict[str, dict] = {}
        for d in self._ch_items:
            code = (d.get("charm_code") or "").strip()
            if not code:
                continue
            lib   = self._charm_library.get(code)
            shop  = (d.get("charm_shop") or "").strip()
            stall = (d.get("stall") or "").strip()
            qty   = int(d.get("qty") or 1)

            # Individual order buy status — keyed exactly like the xlsx round-trip
            order  = d.get("order", "")
            norm   = (d.get("norm_title") or "")[:50]
            status = self._pstatuses.get((order, norm, "charm"), "Pending")

            if code not in agg:
                agg[code] = {
                    "code":          code,
                    "sku":           (lib.sku if lib and lib.sku else ""),
                    "shop":          shop,
                    "stall":         stall,
                    "total_qty":     0,
                    # Charm Library photo is the canonical image (same as gallery tiles)
                    "photo_bytes":   (lib.photo_bytes if lib and lib.photo_bytes else None),
                    "notes":         [],
                    # {status_label: qty} — weighted by order quantity
                    "status_by_qty": {},
                }
            agg[code]["total_qty"] += qty
            sbq = agg[code]["status_by_qty"]
            sbq[status] = sbq.get(status, 0) + qty
            if not agg[code]["shop"] and shop:
                agg[code]["shop"] = shop
            if not agg[code]["stall"] and stall:
                agg[code]["stall"] = stall
            pn = (d.get("private_notes") or "").strip()
            if pn and pn not in agg[code]["notes"]:
                agg[code]["notes"].append(pn)

        return sorted(
            agg.values(),
            key=lambda x: ((x["shop"] or "\uffff"), x["code"]),
        )

    def _show_charm_purchase_summary(self) -> None:
        """Open (or bring-to-front) the Charm Purchase List popup.

        Mirrors the shopping-route Excel's CHARMS TO PURCHASE section:
          • One row per unique charm code — total qty across ALL orders
          • Buy Status column: shows Purchased / Pending / Out of Stock etc.
            When qty > 1, shows a "X / Y Purchased" breakdown so you always
            know how many units are still outstanding.
          • Charm Library photo (the canonical image, not the product photo)
          • Shop + Stall + deduplicated private notes
          • Rows colour-coded by shop group; status badge overrides for done rows
          • Non-blocking: stays open while the user assigns charms in the dashboard
          • Refresh button re-aggregates from the current dashboard data
        """
        # Bring to front if already open
        if self._summary_win and self._summary_win.winfo_exists():
            self._summary_win.lift()
            self._summary_win.focus_force()
            return

        rows = self._build_charm_purchase_data()

        # ── Pre-compute header statistics ─────────────────────────────
        total_qty     = sum(r["total_qty"] for r in rows)
        n_codes       = len(rows)
        n_orders      = len({
            d.get("order", "") for d in self._ch_items if d.get("charm_code")
        })
        purchased_qty = sum(
            r["status_by_qty"].get("Purchased", 0) for r in rows
        )
        purchased_codes = sum(
            1 for r in rows
            if r["status_by_qty"].get("Purchased", 0) == r["total_qty"]
        )

        # ── Window setup ──────────────────────────────────────────────
        win = tk.Toplevel(self._d)
        self._summary_win = win
        win.title(
            "Charm Purchase List" if self._lang == "en"
            else "\u6302\u4ef6\u91c7\u8d2d\u6e05\u5355"
        )
        win.transient(self._d)
        win.configure(bg=COLORS["app"])
        win.geometry("950x540")
        win.minsize(680, 360)
        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(1, weight=1)

        # ── Hero header ───────────────────────────────────────────────
        hdr = tk.Frame(win, bg=COLORS["hero"], highlightthickness=0)
        hdr.grid(row=0, column=0, sticky="ew")

        _en = self._lang == "en"
        if _en:
            title_txt = "\U0001F6D2  Charm Purchase List"
            sub_txt   = (
                f"{n_codes} charm type{'s' if n_codes != 1 else ''}  \u2014  "
                f"{total_qty} unit{'s' if total_qty != 1 else ''} total  \u2014  "
                f"{n_orders} order{'s' if n_orders != 1 else ''}  \u2014  "
                f"{purchased_qty} / {total_qty} purchased  \u2014  "
                "sorted by shop"
            )
        else:
            title_txt = "\U0001F6D2  \u6302\u4ef6\u91c7\u8d2d\u6e05\u5355"
            sub_txt   = (
                f"{n_codes} \u79cd\u6302\u4ef6  \u2014  "
                f"\u5171 {total_qty} \u4e2a  \u2014  "
                f"{n_orders} \u4e2a\u8ba2\u5355  \u2014  "
                f"\u5df2\u8d2d {purchased_qty} / {total_qty} \u4e2a  \u2014  "
                "\u6309\u5e97\u94fa\u6392\u5e8f"
            )
        tk.Label(hdr, text=title_txt,
                 font=("Segoe UI", 14, "bold"), fg="#ffffff",
                 bg=COLORS["hero"]).pack(anchor=tk.W, padx=16, pady=(10, 2))
        tk.Label(hdr, text=sub_txt,
                 font=("Segoe UI", 9), fg="#dbeafe",
                 bg=COLORS["hero"]).pack(anchor=tk.W, padx=16, pady=(0, 10))

        # ── Treeview body ─────────────────────────────────────────────
        body = tk.Frame(win, bg=COLORS["app"])
        body.grid(row=1, column=0, sticky="nsew", padx=12, pady=12)
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(0, weight=1)

        _ROW_H = 52 if self._pil_ok else 26
        try:
            sty = ttk.Style()
            sty.configure("Sum.Treeview", rowheight=_ROW_H,
                          font=("Segoe UI", 10),
                          background="#ffffff", fieldbackground="#ffffff")
            sty.configure("Sum.Treeview.Heading",
                          font=("Segoe UI", 10, "bold"),
                          background="#1e1b4b", foreground="#ffffff",
                          relief="flat", padding=(6, 5))
            sty.map("Sum.Treeview.Heading",
                    background=[("active", "#312e81")])
            sty.map("Sum.Treeview",
                    background=[("selected", "#6d28d9")],
                    foreground=[("selected", "#ffffff")])
        except tk.TclError:
            pass

        tree = ttk.Treeview(
            body, show="tree headings", selectmode="browse",
            style="Sum.Treeview",
            columns=("code", "sku", "shop", "stall", "qty", "status", "notes"),
        )
        vsb = ttk.Scrollbar(body, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        tree.heading("#0",      text="Photo",                                          anchor=tk.CENTER)
        tree.heading("code",    text="Charm Code"    if _en else "\u6302\u4ef6\u7f16\u7801", anchor=tk.CENTER)
        tree.heading("sku",     text="SKU",                                            anchor=tk.CENTER)
        tree.heading("shop",    text="Shop"          if _en else "\u6302\u4ef6\u5e97\u94fa")
        tree.heading("stall",   text="Stall"         if _en else "\u6446\u4f4d",       anchor=tk.CENTER)
        tree.heading("qty",     text="Qty"           if _en else "\u6570\u91cf",       anchor=tk.CENTER)
        tree.heading("status",  text="Buy Status"    if _en else "\u91c7\u8d2d\u72b6\u6001", anchor=tk.CENTER)
        tree.heading("notes",   text="Private Notes" if _en else "\u5907\u6ce8")

        _PH_W = 56 if self._pil_ok else 12
        tree.column("#0",     width=_PH_W, minwidth=_PH_W, stretch=False, anchor=tk.CENTER)
        tree.column("code",   width=92,    minwidth=70,     stretch=False, anchor=tk.CENTER)
        tree.column("sku",    width=70,    minwidth=50,     stretch=False, anchor=tk.CENTER)
        tree.column("shop",   width=130,   minwidth=90,     stretch=False)
        tree.column("stall",  width=60,    minwidth=48,     stretch=False, anchor=tk.CENTER)
        tree.column("qty",    width=46,    minwidth=36,     stretch=False, anchor=tk.CENTER)
        tree.column("status", width=148,   minwidth=100,    stretch=False, anchor=tk.CENTER)
        tree.column("notes",  width=200,   minwidth=100,    stretch=True)

        # ── Row tags ──────────────────────────────────────────────────
        # Shop-group palette (background tints + matching text)
        _PALETTE = [
            ("#f3e8ff", "#5b21b6"),  # violet
            ("#ede9fe", "#4c1d95"),  # deep violet
            ("#e0e7ff", "#3730a3"),  # indigo
            ("#dbeafe", "#1e3a8a"),  # blue
            ("#d1fae5", "#065f46"),  # green
            ("#fef3c7", "#78350f"),  # amber
            ("#ffe4e6", "#9f1239"),  # rose
            ("#f0fdf4", "#14532d"),  # emerald
        ]
        shops_in_order: list[str] = []
        for r in rows:
            if r["shop"] not in shops_in_order:
                shops_in_order.append(r["shop"])
        shop_tag: dict[str, str] = {}
        for i, s in enumerate(shops_in_order):
            tag_name = f"_sp{i}"
            bg, fg   = _PALETTE[i % len(_PALETTE)]
            tree.tag_configure(tag_name, background=bg, foreground=fg)
            shop_tag[s] = tag_name

        # Status-override tags — fully purchased rows get a distinct green finish
        # so they stand out regardless of shop group colour.
        tree.tag_configure("_ps_done",  background="#dcfce7", foreground="#14532d")
        tree.tag_configure("_ps_oos",   background="#fee2e2", foreground="#991b1b")
        tree.tag_configure("_ps_oop",   background="#f3f4f6", foreground="#374151")
        tree.tag_configure("_ps_part",  background="#fef9c3", foreground="#92400e")

        # ── Status formatting helper ──────────────────────────────────
        def _fmt_status(sbq: dict, total: int) -> str:
            """Format status_by_qty into a compact, human-readable badge.

            Single unit  →  plain status text with a leading symbol.
            Multiple     →  "X / Y Purchased" plus a secondary breakdown when
                            units are spread across more than one status bucket.
            """
            if not sbq:
                return "\u23f3 Pending" if _en else "\u23f3 \u5f85\u8d2d"

            purchased = sbq.get("Purchased", 0)
            oos       = sbq.get("Out of Stock", 0)
            oop       = sbq.get("Out of Production", 0)
            pending   = sbq.get("Pending", 0)

            if total == 1:
                # Simple single-unit display
                if purchased:
                    return "\u2713 Purchased"   if _en else "\u2713 \u5df2\u8d2d"
                if oos:
                    return "\u26a0 Out of Stock" if _en else "\u26a0 \u7f3a\u8d27"
                if oop:
                    return "\u2717 Out of Prod." if _en else "\u2717 \u505c\u4ea7"
                return "\u23f3 Pending"          if _en else "\u23f3 \u5f85\u8d2d"

            # Multi-unit: always show X / Y Purchased
            if purchased == total:
                return (f"\u2713 All {total} Purchased"
                        if _en else f"\u2713 \u5168\u90e8 {total} \u4e2a\u5df2\u8d2d")
            if purchased == 0:
                # None purchased — show the dominant non-purchased state
                if oos == total:
                    return (f"\u26a0 All {total} Out of Stock"
                            if _en else f"\u26a0 {total} \u4e2a\u7f3a\u8d27")
                if oop == total:
                    return (f"\u2717 All {total} Out of Prod."
                            if _en else f"\u2717 {total} \u4e2a\u505c\u4ea7")
                # Mixed non-purchased — list each bucket
                parts = []
                if oop:
                    parts.append(f"{oop}\u00d7\u2717" if not _en
                                 else f"{oop}\u00d7Out-of-Prod")
                if oos:
                    parts.append(f"{oos}\u00d7\u26a0" if not _en
                                 else f"{oos}\u00d7OOS")
                if pending:
                    parts.append(f"{pending}\u00d7\u23f3" if not _en
                                 else f"{pending}\u00d7Pending")
                return "  ".join(parts) if parts else f"\u23f3 {total} Pending"

            # Partial purchase — "X / Y Purchased (+details)"
            remaining = total - purchased
            base = (f"{purchased} / {total} Purchased"
                    if _en else f"{purchased} / {total} \u4e2a\u5df2\u8d2d")
            extras = []
            if oos:
                extras.append(f"{oos} OOS" if _en else f"{oos} \u7f3a\u8d27")
            if oop:
                extras.append(f"{oop} OOP" if _en else f"{oop} \u505c\u4ea7")
            if extras:
                return f"{base}  ({', '.join(extras)})"
            return base

        def _status_tag(sbq: dict, total: int) -> str:
            """Return the status-override tag name for a row, or '' for shop colour."""
            purchased = sbq.get("Purchased", 0)
            if purchased == total and total > 0:
                return "_ps_done"
            oos = sbq.get("Out of Stock", 0)
            oop = sbq.get("Out of Production", 0)
            if oop > 0 and oos == 0 and sbq.get("Purchased", 0) == 0 and sbq.get("Pending", 0) == 0:
                return "_ps_oop"
            if oos > 0 and oop == 0 and sbq.get("Purchased", 0) == 0 and sbq.get("Pending", 0) == 0:
                return "_ps_oos"
            purchased = sbq.get("Purchased", 0)
            if purchased > 0 and purchased < total:
                return "_ps_part"
            return ""

        # Keep photo ImageTk refs alive for the window's lifetime
        self._summary_photo_refs.clear()

        for r in rows:
            sbq   = r.get("status_by_qty", {})
            total = r["total_qty"]

            s_tag  = _status_tag(sbq, total)
            sh_tag = shop_tag.get(r["shop"], "")
            # Status tag takes visual priority over shop-group tint
            row_tag = s_tag if s_tag else sh_tag

            ph  = None
            raw = r.get("photo_bytes")
            if raw and self._pil_ok and Image is not None and ImageTk is not None:
                try:
                    im = Image.open(BytesIO(raw))
                    if im.mode == "RGBA":
                        bg_im = Image.new("RGB", im.size, (255, 255, 255))
                        bg_im.paste(im, mask=im.split()[3])
                        im = bg_im
                    elif im.mode != "RGB":
                        im = im.convert("RGB")
                    im.thumbnail((48, 48), Image.Resampling.LANCZOS)
                    ph = ImageTk.PhotoImage(im)
                    self._summary_photo_refs.append(ph)
                except Exception:
                    ph = None

            kw: dict = {}
            if ph:
                kw["image"] = ph
                kw["text"]  = ""
            else:
                kw["text"] = "\U0001F48E"   # gem emoji fallback

            notes_txt = ";  ".join(r["notes"])
            if len(notes_txt) > 80:
                notes_txt = notes_txt[:78] + "\u2026"

            status_txt = _fmt_status(sbq, total)

            tree.insert(
                "", tk.END,
                values=(
                    r["code"],
                    r["sku"] or "",
                    r["shop"] or "\u2014",
                    r["stall"] or "\u2014",
                    total,
                    status_txt,
                    notes_txt,
                ),
                tags=(row_tag,),
                **kw,
            )

        # ── Footer bar ────────────────────────────────────────────────
        foot = tk.Frame(win, bg=COLORS["app"])
        foot.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 10))

        ttk.Button(
            foot,
            text="Close" if self._lang == "en" else "\u5173\u95ed",
            command=win.destroy,
            style="Tool.TButton",
        ).pack(side=tk.RIGHT)

        def _refresh() -> None:
            win.destroy()
            self._summary_win = None
            self._show_charm_purchase_summary()

        ttk.Button(
            foot,
            text="\u27f3  Refresh" if self._lang == "en" else "\u27f3  \u5237\u65b0",
            command=_refresh,
            style="Tool.TButton",
        ).pack(side=tk.RIGHT, padx=(0, 8))

        hint_txt = (
            "Refresh after updating buy statuses or regenerating the route."
            if self._lang == "en" else
            "\u66f4\u65b0\u91c7\u8d2d\u72b6\u6001\u6216\u91cd\u65b0\u751f\u6210\u540e\uff0c\u70b9\u51fb\u5237\u65b0\u3002"
        )
        tk.Label(
            foot, text=hint_txt,
            font=("Segoe UI", 8), fg=COLORS["muted"], bg=COLORS["app"],
        ).pack(side=tk.LEFT)

        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.focus_set()

    # ── Row-mapping refresh ───────────────────────────────────────────

    def _refresh_title_to_row(self) -> None:
        """Re-read the Product Map and rebuild title_to_row from scratch.

        The catalog may be sorted or rebuilt between when the dashboard first
        opens and when the user clicks Save.  A stale mapping can point a
        product title to the wrong row — in the worst case to the merged TOTAL
        row — causing ``'MergedCell' object attribute 'value' is read-only``.

        This is called at the start of every save operation so the row numbers
        are always current without requiring the user to close and reopen the
        dashboard.
        """
        if list_product_map_rows_for_picker is None or not FILE_SUPPLIER_CATALOG.exists():
            return
        try:
            new_map: dict[str, int] = {}
            for pr in list_product_map_rows_for_picker(FILE_SUPPLIER_CATALOG):
                full_key = _normalize(pr.title)
                new_map[full_key] = pr.row_num
                new_map[full_key[:50]] = pr.row_num
            self._title_to_row = new_map
        except Exception:
            pass   # keep the existing mapping if the refresh fails

    # ── Save (single or batch) ────────────────────────────────────────

    def _save(self) -> None:
        if not self._selected_indices or update_product_map_cells is None:
            return
        items = self._cg_items if self._mode == "casegrip" else self._ch_items

        # Always refresh the row mapping before writing so stale row numbers
        # (from catalog sorts / rebuilds that happened since the dashboard opened)
        # never reach update_product_map_cells.
        self._refresh_title_to_row()

        # ── Resolve what to write ──────────────────────────────────────
        if self._mode == "casegrip":
            new_sup   = self._edit_widgets["sup_var"].get().strip()
            new_stall = self._edit_widgets["stall_var"].get().strip()
            new_code  = new_shop = ""
        else:
            new_code = self._selected_charm_code
            new_shop = self._edit_widgets["shop_var"].get().strip()
            new_sup  = new_stall = ""
            if not new_code:
                from tkinter import messagebox
                messagebox.showwarning(
                    "No charm selected" if self._lang == "en" else "\u672a\u9009\u62e9\u6302\u4ef6",
                    "Click a charm tile in the gallery to select a charm code." if self._lang == "en"
                    else "\u8bf7\u5728\u6302\u4ef6\u5e93\u4e2d\u70b9\u51fb\u4e00\u4e2a\u6302\u4ef6\u56fe\u683c\u3002",
                )
                return
            # ── Enforce canonical shop (1:1 rule) ─────────────────────
            # A charm code must always map to the same shop.  Override
            # whatever is in the dropdown with the established canonical
            # shop so this invariant can never be violated from the UI.
            _canonical = self._canonical_charm_shop(new_code)
            if _canonical:
                if new_shop and new_shop != _canonical:
                    # Silently correct the mismatch — the tile-click handler
                    # should have already set the right shop, but guard here
                    # as a second line of defence.
                    new_shop = _canonical
                    sv = self._edit_widgets.get("shop_var")
                    if sv:
                        sv.set(new_shop)
                elif not new_shop:
                    new_shop = _canonical
                    sv = self._edit_widgets.get("shop_var")
                    if sv:
                        sv.set(new_shop)

        # ── Batch loop — write every selected row ──────────────────────
        skipped: list[str] = []
        saved_count = 0
        errors: list[str] = []

        for idx in self._selected_indices:
            if idx < 0 or idx >= len(items):
                continue
            d = items[idx]
            row_num = self._row_num_for_item(d)
            if row_num is None:
                skipped.append(d["title"][:40])
                continue
            try:
                if self._mode == "casegrip":
                    update_product_map_cells(
                        FILE_SUPPLIER_CATALOG, row_num,
                        shop_name=new_sup, stall=new_stall,
                    )
                    d["supplier"] = new_sup
                    d["stall"]    = new_stall
                    d["status"]   = "ready" if (new_sup or new_stall) else d["status"]
                else:
                    update_product_map_cells(
                        FILE_SUPPLIER_CATALOG, row_num,
                        charm_code=new_code, charm_shop=new_shop,
                    )
                    d["charm_code"] = new_code
                    d["charm_shop"] = new_shop
                    if new_code and new_shop:
                        d["status"] = "assigned"
                    elif not new_code:
                        d["status"] = "needs_code"
                    else:
                        d["status"] = "needs_shop"
                saved_count += 1
            except Exception as e:
                errors.append(str(e))

        # ── Feedback ───────────────────────────────────────────────────
        if errors:
            from tkinter import messagebox
            messagebox.showerror(
                "Batch save — errors",
                f"{saved_count} saved, {len(errors)} failed:\n" + "\n".join(errors[:5]),
            )
        elif skipped:
            from tkinter import messagebox
            messagebox.showwarning(
                "Batch save — not found",
                f"{saved_count} saved.\n"
                f"{len(skipped)} product(s) not in catalog:\n" + "\n".join(skipped[:5]),
            )

        if saved_count:
            n = len(self._selected_indices)
            if self._lang == "en":
                msg = (f"\u2713 Saved {saved_count} order{'s' if n > 1 else ''}"
                       if saved_count == n else
                       f"\u2713 Saved {saved_count} / {n}")
            else:
                msg = f"\u2713 \u5df2\u4fdd\u5b58 {saved_count} \u4e2a\u8ba2\u5355"
            self._save_status_lbl.config(text=msg)

            # After a successful charm save, run the catalog normaliser so that
            # any sibling product-map rows sharing the same charm code are also
            # corrected to use the canonical shop — enforcing 1:1 consistency
            # on disk without the user having to do anything extra.
            if self._mode == "charms" and normalize_catalog_charm_shops is not None:
                try:
                    normalize_catalog_charm_shops(FILE_SUPPLIER_CATALOG)
                except Exception:
                    pass   # normalisation is best-effort; never block the UI

            # Auto-refresh the Purchase List popup so the updated quantities
            # are visible immediately, without the user clicking Refresh.
            if (self._mode == "charms"
                    and self._summary_win
                    and self._summary_win.winfo_exists()):
                self._summary_win.destroy()
                self._summary_win = None
                self._show_charm_purchase_summary()

            self._populate_tree()

    def _save_phone(self) -> None:
        """Save the phone model edit to the in-memory items and the order cache."""
        if not self._selected_indices:
            return
        new_phone = self._edit_widgets.get("phone_var", tk.StringVar()).get().strip()
        if not new_phone:
            from tkinter import messagebox
            messagebox.showwarning(
                "No phone model" if self._lang == "en" else "\u672a\u8f93\u5165\u624b\u673a\u578b\u53f7",
                "Enter or select a phone model first." if self._lang == "en"
                else "\u8bf7\u5148\u8f93\u5165\u6216\u9009\u62e9\u624b\u673a\u578b\u53f7\u3002",
            )
            return

        items = self._cg_items
        updated = 0

        for idx in self._selected_indices:
            if idx < 0 or idx >= len(items):
                continue
            d = items[idx]
            old_phone = d.get("phone", "")
            if old_phone == new_phone:
                updated += 1
                continue

            # Update the display dict
            d["phone"] = new_phone

            # Update the corresponding raw ResolvedItem so the cache stays accurate
            order_num  = d.get("order", "")
            norm_title = d.get("norm_title", "")
            for r in self._items:
                if (r.order.order_number == order_num
                        and _normalize is not None
                        and _normalize(r.item.title) == norm_title):
                    r.item.phone_model = new_phone
                    break

            updated += 1

        # Persist the updated phone models to the cache file
        if updated and save_cache is not None:
            try:
                _existing_pdfs: set = set()
                if load_cache is not None and FILE_ORDERS_CACHE.exists():
                    try:
                        _, _existing_pdfs = load_cache(FILE_ORDERS_CACHE)
                    except Exception:
                        pass
                save_cache(FILE_ORDERS_CACHE, self._items, _existing_pdfs)
            except Exception as e:
                from tkinter import messagebox
                messagebox.showerror(
                    "Cache save failed" if self._lang == "en" else "\u7f13\u5b58\u4fdd\u5b58\u5931\u8d25",
                    str(e),
                )
                return

        if updated:
            n = len(self._selected_indices)
            if self._lang == "en":
                msg = (f"\u2713 Phone model updated for {updated} order{'s' if updated > 1 else ''}"
                       if updated == n else
                       f"\u2713 Updated {updated} / {n}")
            else:
                msg = f"\u2713 \u5df2\u66f4\u65b0 {updated} \u4e2a\u8ba2\u5355\u624b\u673a\u578b\u53f7"
            self._save_status_lbl.config(text=msg)
            self._populate_tree()

    # ── Hover zoom ────────────────────────────────────────────────────

    # ── Inline cell status-picker ─────────────────────────────────────

    def _on_tree_cell_click(self, event: tk.Event) -> None:
        """On ButtonRelease-1: if a purchase-status column was clicked, show
        the inline floating status picker anchored to that cell."""
        region = self._tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        col_id = self._tree.identify_column(event.x)   # e.g. "#6"
        iid    = self._tree.identify_row(event.y)
        if not col_id or not iid:
            return

        # Map "#N" (1-based, excluding icon col) to column name
        try:
            col_n = int(col_id.lstrip("#")) - 1     # 0-based index
        except ValueError:
            return
        columns = list(self._tree["columns"])
        if col_n < 0 or col_n >= len(columns):
            return
        col_name = columns[col_n]

        # Inline phone-model picker (Case/Grip mode only)
        if self._mode == "casegrip" and col_name == "phone":
            items = self._cg_items
            try:
                item_idx = int(iid)
            except ValueError:
                return
            if item_idx < 0 or item_idx >= len(items):
                return
            try:
                bbox = self._tree.bbox(iid, col_name)
            except Exception:
                return
            if not bbox:
                return
            self._show_phone_popup(bbox, item_idx, items[item_idx])
            return

        # Only act on purchase-status columns
        if self._mode == "casegrip":
            ps_col_comp = {"case": "case", "grip": "grip"}
        else:
            ps_col_comp = {"buy_status": "charm"}
        comp = ps_col_comp.get(col_name)
        if comp is None:
            return

        items = self._cg_items if self._mode == "casegrip" else self._ch_items
        try:
            item_idx = int(iid)
        except ValueError:
            return
        if item_idx < 0 or item_idx >= len(items):
            return
        d = items[item_idx]

        # Skip N/A cells (component absent from this order)
        if comp == "case" and not d.get("case"):
            return
        if comp == "grip" and not d.get("grip"):
            return

        # Get cell bounding box (relative to the tree widget)
        try:
            bbox = self._tree.bbox(iid, col_name)
        except Exception:
            return
        if not bbox:
            return      # row may be scrolled out of view

        self._show_status_popup(bbox, item_idx, comp, d)

    def _show_status_popup(
        self, bbox: tuple, item_idx: int, comp: str, d: dict
    ) -> None:
        """Minimalist, card-style floating status picker anchored to a cell."""
        bx, by, bw, bh = bbox
        order   = d.get("order", "")
        norm    = (d.get("norm_title") or "")[:50]   # must match xlsx [:50] key format
        current = self._pstatuses.get((order, norm, comp), "Pending")

        # ── Design tokens ─────────────────────────────────────────────
        BG          = "#ffffff"
        BORDER      = "#e2e8f0"
        TEXT        = "#1e293b"
        TEXT_MUTED  = "#94a3b8"
        HOV_BG      = "#f1f5f9"
        SEL_TEXT    = "#1e293b"
        PW          = max(bw, 192)
        ROW_H       = 36
        PAD_X       = 12

        # Dot colour and current-row tint per status
        _DOT: dict[str, str] = {
            "Pending":           "#94a3b8",
            "Purchased":         "#16a34a",
            "Out of Stock":      "#d97706",
            "Out of Production": "#dc2626",
        }
        _TINT: dict[str, str] = {
            "Pending":           "#f8fafc",
            "Purchased":         "#f0fdf4",
            "Out of Stock":      "#fffbeb",
            "Out of Production": "#fef2f2",
        }

        # ── Build window ─────────────────────────────────────────────
        # Two nested frames fake a drop-shadow: dark outer shell sits 2 px
        # below+right of the white card, giving a subtle depth cue.
        shadow = tk.Toplevel(self._d)
        shadow.wm_overrideredirect(True)
        shadow.attributes("-topmost", True)
        shadow.configure(bg="#cbd5e1")          # shadow colour

        popup = tk.Frame(shadow, bg=BORDER)     # 1 px border ring
        popup.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

        card = tk.Frame(popup, bg=BG)
        card.pack(fill=tk.BOTH, expand=True)

        # ── Option rows ───────────────────────────────────────────────
        for i, status in enumerate(self._PURCHASE_STATUSES):
            is_sel   = (status == current)
            row_bg   = _TINT[status] if is_sel else BG
            dot_col  = _DOT[status]
            txt_font = ("Segoe UI", 9, "bold") if is_sel else ("Segoe UI", 9)

            # Thin separator between rows (skip before first)
            if i > 0:
                tk.Frame(card, bg=BORDER, height=1).pack(fill=tk.X)

            row_f = tk.Frame(card, bg=row_bg, cursor="hand2")
            row_f.pack(fill=tk.X)

            # 3 px left accent bar for the current selection
            accent_col = dot_col if is_sel else row_bg
            accent = tk.Frame(row_f, bg=accent_col, width=3)
            accent.pack(side=tk.LEFT, fill=tk.Y)

            # Coloured dot
            dot = tk.Label(
                row_f, text="●",
                font=("Segoe UI", 7),
                fg=dot_col, bg=row_bg,
                padx=PAD_X - 4, pady=0, cursor="hand2",
            )
            dot.pack(side=tk.LEFT)

            # Status label
            lbl = tk.Label(
                row_f, text=status,
                font=txt_font,
                fg=SEL_TEXT, bg=row_bg,
                pady=0, anchor=tk.W, cursor="hand2",
            )
            lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

            # Checkmark for current
            chk = tk.Label(
                row_f,
                text="\u2713" if is_sel else "",
                font=("Segoe UI", 9, "bold"),
                fg=dot_col, bg=row_bg,
                padx=PAD_X, pady=0, cursor="hand2",
            )
            chk.pack(side=tk.RIGHT)

            # Fixed row height via inner padding
            for w in (dot, lbl, chk):
                w.configure(pady=(ROW_H - 18) // 2)

            def _cmd(s=status, sw_=shadow):
                sw_.destroy()
                self._apply_status_from_popup(item_idx, comp, s)

            def _enter(_, rf=row_f, ac=accent, dl=dot, ll=lbl, ck=chk,
                       s=status, dc=dot_col):
                rf.configure(bg=HOV_BG)
                ac.configure(bg=dc)
                for w in (dl, ll, ck):
                    w.configure(bg=HOV_BG)

            def _leave(_, rf=row_f, ac=accent, dl=dot, ll=lbl, ck=chk,
                       s=status, dc=dot_col):
                rb = _TINT[s] if s == current else BG
                ab = dc if s == current else rb
                rf.configure(bg=rb)
                ac.configure(bg=ab)
                for w in (dl, ll, ck):
                    w.configure(bg=rb)

            for w in (row_f, accent, dot, lbl, chk):
                w.bind("<Button-1>", lambda _e, fn=_cmd: fn())
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)

        # ── Position ─────────────────────────────────────────────────
        shadow.update_idletasks()
        ph  = shadow.winfo_reqheight()
        tx  = self._tree.winfo_rootx()
        ty  = self._tree.winfo_rooty()
        # Align left edge with cell; drop 4 px below cell bottom
        px  = tx + bx
        py  = ty + by + bh + 4

        sw_ = shadow.winfo_screenwidth()
        sh_ = shadow.winfo_screenheight()
        if px + PW > sw_:
            px = sw_ - PW - 6
        if py + ph > sh_:
            py = ty + by - ph - 4       # flip above when near screen bottom

        # Shadow offset: the shadow shell is 2 px larger on right+bottom
        shadow.geometry(f"{PW + 2}x{ph + 2}+{px}+{py}")
        popup.place(x=0, y=0, width=PW, height=ph)

        # ── Grab + dismiss ────────────────────────────────────────────
        shadow.grab_set()
        shadow.focus_set()

        def _dismiss_if_outside(event: tk.Event) -> None:
            hit = shadow.winfo_containing(event.x_root, event.y_root)
            if hit is None or not str(hit).startswith(str(shadow)):
                shadow.destroy()

        shadow.bind("<Button-1>", _dismiss_if_outside)
        shadow.bind("<Escape>",   lambda _: shadow.destroy())

    def _apply_status_from_popup(
        self, item_idx: int, comp: str, status: str
    ) -> None:
        """Persist a status selected via the inline cell picker."""
        items = self._cg_items if self._mode == "casegrip" else self._ch_items
        if item_idx < 0 or item_idx >= len(items):
            return
        d = items[item_idx]
        # Truncate norm_title to 50 chars — matches the [:50] key format used
        # by generate_shopping_route.py when writing and reading the xlsx, so
        # that our JSON cache and the xlsx status round-trip use identical keys.
        k = (d.get("order", ""), (d.get("norm_title") or "")[:50], comp)

        if status == "Pending":
            self._pstatuses.pop(k, None)
        else:
            self._pstatuses[k] = status
        self._save_pstatuses_cache()

        # Repopulate and restore selection + scroll position
        prev_sel = set(self._tree.selection())
        self._populate_tree()
        iid = str(item_idx)
        try:
            self._tree.selection_set(iid)
            self._tree.see(iid)
        except Exception:
            for s in prev_sel:
                try:
                    self._tree.selection_set(s)
                except Exception:
                    pass


    # ── Inline phone-model picker ─────────────────────────────────────

    _PHONE_MODEL_LIST = [
        "iPhone 17 Pro Max", "iPhone 17 Pro", "iPhone 17 Plus", "iPhone 17",
        "iPhone 16 Pro Max", "iPhone 16 Pro", "iPhone 16 Plus", "iPhone 16",
        "iPhone 15 Pro Max", "iPhone 15 Pro", "iPhone 15 Plus", "iPhone 15",
        "iPhone 14 Pro Max", "iPhone 14 Pro", "iPhone 14 Plus", "iPhone 14",
        "iPhone 13 Pro Max", "iPhone 13 Pro", "iPhone 13 Mini", "iPhone 13",
        "iPhone 12 Pro Max", "iPhone 12 Pro", "iPhone 12 Mini", "iPhone 12",
        "iPhone 11 Pro Max", "iPhone 11 Pro", "iPhone 11",
    ]

    def _show_phone_popup(self, bbox: tuple, item_idx: int, d: dict) -> None:
        """Floating card for editing the phone model of a single row.

        Design mirrors the status picker: shadow shell + 1 px border card.
        Top section: editable combobox (type or pick).
        Lower section: scrollable list of common models for quick selection.
        """
        bx, by, bw, bh = bbox
        current = d.get("phone", "")

        BG         = "#ffffff"
        BORDER     = "#e2e8f0"
        HOV_BG     = "#f1f5f9"
        TEXT       = "#1e293b"
        MUTED      = "#94a3b8"
        ACCENT     = "#3b82f6"
        PW         = max(bw, 230)
        ROW_H      = 30
        PAD_X      = 10
        MAX_ROWS   = 8          # visible rows before the listbox scrolls

        # ── Shadow + card shell (same as status picker) ───────────────
        shadow = tk.Toplevel(self._d)
        shadow.wm_overrideredirect(True)
        shadow.attributes("-topmost", True)
        shadow.configure(bg="#cbd5e1")

        popup = tk.Frame(shadow, bg=BORDER)
        popup.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

        card = tk.Frame(popup, bg=BG)
        card.pack(fill=tk.BOTH, expand=True)

        def _confirm(value: str, sw_: tk.Toplevel = shadow) -> None:
            v = value.strip()
            sw_.destroy()
            if v:
                self._apply_phone_from_popup(item_idx, v)

        # ── Header row ────────────────────────────────────────────────
        hdr = tk.Frame(card, bg=BG)
        hdr.pack(fill=tk.X, padx=PAD_X, pady=(8, 4))
        tk.Label(
            hdr,
            text="PHONE MODEL" if self._lang == "en" else "\u624b\u673a\u578b\u53f7",
            font=("Segoe UI", 7, "bold"), fg=MUTED, bg=BG,
        ).pack(side=tk.LEFT)
        tk.Label(
            hdr,
            text="\u2328  type or click" if self._lang == "en"
            else "\u2328  \u8f93\u5165\u6216\u70b9\u51fb",
            font=("Segoe UI", 7), fg=MUTED, bg=BG,
        ).pack(side=tk.RIGHT)

        # ── Combobox entry ────────────────────────────────────────────
        entry_frame = tk.Frame(card, bg=BG)
        entry_frame.pack(fill=tk.X, padx=PAD_X, pady=(0, 6))

        var = tk.StringVar(value=current)
        cb = ttk.Combobox(
            entry_frame, textvariable=var,
            values=self._PHONE_MODEL_LIST,
            font=("Segoe UI", 10), width=22,
        )
        cb.pack(side=tk.LEFT, fill=tk.X, expand=True)

        confirm_btn = tk.Label(
            entry_frame,
            text="\u2713",
            font=("Segoe UI", 11, "bold"),
            fg="#16a34a", bg=BG, cursor="hand2",
            padx=8,
        )
        confirm_btn.pack(side=tk.LEFT)
        confirm_btn.bind("<Button-1>", lambda _e: _confirm(var.get()))

        cb.bind("<Return>",   lambda _e: _confirm(var.get()))
        cb.bind("<<ComboboxSelected>>", lambda _e: _confirm(var.get()))

        # ── Thin divider ─────────────────────────────────────────────
        tk.Frame(card, bg=BORDER, height=1).pack(fill=tk.X)

        # ── Scrollable model list ─────────────────────────────────────
        list_frame = tk.Frame(card, bg=BG)
        list_frame.pack(fill=tk.BOTH, expand=True)

        lb = tk.Listbox(
            list_frame,
            font=("Segoe UI", 9),
            fg=TEXT, bg=BG,
            selectbackground=ACCENT,
            selectforeground="#ffffff",
            activestyle="none",
            borderwidth=0, highlightthickness=0,
            relief=tk.FLAT,
            height=MAX_ROWS,
            exportselection=False,
            cursor="hand2",
        )
        sb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for model in self._PHONE_MODEL_LIST:
            lb.insert(tk.END, f"  {model}")

        # Pre-select and scroll to the current model
        for i, model in enumerate(self._PHONE_MODEL_LIST):
            if model == current:
                lb.selection_set(i)
                lb.see(i)
                break

        def _on_lb_select(_e=None) -> None:
            sel = lb.curselection()
            if sel:
                _confirm(self._PHONE_MODEL_LIST[sel[0]])

        lb.bind("<ButtonRelease-1>", _on_lb_select)
        lb.bind("<Return>",          _on_lb_select)

        # Filter list as user types in the combobox
        def _filter(*_) -> None:
            q = var.get().strip().lower()
            lb.delete(0, tk.END)
            for model in self._PHONE_MODEL_LIST:
                if q in model.lower():
                    lb.insert(tk.END, f"  {model}")

        var.trace_add("write", _filter)

        # ── Position (same logic as status picker) ────────────────────
        shadow.update_idletasks()
        ph  = shadow.winfo_reqheight()
        tx  = self._tree.winfo_rootx()
        ty  = self._tree.winfo_rooty()
        px  = tx + bx
        py  = ty + by + bh + 4

        sw_ = shadow.winfo_screenwidth()
        sh_ = shadow.winfo_screenheight()
        if px + PW > sw_:
            px = sw_ - PW - 6
        if py + ph > sh_:
            py = ty + by - ph - 4

        shadow.geometry(f"{PW + 2}x{ph + 2}+{px}+{py}")
        popup.place(x=0, y=0, width=PW, height=ph)

        # ── Focus + dismiss on outside click ─────────────────────────
        shadow.grab_set()
        cb.focus_set()
        cb.icursor(tk.END)

        def _dismiss_if_outside(event: tk.Event) -> None:
            hit = shadow.winfo_containing(event.x_root, event.y_root)
            if hit is None or not str(hit).startswith(str(shadow)):
                shadow.destroy()

        shadow.bind("<Button-1>", _dismiss_if_outside)
        shadow.bind("<Escape>",   lambda _: shadow.destroy())

    def _apply_phone_from_popup(self, item_idx: int, new_phone: str) -> None:
        """Persist a phone model selected via the inline cell picker."""
        items = self._cg_items
        if item_idx < 0 or item_idx >= len(items):
            return
        d = items[item_idx]
        if d.get("phone") == new_phone:
            return

        d["phone"] = new_phone

        # Sync the raw ResolvedItem so the cache is accurate
        order_num  = d.get("order", "")
        norm_title = d.get("norm_title", "")
        for r in self._items:
            if (r.order.order_number == order_num
                    and _normalize is not None
                    and _normalize(r.item.title) == norm_title):
                r.item.phone_model = new_phone
                break

        # Persist to cache
        if save_cache is not None:
            try:
                _existing_pdfs: set = set()
                if load_cache is not None and FILE_ORDERS_CACHE.exists():
                    try:
                        _, _existing_pdfs = load_cache(FILE_ORDERS_CACHE)
                    except Exception:
                        pass
                save_cache(FILE_ORDERS_CACHE, self._items, _existing_pdfs)
            except Exception:
                pass

        # Repopulate and restore selection + scroll
        prev_sel = set(self._tree.selection())
        self._populate_tree()
        iid = str(item_idx)
        try:
            self._tree.selection_set(iid)
            self._tree.see(iid)
        except Exception:
            for s in prev_sel:
                try:
                    self._tree.selection_set(s)
                except Exception:
                    pass

    # ── Hover zoom ────────────────────────────────────────────────────

    def _on_tree_motion(self, _e: tk.Event | None = None) -> None:
        tree = self._tree
        if self._hover_after_id[0] is not None:
            self._d.after_cancel(self._hover_after_id[0])
            self._hover_after_id[0] = None
        try:
            lx = tree.winfo_pointerx() - tree.winfo_rootx()
            ly = tree.winfo_pointery() - tree.winfo_rooty()
        except Exception:
            return
        if tree.identify_column(lx) != "#0":
            self._hide_hover()
            return
        iid = tree.identify_row(ly)
        if not iid:
            self._hide_hover()
            return
        if iid == self._hover_active_iid[0] and self._hover_tip is not None:
            return
        self._hide_hover()
        self._hover_after_id[0] = self._d.after(250, lambda i=iid: self._show_hover(i))

    def _show_hover(self, iid: str) -> None:
        if Image is None or ImageTk is None:
            return
        try:
            idx = int(iid)
        except ValueError:
            return
        items = self._cg_items if self._mode == "casegrip" else self._ch_items
        if idx < 0 or idx >= len(items):
            return
        raw = items[idx].get("photo_bytes")
        if not raw:
            return
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            w, h = im.size
            # Always scale to target — upscale small thumbnails AND downscale large originals.
            # Cap at half the shorter screen dimension so the popup never obscures too much.
            _sw = self._d.winfo_screenwidth()
            _sh = self._d.winfo_screenheight()
            max_dim = min(480, max(320, min(_sw, _sh) // 2))
            ratio = min(max_dim / w, max_dim / h)
            new_w, new_h = max(1, int(w * ratio)), max(1, int(h * ratio))
            if (new_w, new_h) != (w, h):
                resample = (Image.Resampling.LANCZOS if ratio <= 1.0
                            else Image.Resampling.BICUBIC)   # BICUBIC smoother for upscale
                im = im.resize((new_w, new_h), resample)
            ph = ImageTk.PhotoImage(im)
            self._hover_photo.clear()
            self._hover_photo.append(ph)
        except Exception:
            return

        self._hover_active_iid[0] = iid
        tip = tk.Toplevel(self._d)
        tip.overrideredirect(True)
        tip.attributes("-topmost", True)
        tip.configure(bg="#ffffff", highlightthickness=2,
                      highlightbackground=COLORS["accent"])
        tk.Label(tip, image=ph, bg="#ffffff", bd=0).pack()
        tip.update_idletasks()

        px = self._d.winfo_pointerx()
        py = self._d.winfo_pointery()
        sw = self._d.winfo_screenwidth()
        sh = self._d.winfo_screenheight()
        tw = tip.winfo_reqwidth()
        th = tip.winfo_reqheight()
        # Position above cursor so the popup doesn't block the row being hovered
        x = min(max(12, px + 16), sw - tw - 12)
        y = min(max(12, py - th - 12), sh - th - 12)
        # Fall back below cursor if there isn't room above
        if y < 12:
            y = min(py + 16, sh - th - 12)
        tip.geometry(f"+{x}+{y}")
        self._hover_tip = tip

    def _hide_hover(self) -> None:
        if self._hover_after_id[0] is not None:
            self._d.after_cancel(self._hover_after_id[0])
            self._hover_after_id[0] = None
        self._hover_active_iid[0] = None
        if self._hover_tip is not None:
            try:
                self._hover_tip.destroy()
            except Exception:
                pass
            self._hover_tip = None

    # ── Regenerate ────────────────────────────────────────────────────

    def _regen(self) -> None:
        if self._parent._run_busy:
            return
        self._parent._run_busy = True
        self._parent._set_chrome_busy(True)
        self._parent._append_log("Regenerating shopping route...\n")
        # Use _collect_generator_args so every user-configured flag (--chinese,
        # --html, --threshold, etc.) is forwarded.  Previously this built the
        # command manually with only --refresh-catalog, which silently skipped
        # the Chinese file and left shopping_route_zh.xlsx with stale photos
        # even after the user uploaded replacements in the dashboard.
        gen_args = self._parent._collect_generator_args(
            "refresh_catalog", include_charm_steps=False
        )
        cmd = [sys.executable, str(GENERATOR), *gen_args]

        def _work():
            try:
                proc = subprocess.run(
                    cmd, cwd=str(PROJECT_ROOT),
                    capture_output=True, text=True, encoding="utf-8", errors="replace",
                )
                if proc.stdout:
                    self._parent._log_q.put(proc.stdout)
                if proc.stderr:
                    self._parent._log_q.put(proc.stderr)
            except Exception as e:
                self._parent._log_q.put(str(e))
            finally:
                def _done():
                    self._parent._run_busy = False
                    self._parent._set_chrome_busy(False)
                    self._save_status.config(
                        text="\u2713 Route regenerated" if self._lang == "en"
                        else "\u2713 \u5df2\u91cd\u65b0\u751f\u6210"
                    )
                    self._save_status_lbl.config(
                        text="\u2713 Route regenerated" if self._lang == "en"
                        else "\u2713 \u5df2\u91cd\u65b0\u751f\u6210"
                    )
                self._parent.after(0, _done)

        threading.Thread(target=_work, daemon=True).start()

    # ── Open Route file picker ────────────────────────────────────────

    def _open_route(self, event=None) -> None:
        """Show a dropdown menu beneath the button listing all route file variants.

        Each entry shows the file label and a faint availability hint
        ("not generated yet") when the file does not exist on disk.  Clicking
        a present file opens it immediately; clicking a missing file shows the
        standard "File not found" info dialog so the user knows what to do.
        """
        lang = self._lang
        missing_hint = self._parent._t("open_route_missing")

        _routes: list[tuple[str, Path]] = [
            (self._parent._t("open_route_en_detail"),  FILE_SHOPPING_ROUTE),
            (self._parent._t("open_route_en_simple"),  FILE_SHOPPING_ROUTE_SIMPLE),
            (self._parent._t("open_route_zh"),         FILE_SHOPPING_ROUTE_ZH),
            (self._parent._t("open_route_html_en"),    FILE_SHOPPING_HTML),
            (self._parent._t("open_route_html_zh"),    FILE_SHOPPING_HTML_ZH),
        ]

        menu = tk.Menu(self._d, tearoff=False,
                       font=("Segoe UI", 10),
                       bg="#ffffff", fg="#1e293b",
                       activebackground=COLORS["accent"],
                       activeforeground="#ffffff",
                       relief=tk.FLAT, bd=0)

        # Title row (disabled, acts as a visual header)
        menu.add_command(
            label=self._parent._t("open_route_menu_title"),
            state=tk.DISABLED,
            font=("Segoe UI", 9, "bold"),
        )
        menu.add_separator()

        def _make_opener(p: Path):
            def _open():
                if not p.exists():
                    from tkinter import messagebox as _mb
                    _mb.showinfo(
                        self._parent._t("file_missing_title"),
                        self._parent._t("file_missing_body", path=str(p)),
                        parent=self._d,
                    )
                    return
                try:
                    if sys.platform == "win32":
                        os.startfile(p)
                    elif sys.platform == "darwin":
                        subprocess.run(["open", str(p)], check=False)
                    else:
                        subprocess.run(["xdg-open", str(p)], check=False)
                except OSError as e:
                    from tkinter import messagebox as _mb
                    _mb.showerror(
                        self._parent._t("file_open_fail_title"),
                        self._parent._t("file_open_fail_body", err=e),
                        parent=self._d,
                    )
            return _open

        for label, path in _routes:
            exists = path.exists()
            display = label if exists else f"{label}  \u2014 {missing_hint}"
            menu.add_command(
                label=display,
                command=_make_opener(path),
                foreground="#1e293b" if exists else "#94a3b8",
                activeforeground="#ffffff" if exists else "#cbd5e1",
            )

        # Position the menu flush below the button that triggered it
        btn = self._btn_open_route
        bx = btn.winfo_rootx()
        by = btn.winfo_rooty() + btn.winfo_height()
        menu.tk_popup(bx, by)
        menu.grab_release()

    # ── Delete selected orders ────────────────────────────────────────

    def _delete_selected_orders(self) -> None:
        """Remove every selected row (and its backing ResolvedItem) from the cache."""
        from tkinter import messagebox as _mb

        sel = self._tree.selection()
        if not sel:
            _mb.showinfo(
                "No selection" if self._lang == "en" else "未选择",
                "Select one or more rows to delete." if self._lang == "en"
                else "请先选择要删除的行。",
                parent=self._d,
            )
            return

        items = self._cg_items if self._mode == "casegrip" else self._ch_items
        n = len(sel)
        label = (
            f"Delete {n} selected order item{'s' if n != 1 else ''}?\n\n"
            "This cannot be undone — the entries will be permanently removed from the cache."
            if self._lang == "en" else
            f"确认删除所选的 {n} 条订单记录？\n\n此操作无法撤销，数据将从缓存中永久删除。"
        )
        if not _mb.askyesno(
            "Confirm Delete" if self._lang == "en" else "确认删除",
            label,
            icon="warning",
            parent=self._d,
        ):
            return

        # Collect (order_number, norm_title) keys to remove
        keys_to_delete: set[tuple[str, str]] = set()
        for iid in sel:
            try:
                d = items[int(iid)]
            except (ValueError, IndexError):
                continue
            keys_to_delete.add((d["order"], d["norm_title"]))

        if not keys_to_delete:
            return

        # Remove matching ResolvedItems from self._items
        self._items = [
            r for r in self._items
            if (_normalize(r.item.title), r.order.order_number)
            not in {(norm, order) for order, norm in keys_to_delete}
        ]

        # Persist updated cache
        try:
            _existing_pdfs: set[str] = set()
            if load_cache is not None and FILE_ORDERS_CACHE.exists():
                try:
                    _, _existing_pdfs = load_cache(FILE_ORDERS_CACHE)
                except Exception:
                    _existing_pdfs = set()
            if save_cache is not None:
                save_cache(FILE_ORDERS_CACHE, self._items, _existing_pdfs)
        except Exception as e:
            _mb.showwarning(
                "Cache warning" if self._lang == "en" else "缓存警告",
                f"{'Items removed from view but failed to persist to cache' if self._lang == 'en' else '已从视图中移除但缓存保存失败'}:\n{e}",
                parent=self._d,
            )

        # Persist deletion tombstone so the generator never restores these
        # orders from the Excel safety-net or from re-processing PDFs.
        try:
            import json as _json
            FILE_DELETED_ORDERS.parent.mkdir(parents=True, exist_ok=True)
            _tomb: list[dict] = []
            if FILE_DELETED_ORDERS.exists():
                try:
                    _tomb = _json.loads(
                        FILE_DELETED_ORDERS.read_text(encoding="utf-8")
                    ).get("deleted", [])
                except Exception:
                    _tomb = []
            _tomb_set = {(e["order"], e["norm_title"]) for e in _tomb}
            for _ord, _nt in keys_to_delete:
                if (_ord, _nt) not in _tomb_set:
                    _tomb.append({"order": _ord, "norm_title": _nt})
            FILE_DELETED_ORDERS.write_text(
                _json.dumps({"deleted": _tomb}, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass  # tombstone write failure is non-fatal; cache deletion already saved

        # Rebuild + refresh UI
        self._build_item_lists()
        self._populate_tree()
        self._clear_detail()
        removed_label = (
            f"\U0001F5D1 {n} item{'s' if n != 1 else ''} deleted"
            if self._lang == "en" else
            f"\U0001F5D1 已删除 {n} 条记录"
        )
        self._save_status.config(text=removed_label, fg="#dc2626")
        self._d.after(3000, lambda: self._save_status.config(text="", fg="#047857"))

    # ── Manual order add ─────────────────────────────────────────────

    def _open_add_manual_order(self) -> None:
        """Open the Add Manual Order dialog."""
        if Order is None or OrderItem is None or ResolvedItem is None or save_cache is None:
            from tkinter import messagebox as _mb
            _mb.showerror(
                self._t("msg_missing_title") if hasattr(self, "_t") else "Missing imports",
                "Cannot add a manual order: core imports are unavailable.",
            )
            return
        _AddManualOrderDialog(
            self._d,
            lang=self._lang,
            supplier_shops=self._sup_shops,
            supplier_stalls=self._sup_stalls,
            supplier_shop_stalls=self._sup_shop_stalls,
            supplier_stall_shops=self._sup_stall_shops,
            charm_codes=self._charm_codes,
            charm_shop_names=self._charm_shops,
            charm_shop_stalls=self._charm_shop_stalls,
            on_submit=self._commit_manual_order,
        )

    def _commit_manual_order(self, payload: dict) -> None:
        """Build a ResolvedItem from the dialog payload, append to cache + tree."""
        # Build Order / OrderItem / ResolvedItem
        order = Order(
            order_number=payload["order_number"],
            etsy_shop="Manual",
            buyer_name=payload.get("buyer", "").strip(),
            order_date="",
            private_notes=payload.get("private_notes", "").strip(),
        )
        item = OrderItem(
            title=payload["title"],
            quantity=payload["qty"],
            phone_model=payload.get("phone", "").strip(),
            style=payload["style"],
            photo_bytes=payload["photo_bytes"],
        )
        order.items.append(item)

        # Build a CatalogEntry when any of the supplier/charm fields are provided
        supplier = None
        sup_shop   = payload.get("supplier", "").strip()
        sup_stall  = payload.get("stall", "").strip()
        charm_code = payload.get("charm_code", "").strip()
        charm_shop = payload.get("charm_shop", "").strip()
        if sup_shop or sup_stall or charm_code or charm_shop:
            try:
                from generate_shopping_route import CatalogEntry as _CE
                supplier = _CE(
                    product_title=payload["title"],
                    shop_name=sup_shop,
                    stall=sup_stall,
                    charm_code=charm_code,
                    charm_shop=charm_shop,
                )
            except ImportError:
                supplier = None

        resolved = ResolvedItem(order=order, item=item, supplier=supplier)
        self._items.append(resolved)

        # Persist the updated cache
        try:
            _existing_pdfs: set[str] = set()
            if load_cache is not None and FILE_ORDERS_CACHE.exists():
                try:
                    _, _existing_pdfs = load_cache(FILE_ORDERS_CACHE)
                except Exception:
                    _existing_pdfs = set()
            save_cache(FILE_ORDERS_CACHE, self._items, _existing_pdfs)
        except Exception as e:
            from tkinter import messagebox as _mb
            _mb.showwarning(
                "Cache warning" if self._lang == "en" else "\u7f13\u5b58\u8b66\u544a",
                f"Order added in-memory but failed to persist to cache:\n{e}",
            )

        # Rebuild + refresh UI
        self._build_item_lists()
        self._populate_tree()
        self._save_status.config(
            text="\u2713 Order added" if self._lang == "en"
            else "\u2713 \u5df2\u6dfb\u52a0\u8ba2\u5355"
        )
        self._d.after(3000, lambda: self._save_status.config(text=""))


class _AddManualOrderDialog:
    """Modal dialog to manually add an order to the dashboard.

    Mandatory fields: product image + at least one of case/grip/charm.
    All other fields (title, qty, phone, supplier, stall, charm code,
    charm shop, buyer, private notes) are optional.
    Submits a payload dict back via ``on_submit`` callback.
    """

    _MAX_PREVIEW = 160

    def __init__(
        self,
        parent: tk.Misc,
        *,
        lang: str,
        supplier_shops: list[str],
        supplier_stalls: list[str],
        supplier_shop_stalls: dict[str, str],
        supplier_stall_shops: dict[str, str],
        charm_codes: list[str],
        charm_shop_names: list[str],
        charm_shop_stalls: dict[str, str],
        on_submit,
    ) -> None:
        self._lang = lang
        self._sup_shops = supplier_shops
        self._sup_stalls = supplier_stalls
        self._sup_shop_stalls = supplier_shop_stalls
        self._sup_stall_shops = supplier_stall_shops
        self._charm_codes = charm_codes
        self._charm_shops = charm_shop_names
        self._charm_shop_stalls = charm_shop_stalls
        self._on_submit = on_submit
        self._photo_bytes: bytes | None = None
        self._preview_photo: object | None = None
        self._pil_ok = Image is not None and ImageTk is not None

        d = tk.Toplevel(parent)
        self._d = d
        d.title("Add Manual Order" if lang == "en" else "\u6dfb\u52a0\u8ba2\u5355")
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.geometry("560x720")
        d.minsize(520, 640)
        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(1, weight=1)

        # ── Hero ──────────────────────────────────────────────────────
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero,
            text="Add Manual Order" if lang == "en" else "\u6dfb\u52a0\u8ba2\u5355",
            font=("Segoe UI", 13, "bold"), fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=16, pady=(10, 2))
        tk.Label(
            hero,
            text=("Image and at least one of Case / Grip / Charm are required."
                  if lang == "en"
                  else "\u5fc5\u586b\u9879\uff1a\u56fe\u7247 \u4ee5\u53ca "
                       "\u624b\u673a\u58f3 / \u652f\u67b6 / \u6302\u4ef6 \u81f3\u5c11\u4e00\u9879"),
            font=("Segoe UI", 9), fg="#dbeafe", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=16, pady=(0, 10))

        # ── Scrollable body ────────────────────────────────────────────
        body_wrap = tk.Frame(d, bg=COLORS["app"])
        body_wrap.grid(row=1, column=0, sticky="nsew")
        body_wrap.grid_columnconfigure(0, weight=1)
        body_wrap.grid_rowconfigure(0, weight=1)

        canvas = tk.Canvas(body_wrap, bg=COLORS["app"],
                           highlightthickness=0, borderwidth=0)
        vsb = ttk.Scrollbar(body_wrap, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        body = tk.Frame(canvas, bg=COLORS["app"])
        body_win = canvas.create_window((0, 0), window=body, anchor=tk.NW)
        body.bind("<Configure>",
                  lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfigure(body_win, width=e.width))

        def _scroll(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _scroll)
        body.bind("<MouseWheel>", _scroll)

        # ── Section: Image (required) ──────────────────────────────────
        sec_img = self._section(body, "IMAGE *" if lang == "en" else "\u56fe\u7247 *")
        img_row = tk.Frame(sec_img, bg=COLORS["card"])
        img_row.pack(fill=tk.X, padx=14, pady=(0, 10))

        self._photo_card = tk.Frame(
            img_row, bg="#f1f5f9", width=self._MAX_PREVIEW,
            height=self._MAX_PREVIEW, cursor="hand2",
            highlightthickness=1, highlightbackground="#cbd5e1",
        )
        self._photo_card.pack(side=tk.LEFT)
        self._photo_card.pack_propagate(False)
        self._photo_hint = tk.Label(
            self._photo_card,
            text="\U0001F5BC\n\nClick to upload" if lang == "en"
            else "\U0001F5BC\n\n\u70b9\u51fb\u4e0a\u4f20",
            font=("Segoe UI", 9), fg="#94a3b8", bg="#f1f5f9",
            justify=tk.CENTER, cursor="hand2",
        )
        self._photo_hint.pack(expand=True)
        self._photo_card.bind("<Button-1>", lambda _e: self._pick_image())
        self._photo_hint.bind("<Button-1>", lambda _e: self._pick_image())

        img_btns = tk.Frame(img_row, bg=COLORS["card"])
        img_btns.pack(side=tk.LEFT, padx=(14, 0), fill=tk.Y)
        tk.Label(
            img_btns,
            text="Product photo (JPG / PNG / WEBP)" if lang == "en"
            else "\u4ea7\u54c1\u7167\u7247\uff08JPG / PNG / WEBP\uff09",
            font=("Segoe UI", 9, "bold"), fg="#334155", bg=COLORS["card"],
            wraplength=240, justify=tk.LEFT,
        ).pack(anchor=tk.W, pady=(4, 6))
        ttk.Button(
            img_btns,
            text="Browse\u2026" if lang == "en" else "\u9009\u62e9\u6587\u4ef6\u2026",
            command=self._pick_image, style="Tool.TButton",
        ).pack(anchor=tk.W, pady=(0, 4))
        self._clear_img_btn = ttk.Button(
            img_btns,
            text="Clear" if lang == "en" else "\u6e05\u9664",
            command=self._clear_image, style="Tool.TButton",
        )
        self._clear_img_btn.pack(anchor=tk.W)
        self._clear_img_btn.state(["disabled"])

        # ── Section: Product ───────────────────────────────────────────
        sec_prod = self._section(body, "PRODUCT" if lang == "en" else "\u4ea7\u54c1")
        self._title_var = self._text_field(
            sec_prod, "Title" if lang == "en" else "\u6807\u9898", ""
        )
        pg_row = tk.Frame(sec_prod, bg=COLORS["card"])
        pg_row.pack(fill=tk.X, padx=14, pady=(0, 6))
        pg_row.columnconfigure(0, weight=3)
        pg_row.columnconfigure(1, weight=1)
        # Phone model
        phone_col = tk.Frame(pg_row, bg=COLORS["card"])
        phone_col.grid(row=0, column=0, sticky=tk.EW, padx=(0, 6))
        tk.Label(phone_col,
                 text="Phone Model" if lang == "en" else "\u624b\u673a\u578b\u53f7",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).pack(anchor=tk.W, pady=(0, 3))
        self._phone_var = tk.StringVar()
        ttk.Entry(phone_col, textvariable=self._phone_var,
                  font=("Segoe UI", 10)).pack(fill=tk.X)
        # Qty
        qty_col = tk.Frame(pg_row, bg=COLORS["card"])
        qty_col.grid(row=0, column=1, sticky=tk.EW)
        tk.Label(qty_col,
                 text="Qty" if lang == "en" else "\u6570\u91cf",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).pack(anchor=tk.W, pady=(0, 3))
        self._qty_var = tk.StringVar(value="1")
        ttk.Spinbox(qty_col, from_=1, to=999, textvariable=self._qty_var,
                    font=("Segoe UI", 10)).pack(fill=tk.X)
        tk.Frame(sec_prod, bg=COLORS["card"], height=6).pack(fill=tk.X)

        # ── Section: Components (mandatory — need ≥1) ──────────────────
        sec_comp = self._section(
            body,
            "COMPONENTS *" if lang == "en" else "\u7ec4\u4ef6 *",
            hint="Select at least one" if lang == "en" else "\u81f3\u5c11\u9009\u62e9\u4e00\u9879",
        )
        comp_row = tk.Frame(sec_comp, bg=COLORS["card"])
        comp_row.pack(fill=tk.X, padx=14, pady=(0, 10))
        self._case_var  = tk.BooleanVar(value=False)
        self._grip_var  = tk.BooleanVar(value=False)
        self._charm_var = tk.BooleanVar(value=False)
        for i, (var, en, zh) in enumerate([
            (self._case_var,  "Case",  "\u624b\u673a\u58f3"),
            (self._grip_var,  "Grip",  "\u652f\u67b6"),
            (self._charm_var, "Charm", "\u6302\u4ef6"),
        ]):
            cb = tk.Checkbutton(
                comp_row, text=(en if lang == "en" else zh),
                variable=var, font=("Segoe UI", 10),
                bg=COLORS["card"], fg="#1e293b",
                activebackground=COLORS["card"],
                selectcolor="#ffffff",
                padx=8, pady=4, cursor="hand2",
            )
            cb.grid(row=0, column=i, sticky=tk.W, padx=(0, 14))

        # ── Section: Supplier (optional) ───────────────────────────────
        sec_sup = self._section(body, "SUPPLIER" if lang == "en" else "\u4f9b\u5e94\u5546",
                                hint="Optional" if lang == "en" else "\u9009\u586b")
        sup_row = tk.Frame(sec_sup, bg=COLORS["card"])
        sup_row.pack(fill=tk.X, padx=14, pady=(0, 10))
        sup_row.columnconfigure(0, weight=1)
        sup_row.columnconfigure(1, weight=1)
        tk.Label(sup_row, text="Supplier" if lang == "en" else "\u4f9b\u5e94\u5546",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).grid(row=0, column=0, sticky=tk.W, pady=(0, 3))
        tk.Label(sup_row, text="Stall" if lang == "en" else "\u6444\u4f4d",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).grid(row=0, column=1, sticky=tk.W,
                                          padx=(6, 0), pady=(0, 3))
        self._supplier_var = tk.StringVar()
        self._stall_var    = tk.StringVar()
        sup_cb = ttk.Combobox(sup_row, textvariable=self._supplier_var,
                              values=self._sup_shops, font=("Segoe UI", 10))
        sup_cb.grid(row=1, column=0, sticky=tk.EW)
        stall_cb = ttk.Combobox(sup_row, textvariable=self._stall_var,
                                values=self._sup_stalls, font=("Segoe UI", 10))
        stall_cb.grid(row=1, column=1, sticky=tk.EW, padx=(6, 0))
        # Bidirectional autofill
        sup_cb.bind("<<ComboboxSelected>>",
                    lambda _e: self._autofill_pair("sup"))
        stall_cb.bind("<<ComboboxSelected>>",
                      lambda _e: self._autofill_pair("stall"))

        # ── Section: Charm details (shown only if charm is checked) ────
        self._charm_section = self._section(
            body, "CHARM DETAILS" if lang == "en" else "\u6302\u4ef6\u4fe1\u606f",
            hint="Optional" if lang == "en" else "\u9009\u586b",
        )
        chc_row = tk.Frame(self._charm_section, bg=COLORS["card"])
        chc_row.pack(fill=tk.X, padx=14, pady=(0, 10))
        chc_row.columnconfigure(0, weight=1)
        chc_row.columnconfigure(1, weight=1)
        tk.Label(chc_row, text="Charm Code" if lang == "en" else "\u6302\u4ef6\u7f16\u7801",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).grid(row=0, column=0, sticky=tk.W, pady=(0, 3))
        tk.Label(chc_row, text="Charm Shop" if lang == "en" else "\u6302\u4ef6\u5e97\u94fa",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).grid(row=0, column=1, sticky=tk.W,
                                          padx=(6, 0), pady=(0, 3))
        self._charm_code_var = tk.StringVar()
        self._charm_shop_var = tk.StringVar()
        ttk.Combobox(chc_row, textvariable=self._charm_code_var,
                     values=self._charm_codes,
                     font=("Segoe UI", 10)).grid(row=1, column=0, sticky=tk.EW)
        ttk.Combobox(chc_row, textvariable=self._charm_shop_var,
                     values=self._charm_shops,
                     font=("Segoe UI", 10)).grid(row=1, column=1, sticky=tk.EW,
                                                  padx=(6, 0))

        # ── Section: Extras (optional) ─────────────────────────────────
        sec_extra = self._section(body, "EXTRAS" if lang == "en" else "\u5176\u4ed6",
                                  hint="Optional" if lang == "en" else "\u9009\u586b")
        self._buyer_var = self._text_field(
            sec_extra, "Buyer Name" if lang == "en" else "\u4e70\u5bb6\u59d3\u540d", ""
        )
        # Private notes — multi-line
        tk.Label(sec_extra,
                 text="Private Notes" if lang == "en" else "\u79c1\u4eba\u5907\u6ce8",
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).pack(anchor=tk.W, padx=14, pady=(0, 3))
        self._notes_text = tk.Text(sec_extra, height=3, font=("Segoe UI", 10),
                                   wrap=tk.WORD, relief=tk.SOLID, bd=1)
        self._notes_text.pack(fill=tk.X, padx=14, pady=(0, 10))

        # ── Error label ────────────────────────────────────────────────
        self._error_lbl = tk.Label(
            body, text="", font=("Segoe UI", 9, "bold"),
            fg="#dc2626", bg=COLORS["app"], anchor=tk.W,
        )
        self._error_lbl.pack(fill=tk.X, padx=14, pady=(4, 0))

        # ── Footer actions ─────────────────────────────────────────────
        foot = tk.Frame(d, bg=COLORS["card"], highlightthickness=1,
                        highlightbackground=COLORS["border"])
        foot.grid(row=2, column=0, sticky="ew")
        inner = tk.Frame(foot, bg=COLORS["card"])
        inner.pack(fill=tk.X, padx=14, pady=10)
        ttk.Button(
            inner, text="Cancel" if lang == "en" else "\u53d6\u6d88",
            command=d.destroy, style="Tool.TButton",
        ).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(
            inner, text="Add Order" if lang == "en" else "\u6dfb\u52a0\u8ba2\u5355",
            command=self._submit, style="Tool.TButton",
        ).pack(side=tk.RIGHT)

        d.bind("<Escape>", lambda _e: d.destroy())

    # ── Small builders ────────────────────────────────────────────────

    def _section(self, parent: tk.Misc, title: str, *, hint: str = "") -> tk.Frame:
        """Build a card-style section with a header row; return the card frame."""
        outer = tk.Frame(parent, bg=COLORS["app"])
        outer.pack(fill=tk.X, padx=14, pady=(10, 0))
        card = tk.Frame(outer, bg=COLORS["card"],
                        highlightthickness=1,
                        highlightbackground=COLORS["border"])
        card.pack(fill=tk.X)
        hdr = tk.Frame(card, bg=COLORS["card"])
        hdr.pack(fill=tk.X, padx=14, pady=(10, 6))
        tk.Label(hdr, text=title, font=("Segoe UI", 8, "bold"),
                 fg="#475569", bg=COLORS["card"]).pack(side=tk.LEFT)
        if hint:
            tk.Label(hdr, text=f"  {hint}", font=("Segoe UI", 8),
                     fg="#94a3b8", bg=COLORS["card"]).pack(side=tk.LEFT)
        return card

    def _text_field(self, parent: tk.Misc, label: str, default: str) -> tk.StringVar:
        """Label + entry pair; return the StringVar."""
        tk.Label(parent, text=label,
                 font=("Segoe UI", 7, "bold"), fg="#94a3b8",
                 bg=COLORS["card"]).pack(anchor=tk.W, padx=14, pady=(0, 3))
        var = tk.StringVar(value=default)
        ttk.Entry(parent, textvariable=var,
                  font=("Segoe UI", 10)).pack(fill=tk.X, padx=14, pady=(0, 8))
        return var

    # ── Image handling ────────────────────────────────────────────────

    def _pick_image(self) -> None:
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self._d,
            title="Choose product image" if self._lang == "en"
            else "\u9009\u62e9\u4ea7\u54c1\u56fe\u7247",
            filetypes=[
                ("Images", "*.jpg *.jpeg *.png *.webp *.bmp"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        try:
            with open(path, "rb") as fh:
                raw = fh.read()
        except OSError as e:
            self._error_lbl.config(text=f"Could not read file: {e}")
            return

        # Normalize to JPEG bytes when PIL is available so the dashboard
        # thumbnail pipeline reliably decodes the result.
        if self._pil_ok and Image is not None:
            try:
                from io import BytesIO
                im = Image.open(BytesIO(raw))
                if im.mode != "RGB":
                    if "A" in im.mode:
                        bg = Image.new("RGB", im.size, (255, 255, 255))
                        bg.paste(im, mask=im.split()[-1])
                        im = bg
                    else:
                        im = im.convert("RGB")
                buf = BytesIO()
                im.save(buf, format="JPEG", quality=90)
                raw = buf.getvalue()
            except Exception:
                pass   # fall back to original bytes

        self._photo_bytes = raw
        self._render_preview()
        self._clear_img_btn.state(["!disabled"])
        self._error_lbl.config(text="")

    def _render_preview(self) -> None:
        """Show a thumbnail inside the image-upload card."""
        for w in self._photo_card.winfo_children():
            w.destroy()
        if not self._pil_ok or not self._photo_bytes or Image is None or ImageTk is None:
            tk.Label(
                self._photo_card,
                text="[image loaded]" if self._lang == "en" else "[\u5df2\u52a0\u8f7d]",
                font=("Segoe UI", 9), fg="#16a34a", bg="#f1f5f9",
            ).pack(expand=True)
            return
        try:
            from io import BytesIO
            im = Image.open(BytesIO(self._photo_bytes))
            if im.mode != "RGB":
                im = im.convert("RGB")
            im.thumbnail((self._MAX_PREVIEW - 8, self._MAX_PREVIEW - 8),
                         Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._preview_photo = ph
            tk.Label(self._photo_card, image=ph, bg="#f1f5f9",
                     cursor="hand2").pack(expand=True)
        except Exception:
            tk.Label(self._photo_card, text="[image]",
                     bg="#f1f5f9").pack(expand=True)

    def _clear_image(self) -> None:
        self._photo_bytes = None
        self._preview_photo = None
        for w in self._photo_card.winfo_children():
            w.destroy()
        self._photo_hint = tk.Label(
            self._photo_card,
            text="\U0001F5BC\n\nClick to upload" if self._lang == "en"
            else "\U0001F5BC\n\n\u70b9\u51fb\u4e0a\u4f20",
            font=("Segoe UI", 9), fg="#94a3b8", bg="#f1f5f9",
            justify=tk.CENTER, cursor="hand2",
        )
        self._photo_hint.pack(expand=True)
        self._photo_hint.bind("<Button-1>", lambda _e: self._pick_image())
        self._clear_img_btn.state(["disabled"])

    # ── Autofill ──────────────────────────────────────────────────────

    def _autofill_pair(self, source: str) -> None:
        """Autofill the other of (supplier, stall) if mapping is known."""
        if source == "sup":
            shop = self._supplier_var.get().strip()
            if shop and not self._stall_var.get().strip():
                st = self._sup_shop_stalls.get(shop, "")
                if st:
                    self._stall_var.set(st)
        else:
            stall = self._stall_var.get().strip()
            if stall and not self._supplier_var.get().strip():
                sh = self._sup_stall_shops.get(stall, "")
                if sh:
                    self._supplier_var.set(sh)

    # ── Submit ────────────────────────────────────────────────────────

    def _submit(self) -> None:
        """Validate inputs and invoke ``on_submit`` with a payload dict."""
        # Mandatory: image
        if not self._photo_bytes:
            self._error_lbl.config(
                text="Product image is required."
                if self._lang == "en"
                else "\u4ea7\u54c1\u56fe\u7247\u4e3a\u5fc5\u586b\u9879\u3002"
            )
            return

        # Mandatory: at least one component
        has_case  = self._case_var.get()
        has_grip  = self._grip_var.get()
        has_charm = self._charm_var.get()
        if not (has_case or has_grip or has_charm):
            self._error_lbl.config(
                text="Select at least one of Case / Grip / Charm."
                if self._lang == "en"
                else "\u8bf7\u81f3\u5c11\u52fe\u9009 \u624b\u673a\u58f3 / \u652f\u67b6 / \u6302\u4ef6 \u4e4b\u4e00\u3002"
            )
            return

        # Qty validation
        try:
            qty = int(self._qty_var.get().strip() or "1")
            if qty < 1:
                raise ValueError
        except ValueError:
            self._error_lbl.config(
                text="Quantity must be a positive integer."
                if self._lang == "en"
                else "\u6570\u91cf\u5fc5\u987b\u4e3a\u6b63\u6574\u6570\u3002"
            )
            return

        # Build style string (contains "Case"/"Grip"/"Charm" tokens)
        style_parts: list[str] = []
        if has_case:  style_parts.append("Case")
        if has_grip:  style_parts.append("Grip")
        if has_charm: style_parts.append("Charm")
        style = " + ".join(style_parts)

        # Default title when blank
        title = self._title_var.get().strip()
        if not title:
            title = f"Manual order ({style})"

        # Synthesize a unique order number. Pattern MAN-<timestamp><rand>
        import time as _time, random as _random
        order_number = f"MAN{int(_time.time())}{_random.randint(10, 99)}"

        payload = {
            "order_number": order_number,
            "title": title,
            "qty": qty,
            "phone": self._phone_var.get().strip(),
            "style": style,
            "photo_bytes": self._photo_bytes,
            "supplier": self._supplier_var.get().strip(),
            "stall": self._stall_var.get().strip(),
            "charm_code": self._charm_code_var.get().strip() if has_charm else "",
            "charm_shop": self._charm_shop_var.get().strip() if has_charm else "",
            "buyer": self._buyer_var.get().strip(),
            "private_notes": self._notes_text.get("1.0", tk.END).strip(),
        }

        try:
            self._on_submit(payload)
        except Exception as e:
            self._error_lbl.config(text=f"Failed to add: {e}")
            return

        self._d.destroy()


class _ProductMapEditorDialog:
    """
    Modal dialog for editing Shop Name, Stall, Charm Shop, and Charm Code
    on Product Map rows — with photo preview, search filtering, and
    dropdown-style comboboxes populated from the Suppliers and Charm sheets.
    """

    _THUMB = 68
    _ROW_H = 80

    def __init__(
        self,
        parent: App,
        all_rows: list,
        row_photos: dict[int, bytes],
        *,
        supplier_shops: list[str],
        supplier_stalls: list[str],
        charm_shop_names: list[str],
        charm_entries: dict,
    ) -> None:
        self._parent         = parent
        self._all_rows       = all_rows
        self._row_photos     = row_photos
        self._lang           = parent._lang
        self._sup_shops      = supplier_shops
        self._sup_stalls     = supplier_stalls
        self._charm_shops    = charm_shop_names
        self._charm_entries  = charm_entries
        self._charm_codes    = list(charm_entries.keys())
        self._tk_img_refs: list[object] = []
        self._row_by_iid: dict[str, object] = {}
        self._selected_row: object | None = None
        self._preview_photo_ref: list[object] = []
        self._pending_photo_bytes: bytes | None = None

        pil_ok = Image is not None and ImageTk is not None
        self._pil_ok = pil_ok

        d = tk.Toplevel(parent)
        self._d = d
        d.title(self._t("edit_title"))
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.geometry("1380x820")
        d.minsize(1020, 600)

        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(3, weight=1)

        # ── Row 0: hero ──
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero, text=self._t("edit_heading"),
            font=("Segoe UI", 12, "bold"), fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(10, 2))
        tk.Label(
            hero, text=self._t("edit_intro"),
            wraplength=1300, justify=tk.LEFT,
            font=("Segoe UI", 9), fg="#e2e8f0", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(0, 8))

        # ── Row 1: search ──
        sf = tk.Frame(d, bg=COLORS["app"])
        sf.grid(row=1, column=0, sticky="ew", padx=14, pady=(8, 4))
        tk.Label(sf, text=self._t("edit_search_label"),
                 font=("Segoe UI", 10, "bold"), bg=COLORS["app"], fg=COLORS["text"],
                 ).pack(side=tk.LEFT)
        self._filt = tk.StringVar()
        ef = tk.Entry(sf, textvariable=self._filt, font=("Segoe UI", 10))
        ef.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 8))
        tk.Label(sf, text=self._t("edit_search_tip"),
                 font=("Segoe UI", 8), bg=COLORS["app"], fg=COLORS["muted"],
                 ).pack(side=tk.LEFT)

        # ── Row 2: hover tip ──
        tk.Label(
            d, text=self._t("edit_hover_tip"),
            font=("Segoe UI", 8), bg=COLORS["app"], fg=COLORS["muted"],
            wraplength=1300, justify=tk.LEFT,
        ).grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 2))

        # ── Row 3: body (tree + edit panel) ──
        body = tk.Frame(d, bg=COLORS["app"])
        body.grid(row=3, column=0, sticky="nsew", padx=14, pady=(4, 0))
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=1, minsize=360)
        body.grid_rowconfigure(0, weight=1)

        # ── Left: Treeview ──
        tf = tk.Frame(body, bg=COLORS["app"])
        tf.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        sty = ttk.Style(d)
        row_h = self._ROW_H if pil_ok else 24
        try:
            sty.configure("Edit.Treeview", rowheight=row_h, font=("Segoe UI", 9))
            sty.configure("Edit.Treeview.Heading", font=("Segoe UI", 9, "bold"))
            sty.map("Edit.Treeview", background=[("selected", "#dbeafe")])
        except tk.TclError:
            pass

        cols = ("row", "shop", "stall", "charm_shop", "charm_code", "title")
        tree = ttk.Treeview(
            tf, columns=cols, show="tree headings",
            selectmode="browse", style="Edit.Treeview",
        )
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self._tree = tree

        tree.heading("#0", text=self._t("edit_col_photo"))
        tree.column("#0", width=88 if pil_ok else 20, stretch=False, anchor=tk.CENTER)
        tree.heading("row",         text=self._t("edit_col_row"))
        tree.column ("row",         width=40, stretch=False, anchor=tk.CENTER)
        tree.heading("shop",        text=self._t("edit_col_shop"))
        tree.column ("shop",        width=100, stretch=False)
        tree.heading("stall",       text=self._t("edit_col_stall"))
        tree.column ("stall",       width=60, stretch=False, anchor=tk.CENTER)
        tree.heading("charm_shop",  text=self._t("edit_col_charm_shop"))
        tree.column ("charm_shop",  width=100, stretch=False)
        tree.heading("charm_code",  text=self._t("edit_col_charm_code"))
        tree.column ("charm_code",  width=100, stretch=False, anchor=tk.CENTER)
        tree.heading("title",       text=self._t("edit_col_title"))
        tree.column ("title",       width=300, stretch=True)

        # ── Right: edit panel ──
        panel = tk.Frame(body, bg=COLORS["card"],
                         highlightthickness=1, highlightbackground=COLORS["border"])
        panel.grid(row=0, column=1, sticky="nsew")
        panel.grid_columnconfigure(0, weight=1)

        tk.Label(
            panel, text=self._t("edit_preview_title"),
            font=("Segoe UI", 10, "bold"), bg=COLORS["card"], fg=COLORS["text"],
        ).grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 2))

        # Photo preview + upload controls (all in one frame at row 1)
        photo_area = tk.Frame(panel, bg=COLORS["card"])
        photo_area.grid(row=1, column=0, sticky="ew", padx=12, pady=(4, 2))
        photo_area.grid_columnconfigure(0, weight=1)

        self._preview_img_label = tk.Label(
            photo_area, bg=COLORS["card"], bd=0,
            highlightthickness=0, relief=tk.FLAT,
        )
        self._preview_img_label.grid(row=0, column=0, pady=(0, 4))

        self._upload_photo_btn = tk.Button(
            photo_area,
            text=self._t("edit_btn_upload_photo"),
            command=self._upload_photo,
            bg=COLORS["accent_soft"], fg=COLORS["accent"],
            activebackground="#bfdbfe", activeforeground=COLORS["accent"],
            font=("Segoe UI", 8, "bold"),
            padx=10, pady=4,
            cursor="hand2",
            relief=tk.FLAT, borderwidth=0, highlightthickness=0,
            state=tk.DISABLED,
        )
        self._upload_photo_btn.grid(row=1, column=0, sticky="ew", pady=(0, 2))

        self._upload_status_lbl = tk.Label(
            photo_area, text="",
            font=("Segoe UI", 8, "bold"), bg=COLORS["card"], fg="#047857",
            wraplength=310, justify=tk.CENTER,
        )
        self._upload_status_lbl.grid(row=2, column=0, sticky="ew")

        # Title label
        self._preview_title = tk.Label(
            panel, text="", wraplength=320, justify=tk.LEFT,
            font=("Segoe UI", 9), bg=COLORS["card"], fg=COLORS["text"],
        )
        self._preview_title.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 8))

        # Form fields
        form = tk.Frame(panel, bg=COLORS["card"])
        form.grid(row=3, column=0, sticky="ew", padx=12, pady=(0, 4))
        form.grid_columnconfigure(1, weight=1)

        self._fields: dict[str, ttk.Combobox | tk.Entry] = {}
        combo_defs = [
            ("shop_name",  "edit_lbl_shop",       self._sup_shops),
            ("stall",      "edit_lbl_stall",      self._sup_stalls),
            ("charm_shop", "edit_lbl_charm_shop",  self._charm_shops),
        ]
        for i, (key, lbl_key, values) in enumerate(combo_defs):
            tk.Label(
                form, text=self._t(lbl_key),
                font=("Segoe UI", 9, "bold"), bg=COLORS["card"], fg=COLORS["text"],
                anchor=tk.W,
            ).grid(row=i, column=0, sticky="w", padx=(0, 8), pady=3)
            cb = ttk.Combobox(form, values=values, font=("Segoe UI", 9), width=24)
            cb.grid(row=i, column=1, sticky="ew", pady=3)
            self._fields[key] = cb

        # Charm Code — entry + visual browse button
        cc_row = len(combo_defs)
        tk.Label(
            form, text=self._t("edit_lbl_charm_code"),
            font=("Segoe UI", 9, "bold"), bg=COLORS["card"], fg=COLORS["text"],
            anchor=tk.W,
        ).grid(row=cc_row, column=0, sticky="w", padx=(0, 8), pady=3)

        cc_frame = tk.Frame(form, bg=COLORS["card"])
        cc_frame.grid(row=cc_row, column=1, sticky="ew", pady=3)
        cc_frame.grid_columnconfigure(0, weight=1)

        cc_entry = tk.Entry(cc_frame, font=("Segoe UI", 9))
        cc_entry.grid(row=0, column=0, sticky="ew")
        self._fields["charm_code"] = cc_entry

        self._cc_pick_btn = ttk.Button(
            cc_frame, text=self._t("edit_charm_pick"),
            command=self._open_charm_picker,
            style="Tool.TButton", width=5,
        )
        self._cc_pick_btn.grid(row=0, column=1, padx=(4, 0))

        # Charm preview thumbnail (shows next to charm code entry after selection)
        self._charm_preview_frame = tk.Frame(form, bg=COLORS["card"])
        self._charm_preview_frame.grid(
            row=cc_row + 1, column=0, columnspan=2, sticky="ew", pady=(2, 0)
        )
        self._charm_preview_img = tk.Label(
            self._charm_preview_frame, bg=COLORS["card"], bd=0,
            highlightthickness=0, relief=tk.FLAT,
        )
        self._charm_preview_img.pack(side=tk.LEFT, padx=(0, 8))
        self._charm_preview_text = tk.Label(
            self._charm_preview_frame, text="",
            font=("Segoe UI", 8), fg=COLORS["muted"], bg=COLORS["card"],
            wraplength=200, justify=tk.LEFT, anchor=tk.W,
        )
        self._charm_preview_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._charm_preview_photo_ref: list[object] = []

        cc_entry.bind("<KeyRelease>", lambda _e: self._update_charm_preview())
        cc_entry.bind("<FocusOut>", lambda _e: self._update_charm_preview())

        # Save button
        self._save_btn = tk.Button(
            panel, text=self._t("edit_btn_save"),
            command=self._save_current,
            bg=COLORS["run"], fg="#ffffff",
            activebackground=COLORS["run_hover"],
            activeforeground="#ffffff",
            font=("Segoe UI", 10, "bold"),
            padx=18, pady=6,
            cursor="hand2",
            relief=tk.FLAT, borderwidth=0, highlightthickness=0,
        )
        self._save_btn.grid(row=4, column=0, padx=12, pady=(10, 4), sticky="ew")

        placeholder = tk.Label(
            panel, text=self._t("edit_preview_placeholder"),
            font=("Segoe UI", 8), fg=COLORS["muted"], bg=COLORS["card"],
            wraplength=320, justify=tk.LEFT,
        )
        placeholder.grid(row=5, column=0, sticky="ew", padx=12, pady=(0, 6))
        self._placeholder_lbl = placeholder

        # ── Danger zone ──
        ttk.Separator(panel, orient=tk.HORIZONTAL).grid(
            row=6, column=0, sticky="ew", padx=12, pady=(10, 0)
        )
        dz_hdr = tk.Frame(panel, bg=COLORS["card"])
        dz_hdr.grid(row=7, column=0, sticky="ew", padx=12, pady=(6, 2))
        tk.Label(
            dz_hdr,
            text=self._t("edit_danger_zone"),
            font=("Segoe UI", 8, "bold"),
            fg="#b91c1c",
            bg=COLORS["card"],
            anchor=tk.W,
        ).pack(side=tk.LEFT)
        tk.Label(
            panel,
            text=self._t("edit_danger_note"),
            font=("Segoe UI", 8),
            fg=COLORS["muted"],
            bg=COLORS["card"],
            wraplength=310,
            justify=tk.LEFT,
            anchor=tk.W,
        ).grid(row=8, column=0, sticky="ew", padx=12, pady=(0, 4))
        self._disc_btn = tk.Button(
            panel,
            text=self._t("edit_btn_discontinue"),
            command=self._mark_discontinued,
            bg="#dc2626",
            fg="#ffffff",
            activebackground="#b91c1c",
            activeforeground="#ffffff",
            disabledforeground="#fca5a5",
            font=("Segoe UI", 9, "bold"),
            padx=14,
            pady=5,
            cursor="hand2",
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=0,
            state=tk.DISABLED,
        )
        self._disc_btn.grid(row=9, column=0, padx=12, pady=(2, 14), sticky="ew")

        # ── Row 4: footer ──
        foot = tk.Frame(d, bg=COLORS["app"])
        foot.grid(row=4, column=0, sticky="ew", padx=14, pady=(8, 12))
        ttk.Button(
            foot, text=self._t("edit_btn_close"),
            command=d.destroy, style="Tool.TButton",
        ).pack(side=tk.RIGHT)

        # ── Populate + bindings ──
        self._refill_tree()
        self._filt.trace_add("write", lambda *_: self._refill_tree())
        tree.bind("<<TreeviewSelect>>", self._on_select)

        # Hover zoom
        self._hover_tip_win: list[tk.Toplevel | None] = [None]
        self._hover_active_iid: list[str | None] = [None]
        self._hover_after_id: list[str | None] = [None]
        self._hover_photo: list[object] = []
        tree.bind("<Motion>", self._on_tree_motion)
        tree.bind("<Leave>", self._on_tree_leave)

    # ── helpers ──

    def _t(self, key: str, **kw: object) -> str:
        s = CHROME.get(self._lang, CHROME["en"]).get(key, key)
        return s.format(**kw) if kw else s

    def _thumb(self, raw: bytes | None) -> object | None:
        if not self._pil_ok or not raw or Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            im.thumbnail((self._THUMB, self._THUMB), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._tk_img_refs.append(ph)
            return ph
        except Exception:
            return None

    # ── tree ──

    def _refill_tree(self) -> None:
        self._tree.delete(*self._tree.get_children())
        self._row_by_iid.clear()
        self._tk_img_refs.clear()
        q = self._filt.get().strip().lower()
        for r in self._all_rows:
            if q:
                blob = " ".join((r.title, r.shop_name, r.stall)).lower()
                if q not in blob:
                    continue
            iid = f"r{r.row_num}"
            self._row_by_iid[iid] = r
            t_short = r.title if len(r.title) <= 60 else r.title[:57] + "…"
            raw = self._row_photos.get(r.row_num)
            thumb = self._thumb(raw) if self._pil_ok else None
            kw: dict = dict(
                values=(
                    r.row_num,
                    r.shop_name or "—",
                    r.stall or "—",
                    r.charm_shop or "—",
                    r.charm_code or "—",
                    t_short,
                ),
            )
            if thumb is not None:
                kw["image"] = thumb
                kw["text"] = ""
            else:
                kw["text"] = "—"
            self._tree.insert("", tk.END, iid=iid, **kw)

    def _on_select(self, _evt: object = None) -> None:
        sel = self._tree.selection()
        if not sel:
            self._set_edit_panel(None)
            return
        self._set_edit_panel(self._row_by_iid.get(sel[0]))

    def _set_edit_panel(self, r: object | None) -> None:
        self._selected_row = r
        self._pending_photo_bytes = None
        self._upload_status_lbl.config(text="")
        if r is None:
            self._preview_title.config(text="")
            self._preview_img_label.config(image="")
            self._preview_photo_ref.clear()
            for w in self._fields.values():
                if isinstance(w, ttk.Combobox):
                    w.set("")
                    w.config(state="disabled")
                else:
                    w.delete(0, tk.END)
                    w.config(state="disabled")
            self._cc_pick_btn.config(state=tk.DISABLED)
            self._upload_photo_btn.config(state=tk.DISABLED)
            self._save_btn.config(state=tk.DISABLED)
            self._disc_btn.config(state=tk.DISABLED)
            self._clear_charm_preview()
            self._placeholder_lbl.config(text=self._t("edit_preview_placeholder"))
            return

        self._placeholder_lbl.config(text="")
        self._preview_title.config(text=r.title)

        raw = self._row_photos.get(r.row_num)
        self._preview_photo_ref.clear()
        if raw and self._pil_ok and Image is not None and ImageTk is not None:
            try:
                im = Image.open(BytesIO(raw))
                if im.mode not in ("RGB", "RGBA"):
                    im = im.convert("RGBA" if "A" in im.mode else "RGB")
                if im.mode == "RGBA":
                    bg = Image.new("RGB", im.size, (255, 255, 255))
                    bg.paste(im, mask=im.split()[3])
                    im = bg
                im.thumbnail((240, 240), Image.Resampling.LANCZOS)
                ph = ImageTk.PhotoImage(im)
                self._preview_photo_ref.append(ph)
                self._preview_img_label.config(image=ph)
            except Exception:
                self._preview_img_label.config(image="")
        else:
            self._preview_img_label.config(image="")

        for w in self._fields.values():
            if isinstance(w, ttk.Combobox):
                w.config(state="normal")
            else:
                w.config(state="normal")
        self._cc_pick_btn.config(state="normal")
        self._upload_photo_btn.config(state=tk.NORMAL)
        self._fields["shop_name"].set(r.shop_name)
        self._fields["stall"].set(r.stall)
        self._fields["charm_shop"].set(r.charm_shop)
        cc_w = self._fields["charm_code"]
        cc_w.delete(0, tk.END)
        cc_w.insert(0, r.charm_code)
        self._update_charm_preview()
        self._save_btn.config(state=tk.NORMAL)
        self._disc_btn.config(state=tk.NORMAL)

    # ── charm code preview + picker ──

    def _clear_charm_preview(self) -> None:
        self._charm_preview_photo_ref.clear()
        self._charm_preview_img.config(image="")
        self._charm_preview_text.config(text="")

    def _update_charm_preview(self) -> None:
        code = self._fields["charm_code"].get().strip()
        entry = self._charm_entries.get(code)
        if not entry:
            self._clear_charm_preview()
            if code and code not in self._charm_entries:
                self._charm_preview_text.config(
                    text=self._t("edit_charm_unknown"),
                    fg="#b91c1c",
                )
            return
        # Show thumbnail + label
        self._charm_preview_photo_ref.clear()
        if entry.photo_bytes and self._pil_ok and Image is not None and ImageTk is not None:
            try:
                im = Image.open(BytesIO(entry.photo_bytes))
                if im.mode not in ("RGB", "RGBA"):
                    im = im.convert("RGBA" if "A" in im.mode else "RGB")
                if im.mode == "RGBA":
                    bg = Image.new("RGB", im.size, (255, 255, 255))
                    bg.paste(im, mask=im.split()[3])
                    im = bg
                im.thumbnail((52, 52), Image.Resampling.LANCZOS)
                ph = ImageTk.PhotoImage(im)
                self._charm_preview_photo_ref.append(ph)
                self._charm_preview_img.config(image=ph)
            except Exception:
                self._charm_preview_img.config(image="")
        else:
            self._charm_preview_img.config(image="")
        parts = [code]
        if entry.sku:
            parts.append(entry.sku)
        if entry.default_charm_shop:
            parts.append(f"({entry.default_charm_shop})")
        self._charm_preview_text.config(text="  —  ".join(parts), fg=COLORS["text"])

    def _open_charm_picker(self) -> None:
        if not self._charm_entries:
            return
        _CharmPickerPopup(
            self._d,
            self._charm_entries,
            lang=self._lang,
            callback=self._on_charm_picked,
        )

    def _on_charm_picked(self, code: str) -> None:
        w = self._fields["charm_code"]
        w.config(state="normal")
        w.delete(0, tk.END)
        w.insert(0, code)
        self._update_charm_preview()

    # ── photo upload ──

    def _upload_photo(self) -> None:
        """Open a file dialog to choose a new product photo and stage it for saving."""
        path = filedialog.askopenfilename(
            parent=self._d,
            title=self._t("edit_upload_photo_title"),
            filetypes=[
                ("Image files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return

        if Image is None or ImageTk is None:
            messagebox.showerror(
                self._t("file_open_fail_title"),
                "Pillow is required for photo upload.  pip install Pillow",
                parent=self._d,
            )
            return

        try:
            im = Image.open(path)
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            # Encode as JPEG for compact storage in the workbook
            buf = BytesIO()
            im.save(buf, format="JPEG", quality=90)
            self._pending_photo_bytes = buf.getvalue()
        except Exception as exc:
            messagebox.showerror(
                self._t("file_open_fail_title"),
                self._t("edit_photo_save_fail", err=str(exc)),
                parent=self._d,
            )
            return

        # Update the preview in the panel with the new image
        try:
            im_prev = Image.open(BytesIO(self._pending_photo_bytes))
            im_prev.thumbnail((240, 240), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im_prev)
            self._preview_photo_ref.clear()
            self._preview_photo_ref.append(ph)
            self._preview_img_label.config(image=ph)
        except Exception:
            pass

        self._upload_status_lbl.config(text=self._t("edit_upload_photo_pending"))

    # ── save ──

    def _save_current(self) -> None:
        r = self._selected_row
        if r is None:
            messagebox.showinfo(
                self._t("edit_title"), self._t("edit_no_selection"), parent=self._d,
            )
            return

        new_shop       = self._fields["shop_name"].get().strip()
        new_stall      = self._fields["stall"].get().strip()
        new_charm_shop = self._fields["charm_shop"].get().strip()
        new_charm_code = self._fields["charm_code"].get().strip()
        pending_photo  = self._pending_photo_bytes

        prev_cursor = self._d.cget("cursor")
        self._d.config(cursor="watch")
        self._save_btn.config(state=tk.DISABLED)
        self._disc_btn.config(state=tk.DISABLED)
        self._upload_photo_btn.config(state=tk.DISABLED)
        self._d.update_idletasks()

        try:
            update_product_map_cells(
                FILE_SUPPLIER_CATALOG,
                r.row_num,
                shop_name=new_shop,
                stall=new_stall,
                charm_shop=new_charm_shop,
                charm_code=new_charm_code,
            )
            if pending_photo and update_product_map_photo is not None:
                update_product_map_photo(FILE_SUPPLIER_CATALOG, r.row_num, pending_photo)
        except Exception as e:
            messagebox.showerror(
                self._t("file_open_fail_title"), str(e), parent=self._d,
            )
            return
        finally:
            self._d.config(cursor=prev_cursor)
            self._save_btn.config(state=tk.NORMAL)
            self._disc_btn.config(state=tk.NORMAL)
            self._upload_photo_btn.config(state=tk.NORMAL)

        # Clear pending photo state
        self._pending_photo_bytes = None
        self._upload_status_lbl.config(text="")

        # Reload photos + rows so the treeview thumbnail reflects the new photo
        if pending_photo:
            try:
                if extract_photos_from_xlsx is not None:
                    self._row_photos = extract_photos_from_xlsx(
                        FILE_SUPPLIER_CATALOG, sheet_name=CATALOG_SHEET, photo_col_idx=0,
                    )
            except Exception:
                pass
        try:
            self._all_rows = list_product_map_rows_for_picker(FILE_SUPPLIER_CATALOG)
        except Exception:
            pass
        cur_iid = f"r{r.row_num}"
        self._refill_tree()
        if self._tree.exists(cur_iid):
            self._tree.selection_set(cur_iid)
            self._tree.see(cur_iid)
            self._set_edit_panel(self._row_by_iid.get(cur_iid))

        short = r.title if len(r.title) <= 60 else r.title[:57] + "…"
        if pending_photo:
            messagebox.showinfo(
                self._t("edit_title"),
                self._t("edit_photo_saved", title=short),
                parent=self._d,
            )
        else:
            messagebox.showinfo(
                self._t("edit_title"),
                self._t("edit_saved", title=short),
                parent=self._d,
            )

    # ── mark discontinued ──

    def _mark_discontinued(self) -> None:
        r = self._selected_row
        if r is None:
            messagebox.showinfo(
                self._t("edit_title"), self._t("edit_no_selection"), parent=self._d,
            )
            return
        if mark_product_map_discontinued_by_row is None:
            messagebox.showerror(
                self._t("msg_missing_title"), self._t("edit_no_import"), parent=self._d,
            )
            return

        short = r.title if len(r.title) <= 80 else r.title[:77] + "…"
        if not messagebox.askyesno(
            self._t("edit_discontinue_confirm_title"),
            self._t("edit_discontinue_confirm", title=short),
            icon="warning",
            parent=self._d,
        ):
            return

        prev_cursor = self._d.cget("cursor")
        self._d.config(cursor="watch")
        self._save_btn.config(state=tk.DISABLED)
        self._disc_btn.config(state=tk.DISABLED)
        self._d.update_idletasks()

        try:
            mark_product_map_discontinued_by_row(FILE_SUPPLIER_CATALOG, r.row_num)
        except Exception as e:
            messagebox.showerror(
                self._t("file_open_fail_title"), str(e), parent=self._d,
            )
            return
        finally:
            self._d.config(cursor=prev_cursor)

        # Reload the product list — discontinued row is now gone from Product Map
        try:
            self._all_rows = list_product_map_rows_for_picker(FILE_SUPPLIER_CATALOG)
            self._row_photos = {}
            if extract_photos_from_xlsx is not None:
                self._row_photos = extract_photos_from_xlsx(
                    FILE_SUPPLIER_CATALOG, sheet_name=CATALOG_SHEET, photo_col_idx=0,
                )
        except Exception:
            pass

        self._set_edit_panel(None)
        self._refill_tree()

        messagebox.showinfo(
            self._t("edit_discontinue_confirm_title"),
            self._t("edit_discontinue_done"),
            parent=self._d,
        )

    # ── hover zoom (same pattern as discontinue dialog) ──

    def _hide_hover(self) -> None:
        aid = self._hover_after_id[0]
        if aid is not None:
            try:
                self._d.after_cancel(aid)
            except (tk.TclError, ValueError):
                pass
            self._hover_after_id[0] = None
        self._hover_active_iid[0] = None
        tip = self._hover_tip_win[0]
        if tip is not None:
            try:
                tip.destroy()
            except tk.TclError:
                pass
            self._hover_tip_win[0] = None

    def _on_tree_motion(self, _e: tk.Event | None = None) -> None:
        if not self._pil_ok:
            return
        tree = self._tree
        px, py = tree.winfo_pointerx(), tree.winfo_pointery()
        lx = px - tree.winfo_rootx()
        ly = py - tree.winfo_rooty()
        if lx < 0 or ly < 0 or lx >= tree.winfo_width() or ly >= tree.winfo_height():
            self._hide_hover()
            return
        if tree.identify_region(lx, ly) != "tree":
            self._hide_hover()
            return
        iid = tree.identify_row(ly)
        if not iid:
            self._hide_hover()
            return
        r = self._row_by_iid.get(iid)
        if r is None or not self._row_photos.get(r.row_num):
            self._hide_hover()
            return
        if iid == self._hover_active_iid[0]:
            return
        self._hide_hover()
        self._hover_active_iid[0] = iid
        self._hover_after_id[0] = self._d.after(
            350, lambda i=iid: self._show_hover(i)
        )

    def _show_hover(self, iid: str) -> None:
        self._hover_after_id[0] = None
        if Image is None or ImageTk is None:
            return
        tree = self._tree
        px, py = tree.winfo_pointerx(), tree.winfo_pointery()
        lx, ly = px - tree.winfo_rootx(), py - tree.winfo_rooty()
        if lx < 0 or ly < 0 or lx >= tree.winfo_width() or ly >= tree.winfo_height():
            return
        if tree.identify_row(ly) != iid:
            return
        r = self._row_by_iid.get(iid)
        if r is None:
            return
        raw = self._row_photos.get(r.row_num)
        if not raw:
            return
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            w, h = im.size
            max_dim = 320
            if w > max_dim or h > max_dim:
                ratio = min(max_dim / w, max_dim / h)
                im = im.resize((int(w * ratio), int(h * ratio)), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._hover_photo.clear()
            self._hover_photo.append(ph)
        except Exception:
            return
        tip = tk.Toplevel(self._d)
        tip.overrideredirect(True)
        try:
            tip.attributes("-topmost", True)
        except tk.TclError:
            pass
        tip.configure(bg="#ffffff", bd=0, highlightthickness=0)
        tk.Label(tip, image=ph, bg="#ffffff", bd=0, highlightthickness=0, relief=tk.FLAT).pack()
        tip.update_idletasks()
        tw, th = tip.winfo_reqwidth(), tip.winfo_reqheight()
        sw, sh = tip.winfo_screenwidth(), tip.winfo_screenheight()
        x = min(max(12, px + 20), sw - tw - 12)
        y = min(max(12, py + 20), sh - th - 12)
        tip.geometry(f"+{x}+{y}")
        self._hover_tip_win[0] = tip

    def _on_tree_leave(self, _e: tk.Event | None = None) -> None:
        self._hide_hover()


class _CharmPickerPopup:
    """
    Visual charm selector — shows every Charm Library entry with its photo,
    code, and SKU in a scrollable Treeview.  Click a row (or double-click) to
    select that charm and close the popup.  The first row is "(none)" to let
    the user clear the charm code.
    """

    _THUMB = 64
    _ROW_H = 72

    def __init__(
        self,
        parent: tk.Toplevel | tk.Tk,
        charm_entries: dict,
        *,
        lang: str = "en",
        callback: object = None,
    ) -> None:
        self._entries  = charm_entries
        self._lang     = lang
        self._callback = callback
        self._tk_img_refs: list[object] = []

        pil_ok = Image is not None and ImageTk is not None
        self._pil_ok = pil_ok

        d = tk.Toplevel(parent)
        self._d = d
        d.title(self._t("charm_picker_title"))
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.geometry("680x620")
        d.minsize(520, 400)

        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(3, weight=1)

        # Hero
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero, text=self._t("charm_picker_heading"),
            font=("Segoe UI", 11, "bold"), fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(8, 6))

        # Search
        sf = tk.Frame(d, bg=COLORS["app"])
        sf.grid(row=1, column=0, sticky="ew", padx=14, pady=(6, 4))
        tk.Label(sf, text=self._t("charm_picker_search"),
                 font=("Segoe UI", 9, "bold"), bg=COLORS["app"], fg=COLORS["text"],
                 ).pack(side=tk.LEFT)
        self._filt = tk.StringVar()
        ef = tk.Entry(sf, textvariable=self._filt, font=("Segoe UI", 9))
        ef.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 8))
        tk.Label(sf, text=self._t("charm_picker_search_tip"),
                 font=("Segoe UI", 8), bg=COLORS["app"], fg=COLORS["muted"],
                 ).pack(side=tk.LEFT)

        # Hover tip (only when PIL is available so the hover actually works)
        if pil_ok:
            tk.Label(
                d, text=self._t("charm_picker_hover_tip"),
                font=("Segoe UI", 8), bg=COLORS["app"], fg=COLORS["muted"],
                wraplength=640, justify=tk.LEFT,
            ).grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 2))

        # Treeview
        tf = tk.Frame(d, bg=COLORS["app"])
        tf.grid(row=3, column=0, sticky="nsew", padx=14, pady=(2, 10))
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        sty = ttk.Style(d)
        row_h = self._ROW_H if pil_ok else 24
        try:
            sty.configure("CharmPick.Treeview", rowheight=row_h, font=("Segoe UI", 9))
            sty.configure("CharmPick.Treeview.Heading", font=("Segoe UI", 9, "bold"))
            sty.map("CharmPick.Treeview", background=[("selected", COLORS["accent_soft"])])
        except tk.TclError:
            pass

        cols = ("code", "sku", "shop")
        tree = ttk.Treeview(
            tf, columns=cols, show="tree headings",
            selectmode="browse", style="CharmPick.Treeview",
        )
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self._tree = tree

        tree.heading("#0",   text=self._t("charm_picker_col_photo"))
        tree.column ("#0",   width=80 if pil_ok else 20, stretch=False, anchor=tk.CENTER)
        tree.heading("code", text=self._t("charm_picker_col_code"))
        tree.column ("code", width=120, stretch=False, anchor=tk.CENTER)
        tree.heading("sku",  text=self._t("charm_picker_col_sku"))
        tree.column ("sku",  width=260, stretch=True)
        tree.heading("shop", text=self._t("charm_picker_col_shop"))
        tree.column ("shop", width=130, stretch=False)

        # Hover zoom state — must be initialised before _refill() which calls _hide_charm_hover()
        self._hover_after_id: list[object | None] = [None]
        self._hover_tip_win: list[tk.Toplevel | None] = [None]
        self._hover_photo: list[object] = []
        self._hover_active_iid: list[str | None] = [None]

        self._refill()
        self._filt.trace_add("write", lambda *_: self._refill())
        tree.bind("<Double-1>", lambda _e: self._pick())
        tree.bind("<Return>",   lambda _e: self._pick())

        if pil_ok:
            tree.bind("<Motion>", self._on_charm_hover_motion)
            tree.bind("<Leave>",  self._on_charm_hover_leave)

    def _t(self, key: str) -> str:
        return CHROME.get(self._lang, CHROME["en"]).get(key, key)

    def _thumb(self, raw: bytes | None) -> object | None:
        if not self._pil_ok or not raw or Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            im.thumbnail((self._THUMB, self._THUMB), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._tk_img_refs.append(ph)
            return ph
        except Exception:
            return None

    def _refill(self) -> None:
        self._hide_charm_hover()
        self._tree.delete(*self._tree.get_children())
        self._tk_img_refs.clear()
        q = self._filt.get().strip().lower()

        # "(none)" row — lets the user clear the charm code
        if not q:
            self._tree.insert(
                "", tk.END, iid="__none__",
                text="—", values=("", self._t("charm_picker_none"), ""),
            )

        for code, entry in self._entries.items():
            if q:
                blob = f"{code} {entry.sku} {entry.default_charm_shop}".lower()
                if q not in blob:
                    continue
            thumb = self._thumb(entry.photo_bytes)
            kw: dict = dict(
                values=(code, entry.sku or "", entry.default_charm_shop or ""),
            )
            if thumb is not None:
                kw["image"] = thumb
                kw["text"]  = ""
            else:
                kw["text"] = "—"
            self._tree.insert("", tk.END, iid=f"ch_{code}", **kw)

    def _pick(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        if iid == "__none__":
            code = ""
        else:
            code = iid.removeprefix("ch_")
        self._hide_charm_hover()
        if self._callback and callable(self._callback):
            self._callback(code)
        self._d.destroy()

    # ── hover zoom ──

    def _hide_charm_hover(self) -> None:
        aid = self._hover_after_id[0]
        if aid is not None:
            try:
                self._d.after_cancel(aid)
            except (tk.TclError, ValueError):
                pass
            self._hover_after_id[0] = None
        self._hover_active_iid[0] = None
        tip = self._hover_tip_win[0]
        if tip is not None:
            try:
                tip.destroy()
            except tk.TclError:
                pass
            self._hover_tip_win[0] = None

    def _on_charm_hover_motion(self, _e: tk.Event | None = None) -> None:
        if not self._pil_ok:
            return
        tree = self._tree
        px, py = tree.winfo_pointerx(), tree.winfo_pointery()
        lx = px - tree.winfo_rootx()
        ly = py - tree.winfo_rooty()
        if lx < 0 or ly < 0 or lx >= tree.winfo_width() or ly >= tree.winfo_height():
            self._hide_charm_hover()
            return
        # Only activate when the cursor is over the Photo (tree icon) column.
        if tree.identify_column(lx) != "#0":
            self._hide_charm_hover()
            return
        iid = tree.identify_row(ly)
        if not iid or iid == "__none__":
            self._hide_charm_hover()
            return
        code = iid.removeprefix("ch_")
        entry = self._entries.get(code)
        if entry is None or not entry.photo_bytes:
            self._hide_charm_hover()
            return
        if iid == self._hover_active_iid[0]:
            return
        self._hide_charm_hover()
        self._hover_active_iid[0] = iid
        self._hover_after_id[0] = self._d.after(
            220, lambda i=iid: self._show_charm_hover(i)
        )

    def _show_charm_hover(self, iid: str) -> None:
        self._hover_after_id[0] = None
        if Image is None or ImageTk is None:
            return
        tree = self._tree
        px, py = tree.winfo_pointerx(), tree.winfo_pointery()
        lx, ly = px - tree.winfo_rootx(), py - tree.winfo_rooty()
        if lx < 0 or ly < 0 or lx >= tree.winfo_width() or ly >= tree.winfo_height():
            return
        if tree.identify_row(ly) != iid:
            return
        code = iid.removeprefix("ch_")
        entry = self._entries.get(code)
        if entry is None:
            return
        raw = entry.photo_bytes
        if not raw:
            return
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            w, h = im.size
            max_dim = 400
            if w > max_dim or h > max_dim:
                ratio = min(max_dim / w, max_dim / h)
                im = im.resize((int(w * ratio), int(h * ratio)), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._hover_photo.clear()
            self._hover_photo.append(ph)
        except Exception:
            return
        tip = tk.Toplevel(self._d)
        tip.overrideredirect(True)
        try:
            tip.attributes("-topmost", True)
        except tk.TclError:
            pass
        tip.configure(bg="#ffffff", bd=2, highlightthickness=1,
                      highlightbackground="#cccccc")
        tk.Label(tip, image=ph, bg="#ffffff", bd=0,
                 highlightthickness=0, relief=tk.FLAT).pack(padx=2, pady=2)
        tip.update_idletasks()
        tw, th = tip.winfo_reqwidth(), tip.winfo_reqheight()
        sw, sh = tip.winfo_screenwidth(), tip.winfo_screenheight()
        x = min(max(12, px + 24), sw - tw - 12)
        y = min(max(12, py + 24), sh - th - 12)
        tip.geometry(f"+{x}+{y}")
        self._hover_tip_win[0] = tip

    def _on_charm_hover_leave(self, _e: tk.Event | None = None) -> None:
        self._hide_charm_hover()


class _CharmImportSkuDialog:
    """
    Modal dialog shown between photo staging and workbook import.

    Displays a thumbnail + editable SKU entry for each staged photo,
    pre-populated with an auto-suggestion derived from the filename.
    Confirmed SKUs are stored in ``self.result`` as ``{staged_stem: sku_text}``.
    ``self.confirmed`` is True when the user clicked Import, False if cancelled.
    """

    _THUMB = 72

    def __init__(
        self,
        parent: tk.Misc,
        staged_files: list,   # list[Path]
        lang: str,
    ) -> None:
        self.result: dict[str, str] = {}
        self.confirmed: bool = False
        self._lang = lang
        self._sku_vars: list[tuple] = []    # [(Path, tk.StringVar), ...]
        self._tk_img_refs: list[object] = []

        d = tk.Toplevel(parent)
        d.title(
            "Set SKU for Imported Charms" if lang == "en"
            else "\u8bbe\u7f6e\u5bfc\u5165\u6302\u4ef6\u7684SKU"
        )
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.minsize(600, 360)
        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(1, weight=1)

        # ── Hero ──────────────────────────────────────────────────────────
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero,
            text=(
                "Set SKU for Imported Charms" if lang == "en"
                else "\u8bbe\u7f6e\u5bfc\u5165\u6302\u4ef6\u7684SKU"
            ),
            font=("Segoe UI", 12, "bold"),
            fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(10, 2))
        tk.Label(
            hero,
            text=(
                "Review and edit the auto-generated SKU for each photo.  "
                "SKUs are derived from the filename — edit any you want to change.  "
                "Leave blank to fill later in Excel."
            ) if lang == "en" else (
                "\u68c0\u67e5\u5e76\u7f16\u8f91\u6bcf\u5f20\u7167\u7247\u7684\u81ea\u52a8SKU\u3002"
                "SKU\u6839\u636e\u6587\u4ef6\u540d\u751f\u6210\uff0c\u53ef\u81ea\u7531\u4fee\u6539\uff0c"
                "\u7559\u7a7a\u5219\u7a0d\u540e\u5728Excel\u4e2d\u586b\u5199\u3002"
            ),
            wraplength=640, justify=tk.LEFT,
            font=("Segoe UI", 9), fg="#e2e8f0", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(0, 8))

        # ── Scrollable body ────────────────────────────────────────────────
        body_wrap = tk.Frame(d, bg=COLORS["app"])
        body_wrap.grid(row=1, column=0, sticky="nsew")
        body_wrap.grid_rowconfigure(0, weight=1)
        body_wrap.grid_columnconfigure(0, weight=1)

        canvas = tk.Canvas(body_wrap, bg=COLORS["app"], highlightthickness=0)
        vsb = ttk.Scrollbar(body_wrap, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        inner = tk.Frame(canvas, bg=COLORS["app"])
        _win = canvas.create_window((0, 0), window=inner, anchor=tk.NW)

        def _sync(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all") or (0, 0, 0, 0))
            canvas.itemconfig(_win, width=canvas.winfo_width())

        inner.bind("<Configure>", _sync)
        canvas.bind("<Configure>", _sync)

        def _scroll(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _scroll)
        inner.bind("<MouseWheel>", _scroll)

        # Column header strip
        hdr = tk.Frame(inner, bg=COLORS["strip"])
        hdr.pack(fill=tk.X)
        tk.Label(
            hdr, text="Photo" if lang == "en" else "\u7167\u7247",
            font=("Segoe UI", 8, "bold"), bg=COLORS["strip"], width=9,
        ).pack(side=tk.LEFT, padx=(8, 0))
        tk.Label(
            hdr,
            text="Original filename" if lang == "en" else "\u539f\u59cb\u6587\u4ef6\u540d",
            font=("Segoe UI", 8, "bold"), bg=COLORS["strip"], anchor=tk.W,
        ).pack(side=tk.LEFT, padx=8, expand=True, fill=tk.X)
        tk.Label(
            hdr,
            text="SKU (Column C)" if lang == "en" else "SKU\uff08C\u5217\uff09",
            font=("Segoe UI", 8, "bold"), bg=COLORS["strip"], width=26, anchor=tk.W,
        ).pack(side=tk.RIGHT, padx=(0, 8))

        # One row per staged file
        for i, path in enumerate(staged_files):
            row_bg = "#ffffff" if i % 2 == 0 else "#f8f7ff"
            row_f = tk.Frame(
                inner, bg=row_bg,
                highlightthickness=1,
                highlightbackground=COLORS["border"],
            )
            row_f.pack(fill=tk.X, padx=4, pady=2)
            row_f.bind("<MouseWheel>", _scroll)

            # Thumbnail
            ph = self._make_thumb(path)
            if ph:
                self._tk_img_refs.append(ph)
                lbl = tk.Label(row_f, image=ph, bg=row_bg,
                               width=self._THUMB + 6, height=self._THUMB + 6)
            else:
                lbl = tk.Label(row_f, text="\U0001f48e",
                               font=("Segoe UI", 22), fg="#c4b5fd", bg=row_bg,
                               width=5, height=3)
            lbl.pack(side=tk.LEFT, padx=(6, 2), pady=4)
            lbl.bind("<MouseWheel>", _scroll)

            # Original name hint
            display = re.sub(r"^__incoming__[0-9a-fA-F]+_", "", path.stem)
            if len(display) > 34:
                display = display[:33] + "\u2026"
            name_lbl = tk.Label(
                row_f, text=display,
                font=("Segoe UI", 9), fg=COLORS["text"], bg=row_bg,
                anchor=tk.W, justify=tk.LEFT,
            )
            name_lbl.pack(side=tk.LEFT, padx=6, expand=True, fill=tk.X)
            name_lbl.bind("<MouseWheel>", _scroll)

            # Editable SKU entry (pre-filled with auto-suggestion)
            sku_var = tk.StringVar(value=_auto_sku_from_stem(path.stem))
            entry = ttk.Entry(
                row_f, textvariable=sku_var,
                width=28, font=("Segoe UI", 9),
            )
            entry.pack(side=tk.RIGHT, padx=(4, 8), pady=6)
            entry.bind("<MouseWheel>", _scroll)

            self._sku_vars.append((path, sku_var))

        # ── Footer ────────────────────────────────────────────────────────
        foot = tk.Frame(d, bg=COLORS["app"])
        foot.grid(row=2, column=0, sticky="ew", padx=14, pady=10)

        ttk.Button(
            foot,
            text="Cancel" if lang == "en" else "\u53d6\u6d88",
            command=lambda: self._cancel(d),
            style="Tool.TButton",
        ).pack(side=tk.LEFT)

        tk.Button(
            foot,
            text="\u2191 Import \u2192" if lang == "en" else "\u5bfc\u5165 \u2192",
            command=lambda: self._ok(d),
            bg=COLORS["run"], fg="#ffffff",
            activebackground=COLORS.get("run_hover", COLORS["run"]),
            activeforeground="#ffffff",
            font=("Segoe UI", 10, "bold"),
            padx=18, pady=6, cursor="hand2",
            relief=tk.FLAT, borderwidth=0, highlightthickness=0,
        ).pack(side=tk.RIGHT)

        d.bind("<Return>", lambda _e: self._ok(d))
        d.bind("<Escape>", lambda _e: self._cancel(d))

        # Size and centre
        n = len(staged_files)
        h = min(120 + n * (self._THUMB + 18) + 120, 620)
        w = 720
        sw, sh = d.winfo_screenwidth(), d.winfo_screenheight()
        d.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")
        d.wait_window(d)

    # ------------------------------------------------------------------ helpers

    def _make_thumb(self, path) -> object | None:
        if Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(path)
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            im.thumbnail((self._THUMB, self._THUMB), Image.Resampling.LANCZOS)
            return ImageTk.PhotoImage(im)
        except Exception:
            return None

    def _ok(self, d: tk.Toplevel) -> None:
        self.result = {p.stem: var.get().strip() for p, var in self._sku_vars}
        self.confirmed = True
        d.destroy()

    def _cancel(self, d: tk.Toplevel) -> None:
        self.confirmed = False
        d.destroy()


# ======================================================================
class _CharmEditSkuDialog:
    """
    Modal dialog for reviewing and editing SKUs of all existing charm library
    entries.  Shows every charm as a scrollable row:

        [photo thumbnail]  [code]  [SKU entry (pre-filled with current value)]

    Two filter modes let the user work through all charms or only those with
    a blank SKU.  On "Save", the changes are written to column C via
    ``write_charm_library_skus``.  ``self.confirmed`` is True when the user
    saved, False when cancelled.
    """

    _THUMB = 64

    def __init__(
        self,
        parent: tk.Misc,
        entries: dict,          # dict[str, CharmLibraryEntry]
        catalog_path,           # Path
        lang: str,
        on_save_callback=None,  # called(updated_entries) after a successful save
    ) -> None:
        self.confirmed: bool = False
        self._lang = lang
        self._entries = entries
        self._catalog_path = catalog_path
        self._on_save_callback = on_save_callback
        self._sku_vars: list[tuple] = []    # [(code, current_sku, tk.StringVar), ...]
        self._tk_img_refs: list[object] = []
        self._row_frames: list[tk.Frame] = []

        d = tk.Toplevel(parent)
        d.title(
            "Edit Charm SKUs" if lang == "en"
            else "\u7f16\u8f91\u6302\u4ef6SKU"
        )
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.minsize(660, 420)
        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(2, weight=1)
        self._d = d

        # ── Hero ──────────────────────────────────────────────────────────
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero,
            text="Edit Charm SKUs" if lang == "en" else "\u7f16\u8f91\u6302\u4ef6SKU",
            font=("Segoe UI", 12, "bold"),
            fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(10, 2))
        tk.Label(
            hero,
            text=(
                "Edit the SKU (column C) for each charm.  "
                "Changes are written to the Excel workbook when you click Save.  "
                "Leave a field blank to clear its SKU."
            ) if lang == "en" else (
                "\u7f16\u8f91\u6bcf\u4e2a\u6302\u4ef6\u7684SKU\uff08C\u5217\uff09\u3002"
                "\u70b9\u51fb\u300c\u4fdd\u5b58\u300d\u5c06\u66f4\u6539\u5199\u5165Excel\u3002"
                "\u7559\u7a7a\u8868\u793a\u6e05\u9664\u8be5SKU\u3002"
            ),
            wraplength=660, justify=tk.LEFT,
            font=("Segoe UI", 9), fg="#e2e8f0", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(0, 8))

        # ── Filter bar ────────────────────────────────────────────────────
        fbar = tk.Frame(d, bg=COLORS["strip"])
        fbar.grid(row=1, column=0, sticky="ew")

        self._show_all_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            fbar,
            text=(
                "Show only charms with blank SKU" if lang == "en"
                else "\u4ec5\u663e\u793aSKU\u4e3a\u7a7a\u7684\u6302\u4ef6"
            ),
            variable=self._show_all_var,
            command=self._apply_filter,
            bg=COLORS["strip"], font=("Segoe UI", 9),
            fg=COLORS["text"], activebackground=COLORS["strip"],
            cursor="hand2",
        ).pack(side=tk.LEFT, padx=10, pady=4)

        n_blank = sum(1 for e in entries.values() if not (e.sku or "").strip())
        tk.Label(
            fbar,
            text=(
                f"{n_blank} charm(s) with blank SKU  /  {len(entries)} total"
                if lang == "en"
                else f"SKU\u4e3a\u7a7a\uff1a{n_blank}\u4e2a\uff0f\u5171 {len(entries)}\u4e2a"
            ),
            font=("Segoe UI", 8), fg=COLORS["muted"], bg=COLORS["strip"],
        ).pack(side=tk.RIGHT, padx=10)

        # ── Scrollable body ────────────────────────────────────────────────
        body_wrap = tk.Frame(d, bg=COLORS["app"])
        body_wrap.grid(row=2, column=0, sticky="nsew")
        body_wrap.grid_rowconfigure(0, weight=1)
        body_wrap.grid_columnconfigure(0, weight=1)

        canvas = tk.Canvas(body_wrap, bg=COLORS["app"], highlightthickness=0)
        vsb = ttk.Scrollbar(body_wrap, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self._canvas = canvas

        inner = tk.Frame(canvas, bg=COLORS["app"])
        _win = canvas.create_window((0, 0), window=inner, anchor=tk.NW)
        self._inner = inner

        def _sync(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all") or (0, 0, 0, 0))
            canvas.itemconfig(_win, width=canvas.winfo_width())

        inner.bind("<Configure>", _sync)
        canvas.bind("<Configure>", _sync)

        def _scroll(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _scroll)
        inner.bind("<MouseWheel>", _scroll)
        self._scroll_fn = _scroll

        # Column headers
        hdr = tk.Frame(inner, bg=COLORS["strip"])
        hdr.pack(fill=tk.X)
        tk.Label(
            hdr, text="Photo" if lang == "en" else "\u7167\u7247",
            font=("Segoe UI", 8, "bold"), bg=COLORS["strip"], width=8,
        ).pack(side=tk.LEFT, padx=(8, 0))
        tk.Label(
            hdr, text="Code" if lang == "en" else "\u7f16\u7801",
            font=("Segoe UI", 8, "bold"), bg=COLORS["strip"], width=10,
        ).pack(side=tk.LEFT, padx=4)
        tk.Label(
            hdr,
            text="SKU (Column C) — edit here" if lang == "en"
            else "SKU\uff08C\u5217\uff09\u2014\u5728\u6b64\u5904\u7f16\u8f91",
            font=("Segoe UI", 8, "bold"), bg=COLORS["strip"], anchor=tk.W,
        ).pack(side=tk.LEFT, padx=4, expand=True, fill=tk.X)

        # Build one row per entry
        for i, (code, entry) in enumerate(entries.items()):
            row_bg = "#ffffff" if i % 2 == 0 else "#f8f7ff"
            row_f = tk.Frame(
                inner, bg=row_bg,
                highlightthickness=1,
                highlightbackground=COLORS["border"],
            )
            row_f.pack(fill=tk.X, padx=4, pady=2)
            row_f.bind("<MouseWheel>", _scroll)

            # Thumbnail from embedded photo bytes
            ph = self._make_thumb(entry.photo_bytes if hasattr(entry, "photo_bytes") else None)
            if ph:
                self._tk_img_refs.append(ph)
                lbl = tk.Label(row_f, image=ph, bg=row_bg,
                               width=self._THUMB + 6, height=self._THUMB + 6)
            else:
                lbl = tk.Label(row_f, text="\U0001f48e",
                               font=("Segoe UI", 18), fg="#c4b5fd", bg=row_bg,
                               width=4, height=2)
            lbl.pack(side=tk.LEFT, padx=(6, 2), pady=4)
            lbl.bind("<MouseWheel>", _scroll)

            # Code label
            tk.Label(
                row_f, text=code,
                font=("Segoe UI", 9, "bold"), fg="#5b21b6", bg=row_bg,
                width=10, anchor=tk.W,
            ).pack(side=tk.LEFT, padx=4)

            # Editable SKU entry
            current = (entry.sku or "").strip() if hasattr(entry, "sku") else ""
            sku_var = tk.StringVar(value=current)
            entry_widget = ttk.Entry(
                row_f, textvariable=sku_var,
                font=("Segoe UI", 9),
            )
            entry_widget.pack(side=tk.LEFT, padx=(4, 8), pady=6, expand=True, fill=tk.X)
            entry_widget.bind("<MouseWheel>", _scroll)

            self._sku_vars.append((code, current, sku_var))
            self._row_frames.append((row_f, code, current))

        # ── Footer ────────────────────────────────────────────────────────
        foot = tk.Frame(d, bg=COLORS["app"])
        foot.grid(row=3, column=0, sticky="ew", padx=14, pady=10)

        ttk.Button(
            foot,
            text="Cancel" if lang == "en" else "\u53d6\u6d88",
            command=lambda: self._cancel(d),
            style="Tool.TButton",
        ).pack(side=tk.LEFT)

        n_changed_lbl = tk.Label(
            foot, text="",
            font=("Segoe UI", 8), fg=COLORS["muted"], bg=COLORS["app"],
        )
        n_changed_lbl.pack(side=tk.LEFT, padx=12)
        self._n_changed_lbl = n_changed_lbl

        # Wire up trace to show live "N change(s)" count
        for _, _, sv in self._sku_vars:
            sv.trace_add("write", lambda *_: self._update_change_count())

        tk.Button(
            foot,
            text="\U0001f4be  Save SKUs" if lang == "en" else "\U0001f4be  \u4fdd\u5b58SKU",
            command=lambda: self._save(d),
            bg=COLORS["run"], fg="#ffffff",
            activebackground=COLORS.get("run_hover", COLORS["run"]),
            activeforeground="#ffffff",
            font=("Segoe UI", 10, "bold"),
            padx=18, pady=6, cursor="hand2",
            relief=tk.FLAT, borderwidth=0, highlightthickness=0,
        ).pack(side=tk.RIGHT)

        d.bind("<Escape>", lambda _e: self._cancel(d))

        # Initial filter state and sizing
        self._show_all_var.set(True)
        self._apply_filter()

        h = min(160 + len(entries) * (self._THUMB + 14) + 100, 700)
        w = 760
        sw, sh = d.winfo_screenwidth(), d.winfo_screenheight()
        d.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")
        d.wait_window(d)

    # ------------------------------------------------------------------ helpers

    def _make_thumb(self, photo_bytes: bytes | None) -> object | None:
        if not photo_bytes or Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(BytesIO(photo_bytes))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            im.thumbnail((self._THUMB, self._THUMB), Image.Resampling.LANCZOS)
            return ImageTk.PhotoImage(im)
        except Exception:
            return None

    def _apply_filter(self) -> None:
        """Show only blank-SKU rows or all rows based on the checkbox."""
        show_blank_only = self._show_all_var.get()
        for row_f, code, _current in self._row_frames:
            # Find the corresponding sku_var
            sku_now = next(
                (sv.get() for c, _orig, sv in self._sku_vars if c == code), ""
            )
            original = next(
                (orig for c, orig, _sv in self._sku_vars if c == code), ""
            )
            if show_blank_only and original.strip():
                row_f.pack_forget()
            else:
                row_f.pack(fill=tk.X, padx=4, pady=2)

    def _update_change_count(self) -> None:
        n = sum(
            1 for code, orig, sv in self._sku_vars
            if sv.get().strip() != orig.strip()
        )
        if n:
            self._n_changed_lbl.config(
                text=(
                    f"{n} change(s) pending"
                    if self._lang == "en"
                    else f"{n} \u9879\u5f85\u4fdd\u5b58"
                )
            )
        else:
            self._n_changed_lbl.config(text="")

    def _save(self, d: tk.Toplevel) -> None:
        """Collect changed SKUs and write them to the workbook."""
        code_to_sku = {
            code: sv.get().strip()
            for code, orig, sv in self._sku_vars
            if sv.get().strip() != orig.strip()
        }
        if not code_to_sku:
            messagebox.showinfo(
                "No changes" if self._lang == "en" else "\u65e0\u66f4\u6539",
                "No SKUs were changed." if self._lang == "en"
                else "\u6ca1\u6709SKU\u88ab\u4fee\u6539\u3002",
                parent=self._d,
            )
            return

        if write_charm_library_skus is None:
            messagebox.showerror(
                "Unavailable" if self._lang == "en" else "\u4e0d\u53ef\u7528",
                "write_charm_library_skus is not available.",
                parent=self._d,
            )
            return

        try:
            n, lines = write_charm_library_skus(self._catalog_path, code_to_sku)
        except Exception as e:
            messagebox.showerror(
                "Save failed" if self._lang == "en" else "\u4fdd\u5b58\u5931\u8d25",
                str(e), parent=self._d,
            )
            return

        if n:
            messagebox.showinfo(
                "Saved" if self._lang == "en" else "\u5df2\u4fdd\u5b58",
                (
                    f"Updated {n} SKU(s) in the workbook.\n\n"
                    + "\n".join(lines[-6:])
                ) if self._lang == "en" else (
                    f"\u5df2\u66f4\u65b0 {n} \u4e2aSKU\u3002\n\n"
                    + "\n".join(lines[-6:])
                ),
                parent=self._d,
            )
            self.confirmed = True
            if self._on_save_callback is not None:
                self._on_save_callback()
        else:
            messagebox.showinfo(
                "No updates" if self._lang == "en" else "\u65e0\u66f4\u65b0",
                "\n".join(lines) or "Nothing was updated.",
                parent=self._d,
            )

    def _cancel(self, d: tk.Toplevel) -> None:
        self.confirmed = False
        d.destroy()


# ======================================================================
class _CharmReorderDialog:
    """
    Modal dialog that lets the user drag-and-drop (or use arrow buttons) to
    reorder charms.  Calls ``reorder_charm_library_rows`` on Apply.

    Layout
    ------
    Row 0 — hero banner (title + instructions)
    Row 1 — main body: Treeview (left) + arrow buttons (right)
    Row 2 — footer: Close / Apply buttons
    """

    _ROW_H   = 80   # Treeview row height in px when Pillow is available
    _THUMB   = 68   # thumbnail size for Treeview column #0

    def __init__(
        self,
        parent: App,
        entries: dict,   # dict[str, CharmLibraryEntry]
        catalog_path: Path,
        charm_images_dir: Path,
    ) -> None:
        self._parent         = parent
        self._entries        = entries
        self._catalog_path   = catalog_path
        self._charm_images   = charm_images_dir
        self._lang           = parent._lang
        self._tk_img_refs: list[object] = []

        pil_ok = Image is not None and ImageTk is not None
        self._pil_ok = pil_ok

        # Load charm shop names for the shop combobox
        self._charm_shop_names: list[str] = []
        try:
            if load_charm_shops is not None:
                self._charm_shop_names = [
                    cs.shop_name for cs in load_charm_shops(catalog_path)
                    if cs.shop_name
                ]
        except Exception:
            pass

        d = tk.Toplevel(parent)
        self._d = d
        d.title(self._t("reorder_title"))
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.geometry("1020x680")
        d.minsize(760, 500)

        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(1, weight=1)

        # ── Row 0: hero banner ────────────────────────────────────────────
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero,
            text=self._t("reorder_heading"),
            font=("Segoe UI", 12, "bold"),
            fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(10, 2))
        tk.Label(
            hero,
            text=self._t("reorder_intro"),
            wraplength=940, justify=tk.LEFT,
            font=("Segoe UI", 9), fg="#e2e8f0", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(0, 8))

        # ── Row 1: body (tree + arrow buttons) ───────────────────────────
        body = tk.Frame(d, bg=COLORS["app"])
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(10, 0))
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=1)

        tree_frame = tk.Frame(body, bg=COLORS["app"])
        tree_frame.grid(row=0, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Treeview styling
        sty = ttk.Style(d)
        row_h = self._ROW_H if pil_ok else 26
        try:
            sty.configure("Reorder.Treeview",
                          rowheight=row_h, font=("Segoe UI", 9))
            sty.configure("Reorder.Treeview.Heading",
                          font=("Segoe UI", 9, "bold"))
            sty.map("Reorder.Treeview",
                    background=[("selected", COLORS["accent_soft"])])
        except tk.TclError:
            pass

        cols = ("code", "new_code", "sku", "shop")
        tree = ttk.Treeview(
            tree_frame,
            columns=cols,
            show="tree headings",
            selectmode="browse",
            style="Reorder.Treeview",
        )
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self._tree = tree

        thumb_w = self._THUMB + 8 if pil_ok else 20
        tree.heading("#0",       text=self._t("reorder_col_photo"))
        tree.column ("#0",       width=thumb_w, stretch=False, anchor=tk.CENTER)
        tree.heading("code",     text=self._t("reorder_col_code"))
        tree.column ("code",     width=110, stretch=False, anchor=tk.CENTER)
        tree.heading("new_code", text=self._t("reorder_col_new"))
        tree.column ("new_code", width=110, stretch=False, anchor=tk.CENTER)
        tree.heading("sku",      text=self._t("reorder_col_sku"))
        tree.column ("sku",      width=220, stretch=True)
        tree.heading("shop",     text=self._t("reorder_col_shop"))
        tree.column ("shop",     width=130, stretch=False)

        # Arrow-button column (right of tree)
        btn_col = tk.Frame(body, bg=COLORS["app"])
        btn_col.grid(row=0, column=1, sticky="ns", padx=(10, 0), pady=(0, 0))

        btn_cfg = dict(style="Tool.TButton", width=14)
        ttk.Button(btn_col, text=self._t("reorder_btn_top"),
                   command=self._move_top, **btn_cfg).pack(pady=(0, 4))
        ttk.Button(btn_col, text=self._t("reorder_btn_up"),
                   command=self._move_up,  **btn_cfg).pack(pady=(0, 4))
        ttk.Button(btn_col, text=self._t("reorder_btn_down"),
                   command=self._move_down, **btn_cfg).pack(pady=(0, 4))
        ttk.Button(btn_col, text=self._t("reorder_btn_bottom"),
                   command=self._move_bottom, **btn_cfg).pack(pady=(0, 0))

        # Separator + Delete button (danger-styled)
        tk.Frame(btn_col, bg=COLORS["border"], height=1,
                 highlightthickness=0).pack(fill=tk.X, pady=(14, 10))
        sty_d = ttk.Style(d)
        sty_d.configure("ReorderDanger.TButton", font=("Segoe UI", 10), padding=(8, 6))
        sty_d.map("ReorderDanger.TButton",
                  background=[("active", "#fee2e2"), ("disabled", "#f3f4f6")],
                  foreground=[("active", "#b91c1c"), ("disabled", "#9ca3af")])
        self._delete_btn = ttk.Button(
            btn_col,
            text="\u2716 Delete" if self._lang == "en" else "\u2716 \u5220\u9664",
            command=self._delete_charm,
            style="ReorderDanger.TButton",
            width=14,
            state="disabled",
        )
        self._delete_btn.pack(pady=(0, 0))
        # Enable/disable the delete button based on tree selection
        tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        # ── Shop edit section (below delete) ─────────────────────────────
        tk.Frame(btn_col, bg=COLORS["border"], height=1,
                 highlightthickness=0).pack(fill=tk.X, pady=(14, 8))

        tk.Label(btn_col,
                 text="Default Shop" if self._lang == "en" else "\u9ed8\u8ba4\u6302\u4ef6\u5e97",
                 font=("Segoe UI", 8, "bold"), fg=COLORS["muted"],
                 bg=COLORS["app"]).pack(anchor=tk.W)

        self._shop_var = tk.StringVar()
        self._shop_cb  = ttk.Combobox(btn_col, textvariable=self._shop_var,
                                      values=[""] + self._charm_shop_names,
                                      font=("Segoe UI", 9), width=14, state="disabled")
        self._shop_cb.pack(fill=tk.X, pady=(2, 4))

        self._set_shop_btn = ttk.Button(
            btn_col,
            text="Set Shop" if self._lang == "en" else "\u8bbe\u7f6e\u5e97\u94fa",
            command=self._set_charm_shop,
            style="Tool.TButton", width=14, state="disabled",
        )
        self._set_shop_btn.pack(fill=tk.X, pady=(0, 4))

        ttk.Button(btn_col,
                   text="Manage Shops\u2026" if self._lang == "en" else "\u7ba1\u7406\u5e97\u94fa\u2026",
                   command=self._open_shops_manager,
                   style="Tool.TButton", width=14).pack(fill=tk.X, pady=(0, 0))

        # ── SKU edit section ──────────────────────────────────────────────
        tk.Frame(btn_col, bg=COLORS["border"], height=1,
                 highlightthickness=0).pack(fill=tk.X, pady=(14, 10))
        ttk.Button(
            btn_col,
            text="\u270f\ufe0f  Edit SKUs\u2026" if self._lang == "en"
            else "\u270f\ufe0f  \u7f16\u8f91SKU\u2026",
            command=self._open_sku_editor,
            style="Tool.TButton", width=14,
        ).pack(fill=tk.X)

        # ── Row 2: footer buttons ─────────────────────────────────────────
        foot = tk.Frame(d, bg=COLORS["app"])
        foot.grid(row=2, column=0, sticky="ew", padx=14, pady=12)
        close_btn = ttk.Button(
            foot, text=self._t("reorder_btn_close"),
            command=d.destroy, style="Tool.TButton",
        )
        close_btn.pack(side=tk.LEFT)
        self._apply_btn = tk.Button(
            foot,
            text=self._t("reorder_btn_apply"),
            command=self._apply,
            bg=COLORS["run"], fg="#ffffff",
            activebackground=COLORS["run_hover"],
            activeforeground="#ffffff",
            font=("Segoe UI", 10, "bold"),
            padx=18, pady=6,
            cursor="hand2",
            relief=tk.FLAT, borderwidth=0, highlightthickness=0,
        )
        self._apply_btn.pack(side=tk.RIGHT)

        # ── Populate tree ─────────────────────────────────────────────────
        self._iid_to_code: dict[str, str] = {}
        self._populate()

        # ── Drag-and-drop bindings ────────────────────────────────────────
        self._drag: dict[str, object] = {"item": None}
        tree.bind("<ButtonPress-1>",   self._on_drag_start)
        tree.bind("<B1-Motion>",       self._on_drag_motion)
        tree.bind("<ButtonRelease-1>", self._on_drag_release)

        # ── Hover-zoom bindings ───────────────────────────────────────────
        self._hover_tip_win:    list[tk.Toplevel | None] = [None]
        self._hover_after_id:   list[object | None]      = [None]
        self._hover_active_iid: list[str | None]         = [None]
        self._hover_photo:      list[object]             = []
        tree.bind("<Motion>", self._on_hover_motion)
        tree.bind("<Leave>",  self._hide_hover)

        # Block the caller until the dialog is closed so that any post-close
        # logic (e.g. _rebuild_gallery) runs only after the user is done.
        d.wait_window(d)

    # ------------------------------------------------------------------ helpers

    def _t(self, key: str, **kw: object) -> str:
        s = CHROME.get(self._lang, CHROME["en"]).get(key, key)
        return s.format(**kw) if kw else s

    def _thumb(self, raw: bytes | None) -> object | None:
        if not self._pil_ok or not raw or Image is None or ImageTk is None:
            return None
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            im.thumbnail((self._THUMB, self._THUMB), Image.Resampling.LANCZOS)
            ph = ImageTk.PhotoImage(im)
            self._tk_img_refs.append(ph)
            return ph
        except Exception:
            return None

    def _codes_in_order(self) -> list[str]:
        return [self._iid_to_code[iid] for iid in self._tree.get_children()]

    def _new_code_for_pos(self, pos: int, total: int) -> str:
        """Sequential code string for 1-based *pos* out of *total*."""
        import re as _re
        all_codes = list(self._entries.keys())
        pfx = "CH-"
        for code in all_codes:
            m = _re.match(r"^([A-Za-z]+-+)(\d+)$", code.strip())
            if m:
                pfx = m.group(1).upper()
                break
        # Compute width
        widths = []
        for code in all_codes:
            m = _re.match(rf"^{_re.escape(pfx)}(\d+)$", code.strip(), _re.IGNORECASE)
            if m:
                widths.append(len(m.group(1)))
        w = max(widths) if widths else 5
        w = max(w, len(str(total)), 5)
        return f"{pfx}{pos:0{w}d}"

    def _refresh_new_codes(self) -> None:
        """Update the «New Code» column for every row based on current position."""
        items = self._tree.get_children()
        total = len(items)
        for pos, iid in enumerate(items, start=1):
            nc = self._new_code_for_pos(pos, total)
            self._tree.set(iid, "new_code", nc)

    def _populate(self) -> None:
        self._tree.delete(*self._tree.get_children())
        self._iid_to_code.clear()
        self._tk_img_refs.clear()
        total = len(self._entries)
        for pos, (code, entry) in enumerate(self._entries.items(), start=1):
            iid   = f"charm_{code}"
            thumb = self._thumb(entry.photo_bytes)
            nc    = self._new_code_for_pos(pos, total)
            kw: dict = dict(
                values=(code, nc, entry.sku or "", entry.default_charm_shop or ""),
            )
            if thumb is not None:
                kw["image"] = thumb
                kw["text"]  = ""
            else:
                kw["text"] = "—"
            self._tree.insert("", tk.END, iid=iid, **kw)
            self._iid_to_code[iid] = code

    # ------------------------------------------------------------------ movement

    def _move_up(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        item = sel[0]
        idx  = self._tree.index(item)
        if idx > 0:
            self._tree.move(item, "", idx - 1)
            self._refresh_new_codes()

    def _move_down(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        item = sel[0]
        idx  = self._tree.index(item)
        last = len(self._tree.get_children()) - 1
        if idx < last:
            self._tree.move(item, "", idx + 1)
            self._refresh_new_codes()

    def _move_top(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        self._tree.move(sel[0], "", 0)
        self._refresh_new_codes()

    def _move_bottom(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        last = len(self._tree.get_children()) - 1
        self._tree.move(sel[0], "", last)
        self._refresh_new_codes()

    # ------------------------------------------------------------------ drag-drop

    def _on_drag_start(self, event: tk.Event) -> None:
        item = self._tree.identify_row(event.y)
        if item:
            self._tree.selection_set(item)
            self._drag["item"] = item

    def _on_drag_motion(self, event: tk.Event) -> None:
        item = self._drag.get("item")
        if not item:
            return
        target = self._tree.identify_row(event.y)
        if not target or target == item:
            return
        drag_idx   = self._tree.index(item)
        target_idx = self._tree.index(target)
        # Determine insert-before or insert-after from mouse Y within target row
        bbox = self._tree.bbox(target, "#0")
        if bbox:
            mid = bbox[1] + bbox[3] // 2
            if event.y < mid:
                dest = target_idx if drag_idx > target_idx else target_idx - 1
            else:
                dest = target_idx if drag_idx < target_idx else target_idx + 1
        else:
            dest = target_idx
        self._tree.move(item, "", max(0, dest))
        self._refresh_new_codes()

    def _on_drag_release(self, _event: tk.Event) -> None:
        self._drag["item"] = None

    # ------------------------------------------------------------------ apply

    # ── Hover-zoom ────────────────────────────────────────────────────

    def _hide_hover(self, _event=None) -> None:
        if self._hover_after_id[0] is not None:
            try:
                self._d.after_cancel(self._hover_after_id[0])
            except (tk.TclError, ValueError):
                pass
            self._hover_after_id[0] = None
        self._hover_active_iid[0] = None
        tip = self._hover_tip_win[0]
        if tip is not None:
            try:
                tip.destroy()
            except tk.TclError:
                pass
            self._hover_tip_win[0] = None

    def _on_hover_motion(self, _e: tk.Event | None = None) -> None:
        if not self._pil_ok:
            return
        tree = self._tree
        px, py = tree.winfo_pointerx(), tree.winfo_pointery()
        lx, ly = px - tree.winfo_rootx(), py - tree.winfo_rooty()
        if lx < 0 or ly < 0 or lx >= tree.winfo_width() or ly >= tree.winfo_height():
            self._hide_hover()
            return
        # Only activate over the photo column (#0)
        if tree.identify_column(lx) != "#0":
            self._hide_hover()
            return
        iid = tree.identify_row(ly)
        if not iid:
            self._hide_hover()
            return
        code = iid.removeprefix("charm_")
        entry = self._entries.get(code)
        if entry is None or not entry.photo_bytes:
            self._hide_hover()
            return
        if iid == self._hover_active_iid[0]:
            return   # already showing
        self._hide_hover()
        self._hover_active_iid[0] = iid
        self._hover_after_id[0] = self._d.after(
            220, lambda i=iid: self._show_hover(i)
        )

    def _show_hover(self, iid: str) -> None:
        self._hover_after_id[0] = None
        if Image is None or ImageTk is None:
            return
        tree = self._tree
        px, py = tree.winfo_pointerx(), tree.winfo_pointery()
        lx, ly = px - tree.winfo_rootx(), py - tree.winfo_rooty()
        if (lx < 0 or ly < 0 or lx >= tree.winfo_width()
                or ly >= tree.winfo_height()
                or tree.identify_row(ly) != iid):
            return
        code  = iid.removeprefix("charm_")
        entry = self._entries.get(code)
        raw   = entry.photo_bytes if entry else None
        if not raw:
            return
        try:
            im = Image.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA" if "A" in im.mode else "RGB")
            if im.mode == "RGBA":
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                im = bg
            w, h = im.size
            # Always scale to a useful preview size (up- and down-scale)
            _sw = self._d.winfo_screenwidth()
            _sh = self._d.winfo_screenheight()
            max_dim = min(400, max(260, min(_sw, _sh) // 3))
            ratio   = min(max_dim / w, max_dim / h)
            nw, nh  = max(1, int(w * ratio)), max(1, int(h * ratio))
            if (nw, nh) != (w, h):
                resample = (Image.Resampling.LANCZOS if ratio <= 1.0
                            else Image.Resampling.BICUBIC)
                im = im.resize((nw, nh), resample)
            ph = ImageTk.PhotoImage(im)
            self._hover_photo.clear()
            self._hover_photo.append(ph)
        except Exception:
            return
        tip = tk.Toplevel(self._d)
        tip.overrideredirect(True)
        try:
            tip.attributes("-topmost", True)
        except tk.TclError:
            pass
        tip.configure(bg="#ffffff", highlightthickness=2,
                      highlightbackground=COLORS["accent"])
        tk.Label(tip, image=ph, bg="#ffffff", bd=0,
                 highlightthickness=0, relief=tk.FLAT).pack()
        tip.update_idletasks()
        sw, sh = tip.winfo_screenwidth(), tip.winfo_screenheight()
        tw, th = tip.winfo_reqwidth(), tip.winfo_reqheight()
        # Position above cursor so popup doesn't cover the row being hovered
        x = min(max(12, px + 16), sw - tw - 12)
        y = min(max(12, py - th - 8), sh - th - 12)
        if y < 12:
            y = min(py + 16, sh - th - 12)
        tip.geometry(f"+{x}+{y}")
        self._hover_tip_win[0] = tip

    def _on_tree_select(self, _event=None) -> None:
        """Enable action buttons when a row is selected; pre-fill the shop combobox."""
        sel = self._tree.selection()
        has_sel = bool(sel)
        self._delete_btn.config(state="normal" if has_sel else "disabled")
        self._shop_cb.config(state="readonly" if has_sel else "disabled")
        self._set_shop_btn.config(state="normal" if has_sel else "disabled")
        if has_sel:
            code  = self._iid_to_code.get(sel[0], "")
            entry = self._entries.get(code)
            self._shop_var.set(entry.default_charm_shop if entry and entry.default_charm_shop else "")

    def _set_charm_shop(self) -> None:
        """Update the Default Shop for the selected charm in memory and in the XLSX."""
        sel = self._tree.selection()
        if not sel:
            return
        iid  = sel[0]
        code = self._iid_to_code.get(iid, "")
        if not code:
            return
        new_shop = self._shop_var.get().strip()

        # Update in-memory entry
        entry = self._entries.get(code)
        if entry is not None:
            entry.default_charm_shop = new_shop
        # Reflect in the tree
        self._tree.set(iid, "shop", new_shop or "\u2014")

        # Write col D of the matching Charm Library row
        try:
            if backup_supplier_catalog_before_write is not None:
                backup_supplier_catalog_before_write(self._catalog_path, "set_charm_shop")
            import openpyxl as _xl
            wb = _xl.load_workbook(self._catalog_path)
            if CHARM_LIBRARY_SHEET in wb.sheetnames:
                ws = wb[CHARM_LIBRARY_SHEET]
                for r in range(2, ws.max_row + 1):
                    if str(ws.cell(r, 2).value or "").strip() == code:
                        ws.cell(r, 4).value = new_shop or None
                        break
            wb.save(self._catalog_path)
            wb.close()
        except Exception as e:
            messagebox.showerror(
                "Save failed" if self._lang == "en" else "\u4fdd\u5b58\u5931\u8d25",
                str(e), parent=self._d,
            )

    def _reload_shop_combobox(self) -> None:
        """Refresh the shop combobox values after the Charm Shops list changes."""
        try:
            if load_charm_shops is not None:
                self._charm_shop_names = [
                    cs.shop_name for cs in load_charm_shops(self._catalog_path)
                    if cs.shop_name
                ]
        except Exception:
            pass
        self._shop_cb["values"] = [""] + self._charm_shop_names

    def _open_shops_manager(self) -> None:
        """Open the Charm Shops manager dialog (add / edit / remove shops)."""
        _CharmShopsManagerDialog(self._d, self._catalog_path,
                                 self._lang, on_close=self._reload_shop_combobox)

    def _open_sku_editor(self) -> None:
        """Open the SKU editor dialog so the user can fill/update charm SKUs."""
        def _after_save() -> None:
            # Reload entries from disk to pick up the freshly written SKUs,
            # then repopulate the tree so column C reflects the new values.
            if load_charm_library is not None:
                try:
                    updated = load_charm_library(self._catalog_path)
                    if updated:
                        self._entries = updated
                except Exception:
                    pass
            self._populate()

        _CharmEditSkuDialog(
            self._d,
            self._entries,
            self._catalog_path,
            self._lang,
            on_save_callback=_after_save,
        )

    def _delete_charm(self) -> None:
        """Delete the selected charm from the Charm Library and clear Product Map refs."""
        sel = self._tree.selection()
        if not sel:
            return
        iid  = sel[0]
        code = self._iid_to_code.get(iid, "")
        if not code:
            return

        entry      = self._entries.get(code)
        sku_txt    = f"  ({entry.sku})" if entry and entry.sku else ""

        if self._lang == "en":
            msg = (
                f"Permanently delete charm {code}{sku_txt}?\n\n"
                f"This will:\n"
                f"  \u2022 Remove its row and photo from the Charm Library\n"
                f"  \u2022 Clear this Charm Code from every Product Map row that references it\n\n"
                f"A backup is created automatically before any change is made."
            )
            title = "Delete Charm"
        else:
            msg = (
                f"\u6c38\u4e45\u5220\u9664\u6302\u4ef6 {code}{sku_txt}\uff1f\n\n"
                f"\u64cd\u4f5c\u5185\u5bb9\uff1a\n"
                f"  \u2022 \u5220\u9664\u6302\u4ef6\u5e93\u4e2d\u7684\u884c\u4e0e\u7167\u7247\n"
                f"  \u2022 \u6e05\u9664\u5546\u54c1\u8868\u4e2d\u6240\u6709\u5f15\u7528\u6b64\u6302\u4ef6\u7f16\u7801\u7684\u884c\n\n"
                f"\u5199\u5165\u524d\u81ea\u52a8\u521b\u5efa\u5907\u4efd\u3002"
            )
            title = "\u5220\u9664\u6302\u4ef6"

        if not messagebox.askyesno(title, msg, icon="warning", parent=self._d):
            return

        # Backup
        try:
            if backup_supplier_catalog_before_write is not None:
                backup_supplier_catalog_before_write(self._catalog_path, "delete_charm")
        except Exception:
            pass

        # Write changes via openpyxl
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(self._catalog_path)

            # 1. Clear the charm's row in Charm Library (col B = charm code).
            #    We intentionally do NOT use delete_rows() because openpyxl
            #    shifts cell data but does NOT reliably shift image anchors,
            #    corrupting the photo-to-row mapping for every row below.
            #    Instead we clear cell values + remove the image.  The empty
            #    row is invisible to load_charm_library (skips empty col B)
            #    and gets compacted by the next reorder_charm_library_rows.
            if CHARM_LIBRARY_SHEET in wb.sheetnames:
                ws_lib = wb[CHARM_LIBRARY_SHEET]
                target_row: int | None = None
                for r in range(2, ws_lib.max_row + 1):
                    if str(ws_lib.cell(r, 2).value or "").strip() == code:
                        target_row = r
                        break
                if target_row is not None:
                    for c in range(1, 6):
                        ws_lib.cell(target_row, c).value = None
                    ws_lib._images = [
                        img for img in list(getattr(ws_lib, "_images", []) or [])
                        if not (
                            hasattr(img, "anchor")
                            and hasattr(img.anchor, "_from")
                            and getattr(img.anchor._from, "row", None) == target_row - 1
                        )
                    ]

            # 2. Clear Charm Code (col G) in Product Map for all matching rows
            if CATALOG_SHEET in wb.sheetnames:
                ws_cat = wb[CATALOG_SHEET]
                for r in range(2, ws_cat.max_row + 1):
                    if str(ws_cat.cell(r, 7).value or "").strip() == code:
                        ws_cat.cell(r, 7).value = None

            wb.save(self._catalog_path)
            wb.close()
        except Exception as e:
            messagebox.showerror(
                "Delete failed" if self._lang == "en" else "\u5220\u9664\u5931\u8d25",
                str(e), parent=self._d,
            )
            return

        # Remove disk image files for this charm code
        if self._charm_images and self._charm_images.is_dir():
            for ext in (".png", ".jpg", ".jpeg", ".webp"):
                img_path = self._charm_images / f"{code}{ext}"
                try:
                    if img_path.exists():
                        img_path.unlink()
                except OSError:
                    pass

        # Remove from in-memory entries and refresh the tree
        self._entries.pop(code, None)
        self._tree.delete(iid)
        self._iid_to_code.pop(iid, None)
        self._refresh_new_codes()
        self._delete_btn.config(state="disabled")
        self._shop_cb.config(state="disabled")
        self._set_shop_btn.config(state="disabled")
        self._shop_var.set("")

        # Briefly show confirmation in the footer area
        info_lbl = tk.Label(
            self._d,
            text=f"\u2716  {code} deleted" if self._lang == "en"
                 else f"\u2716  {code} \u5df2\u5220\u9664",
            font=("Segoe UI", 9, "bold"), fg="#b91c1c",
            bg=COLORS["app"],
        )
        info_lbl.grid(row=2, column=0, sticky="w", padx=20)
        self._d.after(3000, info_lbl.destroy)

    def _apply(self) -> None:
        if not messagebox.askyesno(
            self._t("reorder_confirm_title"),
            self._t("reorder_confirm_body"),
            parent=self._d,
        ):
            return

        new_order = self._codes_in_order()

        self._apply_btn.config(state=tk.DISABLED)
        prev_cursor = self._d.cget("cursor")
        self._d.config(cursor="watch")
        self._d.update_idletasks()

        try:
            n, lines = reorder_charm_library_rows(
                self._catalog_path,
                new_order,
                charm_images_dir=self._charm_images,
                caller_entries=self._entries,
            )
            summary = "\n".join(lines)
            msg = f"{self._t('reorder_done')}\n\n{summary}"
            messagebox.showinfo(self._t("reorder_title"), msg, parent=self._d)

            self._refresh_entries_after_apply(new_order)
            self._populate()
        except Exception as e:
            messagebox.showerror(self._t("file_open_fail_title"), str(e), parent=self._d)
        finally:
            self._d.config(cursor=prev_cursor)
            self._apply_btn.config(state=tk.NORMAL)

    def _refresh_entries_after_apply(self, new_order: list[str]) -> None:
        """Rebuild ``self._entries`` to reflect the just-saved reorder.

        Primary strategy: reload from disk (authoritative source, includes
        refreshed embedded photos).  Retry once after a short delay to
        handle Windows/OneDrive file-sync latency.

        Fallback: build in-memory from the pre-apply entries using the
        same sequential-code logic that ``reorder_charm_library_rows``
        applied, so the tree is always correct even if disk I/O fails.
        """
        import re as _re2, time, copy

        # -- Fallback: in-memory rebuild from pre-apply state ---------------
        old_entries = dict(self._entries)
        all_old_codes = list(old_entries.keys())

        pfx = "CH-"
        for c in all_old_codes:
            m = _re2.match(r"^([A-Za-z]+-+)(\d+)$", c.strip())
            if m:
                pfx = m.group(1).upper()
                break

        widths: list[int] = []
        for c in all_old_codes:
            m = _re2.match(
                rf"^{_re2.escape(pfx)}(\d+)$", c.strip(), _re2.IGNORECASE
            )
            if m:
                widths.append(len(m.group(1)))
        w = max(widths) if widths else 5
        w = max(w, len(str(len(new_order))), 5)

        rebuilt: dict = {}
        for seq, old_code in enumerate(new_order, start=1):
            if old_code in old_entries:
                new_code = f"{pfx}{seq:0{w}d}"
                ent = copy.copy(old_entries[old_code])
                ent.code = new_code
                rebuilt[new_code] = ent

        self._entries = rebuilt

        # -- Primary: reload from disk (with one retry) --------------------
        for attempt in range(2):
            try:
                if attempt > 0:
                    time.sleep(0.3)
                disk = load_charm_library(self._catalog_path)
                if disk and len(disk) == len(rebuilt):
                    self._entries = disk
                    return
            except Exception:
                pass


# ======================================================================
# Charm Shops Manager Dialog — add, edit, and remove charm shop entries
# ======================================================================

class _CharmShopsManagerDialog:
    """Simple modal dialog for managing the Charm Shops reference list."""

    def __init__(
        self,
        parent: tk.Misc,
        catalog_path: Path,
        lang: str,
        on_close: object = None,
    ) -> None:
        self._catalog_path = catalog_path
        self._lang         = lang
        self._on_close     = on_close
        self._shops: list[dict] = []   # [{name, stall, notes}]

        d = tk.Toplevel(parent)
        self._d = d
        d.title("Manage Charm Shops" if lang == "en" else "\u7ba1\u7406\u6302\u4ef6\u5e97\u94fa")
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.geometry("560x420")
        d.minsize(480, 340)
        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(1, weight=1)

        # Hero
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(hero,
                 text="Charm Shops" if lang == "en" else "\u6302\u4ef6\u5e97\u94fa",
                 font=("Segoe UI", 12, "bold"), fg="#ffffff",
                 bg=COLORS["hero"]).pack(anchor=tk.W, padx=14, pady=(10, 2))
        tk.Label(hero,
                 text="Add, remove or edit rows that appear in the 'Charm Shop' dropdowns."
                      if lang == "en" else
                      "\u6dfb\u52a0\u3001\u5220\u9664\u6216\u7f16\u8f91\u5728\u300c\u6302\u4ef6\u5e97\u94fa\u300d\u4e0b\u62c9\u5217\u8868\u4e2d\u663e\u793a\u7684\u884c\u3002",
                 font=("Segoe UI", 9), fg="#e2e8f0",
                 bg=COLORS["hero"]).pack(anchor=tk.W, padx=14, pady=(0, 8))

        # Body: Treeview + buttons
        body = tk.Frame(d, bg=COLORS["app"])
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(10, 0))
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=1)

        # Treeview
        tf = tk.Frame(body, bg=COLORS["app"])
        tf.grid(row=0, column=0, sticky="nsew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        cols = ("name", "stall", "notes")
        self._tree = ttk.Treeview(tf, columns=cols, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self._tree.heading("name",  text="Shop Name" if lang == "en" else "\u5e97\u540d")
        self._tree.heading("stall", text="Stall"      if lang == "en" else "\u6444\u4f4d")
        self._tree.heading("notes", text="Notes"      if lang == "en" else "\u5907\u6ce8")
        self._tree.column("name",  width=180, stretch=False)
        self._tree.column("stall", width=80,  stretch=False, anchor=tk.CENTER)
        self._tree.column("notes", width=240, stretch=True)
        self._tree.bind("<<TreeviewSelect>>", self._on_select)

        # Side buttons
        btn_col = tk.Frame(body, bg=COLORS["app"])
        btn_col.grid(row=0, column=1, sticky="n", padx=(8, 0))
        btn_w = dict(style="Tool.TButton", width=12)
        self._add_btn = ttk.Button(btn_col,
                                   text="+ Add" if lang == "en" else "+ \u6dfb\u52a0",
                                   command=self._add_shop, **btn_w)
        self._add_btn.pack(pady=(0, 4))
        self._edit_btn = ttk.Button(btn_col,
                                    text="Edit" if lang == "en" else "\u7f16\u8f91",
                                    command=self._edit_shop, state="disabled", **btn_w)
        self._edit_btn.pack(pady=(0, 4))

        sty_r = ttk.Style(d)
        sty_r.configure("ShopDanger.TButton", font=("Segoe UI", 10), padding=(8, 6))
        sty_r.map("ShopDanger.TButton",
                  background=[("active", "#fee2e2"), ("disabled", "#f3f4f6")],
                  foreground=[("active", "#b91c1c"), ("disabled", "#9ca3af")])
        self._remove_btn = ttk.Button(btn_col,
                                      text="Remove" if lang == "en" else "\u5220\u9664",
                                      command=self._remove_shop,
                                      style="ShopDanger.TButton",
                                      width=12, state="disabled")
        self._remove_btn.pack(pady=(12, 0))

        # Footer
        foot = tk.Frame(d, bg=COLORS["app"])
        foot.grid(row=2, column=0, sticky="ew", padx=14, pady=10)
        ttk.Button(foot,
                   text="Close" if lang == "en" else "\u5173\u95ed",
                   command=self._close, style="Tool.TButton").pack(side=tk.RIGHT)

        self._load()
        d.protocol("WM_DELETE_WINDOW", self._close)

    # ── Helpers ────────────────────────────────────────────────────────

    def _load(self) -> None:
        """Load current Charm Shops from the catalog."""
        self._shops.clear()
        try:
            if load_charm_shops is not None:
                for cs in load_charm_shops(self._catalog_path):
                    self._shops.append(
                        {"name": cs.shop_name, "stall": cs.stall, "notes": cs.notes}
                    )
        except Exception:
            pass
        self._repopulate()

    def _repopulate(self) -> None:
        self._tree.delete(*self._tree.get_children())
        for s in self._shops:
            self._tree.insert("", tk.END, values=(
                s["name"], s["stall"], s["notes"],
            ))

    def _on_select(self, _event=None) -> None:
        has = bool(self._tree.selection())
        self._edit_btn.config(state="normal" if has else "disabled")
        self._remove_btn.config(state="normal" if has else "disabled")

    def _write(self) -> None:
        """Overwrite the Charm Shops sheet with current in-memory list."""
        try:
            if backup_supplier_catalog_before_write is not None:
                backup_supplier_catalog_before_write(self._catalog_path, "manage_charm_shops")
            import openpyxl as _xl
            wb = _xl.load_workbook(self._catalog_path)
            if CHARM_SHOPS_SHEET not in wb.sheetnames:
                wb.close()
                return
            ws = wb[CHARM_SHOPS_SHEET]
            # Clear all data rows (keep row 1 headers or instruction row)
            for r in range(ws.max_row, 1, -1):
                row_val = str(ws.cell(r, 1).value or "").strip().lower()
                if "charm shops" in row_val or "quick guide" in row_val:
                    continue
                ws.delete_rows(r, 1)
            # Find insert position (first empty after headers)
            insert_row = 2
            ws.cell(insert_row - 1, 1)   # ensure we're at row 2
            for r in range(2, ws.max_row + 2):
                insert_row = r
                if ws.cell(r, 1).value is None:
                    break
            for s in self._shops:
                ws.cell(insert_row, 1).value = s["name"]  or None
                ws.cell(insert_row, 2).value = s["stall"] or None
                ws.cell(insert_row, 3).value = s["notes"] or None
                insert_row += 1
            wb.save(self._catalog_path)
            wb.close()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror(
                "Save failed" if self._lang == "en" else "\u4fdd\u5b58\u5931\u8d25",
                str(e), parent=self._d,
            )

    # ── Add / Edit ─────────────────────────────────────────────────────

    def _shop_form(self, title: str, name: str = "", stall: str = "",
                   notes: str = "") -> dict | None:
        """Show a modal form for a shop entry; returns dict or None if cancelled."""
        top = tk.Toplevel(self._d)
        top.title(title)
        top.transient(self._d)
        top.grab_set()
        top.configure(bg=COLORS["app"])
        top.resizable(False, False)
        result: list[dict | None] = [None]

        f = tk.Frame(top, bg=COLORS["app"])
        f.pack(fill=tk.BOTH, expand=True, padx=18, pady=16)
        f.columnconfigure(1, weight=1)

        def _lbl(row: int, text: str):
            tk.Label(f, text=text, font=("Segoe UI", 9, "bold"),
                     fg=COLORS["muted"], bg=COLORS["app"]).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=4)

        _lbl(0, "Shop Name *" if self._lang == "en" else "\u5e97\u540d *")
        name_var = tk.StringVar(value=name)
        name_ent = ttk.Entry(f, textvariable=name_var, font=("Segoe UI", 10), width=24)
        name_ent.grid(row=0, column=1, sticky="ew", pady=4)

        _lbl(1, "Stall" if self._lang == "en" else "\u6444\u4f4d")
        stall_var = tk.StringVar(value=stall)
        ttk.Entry(f, textvariable=stall_var, font=("Segoe UI", 10), width=24).grid(
            row=1, column=1, sticky="ew", pady=4)

        _lbl(2, "Notes" if self._lang == "en" else "\u5907\u6ce8")
        notes_var = tk.StringVar(value=notes)
        ttk.Entry(f, textvariable=notes_var, font=("Segoe UI", 10), width=24).grid(
            row=2, column=1, sticky="ew", pady=4)

        err_lbl = tk.Label(f, text="", font=("Segoe UI", 8),
                           fg="#b91c1c", bg=COLORS["app"])
        err_lbl.grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(4, 0))

        btn_row = tk.Frame(f, bg=COLORS["app"])
        btn_row.grid(row=4, column=0, columnspan=2, sticky="e", pady=(12, 0))

        def _ok():
            n = name_var.get().strip()
            if not n:
                err_lbl.config(text="Shop name is required." if self._lang == "en"
                                else "\u5e97\u540d\u4e0d\u80fd\u4e3a\u7a7a\u3002")
                return
            result[0] = {"name": n, "stall": stall_var.get().strip(),
                         "notes": notes_var.get().strip()}
            top.destroy()

        ttk.Button(btn_row, text="Cancel" if self._lang == "en" else "\u53d6\u6d88",
                   command=top.destroy, style="Tool.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="OK",
                   command=_ok, style="Tool.TButton").pack(side=tk.LEFT)

        name_ent.focus_set()
        top.bind("<Return>", lambda _: _ok())
        top.bind("<Escape>", lambda _: top.destroy())
        top.wait_window()
        return result[0]

    def _add_shop(self) -> None:
        data = self._shop_form(
            "Add Charm Shop" if self._lang == "en" else "\u6dfb\u52a0\u6302\u4ef6\u5e97\u94fa"
        )
        if data is None:
            return
        self._shops.append(data)
        self._repopulate()
        self._write()

    def _edit_shop(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        idx = self._tree.index(sel[0])
        s   = self._shops[idx]
        data = self._shop_form(
            "Edit Charm Shop" if self._lang == "en" else "\u7f16\u8f91\u6302\u4ef6\u5e97\u94fa",
            name=s["name"], stall=s["stall"], notes=s["notes"],
        )
        if data is None:
            return
        self._shops[idx] = data
        self._repopulate()
        self._write()

    def _remove_shop(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        idx  = self._tree.index(sel[0])
        name = self._shops[idx]["name"]
        from tkinter import messagebox
        if not messagebox.askyesno(
            "Remove shop" if self._lang == "en" else "\u5220\u9664\u5e97\u94fa",
            f"Remove '{name}' from the Charm Shops list?"
            if self._lang == "en" else
            f"\u5220\u9664\u300c{name}\u300d\uff1f",
            parent=self._d,
        ):
            return
        del self._shops[idx]
        self._repopulate()
        self._write()
        self._edit_btn.config(state="disabled")
        self._remove_btn.config(state="disabled")

    def _close(self) -> None:
        self._d.destroy()
        if callable(self._on_close):
            self._on_close()


class _SuppliersManagerDialog:
    """
    Modal dialog for managing the Suppliers reference list.

    Reads and writes the *Suppliers* sheet in supplier_catalog.xlsx
    (columns: ID | Shop Name | Mall | Floor | Stall | Address | Contact | Notes).

    Rows are always displayed and persisted in stall-ascending order
    (floor → stall code → shop name), mirroring the shopping-route sort.

    Calling pattern mirrors ``_CharmShopsManagerDialog``:
        _SuppliersManagerDialog(parent, catalog_path, lang, on_close=callback)
    ``on_close`` is called (no args) after any write so the dashboard can
    refresh its supplier dropdowns.
    """

    # ── Column indices in the Suppliers sheet (1-based) ───────────────
    _C_ID      = 1   # A
    _C_SHOP    = 2   # B  — Shop Name
    _C_MALL    = 3   # C  — Mall
    _C_FLOOR   = 4   # D  — Floor
    _C_STALL   = 5   # E  — Stall
    _C_ADDR    = 6   # F  — Address
    _C_CONTACT = 7   # G  — Contact
    _C_NOTES   = 8   # H  — Notes

    @staticmethod
    def _stall_sort_key(row: dict) -> tuple:
        """
        Sort key: floor-ascending → stall code → shop name.
        Mirrors ``_stall_floor`` from generate_shopping_route.py without
        requiring an import of that private function.
        """
        import re as _re
        stall = (row.get("stall") or "").strip()
        shop  = (row.get("shop")  or "").strip().lower()

        if not stall or stall in ("\u2014", "???"):
            floor_n = 999
        elif _re.match(r"^A2", stall, _re.IGNORECASE):
            floor_n = 2
        else:
            m = _re.match(r"^(\d)", stall)
            if m:
                floor_n = int(m.group(1))
            else:
                m = _re.search(r"(\d)[A-Za-z]", stall)
                floor_n = int(m.group(1)) if m else 999

        return (floor_n, stall.lower(), shop)

    def __init__(
        self,
        parent: tk.Misc,
        catalog_path,           # Path
        lang: str,
        on_close: object = None,
    ) -> None:
        self._catalog_path = catalog_path
        self._lang         = lang
        self._on_close     = on_close
        self._rows: list[dict] = []
        # {id, shop, mall, floor, stall, addr, contact, notes}

        d = tk.Toplevel(parent)
        self._d = d
        d.title(
            "Manage Suppliers" if lang == "en"
            else "\u7ba1\u7406\u4f9b\u5e94\u5546"
        )
        d.transient(parent)
        d.grab_set()
        d.configure(bg=COLORS["app"])
        d.geometry("680x460")
        d.minsize(560, 360)
        d.grid_columnconfigure(0, weight=1)
        d.grid_rowconfigure(1, weight=1)

        # ── Hero ──────────────────────────────────────────────────────
        hero = tk.Frame(d, bg=COLORS["hero"], highlightthickness=0)
        hero.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hero,
            text="Manage Suppliers" if lang == "en" else "\u7ba1\u7406\u4f9b\u5e94\u5546",
            font=("Segoe UI", 12, "bold"), fg="#ffffff", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(10, 2))
        tk.Label(
            hero,
            text=(
                "Add, edit or remove supplier entries that appear in the "
                "\u2018Supplier\u2019 and \u2018Stall\u2019 dropdowns."
            ) if lang == "en" else (
                "\u6dfb\u52a0\u3001\u7f16\u8f91\u6216\u5220\u9664\u5728\u300c\u4f9b\u5e94\u5546"
                "\u300d\u548c\u300c\u6444\u4f4d\u300d\u4e0b\u62c9\u5217\u8868\u4e2d\u663e\u793a"
                "\u7684\u4f9b\u5e94\u5546\u6761\u76ee\u3002"
            ),
            font=("Segoe UI", 9), fg="#e2e8f0", bg=COLORS["hero"],
        ).pack(anchor=tk.W, padx=14, pady=(0, 8))

        # ── Body: treeview + side buttons ─────────────────────────────
        body = tk.Frame(d, bg=COLORS["app"])
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(10, 0))
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=1)

        tf = tk.Frame(body, bg=COLORS["app"])
        tf.grid(row=0, column=0, sticky="nsew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        cols = ("shop", "stall", "mall", "notes")
        self._tree = ttk.Treeview(
            tf, columns=cols, show="headings", selectmode="browse"
        )
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        _T = lambda en, zh: en if lang == "en" else zh
        self._tree.heading("shop",  text=_T("Shop Name", "\u5e97\u540d"))
        self._tree.heading("stall", text=_T("Stall",     "\u6444\u4f4d"))
        self._tree.heading("mall",  text=_T("Mall",      "\u5546\u573a"))
        self._tree.heading("notes", text=_T("Notes",     "\u5907\u6ce8"))
        self._tree.column("shop",  width=200, stretch=False)
        self._tree.column("stall", width=80,  stretch=False, anchor=tk.CENTER)
        self._tree.column("mall",  width=120, stretch=False)
        self._tree.column("notes", width=200, stretch=True)
        self._tree.bind("<<TreeviewSelect>>", self._on_select)
        self._tree.bind("<Double-Button-1>",  lambda _: self._edit_entry())

        # Side buttons
        btn_col = tk.Frame(body, bg=COLORS["app"])
        btn_col.grid(row=0, column=1, sticky="n", padx=(8, 0))
        _bw = dict(style="Tool.TButton", width=14)
        self._add_btn  = ttk.Button(
            btn_col,
            text=_T("+ Add",   "+ \u6dfb\u52a0"),
            command=self._add_entry, **_bw,
        )
        self._add_btn.pack(pady=(0, 4))
        self._edit_btn = ttk.Button(
            btn_col,
            text=_T("Edit",    "\u7f16\u8f91"),
            command=self._edit_entry, state="disabled", **_bw,
        )
        self._edit_btn.pack(pady=(0, 4))

        sty = ttk.Style(d)
        sty.configure("SupDanger.TButton", font=("Segoe UI", 10), padding=(8, 6))
        sty.map(
            "SupDanger.TButton",
            background=[("active", "#fee2e2"), ("disabled", "#f3f4f6")],
            foreground=[("active", "#b91c1c"), ("disabled", "#9ca3af")],
        )
        self._remove_btn = ttk.Button(
            btn_col,
            text=_T("Remove",  "\u5220\u9664"),
            command=self._remove_entry,
            style="SupDanger.TButton", width=14, state="disabled",
        )
        self._remove_btn.pack(pady=(12, 0))

        # ── Footer ────────────────────────────────────────────────────
        foot = tk.Frame(d, bg=COLORS["app"])
        foot.grid(row=2, column=0, sticky="ew", padx=14, pady=10)
        ttk.Button(
            foot,
            text=_T("Close", "\u5173\u95ed"),
            command=self._close,
            style="Tool.TButton",
        ).pack(side=tk.RIGHT)

        self._load()
        d.protocol("WM_DELETE_WINDOW", self._close)

    # ── Data helpers ──────────────────────────────────────────────────

    def _load(self) -> None:
        """Read the Suppliers sheet into ``self._rows``, sorted by stall."""
        self._rows.clear()
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(self._catalog_path, read_only=True, data_only=True)
            if "Suppliers" not in wb.sheetnames:
                wb.close()
                self._repopulate()
                return
            ws = wb["Suppliers"]
            # Detect header row column positions dynamically for robustness
            ci_shop = ci_stall = ci_mall = ci_floor = None
            ci_addr = ci_contact = ci_notes = ci_id = None
            for ci in range(1, 15):
                h = str(ws.cell(1, ci).value or "").strip().lower()
                if h == "id":             ci_id      = ci
                elif h == "shop name":    ci_shop    = ci
                elif h == "mall":         ci_mall    = ci
                elif h == "floor":        ci_floor   = ci
                elif h == "stall":        ci_stall   = ci
                elif h == "address":      ci_addr    = ci
                elif h == "contact":      ci_contact = ci
                elif h == "notes":        ci_notes   = ci
            # Fall back to fixed positions from _SUPPLIERS_HEADER_ROW
            ci_id      = ci_id      or 1
            ci_shop    = ci_shop    or 2
            ci_mall    = ci_mall    or 3
            ci_floor   = ci_floor   or 4
            ci_stall   = ci_stall   or 5
            ci_addr    = ci_addr    or 6
            ci_contact = ci_contact or 7
            ci_notes   = ci_notes   or 8

            def _cv(row, ci): return str(row[ci - 1].value or "").strip()

            for row in ws.iter_rows(min_row=2, values_only=False):
                shop = _cv(row, ci_shop)
                if not shop:
                    continue
                self._rows.append({
                    "id":      _cv(row, ci_id),
                    "shop":    shop,
                    "mall":    _cv(row, ci_mall),
                    "floor":   _cv(row, ci_floor),
                    "stall":   _cv(row, ci_stall),
                    "addr":    _cv(row, ci_addr),
                    "contact": _cv(row, ci_contact),
                    "notes":   _cv(row, ci_notes),
                })
            wb.close()
        except Exception:
            pass
        # Always display in stall-ascending order
        self._rows.sort(key=self._stall_sort_key)
        self._repopulate()

    def _repopulate(self) -> None:
        self._tree.delete(*self._tree.get_children())
        for row in self._rows:
            self._tree.insert("", tk.END, values=(
                row["shop"], row["stall"], row["mall"], row["notes"],
            ))

    def _on_select(self, _event=None) -> None:
        has = bool(self._tree.selection())
        self._edit_btn.config(state="normal" if has else "disabled")
        self._remove_btn.config(state="normal" if has else "disabled")

    def _write(self) -> None:
        """Sort rows by stall, overwrite Suppliers data rows, preserve header row 1."""
        # Sort in memory first so the dialog display and file stay in sync
        self._rows.sort(key=self._stall_sort_key)
        self._repopulate()

        try:
            if backup_supplier_catalog_before_write is not None:
                backup_supplier_catalog_before_write(
                    self._catalog_path, "manage_suppliers"
                )
            import openpyxl as _xl
            wb = _xl.load_workbook(self._catalog_path)
            if "Suppliers" not in wb.sheetnames:
                # Create the sheet with standard headers
                ws = wb.create_sheet("Suppliers", 0)
                for ci, h in enumerate(
                    ("ID", "Shop Name", "Mall", "Floor", "Stall",
                     "Address", "Contact", "Notes"), 1
                ):
                    ws.cell(1, ci).value = h
            else:
                ws = wb["Suppliers"]
                # Clear all data rows (keep row 1 = header)
                for r in range(ws.max_row, 1, -1):
                    ws.delete_rows(r, 1)

            # Write rows in sorted order with fresh sequential IDs
            for seq, row in enumerate(self._rows, start=1):
                ws.cell(seq + 1, self._C_ID).value      = str(seq)
                ws.cell(seq + 1, self._C_SHOP).value    = row["shop"]    or None
                ws.cell(seq + 1, self._C_MALL).value    = row["mall"]    or None
                ws.cell(seq + 1, self._C_FLOOR).value   = row["floor"]   or None
                ws.cell(seq + 1, self._C_STALL).value   = row["stall"]   or None
                ws.cell(seq + 1, self._C_ADDR).value    = row["addr"]    or None
                ws.cell(seq + 1, self._C_CONTACT).value = row["contact"] or None
                ws.cell(seq + 1, self._C_NOTES).value   = row["notes"]   or None

            wb.save(self._catalog_path)
            wb.close()
        except Exception as e:
            messagebox.showerror(
                "Save failed" if self._lang == "en" else "\u4fdd\u5b58\u5931\u8d25",
                str(e), parent=self._d,
            )

    # ── Entry form ────────────────────────────────────────────────────

    def _entry_form(
        self, title: str,
        shop: str = "", mall: str = "", floor_: str = "", stall: str = "",
        addr: str = "", contact: str = "", notes: str = "",
    ) -> dict | None:
        """Modal inline form; returns a dict of values or None if cancelled."""
        top = tk.Toplevel(self._d)
        top.title(title)
        top.transient(self._d)
        top.grab_set()
        top.configure(bg=COLORS["app"])
        top.resizable(False, False)
        result: list[dict | None] = [None]

        f = tk.Frame(top, bg=COLORS["app"])
        f.pack(fill=tk.BOTH, expand=True, padx=18, pady=16)
        f.columnconfigure(1, weight=1)

        _T = lambda en, zh: en if self._lang == "en" else zh

        def _lbl(row: int, text: str) -> None:
            tk.Label(
                f, text=text, font=("Segoe UI", 9, "bold"),
                fg=COLORS["muted"], bg=COLORS["app"],
            ).grid(row=row, column=0, sticky=tk.W, padx=(0, 10), pady=(6, 2))

        def _entry(row: int, val: str, width: int = 26) -> tk.StringVar:
            sv = tk.StringVar(value=val)
            ttk.Entry(f, textvariable=sv, font=("Segoe UI", 10), width=width).grid(
                row=row, column=1, sticky="ew", pady=(6, 2),
            )
            return sv

        _lbl(0, _T("Shop Name *", "\u5e97\u540d *"))
        shop_v    = _entry(0, shop)
        _lbl(1, _T("Stall",       "\u6444\u4f4d"))
        stall_v   = _entry(1, stall)
        _lbl(2, _T("Mall",        "\u5546\u573a"))
        mall_v    = _entry(2, mall)
        _lbl(3, _T("Floor",       "\u697c\u5c42"))
        floor_v   = _entry(3, floor_)
        _lbl(4, _T("Address",     "\u5730\u5740"))
        addr_v    = _entry(4, addr)
        _lbl(5, _T("Contact",     "\u8054\u7cfb\u65b9\u5f0f"))
        contact_v = _entry(5, contact)
        _lbl(6, _T("Notes",       "\u5907\u6ce8"))
        notes_v   = _entry(6, notes)

        err_lbl = tk.Label(
            f, text="", font=("Segoe UI", 8), fg="#b91c1c", bg=COLORS["app"],
        )
        err_lbl.grid(row=7, column=0, columnspan=2, sticky=tk.W, pady=(4, 0))

        btn_row = tk.Frame(f, bg=COLORS["app"])
        btn_row.grid(row=8, column=0, columnspan=2, sticky="e", pady=(14, 0))

        def _ok() -> None:
            n = shop_v.get().strip()
            if not n:
                err_lbl.config(
                    text=_T("Shop name is required.", "\u5e97\u540d\u4e0d\u80fd\u4e3a\u7a7a\u3002")
                )
                return
            result[0] = {
                "shop":    n,
                "stall":   stall_v.get().strip(),
                "mall":    mall_v.get().strip(),
                "floor":   floor_v.get().strip(),
                "addr":    addr_v.get().strip(),
                "contact": contact_v.get().strip(),
                "notes":   notes_v.get().strip(),
            }
            top.destroy()

        ttk.Button(
            btn_row,
            text=_T("Cancel", "\u53d6\u6d88"),
            command=top.destroy, style="Tool.TButton",
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            btn_row, text="OK", command=_ok, style="Tool.TButton",
        ).pack(side=tk.LEFT)

        # Focus first field and bind keyboard shortcuts
        f.winfo_children()[1].focus_set()   # the Shop Name Entry
        top.bind("<Return>", lambda _: _ok())
        top.bind("<Escape>", lambda _: top.destroy())
        top.wait_window()
        return result[0]

    # ── CRUD actions ──────────────────────────────────────────────────

    def _add_entry(self) -> None:
        data = self._entry_form(
            "Add Supplier" if self._lang == "en" else "\u6dfb\u52a0\u4f9b\u5e94\u5546"
        )
        if data is None:
            return
        # Preserve id as empty; _write will re-sequence on save
        data["id"] = ""
        self._rows.append(data)
        self._repopulate()
        # Select the newly added row
        children = self._tree.get_children()
        if children:
            self._tree.selection_set(children[-1])
            self._tree.see(children[-1])
        self._write()

    def _edit_entry(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        idx = self._tree.index(sel[0])
        row = self._rows[idx]
        data = self._entry_form(
            "Edit Supplier" if self._lang == "en" else "\u7f16\u8f91\u4f9b\u5e94\u5546",
            shop=row["shop"],    mall=row["mall"],  floor_=row["floor"],
            stall=row["stall"],  addr=row["addr"],  contact=row["contact"],
            notes=row["notes"],
        )
        if data is None:
            return
        data["id"] = row["id"]  # keep original ID
        self._rows[idx] = data
        self._repopulate()
        self._write()

    def _remove_entry(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        idx  = self._tree.index(sel[0])
        name = self._rows[idx]["shop"]
        if not messagebox.askyesno(
            "Remove supplier" if self._lang == "en" else "\u5220\u9664\u4f9b\u5e94\u5546",
            (f"Remove \u2018{name}\u2019 from the Suppliers list?"
             if self._lang == "en" else
             f"\u5220\u9664\u300c{name}\u300d\uff1f"),
            parent=self._d,
        ):
            return
        del self._rows[idx]
        self._repopulate()
        self._edit_btn.config(state="disabled")
        self._remove_btn.config(state="disabled")
        self._write()

    def _close(self) -> None:
        self._d.destroy()
        if callable(self._on_close):
            self._on_close()


def main() -> None:
    if not GENERATOR.is_file():
        print(f"Missing {GENERATOR}", file=sys.stderr)
        sys.exit(1)
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
